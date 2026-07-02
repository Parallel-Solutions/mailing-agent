from __future__ import annotations

import json
import random
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from src.jobs import create_job_id, resolve_job_paths, save_agent_state


LOAD_TEST_HEADERS = [
    "ID",
    "SUB_RF",
    "MUN_R_NAME",
    "MUN_NAME",
    "ADM_NAME",
    "ADRES",
    "HEAD_FIO",
    "POPULATION",
    "EMAIL_OSN",
    "EMAIL_DOP",
    "TEL_OSN",
    "TEL_DOP",
    "REQUISITES_INN",
    "REQUISITES_KPP",
    "REQUISITES_OGRN",
    "REQUISITES_OKPO",
    "REQUISITES_OKTNO",
    "STATUS",
]

REGIONS = [
    "Республика Адыгея",
    "Алтайский край",
    "Краснодарский край",
    "Московская область",
    "Нижегородская область",
    "Новосибирская область",
    "Пермский край",
    "Республика Татарстан",
    "Самарская область",
    "Тверская область",
]

DISTRICT_TAILS = [
    "Александровский",
    "Березовский",
    "Верхневолжский",
    "Дмитровский",
    "Ельнинский",
    "Заволжский",
    "Красноармейский",
    "Никольский",
    "Орловский",
    "Сосновский",
]

LOCALITY_TAILS = [
    "Березовка",
    "Васильевское",
    "Дубровка",
    "Заречный",
    "Ильинское",
    "Красный Яр",
    "Лесное",
    "Михайловка",
    "Новая Слобода",
    "Покровское",
    "Солнечный",
    "Троицкое",
]

LAST_NAMES = [
    "Иванова",
    "Петрова",
    "Сидорова",
    "Смирнова",
    "Кузнецова",
    "Васильева",
    "Попова",
    "Новикова",
    "Федорова",
    "Морозова",
]
FIRST_NAMES = ["Анна", "Мария", "Елена", "Ольга", "Наталья", "Татьяна", "Ирина", "Светлана"]
PATRONYMICS = ["Ивановна", "Петровна", "Сергеевна", "Александровна", "Владимировна", "Николаевна"]
LOAD_TEST_MARKER_FILENAME = "load_test.json"


def _copy_templates(source_job_id: str | None, target_job_id: str) -> list[str]:
    target_paths = resolve_job_paths(target_job_id)
    copied: list[str] = []
    candidates = []
    if source_job_id:
        candidates.append(resolve_job_paths(source_job_id).templates_dir)
    candidates.append(resolve_job_paths(None).templates_dir)

    target_paths.templates_dir.mkdir(parents=True, exist_ok=True)
    for source_dir in candidates:
        if not source_dir.exists():
            continue
        for source_path in source_dir.iterdir():
            if not source_path.is_file():
                continue
            target_path = target_paths.templates_dir / source_path.name
            if target_path.exists():
                continue
            shutil.copy2(source_path, target_path)
            copied.append(source_path.name)
    return copied


def _missing_generator_templates(job_id: str) -> list[str]:
    templates_dir = resolve_job_paths(job_id).templates_dir
    missing: list[str] = []
    if not any((templates_dir / name).exists() for name in ("kp_template_source.docx", "kp_template_source.pdf")):
        missing.append("шаблон КП")
    if not (templates_dir / "contract_template_source.docx").exists():
        missing.append("шаблон договора")
    return missing


def _load_test_row(index: int, rng: random.Random) -> list[Any]:
    region = rng.choice(REGIONS)
    district = f"{rng.choice(DISTRICT_TAILS)} муниципальный район"
    locality = rng.choice(LOCALITY_TAILS)
    settlement_kind = rng.choice(["Сельское поселение", "Городское поселение"])
    mun_name = f"{settlement_kind} {locality}"
    adm_name = f'АДМИНИСТРАЦИЯ МУНИЦИПАЛЬНОГО ОБРАЗОВАНИЯ "{mun_name.upper()}"'
    fio = f"{rng.choice(LAST_NAMES)} {rng.choice(FIRST_NAMES)} {rng.choice(PATRONYMICS)}"
    inn_base = 1000000000 + index
    return [
        index,
        region,
        district,
        mun_name,
        adm_name,
        f"{100000 + index}, {region}, {district}, {locality}, ул. Центральная, д. {rng.randint(1, 120)}",
        fio,
        rng.randint(900, 85000),
        f"loadtest{index:04d}@example.test",
        "",
        f"+79{rng.randint(100000000, 999999999)}",
        "",
        str(inn_base),
        f"{rng.randint(100000000, 999999999)}",
        f"1{rng.randint(100000000000, 999999999999)}",
        f"{rng.randint(10000000, 99999999)}",
        f"{rng.randint(10000000, 99999999)}",
        "",
    ]


def _write_load_test_workbook(path: Path, *, row_count: int, seed: int) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    rng = random.Random(seed)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "load-test"
    worksheet.cell(row=1, column=1).value = f"Нагрузочный тест: {row_count} синтетических клиентов"
    worksheet.cell(row=1, column=1).font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E8F1DB")
    for column_index, header in enumerate(LOAD_TEST_HEADERS, start=1):
        cell = worksheet.cell(row=2, column=column_index)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for index in range(1, row_count + 1):
        for column_index, value in enumerate(_load_test_row(index, rng), start=1):
            worksheet.cell(row=index + 2, column=column_index).value = value
    worksheet.freeze_panes = "A3"
    workbook.save(path)


def _prime_parser_completed(job_id: str, *, row_count: int) -> None:
    now = datetime.now().isoformat(timespec="seconds")
    save_agent_state(
        "parser",
        {
            "status": "completed",
            "started_at": now,
            "completed_at": now,
            "summary_text": f"Создана синтетическая таблица для нагрузочного теста: {row_count} строк.",
            "row_count": row_count,
            "municipality_name_verification": {
                "status": "ok",
                "total_rows": row_count,
                "verified_rows": 0,
                "updated_rows": 0,
                "kept_rows": row_count,
                "missing_rows": 0,
                "table_mode": "load_test",
            },
            "municipality_name_verification_state": {
                "status": "completed",
                "source": "load_test",
                "started_at": now,
                "completed_at": now,
                "summary_text": "Проверка пропущена для синтетической нагрузочной таблицы.",
            },
        },
        job_id=job_id,
    )


def _load_test_marker_path(job_id: str | None) -> Path:
    return resolve_job_paths(job_id).root_dir / "state" / LOAD_TEST_MARKER_FILENAME


def is_load_test_job(job_id: str | None) -> bool:
    if not job_id:
        return False
    marker_path = _load_test_marker_path(job_id)
    try:
        if not marker_path.exists():
            return False
        payload = json.loads(marker_path.read_text(encoding="utf-8-sig"))
    except Exception:
        return False
    return bool(isinstance(payload, dict) and payload.get("load_test"))


def _write_load_test_marker(job_id: str, *, row_count: int, seed: int) -> None:
    marker_path = _load_test_marker_path(job_id)
    marker_path.parent.mkdir(parents=True, exist_ok=True)
    marker_path.write_text(
        json.dumps(
            {
                "load_test": True,
                "kind": "documents_generation",
                "row_count": row_count,
                "seed": seed,
                "created_at": datetime.now().isoformat(timespec="seconds"),
                "send_disabled": True,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )


def create_documents_load_test_job(
    *,
    row_count: int = 500,
    source_job_id: str | None = None,
    seed: int | None = None,
) -> dict[str, Any]:
    safe_row_count = max(1, min(int(row_count or 500), 2000))
    actual_seed = int(seed if seed is not None else datetime.now().strftime("%Y%m%d%H%M%S"))
    job_id = create_job_id()
    paths = resolve_job_paths(job_id)
    paths.ensure_dirs()
    copied_templates = _copy_templates(source_job_id, job_id)
    _write_load_test_workbook(paths.data_xlsx, row_count=safe_row_count, seed=actual_seed)
    _write_load_test_marker(job_id, row_count=safe_row_count, seed=actual_seed)
    _prime_parser_completed(job_id, row_count=safe_row_count)
    missing_templates = _missing_generator_templates(job_id)
    return {
        "job_id": job_id,
        "row_count": safe_row_count,
        "seed": actual_seed,
        "data_path": str(paths.data_xlsx),
        "copied_templates": copied_templates,
        "missing_templates": missing_templates,
        "generator_ready": not missing_templates,
        "load_test": True,
        "send_disabled": True,
    }
