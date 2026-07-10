from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.web.download_sources import PreviewMode


MAX_TABLE_ROWS = 200
MAX_TABLE_COLUMNS = 50
MAX_TEXT_LINES = 500
MAX_TEXT_BYTES = 512 * 1024
MAX_ARCHIVE_PAGE = 100


def _clamp(value: int, *, minimum: int, maximum: int) -> int:
    return max(minimum, min(maximum, value))


def _cell_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def read_xlsx_preview(
    path: Path,
    *,
    sheet_index: int = 0,
    offset: int = 0,
    limit: int = 100,
) -> dict[str, Any]:
    safe_limit = _clamp(limit, minimum=1, maximum=MAX_TABLE_ROWS)
    safe_offset = max(0, offset)
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_names = list(workbook.sheetnames)
    if not sheet_names:
        workbook.close()
        return {"columns": [], "rows": [], "total_rows": 0, "sheet_names": [], "sheet_index": 0}
    safe_sheet_index = _clamp(sheet_index, minimum=0, maximum=len(sheet_names) - 1)
    worksheet = workbook[sheet_names[safe_sheet_index]]

    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    columns = [_cell_value(value) or f"column_{index + 1}" for index, value in enumerate(header_row[:MAX_TABLE_COLUMNS])]
    if not columns:
        columns = [f"column_{index + 1}" for index in range(min(worksheet.max_column or 0, MAX_TABLE_COLUMNS))]

    data_rows: list[list[str]] = []
    total_rows = 0
    for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=1):
        if not any(cell not in (None, "") for cell in row):
            continue
        total_rows += 1
        if total_rows <= safe_offset:
            continue
        if len(data_rows) >= safe_limit:
            continue
        data_rows.append([_cell_value(row[index]) if index < len(row) else "" for index in range(len(columns))])

    workbook.close()
    return {
        "columns": columns,
        "rows": data_rows,
        "total_rows": total_rows,
        "sheet_names": sheet_names,
        "sheet_index": safe_sheet_index,
    }


def read_csv_preview(path: Path, *, offset: int = 0, limit: int = 100) -> dict[str, Any]:
    safe_limit = _clamp(limit, minimum=1, maximum=MAX_TABLE_ROWS)
    safe_offset = max(0, offset)
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        rows = list(reader)
    if not rows:
        return {"columns": [], "rows": [], "total_rows": 0, "sheet_names": [], "sheet_index": 0}
    columns = rows[0][:MAX_TABLE_COLUMNS]
    data_rows = rows[1:]
    total_rows = len(data_rows)
    page_rows = data_rows[safe_offset : safe_offset + safe_limit]
    normalized_rows = [
        [row[index] if index < len(row) else "" for index in range(len(columns))]
        for row in page_rows
    ]
    return {
        "columns": columns,
        "rows": normalized_rows,
        "total_rows": total_rows,
        "sheet_names": [],
        "sheet_index": 0,
    }


def read_table_preview(
    path: Path,
    *,
    preview_mode: PreviewMode,
    sheet_index: int = 0,
    offset: int = 0,
    limit: int = 100,
) -> dict[str, Any]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return read_xlsx_preview(path, sheet_index=sheet_index, offset=offset, limit=limit)
    if suffix == ".csv" or preview_mode == PreviewMode.TABLE and suffix == ".csv":
        return read_csv_preview(path, offset=offset, limit=limit)
    if suffix == ".csv":
        return read_csv_preview(path, offset=offset, limit=limit)
    raise ValueError(f"Unsupported table preview for {path.name}")


def read_text_preview(path: Path, *, offset: int = 0, limit: int = MAX_TEXT_LINES) -> dict[str, Any]:
    safe_limit = _clamp(limit, minimum=1, maximum=MAX_TEXT_LINES)
    safe_offset = max(0, offset)
    raw = path.read_bytes()[:MAX_TEXT_BYTES]
    text = raw.decode("utf-8", errors="replace")
    lines = text.splitlines()
    total_lines = len(lines)
    page_lines = lines[safe_offset : safe_offset + safe_limit]
    truncated = len(path.read_bytes()) > MAX_TEXT_BYTES or safe_offset + safe_limit < total_lines
    return {
        "content": "\n".join(page_lines),
        "total_lines": total_lines,
        "offset": safe_offset,
        "limit": safe_limit,
        "truncated": truncated,
    }


def read_ndjson_preview(path: Path, *, offset: int = 0, limit: int = 100) -> dict[str, Any]:
    safe_limit = _clamp(limit, minimum=1, maximum=MAX_TABLE_ROWS)
    safe_offset = max(0, offset)
    columns: list[str] = []
    rows: list[list[str]] = []
    total_rows = 0
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            total_rows += 1
            if total_rows <= safe_offset:
                continue
            if len(rows) >= safe_limit:
                continue
            try:
                payload = json.loads(line)
            except json.JSONDecodeError:
                payload = {"raw": line}
            if isinstance(payload, dict):
                for key in payload.keys():
                    key_text = str(key)
                    if key_text not in columns:
                        columns.append(key_text)
                rows.append([_cell_value(payload.get(column)) for column in columns])
            else:
                if "value" not in columns:
                    columns = ["value"]
                rows.append([_cell_value(payload)])
    return {
        "columns": columns[:MAX_TABLE_COLUMNS],
        "rows": [row[:MAX_TABLE_COLUMNS] for row in rows],
        "total_rows": total_rows,
        "sheet_names": [],
        "sheet_index": 0,
    }
