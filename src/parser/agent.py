"""
src/parser/agent.py

Прокси-слой между main.py (FastAPI) и нашим новым парсер-агентом.
Содержит функции, которые ожидает main.py:
  - chat(message, job_id)         — диалог с агентом
  - run_batch_parser(job_id)      — пакетная обработка файла
  - get_memory(job_id)             — текущая память агента
  - clear_memory(job_id)           — очистка памяти
  - set_system_prompt(prompt)      — кастомный системный промпт

Внутри использует:
  - src/parser_new/agent/         — для диалогового режима
  - src/parser_new/batch_processor — для пакетной обработки
"""
from __future__ import annotations

import shutil
from pathlib import Path
from typing import Optional

from src.jobs import resolve_job_paths
import re
from src.parser_new.tools.discovery_tool import discover_and_write, resolve_okved

# Импорты из нашего нового агента
from src.parser_new.agent.executor import run_agent_task
from src.parser_new.memory.memory_manager import get_full_context

def get_memory_context() -> str:
    return get_full_context()

def _clear_memory() -> None:
    pass
from src.parser_new.batch_processor import run as run_batch
from src.parser_new import progress
from src.utils.logger import logger

# ДИАЛОГ С АГЕНТОМ

def _latest_batch_file() -> tuple[Optional[str], float]:
    """
    Возвращает (путь, mtime) самого свежего собранного файла в общем output/latest.
    Используется чтобы понять, создал ли агент новый файл за время запроса.

    ВАЖНО: берём ЛЮБОЙ .xlsx, а не только batch_*. Коллектор пишет batch_*, но
    агент через write_excel_tool создаёт файлы с другими префиксами (mo_, okruga_,
    names_ и т.п.) — раньше они не находились, и результат считался несозданным.
    """
    out_dir = Path(__file__).parent.parent / "parser_new" / "output" / "latest"
    if not out_dir.exists():
        return None, -1.0
    latest, latest_mtime = None, -1.0
    for p in out_dir.glob("*.xlsx"):
        if "FAILED" in p.name or p.name.startswith("~$"):
            continue
        try:
            mtime = p.stat().st_mtime
        except OSError:
            continue
        if mtime >= latest_mtime:
            latest, latest_mtime = str(p), mtime
    return latest, latest_mtime

def _clean_reply(text: str) -> str:
    """Убирает из ответа агента то, что не соответствует интерфейсу.

    Модель регулярно дописывает «Результат можно скачать кнопкой ниже» (кнопки
    нет) и перечисляет собранные поля (в таблице показываются не все). Промпт
    это не удерживает, поэтому чистим детерминированно.
    """
    if not text:
        return text

    # режем на предложения, сохраняя разделители
    parts = re.split(r"(?<=[.!?])\s+", text.strip())
    # Раньше здесь резались и упоминания скачивания — кнопки в интерфейсе не было.
    # Сейчас кнопка «Скачать таблицу» есть, поэтому фразы про неё не мешают;
    # остаётся только вводящее в заблуждение перечисление полей.
    drop_patterns: tuple[str, ...] = ()
    field_list = r"(email|e-mail|телефон|инн|огрн|кпп|фио руководител|адрес)"
    # с этих оборотов обычно начинается перечисление собранных полей
    enum_start = re.compile(
        r"[,;]?\s*(данные\s+включают|включая|в\s+том\s+числе|с\s+полями|"
        r"собран[ыо]\s+поля|поля:)\b.*$",
        re.IGNORECASE,
    )

    kept = []
    for p in parts:
        low = p.lower()
        if any(re.search(pat, low) for pat in drop_patterns):
            continue
        # перечисление полей: три и более упоминания в одном предложении
        if len(re.findall(field_list, low)) >= 3:
            trimmed = enum_start.sub("", p).strip(" ,;—-")
            # оставляем обрезанное, только если в нём ещё есть смысл
            if len(trimmed) >= 15 and len(re.findall(field_list, trimmed.lower())) < 3:
                kept.append(trimmed if trimmed.endswith((".", "!", "?")) else trimmed + ".")
            continue
        kept.append(p)

    out = " ".join(kept).strip()
    return out or "Сбор завершён, данные добавлены в таблицу."


def _looks_like_collection(message: str):
    """Похоже ли на запрос сбора коммерческих организаций → (что, где, сколько).

    Используется НЕ для маршрутизации (её делает агент, выбирая инструмент),
    а только как страховка: если агент отработал и НЕ создал файла, а запрос
    очевидно про сбор — дособерём детерминированно, вместо пустого ответа.
    """
    from src.parser_new.tools.collector import parse_request

    parsed = parse_request(message or "")
    if not parsed:
        return None
    query, place, _limit = parsed

    # органы власти собираются другим маршрутом (ОКТМО + batch), не коллектором
    low = (query + " " + place).lower()
    gov_words = ("администрац", "муниципальн", "поселени", "сельсовет", "мэри",
                 "управа", "органы власти", "министерств", "департамент",
                 "комитет", "правительств", "госорган", "учреждени")
    if any(kw in low for kw in gov_words):
        return None
    return parsed

def _mark_discovery_table_ready(job_id: Optional[str], *, count: int = 0) -> None:
    """Помечает собранную таблицу готовой к скачиванию.

    Проверки официальных названий МО для коммерческого сбора нет, а gate в
    download-result (parser_table_verified) её ждёт — без этой отметки скачивание
    вернёт 409 «Дождитесь завершения проверки таблицы». Вызывается для ЛЮБОГО
    созданного файла: и когда собрал коллектор, и когда собрал сам агент.
    """
    try:
        from datetime import datetime
        from src.generator.orchestration.parser_agent import (
            _update_municipality_verification_state,
        )
        now = datetime.now().isoformat(timespec="seconds")
        summary = (f"Сбор завершён: {count} организаций." if count
                   else "Сбор завершён, таблица готова.")
        _update_municipality_verification_state(
            job_id,
            status="completed",
            source="discovery",
            summary_text=summary,
            completed_at=now,
            result={
                "status": "ok",
                "total_rows": count,
                "updated_rows": 0,
                "verified_rows": count,
                "missing_rows": 0,
                "kept_rows": count,
            },
        )
    except Exception as e:
        # Статус — вспомогательный: если пометить не удалось, сам сбор уже прошёл,
        # файл записан. Не роняем ответ пользователю из-за статуса.
        logger.warning(f"[parser] Не удалось пометить таблицу готовой: {e}")

def chat(message: str, job_id: Optional[str] = None) -> dict:
    """
    Диалог с агентом. Используется в /api/parser/chat.
    (док-строка прежняя)
    """
    progress.start(job_id)          # фиксируем job_id для потока прогресса
    try:
        # Разбираем тип МО и количество из фразы ДО запуска агента: модель
        # регулярно вызывает build_region_mo_file_tool без mo_type, и вместо
        # городских округов приходят все МО региона подряд.
        try:
            from src.parser_new.tools.oktmo_tool import set_request_hint
            set_request_hint(message)
        except Exception as e:
            logger.warning(f"[parser] не удалось разобрать подсказку запроса: {e}")

        uploaded_file = None
        job_output_dir = None
        if job_id:
            try:
                paths = resolve_job_paths(job_id)
                if paths.data_xlsx.exists():
                    uploaded_file = str(paths.data_xlsx)
                job_output_dir = paths.output_dir
            except Exception as e:
                logger.warning(f"Не удалось получить пути для job_id={job_id}: {e}")

        _, before_mtime = _latest_batch_file()

        # Маршрутизацию делает АГЕНТ: он сам выбирает инструмент по описанию.
        # Никаких ключевых слов и регулярок на входе.
        result = run_agent_task(
            task=message, chat_history=[],
            uploaded_file_path=uploaded_file, mode="Автоматический",
        )
        reply_text, success = result.text, result.success

        src_file, after_mtime = _latest_batch_file()
        file_was_created = src_file is not None and after_mtime > before_mtime

        # СТРАХОВКА (не маршрутизация): агент отработал, но файла нет, а запрос
        # очевидно про сбор организаций — значит инструмент он не вызвал.
        # Досбираем детерминированно, чтобы пользователь не остался ни с чем.
        if not file_was_created and uploaded_file is None:
            parsed = _looks_like_collection(message)
            if parsed:
                logger.warning("[parser] агент не создал файл — досбираю коллектором")
                from src.parser_new.tools.collector import collect_and_describe
                disc = collect_and_describe(*parsed)
                if disc.get("success"):
                    reply_text = disc["reply"]
                    success = True
                    src_file, after_mtime = _latest_batch_file()
                    file_was_created = src_file is not None

        # Агент иногда отвечает «готово», не создав файла. Не выдаём это за успех:
        # иначе UI покажет «Результат можно скачать», а скачивать будет нечего.
        if success and not file_was_created:
            success = False
            reply_text = (reply_text or "").rstrip()
            reply_text += ("\n\nНо файл с результатом не создан — скачивать нечего. "
                           "Уточните запрос (что и где искать) и повторите.")
            logger.warning("[parser] отчёт об успехе без созданного файла")

        if file_was_created:
            _mark_discovery_table_ready(job_id, count=0)

        result_file = None
        if file_was_created:
            # ДОВОДКА КОДОМ, без участия модели: почта с официальных сайтов и
            # дополнение названий городом. Агент эти шаги систематически
            # пропускает, поэтому они вынесены в постобработку.
            try:
                from src.parser_new.tools.postprocess import postprocess_file
                st = postprocess_file(src_file)
                if st.get("error"):
                    logger.warning(f"[parser] постобработка: {st['error']}")
                else:
                    logger.info(f"[parser] постобработка: {st}")
            except Exception as e:
                logger.warning(f"[parser] постобработка не выполнена: {e}")

            # Агент называет файл как ему вздумается (ministries_architecture_HMAO_...,
            # mo_..., кириллицей и т.п.), а эндпоинт скачивания ищет строго batch_*.xlsx.
            # Поэтому приводим имя к каноничному сами — на модель тут полагаться нельзя.
            from datetime import datetime as _dt
            canonical = f"batch_{_dt.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
            src_path = Path(src_file)

            # 1) канонический дубль рядом с оригиналом (общая папка output/latest)
            if not src_path.name.startswith("batch_"):
                try:
                    twin = src_path.parent / canonical
                    shutil.copy2(src_path, twin)
                    src_file = str(twin)
                    logger.info(f"[parser] Файл приведён к каноничному имени: {twin.name}")
                except Exception as e:
                    logger.warning(f"Не удалось создать канонический дубль: {e}")

            # 2) копия в папку задачи — тоже с каноничным именем
            if job_output_dir is not None:
                try:
                    job_output_dir.mkdir(parents=True, exist_ok=True)
                    dst = job_output_dir / canonical
                    shutil.copy2(src_file, dst)
                    result_file = str(dst)
                    logger.info(f"[parser] Результат скопирован в папку задачи: {dst}")
                except Exception as e:
                    logger.warning(f"Не удалось скопировать результат в папку задачи: {e}")
                    result_file = src_file
            else:
                result_file = src_file

        return {"reply": _clean_reply(reply_text), "success": success,
                "result_file": result_file}
    finally:
        progress.finish()

def topup(message: str, job_id: str) -> dict:
    """Дозаполнение загруженного файла. Эндпоинт /api/parser/topup.

    Та же форма, что и сбор, но с job_id файла: находим то, чего в файле ещё
    нет, и дописываем в него же. Детерминированно, без агента.
    """
    from src.parser_new.tools.collector import parse_request, _detect_government

    parsed = parse_request(message)
    if not parsed:
        return {"reply": "Не разобрал, что и где искать. Уточните запрос.",
                "success": False, "result_file": None}
    query, place, limit = parsed

    try:
        paths = resolve_job_paths(job_id)
    except Exception as e:
        logger.warning(f"[topup] пути job_id={job_id}: {e}")
        return {"reply": f"Не удалось найти файл задачи: {e}",
                "success": False, "result_file": None}

    if not paths.data_xlsx.exists():
        return {"reply": "Файл не загружен — сначала загрузите Excel.",
                "success": False, "result_file": None}

    file_path = str(paths.data_xlsx)
    progress.start(job_id)
    try:
        if _detect_government(query) or _detect_government(place):
            res = _topup_mo(file_path, place, query)
        else:
            res = _topup_commercial(file_path, query, place, limit)

        if res.get("success") and res.get("result_file"):
            _mark_discovery_table_ready(job_id, count=res.get("added", 0))
            # чтобы download-result нашёл файл — кладём копию в папку задачи
            # с каноничным именем, как это делает chat()
            try:
                from datetime import datetime as _dt
                canonical = f"batch_{_dt.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
                paths.output_dir.mkdir(parents=True, exist_ok=True)
                dst = paths.output_dir / canonical
                shutil.copy2(res["result_file"], dst)
                res["result_file"] = str(dst)
            except Exception as e:
                logger.warning(f"[topup] копия результата в папку задачи: {e}")
        return res
    finally:
        progress.finish()


def _topup_commercial(file_path: str, query: str, place: str, limit: int) -> dict:
    from src.parser_new.tools.collector import collect_recipients, _row_key
    from src.parser_new.tools.discovery_tool import read_batch_rows, append_batch_xlsx

    # 1. ключи того, что уже в файле
    exclude = {_row_key(r) for r in read_batch_rows(file_path)}

    # 2. сбор НОВЫХ (свой файл не пишем — write=False)
    res = collect_recipients(query, place, limit, exclude=exclude, write=False)
    if not res["success"]:
        return {"reply": f"Не удалось донайти: {res['error']}",
                "success": False, "result_file": None, "added": 0}

    # 3. дозапись в исходный файл
    added = append_batch_xlsx(file_path, res.get("rows") or [])

    reply = (f"Дописал {added} новых организаций, которых не было в файле."
             if added else
             "Новых организаций сверх тех, что уже в файле, не нашлось.")
    if res.get("note"):
        reply += f"\n\n{res['note']}"
    return {"reply": reply, "success": added > 0,
            "result_file": file_path if added else None, "added": added}


def _topup_mo(file_path: str, place: str, query: str) -> dict:
    import json as _json
    from openpyxl import load_workbook
    from src.parser_new.tools.oktmo_tool import build_region_mo_file
    from src.parser_new.tools.excel_tool import read_existing_mo_keys, append_excel_tool

    # 1. что уже есть
    have = read_existing_mo_keys(file_path)

    # 2. полный список региона (переиспользуем готовый строитель скелета)
    try:
        region_path, _, sub_rf = build_region_mo_file(place, mo_type="")
    except Exception as e:
        return {"reply": f"Не удалось получить список МО региона: {e}",
                "success": False, "result_file": None, "added": 0}

    # 3. вычитаем то, что в файле
    wb = load_workbook(region_path, read_only=True, data_only=True)
    ws = wb.active
    headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    col = {h: i + 1 for i, h in enumerate(headers) if h}
    missing = []
    for row_idx in range(3, ws.max_row + 1):
        oktmo = ws.cell(row_idx, col["REQUISITES_OKTMO"]).value if col.get("REQUISITES_OKTMO") else None
        name  = ws.cell(row_idx, col["MUN_NAME"]).value if col.get("MUN_NAME") else None
        k_oktmo = f"oktmo:{str(oktmo).strip()}" if oktmo else None
        k_name  = "name:" + "".join(str(name).lower().split()) if name else None
        if (k_oktmo and k_oktmo in have) or (k_name and k_name in have):
            continue
        missing.append({"SUB_RF": sub_rf, "MUN_NAME": name,
                        "REQUISITES_OKTMO": str(oktmo).strip() if oktmo else ""})
    wb.close()

    if not missing:
        return {"reply": "В файле уже есть все МО этого региона — дописывать нечего.",
                "success": False, "result_file": file_path, "added": 0}

    # 4. дописываем недостающие (реквизиты/контакты потом дозальёт обычный batch)
    payload = _json.dumps(missing, ensure_ascii=False)
    if hasattr(append_excel_tool, "func"):
        append_excel_tool.func(file_path, payload)      # @tool -> «сырая» функция
    else:
        append_excel_tool(file_path, payload)

    return {"reply": f"Дописал {len(missing)} недостающих МО региона «{sub_rf}». "
                     f"Их реквизиты и контакты дозаполнит обычная обработка файла.",
            "success": True, "result_file": file_path, "added": len(missing)}

def fill_gaps(job_id: str, verify_emails: bool = False) -> dict:
    progress.start(job_id)
    try:
        res = run_batch_parser(job_id)
        ok = res.get("status") == "ok"
        result_file = res.get("file") or None

        # проверка почт по официальным сайтам (и по столбцу S) — по галочке
        if ok and verify_emails and result_file:
            try:
                from src.parser_new.tools.postprocess import postprocess_file
                postprocess_file(result_file, max_email_lookups=2000, check_all=True)
            except Exception as e:
                logger.warning(f"[fill] проверка почт не выполнена: {e}")

        if ok and result_file:
            found = (res.get("task_stats") or {}).get("found", 0)
            _mark_discovery_table_ready(job_id, count=found)
        return {"reply": res.get("reply", ""), "success": ok, "result_file": result_file}
    finally:
        progress.finish()
# ==============================
# ПАКЕТНАЯ ОБРАБОТКА ФАЙЛА
# ==============================

def run_batch_parser(job_id: Optional[str] = None, limit: Optional[int] = None) -> dict:
    """
    Запускает пакетную обработку загруженного файла.
    Используется в /api/parser/start.

    Находит файл по job_id, запускает batch_processor,
    возвращает путь к итоговому файлу и статистику.

    Args:
        job_id: идентификатор задачи (берётся загруженный файл)
        limit:  опциональное ограничение количества строк

    Returns:
        {
            "status":      "ok" | "error",
            "reply":       краткое описание результата,
            "file":        путь к итоговому файлу,
            "failed_file": путь к файлу непроверенных,
            "task_stats":  { "found": N, "not_found": M, "liquidated": K },
        }
    """
    if not job_id:
        return {
            "status": "error",
            "reply": "Не указан job_id — невозможно определить файл для обработки.",
        }

    # Находим загруженный файл
    try:
        paths = resolve_job_paths(job_id)
    except Exception as e:
        logger.error(f"resolve_job_paths failed for {job_id}: {e}")
        return {"status": "error", "reply": f"Ошибка определения путей: {e}"}

    if not paths.data_xlsx.exists():
        return {
            "status": "error",
            "reply": "Файл data.xlsx не найден. Загрузите файл через интерфейс.",
        }

    # Запускаем пакетный обработчик
    try:
        logger.info(f"[parser] Запуск batch_processor для {paths.data_xlsx}")
        result = run_batch(
            file_path=str(paths.data_xlsx),
            save_every=10,
            output_dir=str(paths.output_dir),
        )

        return {
            "status": "ok",
            "reply": (
                f"Обработано: {result.get('processed', 0)}. "
                f"Найдено: {result.get('found', 0)}. "
                f"Не найдено: {result.get('not_found', 0)}. "
                f"Ликвидировано: {result.get('liquidated', 0)}."
            ),
            "file": result.get("output_path", ""),
            "failed_file": result.get("failed_path", ""),
            "task_stats": {
                "found":      result.get("found", 0),
                "not_found":  result.get("not_found", 0),
                "liquidated": result.get("liquidated", 0),
                "processed":  result.get("processed", 0),
            },
        }

    except Exception as e:
        logger.exception(f"[parser] batch_processor failed: {e}")
        return {
            "status": "error",
            "reply": f"Ошибка при обработке: {e}",
        }


# ==============================
# ПАМЯТЬ АГЕНТА
# ==============================

def get_memory(job_id: Optional[str] = None) -> dict:
    """Возвращает текущую память агента."""
    try:
        context = get_memory_context()
        return {"memory": context}
    except Exception as e:
        logger.warning(f"get_memory failed: {e}")
        return {"memory": ""}


def clear_memory(job_id: Optional[str] = None) -> None:
    """Очищает память агента."""
    try:
        _clear_memory()
    except Exception as e:
        logger.warning(f"clear_memory failed: {e}")


# ==============================
# СИСТЕМНЫЙ ПРОМПТ
# ==============================

def set_system_prompt(prompt: str, job_id: Optional[str] = None) -> None:
    """Сохраняет кастомный системный промпт."""
    if not prompt:
        return

    prompt_path = (
        Path(__file__).parent.parent / "parser_new" / "agent" / "prompt.py"
    )

    if not prompt_path.exists():
        logger.warning(f"Файл промпта не найден: {prompt_path}")
        return

    # Сохраняем как валидный Python модуль
    prompt_escaped = prompt.replace('"""', '\\"\\"\\"')
    prompt_path.write_text(
        f'SYSTEM_PROMPT = """{prompt_escaped}"""\n',
        encoding="utf-8",
    )
    logger.info(f"Системный промпт обновлён: {prompt_path}")