from pathlib import Path
from dataclasses import dataclass

from openpyxl import load_workbook

from src.utils.logger import logger


@dataclass
class MoRow:
    row_index: int        # номер строки в файле (для чекпоинтинга)
    sub_rf: str           # Субъект РФ
    mun_r_name: str       # Муниципальный район
    mun_name: str         # Муниципальное образование
    population: int | None  # Население


def read_base_mo(path: Path) -> list[MoRow]:
    """
    Читает файл «База МО» и возвращает список строк.

    Ожидаемая структура файла:
    - Строка 1: шапка (Субъект РФ | Муниципальный район | Муниципальное образование | Население)
    - Строки 2+: данные

    Args:
        path: путь к файлу base.xlsx

    Returns:
        Список MoRow, отсортированный по порядку строк в файле.

    Raises:
        FileNotFoundError: если файл не найден
        ValueError: если структура файла не соответствует ожидаемой
    """
    if not path.exists():
        raise FileNotFoundError(f"Файл не найден: {path}")

    logger.info("base_mo_read_start", path=str(path))

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active

    # Читаем шапку из первой строки и проверяем структуру
    headers = [
        str(cell.value).strip() if cell.value else ""
        for cell in next(worksheet.iter_rows(min_row=1, max_row=1))
    ]

    expected = ["Субъект РФ", "Муниципальный район", "Муниципальное образование", "Население"]
    for expected_col, actual_col in zip(expected, headers):
        if expected_col.lower() not in actual_col.lower():
            raise ValueError(
                f"Неожиданная структура файла. "
                f"Ожидался столбец '{expected_col}', найден '{actual_col}'"
            )

    # Определяем индексы столбцов по шапке (на случай если порядок отличается)
    col_index = {name: idx for idx, name in enumerate(headers)}
    sub_rf_idx = _find_col(col_index, "Субъект РФ")
    mun_r_idx = _find_col(col_index, "Муниципальный район")
    mun_name_idx = _find_col(col_index, "Муниципальное образование")
    population_idx = _find_col(col_index, "Население")

    rows: list[MoRow] = []
    skipped = 0

    for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        sub_rf = _clean(row[sub_rf_idx])
        mun_r_name = _clean(row[mun_r_idx])
        mun_name = _clean(row[mun_name_idx])
        population_raw = row[population_idx] if population_idx < len(row) else None

        # Пропускаем полностью пустые строки
        if not sub_rf and not mun_r_name and not mun_name:
            skipped += 1
            continue

        # Если нет названия МО — строка бесполезна для парсинга
        if not mun_name:
            logger.warning("base_mo_skip_no_mun_name", row_index=row_index, sub_rf=sub_rf)
            skipped += 1
            continue

        population = _parse_population(population_raw)

        rows.append(MoRow(
            row_index=row_index,
            sub_rf=sub_rf,
            mun_r_name=mun_r_name,
            mun_name=mun_name,
            population=population,
        ))

    workbook.close()

    logger.info(
        "base_mo_read_done",
        total=len(rows),
        skipped=skipped,
        path=str(path),
    )

    return rows


def _find_col(col_index: dict, name: str) -> int:
    for key, idx in col_index.items():
        if name.lower() in key.lower():
            return idx
    raise ValueError(f"Столбец '{name}' не найден в файле")


def _clean(value) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split())


def _parse_population(value) -> int | None:
    if value is None:
        return None
    try:
        return int(str(value).replace(" ", "").replace("\xa0", ""))
    except (ValueError, TypeError):
        return None