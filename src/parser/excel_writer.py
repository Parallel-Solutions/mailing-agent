from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.utils.logger import logger


# Порядок столбцов и их технические имена (строка 2 шапки)
COLUMNS: list[tuple[str, str]] = [
    ("№",                           "ID"),
    ("Субъект РФ",                  "SUB_RF"),
    ("Муниципальный район",         "MUN_R_NAME"),
    ("Муниципальное образование",   "MUN_NAME"),
    ("Полное название администрации","ADM_NAME"),
    ("Адрес",                       "ADRES"),
    ("Глава МО",                    "HEAD_FIO"),
    ("Численность населения",       "POPULATION"),
    ("Эл. Адрес (основной)",        "EMAIL_OSN"),
    ("Эл. Адрес (доп)",             "EMAIL_DOP"),
    ("Телефон",                     "TEL_OSN"),
    ("Телефон (доп)",               "TEL_DOP"),
    ("Реквизиты",                   "REQUISITES_INN"),
    ("",                            "REQUISITES_KPP"),
    ("",                            "REQUISITES_OGRN"),
    ("",                            "REQUISITES_OKPO"),
    ("",                            "REQUISITES_OKTNO"),
    ("Статус отправки",             "STATUS"),
]

REQUISITES_START_COL = next(i for i, (_, key) in enumerate(COLUMNS) if key == "REQUISITES_INN")
REQUISITES_END_COL = next(i for i, (_, key) in enumerate(COLUMNS) if key == "REQUISITES_OKTNO")

# Технические ключи для быстрого доступа
COL_KEYS = [key for _, key in COLUMNS]
COL_COUNT = len(COLUMNS)

HEADER_ROW_1 = 1
HEADER_ROW_2 = 2
DATA_START_ROW = 3


@dataclass
class MoRecord:
    """Одна запись об МО для записи в data.xlsx."""
    id: int
    sub_rf: str = ""
    mun_r_name: str = ""
    mun_name: str = ""
    adm_name: str = ""
    adres: str = ""
    head_fio: str = ""
    population: Optional[int] = None
    email_osn: str = ""
    email_dop: str = ""
    tel_osn: str = ""
    tel_dop: str = ""
    requisites_inn: str = ""
    requisites_kpp: str = ""
    requisites_ogrn: str = ""
    requisites_okpo: str = ""
    requisites_oktno: str = ""
    status: str = ""

    def to_row(self) -> list:
        return [
            self.id,
            self.sub_rf,
            self.mun_r_name,
            self.mun_name,
            self.adm_name,
            self.adres,
            self.head_fio,
            self.population,
            self.email_osn,
            self.email_dop,
            self.tel_osn,
            self.tel_dop,
            self.requisites_inn,
            self.requisites_kpp,
            self.requisites_ogrn,
            self.requisites_okpo,
            self.requisites_oktno,
            self.status,
        ]


class ExcelWriter:

    def __init__(self, path: Path):
        self.path = path
        self.path.parent.mkdir(parents=True, exist_ok=True)

        if self.path.exists():
            self._workbook = load_workbook(self.path)
            self._worksheet = self._workbook.active
            logger.info("excel_writer_opened_existing", path=str(path))
        else:
            self._workbook = Workbook()
            self._worksheet = self._workbook.active
            self._build_header()
            self._workbook.save(self.path)
            logger.info("excel_writer_created_new", path=str(path))

    def append_record(self, record: MoRecord) -> None:
        """Добавляет одну запись в конец таблицы."""
        self._worksheet.append(record.to_row())
        last_row = self._worksheet.max_row
        for key in ("TEL_OSN", "TEL_DOP"):
            col_idx = COL_KEYS.index(key) + 1
            cell = self._worksheet.cell(row=last_row, column=col_idx)
            cell.number_format = "@"

    def append_records(self, records: list[MoRecord], checkpoint_every: int = 50) -> None:
        existing_ids = self._get_existing_ids()
        new_records = [r for r in records if r.id not in existing_ids]

        if len(existing_ids) > 0:
            logger.info(
                "excel_writer_checkpoint_resume",
                already_written=len(existing_ids),
                remaining=len(new_records),
            )

        for i, record in enumerate(new_records, start=1):
            self.append_record(record)
            if i % checkpoint_every == 0:
                self.save()
                logger.info("excel_writer_checkpoint", written=i, total=len(new_records))

        self.save()
        logger.info("excel_writer_append_done", total_written=len(new_records))

    def update_status(self, record_id: int, status: str) -> bool:
        status_col = COL_KEYS.index("STATUS") + 1
        id_col = COL_KEYS.index("ID") + 1

        for row in self._worksheet.iter_rows(min_row=DATA_START_ROW):
            if row[id_col - 1].value == record_id:
                row[status_col - 1].value = status
                return True
        return False

    def get_processed_ids(self) -> set[int]:
        id_col = COL_KEYS.index("ID") + 1
        status_col = COL_KEYS.index("STATUS") + 1
        result = set()
        for row in self._worksheet.iter_rows(min_row=DATA_START_ROW, values_only=True):
            row_id = row[id_col - 1]
            row_status = row[status_col - 1]
            if row_id is not None and row_status not in (None, ""):
                result.add(int(row_id))
        return result

    def save(self) -> None:
        self._workbook.save(self.path)

    def close(self) -> None:
        self._workbook.close()

    def _build_header(self) -> None:
        ws = self._worksheet

        for col_idx, (human_name, _) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=HEADER_ROW_1, column=col_idx, value=human_name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor="D9E1F2")

        req_start = get_column_letter(REQUISITES_START_COL + 1)
        req_end = get_column_letter(REQUISITES_END_COL + 1)
        ws.merge_cells(f"{req_start}{HEADER_ROW_1}:{req_end}{HEADER_ROW_1}")

        for col_idx, (_, tech_key) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=HEADER_ROW_2, column=col_idx, value=tech_key)
            cell.font = Font(bold=True, color="595959")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill("solid", fgColor="EDF0F8")

        column_widths = {
            "ID": 6, "SUB_RF": 25, "MUN_R_NAME": 30, "MUN_NAME": 30,
            "ADM_NAME": 40, "ADRES": 35, "HEAD_FIO": 25, "POPULATION": 12,
            "EMAIL_OSN": 28, "EMAIL_DOP": 28, "TEL_OSN": 16, "TEL_DOP": 16,
            "REQUISITES_INN": 14, "REQUISITES_KPP": 12, "REQUISITES_OGRN": 16,
            "REQUISITES_OKPO": 12, "REQUISITES_OKTNO": 14, "STATUS": 18,
        }
        for col_idx, (_, key) in enumerate(COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = column_widths.get(key, 15)

        ws.freeze_panes = "E3"

        logger.info("excel_writer_header_built")

    def _get_existing_ids(self) -> set[int]:
        """Возвращает все ID которые уже есть в файле."""
        id_col = COL_KEYS.index("ID") + 1
        result = set()
        for row in self._worksheet.iter_rows(min_row=DATA_START_ROW, values_only=True):
            row_id = row[id_col - 1]
            if row_id is not None:
                result.add(int(row_id))
        return result