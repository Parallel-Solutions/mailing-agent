"""
Инструменты агента-парсера.

Каждый инструмент — это:
1. Python-функция которая реально что-то делает
2. Описание в формате OpenAI function calling (TOOL_DEFINITIONS)

Агент сам решает какой инструмент вызвать и с какими аргументами.
"""

import json
import re
from pathlib import Path

import httpx
from tenacity import retry, stop_after_attempt, wait_exponential

from src.parser.excel_reader import MoRow, read_base_mo
from src.parser.excel_writer import ExcelWriter, MoRecord
from src.parser.region_codes import get_region_code
from src.parser.validator import MoValidator
from src.utils.config import settings
from src.utils.logger import logger


# ------------------------------------------------------------------
# Константы
# ------------------------------------------------------------------

BASE_MO_PATH = Path("service_docs/base.xlsx")
DATA_XLSX_PATH = Path("data/data.xlsx")

OKVED_RURAL = "84.11.31"   # сельские поселения
OKVED_URBAN = "84.11.32"   # городские поселения

# Ключевые слова для проверки что организация — это администрация МО
_ADM_KEYWORDS = ("администрация", "адм.")
_SETTLEMENT_KEYWORDS = (
    "поселени", "округ", "район", "поссовет", "сельсовет",
    "муниципальн", "городской", "сельский",
)


# ------------------------------------------------------------------
# 1. Чтение файла База МО
# ------------------------------------------------------------------

def tool_read_base_mo() -> dict:
    """
    Читает файл База МО с сервера.
    Возвращает список МО с полями: sub_rf, mun_r_name, mun_name, population.
    """
    if not BASE_MO_PATH.exists():
        return {"error": "Файл База МО не загружен. Загрузите файл через интерфейс."}

    try:
        rows = read_base_mo(BASE_MO_PATH)
        return {
            "total": len(rows),
            "rows": [
                {
                    "row_index": r.row_index,
                    "sub_rf": r.sub_rf,
                    "mun_r_name": r.mun_r_name,
                    "mun_name": r.mun_name,
                    "population": r.population,
                    "region_code": get_region_code(r.sub_rf) or "не определён",
                }
                for r in rows[:50]  # первые 50 для превью агенту
            ],
            "note": "Показаны первые 50 строк. Полный список доступен при парсинге.",
        }
    except Exception as e:
        logger.exception("tool_read_base_mo_error")
        return {"error": str(e)}


# ------------------------------------------------------------------
# 2. Checko: поиск по названию МО
# ------------------------------------------------------------------

@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=1, max=10))
def tool_checko_search_by_name(mun_name: str, region_code: str) -> dict:
    """
    Ищет организацию в Checko по названию МО.
    Возвращает список найденных организаций с базовыми реквизитами.
    """
    url = "https://api.checko.ru/v2/search"
    params = {
        "key": settings.checko_api_key,
        "by": "name",
        "obj": "org",
        "query": mun_name,
        "region": region_code,
        "limit": 10,
    }

    try:
        response = httpx.get(url, params=params, timeout=15)
        response.raise_for_status()
        data = response.json()
    except Exception as e:
        logger.exception("checko_search_by_name_error", mun_name=mun_name)
        return {"error": str(e)}

    raw_items = data.get("data", [])
    if not raw_items:
        return {"found": 0, "items": []}

    items = [_parse_search_item(item) for item in raw_items]
    items = [i for i in items if i]
    # Фильтруем — оставляем только администрации МО
    filtered = [item for item in items if _is_administration(item.get("full_name", ""))]

    logger.info(
        "checko_search_by_name",
        mun_name=mun_name,
        region_code=region_code,
        total=len(items),
        filtered=len(filtered),
    )

    return {"found": len(filtered), "items": filtered}


# ------------------------------------------------------------------
# 3. Checko: поиск по ОКВЭД
# ------------------------------------------------------------------

@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=1, max=10))
def tool_checko_search_by_okved(region_code: str, okved_code: str) -> dict:
    """
    Ищет организации в Checko по коду ОКВЭД и региону.
    Используй для поиска всех МО в регионе.
    okved_code: '84.11.31' (сельские) или '84.11.32' (городские)
    """
    url = "https://api.checko.ru/v2/search"
    params = {
        "key": settings.checko_api_key,
        "by": "okved",
        "obj": "org",
        "query": okved_code,
        "active": "true",
        "region": region_code,
    }

    try:
        response = httpx.get(url, params=params, timeout=15)
        response.raise_for_status()
        data = response.json()
    except Exception as e:
        logger.exception("checko_search_by_okved_error", region=region_code)
        return {"error": str(e)}

    raw_items = data.get("data", [])
    items = [_parse_search_item(item) for item in raw_items]

    logger.info(
        "checko_search_by_okved",
        region_code=region_code,
        okved_code=okved_code,
        total=len(items),
    )

    return {"found": len(items), "items": items}


# ------------------------------------------------------------------
# 4. Checko: детали компании по ИНН
# ------------------------------------------------------------------

@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=1, max=10))
def tool_checko_get_details(inn: str) -> dict:
    """
    Получает полную карточку организации по ИНН из Checko.
    Возвращает контакты, ОКПО, ОКТМО и другие реквизиты.
    """
    url = "https://api.checko.ru/v2/company"
    params = {
        "key": settings.checko_api_key,
        "inn": inn,
    }

    try:
        response = httpx.get(url, params=params, timeout=15)
        response.raise_for_status()
        data = response.json()
    except Exception as e:
        logger.exception("checko_get_details_error", inn=inn)
        return {"error": str(e)}

    company = data.get("data", {})
    if not company:
        return {"error": f"Организация с ИНН {inn} не найдена"}

    contacts = company.get("Контакты", {})
    emails = contacts.get("Емэйл", [])
    phones = contacts.get("Тел", [])
    oktmo = company.get("ОКТМО", {})

    result = {
        "inn": inn,
        "okpo": company.get("ОКПО", ""),
        "email_osn": emails[0] if emails else "",
        "email_dop": ", ".join(emails[1:]) if len(emails) > 1 else "",
        "tel_osn": phones[0] if phones else "",
        "tel_dop": ", ".join(phones[1:]) if len(phones) > 1 else "",
        "oktmo_code": oktmo.get("Код", "") if isinstance(oktmo, dict) else "",
        "website": contacts.get("ВебСайт", ""),
    }

    logger.info("checko_get_details", inn=inn, has_email=bool(result["email_osn"]))
    return result


# ------------------------------------------------------------------
# 5. Поиск в интернете через Tavily
# ------------------------------------------------------------------

def tool_tavily_search(query: str) -> dict:
    """
    Ищет информацию в интернете через Tavily.
    Используй когда Checko не нашёл организацию.
    Возвращает текстовые фрагменты из найденных страниц.
    """
    try:
        from tavily import TavilyClient
        client = TavilyClient(api_key=settings.tavily_api_key)
        response = client.search(
            query=query,
            search_depth="basic",
            max_results=5,
            include_answer=True,
        )
        results = []
        for r in response.get("results", []):
            results.append({
                "title": r.get("title", ""),
                "url": r.get("url", ""),
                "content": r.get("content", "")[:500],  # первые 500 символов
            })
        return {
            "answer": response.get("answer", ""),
            "results": results,
        }
    except Exception as e:
        logger.exception("tavily_search_error", query=query)
        return {"error": str(e)}


# ------------------------------------------------------------------
# 6. Запись в data.xlsx
# ------------------------------------------------------------------
def tool_validate_matches(suspicious_matches: list[dict]) -> dict:
    """
    Принимает список спорных совпадений и проверяет каждое.
    Агент сам вызывает этот инструмент после merge-rmz.
    """
    # Просто возвращает список для анализа агентом через LLM
    return {"matches_to_review": suspicious_matches}


def tool_write_to_excel(records: list[dict], output_filename: str = "data.xlsx") -> dict:
    path = Path("data") / output_filename
    validator = MoValidator()
    writer = ExcelWriter(path)

    written = 0
    skipped = 0
    errors = []

    for i, raw in enumerate(records):
        try:
            validated, result = validator.validate_and_normalize(raw)
            if not result.is_valid:
                skipped += 1
                errors.append(f"Строка {i}: {result.errors}")
                continue

            record = MoRecord(
                id=raw.get("ID", i + 1),
                sub_rf=validated.get("SUB_RF", ""),
                mun_r_name=validated.get("MUN_R_NAME", ""),
                mun_name=validated.get("MUN_NAME", ""),
                adm_name=validated.get("ADM_NAME", ""),
                adres=validated.get("ADRES", ""),
                head_fio=validated.get("HEAD_FIO", ""),
                population=validated.get("POPULATION"),
                email_osn=validated.get("EMAIL_OSN", ""),
                email_dop=validated.get("EMAIL_DOP", ""),
                tel_osn=validated.get("TEL_OSN", ""),
                tel_dop=validated.get("TEL_DOP", ""),
                requisites_inn=validated.get("REQUISITES_INN", ""),
                requisites_kpp=validated.get("REQUISITES_KPP", ""),
                requisites_ogrn=validated.get("REQUISITES_OGRN", ""),
                requisites_okpo=validated.get("REQUISITES_OKPO", ""),
                requisites_oktno=validated.get("REQUISITES_OKTNO", ""),
                status=validated.get("STATUS", "найдено"),
            )
            writer.append_record(record)
            written += 1
        except Exception as e:
            errors.append(f"Строка {i}: {e}")
            skipped += 1

    writer.save()
    writer.close()

    logger.info("tool_write_to_excel", written=written, skipped=skipped)
    return {"written": written, "skipped": skipped, "errors": errors[:10]}

def tool_search_in_base_mo(mun_name: str, sub_rf: str, return_list: bool = False) -> dict:
    """
    Ищет МО в локальном файле base.xlsx.
    Если return_list=True — возвращает все МО указанного субъекта.
    """
    path = Path("service_docs/base.xlsx")
    if not path.exists():
        return {"error": "Файл base.xlsx не найден в service_docs/"}
    try:
        rows = read_base_mo(path)
        sub_norm = sub_rf.lower().strip()

        # Фильтруем по субъекту
        regional = [r for r in rows if sub_norm in r.sub_rf.lower()]
        if not regional:
            return {"found": False, "message": f"Субъект '{sub_rf}' не найден в базе МО"}

        if return_list:
            return {
                "found": True,
                "total": len(regional),
                "rows": [{"sub_rf": r.sub_rf, "mun_r_name": r.mun_r_name,
                          "mun_name": r.mun_name, "population": r.population,
                          "region_code": get_region_code(r.sub_rf) or "?"} for r in regional]
            }

        # Ищем конкретное МО
        name_norm = mun_name.lower().strip()
        found = [r for r in regional if name_norm in r.mun_name.lower() or r.mun_name.lower() in name_norm]
        if not found:
            return {"found": False, "message": f"МО '{mun_name}' не найдено в субъекте '{sub_rf}'"}

        r = found[0]
        return {
            "found": True,
            "sub_rf": r.sub_rf,
            "mun_r_name": r.mun_r_name,
            "mun_name": r.mun_name,
            "population": r.population,
            "region_code": get_region_code(r.sub_rf) or "?",
        }
    except Exception as e:
        logger.exception("tool_search_in_base_mo_error")
        return {"error": str(e)}

def tool_find_missing_mo() -> dict:
    """
    Сравнивает base.xlsx и data.xlsx.
    Возвращает список МО которые есть в базе но отсутствуют в data.xlsx.
    Совпадение точное — по субъекту, району и названию МО.
    """
    import openpyxl as _xl
    base_path = Path("service_docs/base.xlsx")
    data_path = Path("data/data.xlsx")

    if not base_path.exists():
        return {"error": "Файл base.xlsx не найден"}

    # Читаем base.xlsx — col A=sub_rf, B=mun_r_name, C=mun_name
    wb_base = _xl.load_workbook(base_path, data_only=True, read_only=True)
    base_rows = []
    for row in wb_base.worksheets[0].iter_rows(min_row=2, values_only=True):
        sub = str(row[0] or "").strip()
        dist = str(row[1] or "").strip()
        name = str(row[2] or "").strip()
        pop = row[3]
        if name:
            base_rows.append({"sub_rf": sub, "mun_r_name": dist,
                              "mun_name": name, "population": pop,
                              "region_code": get_region_code(sub) or "?"})
    wb_base.close()

    # Читаем data.xlsx — col B=sub_rf, C=mun_r_name, D=mun_name (с 3й строки)
    existing = set()
    if data_path.exists():
        wb_data = _xl.load_workbook(data_path, data_only=True, read_only=True)
        for row in wb_data.worksheets[0].iter_rows(min_row=3, values_only=True):
            sub = str(row[1] or "").strip().lower()
            dist = str(row[2] or "").strip().lower()
            name = str(row[3] or "").strip().lower()
            if name:
                existing.add((sub, dist, name))
        wb_data.close()

    missing = [
        r for r in base_rows
        if (r["sub_rf"].lower(), r["mun_r_name"].lower(), r["mun_name"].lower()) not in existing
    ]

    logger.info("find_missing_mo", total_base=len(base_rows),
                existing=len(existing), missing=len(missing))

    return {
        "total_base": len(base_rows),
        "already_in_data": len(base_rows) - len(missing),
        "missing_count": len(missing),
        "missing": missing[:100],  # первые 100 для агента
        "note": f"Показаны первые 100 из {len(missing)} пропущенных МО" if len(missing) > 100 else ""
    }


def tool_search_in_rmz(mun_name: str, sub_rf: str = "") -> dict:
    """
    Ищет администрацию МО в локальном файле RMZ7KH.xlsx.
    Ищет в столбцах A (сокр. наим) и B (полн. наим).
    Проверяет субъект по столбцу K.
    """
    import openpyxl as _xl
    path = Path("service_docs/RMZ7KH.xlsx")
    if not path.exists():
        return {"error": "Файл RMZ7KH.xlsx не найден в service_docs/"}
    try:
        wb = _xl.load_workbook(path, data_only=True, read_only=True)
        sheet = wb.worksheets[0]

        name_norm = mun_name.lower().strip()
        sub_norm = sub_rf.lower().strip() if sub_rf else ""

        # Извлекаем ключевые слова из названия МО
        _noise = {
            "сельское","поселение","поселения","сельской","городское","городской",
            "муниципальное","муниципального","образование","образования",
            "администрация","округ","район","рабочего","рабочий","пгт","гп",
        }
        keywords = [w for w in name_norm.split() if w not in _noise and len(w) > 3]
        if not keywords:
            return {"error": f"Не удалось извлечь ключевые слова из '{mun_name}'"}

        results = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            col_a = str(row[0] or "").strip().lower()   # сокр. наим
            col_b = str(row[1] or "").strip()           # полн. наим
            col_b_lower = col_b.lower()
            col_k = str(row[10] or "").strip().lower()  # регион

            if not col_b:
                continue

            # Проверяем что это администрация
            if "администрация" not in col_b_lower:
                continue
            banned = ["совет депутатов", "совет народных депутатов",
                      "бюджетное учреждение", "казенное учреждение"]
            if any(b in col_b_lower for b in banned):
                continue

            # Проверяем субъект если указан
            if sub_norm:
                sub_words = [w for w in sub_norm.split() if len(w) > 3]
                if sub_words and not any(w in col_k for w in sub_words):
                    continue

            # Ищем ключевые слова в столбцах A и B
            search_text = col_a + " " + col_b_lower
            if not all(kw[:7] in search_text for kw in keywords):
                continue

            emails = str(row[6] or "").strip()
            phones = str(row[5] or "").strip()
            email1 = emails.split(",")[0].strip() if emails else ""
            email_rest = ", ".join(emails.split(",")[1:]).strip() if "," in emails else ""
            phone1 = phones.split(",")[0].strip() if phones else ""
            phone_rest = ", ".join(phones.split(",")[1:]).strip() if "," in phones else ""

            results.append({
                "full_name": col_b,
                "ogrn": str(row[2] or ""),
                "inn": str(row[3] or ""),
                "kpp": str(row[4] or ""),
                "address": str(row[11] or ""),
                "head_fio": str(row[15] or ""),
                "email_osn": email1,
                "email_dop": email_rest,
                "tel_osn": phone1,
                "tel_dop": phone_rest,
                "region": str(row[10] or ""),
            })

        wb.close()
        if not results:
            return {"found": False, "message": f"МО '{mun_name}' не найдено в RMZ7KH"}
        return {"found": True, "total": len(results), "items": results[:5]}
    except Exception as e:
        logger.exception("tool_search_in_rmz_error")
        return {"error": str(e)}
    

    def tool_find_missing_mo() -> dict:
        """
        Сравнивает base.xlsx и data.xlsx.
        Возвращает список МО которые есть в базе но отсутствуют в data.xlsx.
        Совпадение точное — по субъекту, району и названию МО.
        """
        import openpyxl as _xl
        base_path = Path("service_docs/base.xlsx")
        data_path = Path("data/data.xlsx")

        if not base_path.exists():
            return {"error": "Файл base.xlsx не найден"}

        # Читаем base.xlsx — col A=sub_rf, B=mun_r_name, C=mun_name
        wb_base = _xl.load_workbook(base_path, data_only=True, read_only=True)
        base_rows = []
        for row in wb_base.worksheets[0].iter_rows(min_row=2, values_only=True):
            sub = str(row[0] or "").strip()
            dist = str(row[1] or "").strip()
            name = str(row[2] or "").strip()
            pop = row[3]
            if name:
                base_rows.append({"sub_rf": sub, "mun_r_name": dist,
                                "mun_name": name, "population": pop,
                                "region_code": get_region_code(sub) or "?"})
        wb_base.close()

        # Читаем data.xlsx — col B=sub_rf, C=mun_r_name, D=mun_name (с 3й строки)
        existing = set()
        if data_path.exists():
            wb_data = _xl.load_workbook(data_path, data_only=True, read_only=True)
            for row in wb_data.worksheets[0].iter_rows(min_row=3, values_only=True):
                sub = str(row[1] or "").strip().lower()
                dist = str(row[2] or "").strip().lower()
                name = str(row[3] or "").strip().lower()
                if name:
                    existing.add((sub, dist, name))
            wb_data.close()

        missing = [
            r for r in base_rows
            if (r["sub_rf"].lower(), r["mun_r_name"].lower(), r["mun_name"].lower()) not in existing
        ]

        logger.info("find_missing_mo", total_base=len(base_rows),
                    existing=len(existing), missing=len(missing))

        return {
            "total_base": len(base_rows),
            "already_in_data": len(base_rows) - len(missing),
            "missing_count": len(missing),
            "missing": missing[:100],  # первые 100 для агента
            "note": f"Показаны первые 100 из {len(missing)} пропущенных МО" if len(missing) > 100 else ""
        }

# ------------------------------------------------------------------
# Вспомогательные функции
# ------------------------------------------------------------------

def _parse_search_item(item: dict) -> dict:
    """Извлекает нужные поля из одного элемента ответа Checko /search."""
    if not isinstance(item, dict):
        return {}
    rukovod = item.get("Руковод", [])
    head_fio = rukovod[0].get("ФИО", "") if rukovod else ""

    adres = item.get("ЮрАдрес", "")
    mun_r_name = _extract_district_from_address(adres)

    return {
        "inn": item.get("ИНН", ""),
        "kpp": item.get("КПП", ""),
        "ogrn": item.get("ОГРН", ""),
        "full_name": item.get("НаимПолн", ""),
        "short_name": _to_sentence_case(item.get("НаимСокр", "")),
        "address": adres,
        "head_fio": head_fio,
        "mun_r_name_from_address": mun_r_name,
    }


def _is_administration(full_name: str) -> bool:
    """
    Проверяет что организация — это администрация МО,
    а не какая-то посторонняя организация.
    """
    name_lower = full_name.lower()
    has_adm = any(kw in name_lower for kw in _ADM_KEYWORDS)
    has_settlement = any(kw in name_lower for kw in _SETTLEMENT_KEYWORDS)
    return has_adm and has_settlement


def _extract_district_from_address(address: str) -> str:
    """
    Пытается извлечь название района из юридического адреса.
    Например: '385773, Республика Адыгея, Майкопский район, п. Совхозный...'
    → 'Майкопский район'
    """
    match = re.search(r"([А-ЯЁа-яё\s\-]+\s+район)", address)
    return match.group(1).strip() if match else ""


def _to_sentence_case(text: str) -> str:
    """ПОБЕДЕНСКОЕ СЕЛЬСКОЕ ПОСЕЛЕНИЕ → Победенское сельское поселение"""
    if not text:
        return text
    if text.isupper():
        return text[0].upper() + text[1:].lower()
    return text


# ------------------------------------------------------------------
# Описания инструментов для OpenAI function calling
# ------------------------------------------------------------------
TOOL_DEFINITIONS = [
    {
        "type": "function",
        "function": {
            "name": "read_base_mo",
            "description": (
                "Читает файл 'База МО' с сервера и возвращает список муниципальных образований. "
                "Используй в начале работы чтобы узнать какие МО нужно обработать."
            ),
            "parameters": {
                "type": "object",
                "properties": {},
                "required": [],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "checko_search_by_name",
            "description": (
                "Ищет администрацию МО в Checko по названию. "
                "Возвращает ИНН, КПП, ОГРН, адрес, ФИО руководителя. "
                "Используй когда знаешь конкретное название МО."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "mun_name": {
                        "type": "string",
                        "description": "Название муниципального образования, например 'Айрюмовское сельское поселение'",
                    },
                    "region_code": {
                        "type": "string",
                        "description": "Двузначный код региона для Checko, например '01' для Адыгеи",
                    },
                },
                "required": ["mun_name", "region_code"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "find_missing_mo",
            "description": (
                "Сравнивает base.xlsx и data.xlsx, находит МО которых нет в data.xlsx. "
                "Используй когда пользователь просит дополнить data.xlsx недостающими МО."
            ),
            "parameters": {"type": "object", "properties": {}, "required": []},
        },
    },
    {
        "type": "function",
        "function": {
            "name": "checko_search_by_okved",
            "description": (
                "Ищет все МО в регионе по коду ОКВЭД. "
                "84.11.31 — сельские поселения, 84.11.32 — городские. "
                "Используй когда пользователь просит найти все МО в регионе."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "region_code": {
                        "type": "string",
                        "description": "Двузначный код региона, например '39' для Калининградской области",
                    },
                    "okved_code": {
                        "type": "string",
                        "description": "Код ОКВЭД: '84.11.31' или '84.11.32'",
                        "enum": ["84.11.31", "84.11.32"],
                    },
                },
                "required": ["region_code", "okved_code"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "checko_get_details",
            "description": (
                "Получает полную карточку организации по ИНН: "
                "email, телефоны, ОКПО, ОКТМО. "
                "Используй после checko_search_by_name чтобы дозаполнить контактные данные."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "inn": {
                        "type": "string",
                        "description": "ИНН организации (10 цифр)",
                    },
                },
                "required": ["inn"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "find_missing_mo",
            "description": (
                "Сравнивает base.xlsx и data.xlsx, находит МО которых нет в data.xlsx. "
                "Используй когда пользователь просит дополнить data.xlsx недостающими МО."
            ),
            "parameters": {"type": "object", "properties": {}, "required": []},
        },
    },
    {
        "type": "function",
        "function": {
            "name": "tavily_search",
            "description": (
                "Ищет информацию в интернете. "
                "Используй когда Checko не нашёл организацию. "
                "Например: 'администрация Айрюмовского сельского поселения контакты реквизиты'"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "Поисковый запрос",
                    },
                },
                "required": ["query"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "write_to_excel",
            "description": (
                "Записывает найденные данные об МО в Excel файл. "
                "По умолчанию пишет в data.xlsx. "
                "Если пользователь попросил создать новый файл — передай имя в output_filename, "
                "например 'data_Kaliningrad.xlsx'."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "records": {
                        "type": "array",
                        "description": "Список записей для записи в Excel",
                        "items": {
                            "type": "object",
                            "properties": {
                                "ID": {"type": "integer"},
                                "SUB_RF": {"type": "string"},
                                "MUN_R_NAME": {"type": "string"},
                                "MUN_NAME": {"type": "string"},
                                "ADM_NAME": {"type": "string"},
                                "ADRES": {"type": "string"},
                                "HEAD_FIO": {"type": "string"},
                                "POPULATION": {"type": "integer"},
                                "EMAIL_OSN": {"type": "string"},
                                "EMAIL_DOP": {"type": "string"},
                                "TEL_OSN": {"type": "string"},
                                "TEL_DOP": {"type": "string"},
                                "REQUISITES_INN": {"type": "string"},
                                "REQUISITES_KPP": {"type": "string"},
                                "REQUISITES_OGRN": {"type": "string"},
                                "REQUISITES_OKPO": {"type": "string"},
                                "REQUISITES_OKTNO": {"type": "string"},
                                "STATUS": {"type": "string"},
                            },
                        },
                    },
                    "output_filename": {
                        "type": "string",
                        "description": (
                            "Имя файла для записи. По умолчанию 'data.xlsx'. "
                            "Для нового файла используй имя вида 'data_Kaliningrad.xlsx'."
                        ),
                    },
                },
                "required": ["records"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "search_in_base_mo",
            "description": (
                "Ищет МО в локальном файле base.xlsx. "
                "Используй ПЕРВЫМ при любом запросе про МО. "
                "Если return_list=true — возвращает все МО субъекта."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "mun_name": {"type": "string", "description": "Название МО"},
                    "sub_rf": {"type": "string", "description": "Субъект РФ"},
                    "return_list": {"type": "boolean", "description": "true — вернуть все МО субъекта"},
                },
                "required": ["mun_name", "sub_rf"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "search_in_rmz",
            "description": (
                "Ищет администрацию МО в локальном файле RMZ7KH.xlsx. "
                "Используй если в base.xlsx нет нужных контактных данных. "
                "Обязательно проверяет совпадение субъекта."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "mun_name": {"type": "string", "description": "Название МО"},
                    "sub_rf": {"type": "string", "description": "Субъект РФ для проверки"},
                },
                "required": ["mun_name", "sub_rf"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "validate_matches",
            "description": (
                "Проверяет список спорных совпадений из слияния RMZ7KH. "
                "Для каждого совпадения определяет верно ли что org_name является "
                "администрацией mo_name с учётом района и субъекта."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "suspicious_matches": {
                        "type": "array",
                        "description": "Список спорных совпадений для проверки",
                        "items": {
                            "type": "object",
                            "properties": {
                                "mo_name": {"type": "string"},
                                "org_name": {"type": "string"},
                                "sub_rf": {"type": "string"},
                                "mun_r_name": {"type": "string"},
                                "reason": {"type": "string"},
                            },
                        },
                    },
                },
                "required": ["suspicious_matches"],
            },
        },
    },
]

# Маппинг имени инструмента → функция (для вызова агентом)
TOOL_FUNCTIONS = {
    "read_base_mo": tool_read_base_mo,
    "checko_search_by_name": tool_checko_search_by_name,
    "checko_search_by_okved": tool_checko_search_by_okved,
    "checko_get_details": tool_checko_get_details,
    "tavily_search": tool_tavily_search,
    "write_to_excel": tool_write_to_excel,
    "validate_matches": tool_validate_matches,
    "search_in_base_mo": tool_search_in_base_mo,
    "search_in_rmz": tool_search_in_rmz,
    "find_missing_mo": tool_find_missing_mo,
    "search_in_rmz": tool_search_in_rmz,
}


def call_tool(name: str, arguments: dict) -> str:
    """
    Вызывает инструмент по имени и возвращает результат как JSON-строку.
    Используется агентом для выполнения tool_call от LLM.
    """
    func = TOOL_FUNCTIONS.get(name)
    if not func:
        return json.dumps({"error": f"Инструмент '{name}' не найден"}, ensure_ascii=False)

    try:
        result = func(**arguments)
        return json.dumps(result, ensure_ascii=False, default=str)
    except Exception as e:
        logger.exception("tool_call_error", name=name)
        return json.dumps({"error": str(e)}, ensure_ascii=False)