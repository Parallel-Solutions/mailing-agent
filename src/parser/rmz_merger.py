"""
Слияние данных из трёх источников:
- service_docs/base.xlsx  — база МО
- data/data.xlsx          — уже обработанные записи (пропускаем)
- service_docs/RMZ7KH.xlsx — справочник организаций

Ключевое отличие от предыдущей версии:
- Ключевые слова МО сравниваются ТОЛЬКО с частью названия поселения в строке организации,
  а не со всей строкой. Это устраняет ложные срабатывания типа
  "Лямбирское СП" → "...ЛЯМБИРСКОГО МУНИЦИПАЛЬНОГО РАЙОНА".
- ПГТ сопоставляются только с городскими поселениями/округами.
- Проверка уникальности ИНН перед записью.
"""

import re
from dataclasses import dataclass
from pathlib import Path

import openpyxl

from src.jobs import resolve_job_paths
from src.parser.excel_writer import ExcelWriter, MoRecord
from src.utils.logger import logger

BASE_MO_PATH = Path("service_docs/base.xlsx")
RMZ_PATH = Path("service_docs/RMZ7KH.xlsx")
DATA_XLSX_PATH = Path("data/data.xlsx")

_MERGE_PATHS = {
    "job_id": None,
    "base_mo_path": BASE_MO_PATH,
    "rmz_path": RMZ_PATH,
    "data_xlsx_path": DATA_XLSX_PATH,
}

BRACKET_RE = re.compile(r'\s*\(.*?\)\s*')
QUOTE_RE = re.compile(r'[«"\'](.*?)[»"\']')

NOISE_WORDS = {
    "сельское", "поселение", "поселения", "сельской",
    "поселок", "посёлок", "рабочего", "рабочий",
    "пгт", "гп", "рп", "кп", "город",
    "муниципальное", "муниципального", "образование", "образования",
}

REGION_NOISE = {
    "республика", "область", "край", "округ", "автономная", "автономный",
    "автономной", "федерального", "значения", "город", "г", "и", "ао",
    "федеральный",
}

DISTRICT_NOISE = {
    "муниципальный", "муниципального", "район", "административный",
    "административного", "городской", "городского", "округ",
}

ALLOWED_MARKERS = ["администрация", "муниципальное образование"]
BANNED_MARKERS = [
    "совет депутатов", "совет народных депутатов",
    "бюджетное учреждение", "казенное учреждение", "казённое учреждение",
    "автономное учреждение", "школа", "детский", "дошкольное",
]

RURAL_MARKERS = ["сельск", "сельсовет"]
URBAN_MARKERS = ["городск", "городское", "пгт", "посёлок городского", "поселок городского"]

DISTRICT_START_MARKERS = [
    " муниципального района",
    " муниципального округа",
    " городского района",
    " административного района",
]


@dataclass
class SuspiciousMatch:
    mo_name: str
    org_name: str
    sub_rf: str
    mun_r_name: str
    reason: str


@dataclass
class MergeResult:
    written: int
    skipped_existing: int
    not_found: int
    suspicious: list[SuspiciousMatch]
    log_lines: list[str]
    not_found_list: list[dict]


def configure_merge_paths(job_id: str | None = None) -> dict:
    job_paths = resolve_job_paths(job_id)
    rmz_candidate = job_paths.root_dir / "service_docs" / "RMZ7KH.xlsx"
    _MERGE_PATHS["job_id"] = job_paths.job_id
    _MERGE_PATHS["base_mo_path"] = job_paths.base_xlsx if job_paths.base_xlsx.exists() else BASE_MO_PATH
    _MERGE_PATHS["rmz_path"] = rmz_candidate if rmz_candidate.exists() else RMZ_PATH
    _MERGE_PATHS["data_xlsx_path"] = job_paths.data_xlsx if job_paths.data_xlsx.exists() else DATA_XLSX_PATH
    return dict(_MERGE_PATHS)


def _current_base_mo_path() -> Path:
    return Path(_MERGE_PATHS["base_mo_path"])


def _current_rmz_path() -> Path:
    return Path(_MERGE_PATHS["rmz_path"])


def _current_data_xlsx_path() -> Path:
    return Path(_MERGE_PATHS["data_xlsx_path"])


def run_merge(job_id: str | None = None) -> MergeResult:
    configure_merge_paths(job_id)
    base_mo_path = _current_base_mo_path()
    rmz_path = _current_rmz_path()
    data_xlsx_path = _current_data_xlsx_path()

    if not base_mo_path.exists():
        raise FileNotFoundError("Файл base.xlsx не найден в service_docs/")
    if not rmz_path.exists():
        raise FileNotFoundError("Файл RMZ7KH.xlsx не найден в service_docs/")

    logger.info("rmz_merge_start")

    wb_base = openpyxl.load_workbook(base_mo_path, data_only=True, read_only=True)
    wb_rmz = openpyxl.load_workbook(rmz_path, data_only=True, read_only=True)

    existing_keys, existing_inns = _get_existing_data()
    base_rows = _load_base_rows(wb_base.worksheets[0])
    rmz_rows = _load_rmz_rows(wb_rmz.worksheets[0])

    logger.info("rmz_merge_loaded", base=len(base_rows), rmz=len(rmz_rows))

    writer = ExcelWriter(data_xlsx_path)
    start_id = _get_next_id()

    written = 0
    skipped_existing = 0
    not_found = 0
    suspicious: list[SuspiciousMatch] = []
    log_lines: list[str] = []
    not_found_list: list[dict] = []
    record_id = start_id

    for mo in base_rows:
        key = (mo["sub_rf"].lower(), mo["mun_r_name"].lower(), mo["mun_name"].lower())
        if key in existing_keys:
            skipped_existing += 1
            continue

        if not mo["words"]:
            log_lines.append(f"НЕТ КЛЮЧА | МО: {mo['mun_name']}")
            not_found += 1
            not_found_list.append(mo)
            continue

        candidates, level = _find_candidates(mo, rmz_rows)

        if not candidates:
            log_lines.append(
                f"НЕ НАЙДЕНО | МО: {mo['mun_name']} | Субъект: {mo['sub_rf']} "
                f"| Район: {mo['mun_r_name']} | Причина: {level}"
            )
            not_found += 1
            not_found_list.append(mo)
            continue

        match = candidates[0]

        inn = str(match["D"] or "").strip()
        if inn and inn in existing_inns:
            log_lines.append(f"ДУБЛЬ ИНН | МО: {mo['mun_name']} | ИНН: {inn}")
            skipped_existing += 1
            continue
        if inn:
            existing_inns.add(inn)

        if len(candidates) > 1:
            suspicious.append(SuspiciousMatch(
                mo_name=mo["mun_name"],
                org_name=match["B"],
                sub_rf=mo["sub_rf"],
                mun_r_name=mo["mun_r_name"],
                reason=f"multiple({len(candidates)})",
            ))

        email1, email_rest = _split_first(match["G"])
        phone1, phone_rest = _split_first(match["F"])

        record = MoRecord(
            id=record_id,
            sub_rf=mo["sub_rf"],
            mun_r_name=mo["mun_r_name"],
            mun_name=mo["mun_name"],
            adm_name=match["B"],
            adres=match["L"],
            head_fio=match["P"],
            population=mo["population"],
            email_osn=email1 or "",
            email_dop=email_rest or "",
            tel_osn=phone1 or "",
            tel_dop=phone_rest or "",
            requisites_inn=inn,
            requisites_kpp=str(match["E"] or ""),
            requisites_ogrn=str(match["C"] or ""),
            status="rmz",
        )
        writer.append_record(record)

        if written > 0 and written % 100 == 0:
            writer.save()
            logger.info("rmz_checkpoint", written=written)

        written += 1
        record_id += 1

    writer.save()
    writer.close()

    logger.info("rmz_merge_done", written=written, skipped=skipped_existing,
                not_found=not_found, suspicious=len(suspicious))

    return MergeResult(
        written=written,
        skipped_existing=skipped_existing,
        not_found=not_found,
        suspicious=suspicious,
        log_lines=log_lines,
        not_found_list=not_found_list,
    )


# ------------------------------------------------------------------
# Извлечение части с названием поселения
# ------------------------------------------------------------------

def _extract_settlement_part(org_name: str) -> str:
    """
    Извлекает ТОЛЬКО часть названия организации относящуюся к поселению.

    "АДМИНИСТРАЦИЯ БЕРСЕНЕВСКОГО СЕЛЬСКОГО ПОСЕЛЕНИЯ ЛЯМБИРСКОГО МУНИЦИПАЛЬНОГО РАЙОНА"
    → "БЕРСЕНЕВСКОГО СЕЛЬСКОГО ПОСЕЛЕНИЯ"

    "АДМИНИСТРАЦИЯ МО «АЙРЮМОВСКОЕ СЕЛЬСКОЕ ПОСЕЛЕНИЕ»"
    → "АЙРЮМОВСКОЕ СЕЛЬСКОЕ ПОСЕЛЕНИЕ"
    """
    # 1. Текст в кавычках
    q = QUOTE_RE.search(org_name)
    if q:
        return q.group(1).strip()

    # 2. Убираем административный префикс
    lower = org_name.lower()
    rest = org_name
    for p in ["администрация муниципального образования", "администрация мо", "администрация"]:
        if lower.startswith(p):
            rest = org_name[len(p):].strip().lstrip(",").strip()
            break

    # 3. Обрезаем по маркеру начала района
    rest_lower = rest.lower()
    end_pos = len(rest)
    for marker in DISTRICT_START_MARKERS:
        idx = rest_lower.find(marker)
        if idx != -1 and idx < end_pos:
            end_pos = idx

    # 4. Дополнительно ищем паттерн " XXXСКОГО РАЙОНА"
    m = re.search(r'\s+\w{5,}ского\s+района\b', rest_lower)
    if m and m.start() < end_pos:
        end_pos = m.start()

    return rest[:end_pos].strip()


def _settlement_matches_mo(mo_name: str, org_name: str) -> bool:
    """
    Проверяет что название поселения в организации соответствует МО.
    Сравниваем ключевые слова МО ТОЛЬКО с частью про поселение.
    """
    settlement_part = _extract_settlement_part(org_name)
    if not settlement_part:
        return False

    settlement_lower = settlement_part.lower()
    mo_lower = BRACKET_RE.sub(' ', mo_name).lower()
    mo_words = [w for w in mo_lower.split() if w not in NOISE_WORDS and len(w) > 3]

    if not mo_words:
        return False

    for w in mo_words:
        prefix = _word_prefix(w)
        if prefix not in settlement_lower:
            return False
    return True


# ------------------------------------------------------------------
# Тип МО
# ------------------------------------------------------------------

def _mo_type(mun_name: str) -> str:
    lower = mun_name.lower()
    if lower.startswith("пгт") or "посёлок городского" in lower or "поселок городского" in lower:
        return "pgt"
    if any(m in lower for m in ["городск", "городское", "городской"]):
        return "urban"
    if any(m in lower for m in ["сельск", "сельсовет"]):
        return "rural"
    return "unknown"


def _org_type(org_name: str) -> str:
    lower = org_name.lower()
    if any(m in lower for m in RURAL_MARKERS):
        return "rural"
    if any(m in lower for m in URBAN_MARKERS):
        return "urban"
    return "unknown"


def _types_compatible(mo_type: str, org_type: str) -> bool:
    if mo_type == "rural":
        return org_type in ("rural", "unknown")
    if mo_type in ("pgt", "urban"):
        return org_type in ("urban", "unknown")
    return True


# ------------------------------------------------------------------
# Поиск кандидатов
# ------------------------------------------------------------------

def _find_candidates(mo: dict, rmz_rows: list) -> tuple:
    words = mo["words"]
    subj_norm = mo["sub_rf_norm"]
    district = mo["mun_r_name"]
    mo_name = mo["mun_name"]
    mo_t = _mo_type(mo_name)

    def preliminary_match(b_lower: str) -> bool:
        for w in words:
            variants = [_word_prefix(w)]
            if w == "новый": variants.append("ново")
            elif w == "нижний": variants.append("нижне")
            elif w == "верхний": variants.append("верхне")
            elif w == "старый": variants.append("старо")
            elif w == "большой": variants.extend(["больше", "больш"])
            if not any(v in b_lower for v in variants):
                return False
        return True

    # Шаг 1: субъект + район + слова (предварительная фильтрация по всей строке)
    level1 = [
        r for r in rmz_rows
        if _regions_match(subj_norm, r["region_norm"])
        and _district_match(district, r["B"])
        and preliminary_match(r["B_lower"])
    ]

    if not level1:
        return [], "not_found"

    # Шаг 2: слова МО только в части про поселение
    level2 = [r for r in level1 if _settlement_matches_mo(mo_name, r["B"])]

    # Шаг 3: тип МО
    if level2 and mo_t in ("rural", "pgt", "urban"):
        typed = [r for r in level2 if _types_compatible(mo_t, _org_type(r["B"]))]
        if typed:
            level2 = typed

    if not level2:
        return [], "settlement_name_mismatch"

    if len(level2) == 1:
        return level2, "exact"

    return level2, f"multiple({len(level2)})"


# ------------------------------------------------------------------
# Загрузка
# ------------------------------------------------------------------

def _load_base_rows(sheet) -> list[dict]:
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        sub_rf = _clean(row[0])
        mun_r_name = _clean(row[1])
        mun_name = _clean(row[2])
        population = row[3]
        if not mun_name:
            continue
        rows.append({
            "sub_rf": sub_rf,
            "sub_rf_norm": _normalize_region(sub_rf),
            "mun_r_name": mun_r_name,
            "mun_name": mun_name,
            "population": population,
            "words": _extract_keywords(mun_name),
        })
    return rows


def _load_rmz_rows(sheet) -> list[dict]:
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        b = row[1] if len(row) > 1 else None
        if not b:
            continue
        b_str = _clean(str(b))
        if not _is_administration(b_str):
            continue
        region_k = row[10] if len(row) > 10 else None
        rows.append({
            "B": b_str,
            "B_lower": b_str.lower(),
            "C": row[2] if len(row) > 2 else None,
            "D": row[3] if len(row) > 3 else None,
            "E": row[4] if len(row) > 4 else None,
            "F": row[5] if len(row) > 5 else None,
            "G": row[6] if len(row) > 6 else None,
            "L": row[11] if len(row) > 11 else None,
            "P": row[15] if len(row) > 15 else None,
            "region_norm": _normalize_region(str(region_k) if region_k else ""),
        })
    return rows


def _get_existing_data() -> tuple[set, set]:
    keys: set = set()
    inns: set = set()
    data_xlsx_path = _current_data_xlsx_path()
    if not data_xlsx_path.exists():
        return keys, inns
    try:
        wb = openpyxl.load_workbook(data_xlsx_path, data_only=True, read_only=True)
        sheet = wb.worksheets[0]
        for row in sheet.iter_rows(min_row=3, values_only=True):
            sub = str(row[1] or "").strip().lower()
            dist = str(row[2] or "").strip().lower()
            name = str(row[3] or "").strip().lower()
            inn = str(row[12] or "").strip()
            if name:
                keys.add((sub, dist, name))
            if inn:
                inns.add(inn)
        wb.close()
    except Exception as e:
        logger.warning("get_existing_data_failed", error=str(e))
    return keys, inns


def _get_next_id() -> int:
    data_xlsx_path = _current_data_xlsx_path()
    if not data_xlsx_path.exists():
        return 1
    try:
        wb = openpyxl.load_workbook(data_xlsx_path, data_only=True, read_only=True)
        sheet = wb.worksheets[0]
        last_id = 0
        for row in sheet.iter_rows(min_row=3, values_only=True):
            val = row[0]
            if val is not None:
                try:
                    last_id = max(last_id, int(val))
                except (ValueError, TypeError):
                    pass
        wb.close()
        return last_id + 1
    except Exception:
        return 1


# ------------------------------------------------------------------
# Утилиты
# ------------------------------------------------------------------

def _is_administration(name: str) -> bool:
    lower = name.lower()
    return (
        any(m in lower for m in ALLOWED_MARKERS)
        and not any(m in lower for m in BANNED_MARKERS)
    )


def _extract_keywords(name: str) -> list[str]:
    s = BRACKET_RE.sub(' ', name).strip().lower()
    return [w for w in s.split() if w not in NOISE_WORDS and len(w) > 2]


def _word_prefix(word: str) -> str:
    return word[:10] if len(word) > 13 else word[:7]


def _normalize_region(text: str) -> set[str]:
    if not text:
        return set()
    s = text.lower().replace(",", " ").replace("-", " ").replace(".", " ")
    return {w for w in s.split() if w not in REGION_NOISE and len(w) > 2}


def _regions_match(a: set, b: set) -> bool:
    return bool(a and b and a & b)


def _district_match(district_from_mo: str, org_name: str) -> bool:
    if not district_from_mo:
        return False
    prefix = _get_district_prefix(district_from_mo)
    if not prefix:
        return False

    lower = org_name.lower()
    marker = "муниципального района"
    idx = lower.find(marker)
    if idx != -1:
        before = lower[:idx].strip().split()
        if before and before[-1].startswith(prefix):
            return True

    return prefix in lower


def _get_district_prefix(district: str) -> str:
    if not district:
        return ""
    s = district.lower()
    for noise in DISTRICT_NOISE:
        s = s.replace(noise, " ")
    words = s.split()
    if not words:
        return ""
    return _word_prefix(words[0])


def _split_first(value) -> tuple[str | None, str | None]:
    if not value:
        return None, None
    s = str(value).strip()
    if "," in s:
        idx = s.index(",")
        return s[:idx].strip() or None, s[idx + 1:].strip() or None
    return s or None, None


def _clean(value) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split()).strip()
