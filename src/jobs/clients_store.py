from __future__ import annotations

import os
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from sqlalchemy import delete, select

from src.infra.db import session_scope
from src.infra.models import Client
from src.jobs.storage import normalize_job_id, resolve_job_paths


def _human_label_map() -> dict[str, str]:
    """Human-readable header (row 1) per technical column key (row 2)."""
    from src.parser.excel_writer import COLUMNS

    return {tech: human for human, tech in COLUMNS if tech}


def prepare_data_xlsx(job_id: str | None, local_path: Path | None = None) -> Path:
    paths = resolve_job_paths(job_id)
    target = Path(local_path) if local_path is not None else paths.data_xlsx
    if normalize_job_id(job_id):
        return materialize_xlsx(job_id, target)
    return target


def sync_client_row(job_id: str | None, row: dict[str, Any]) -> None:
    row_index = row.get("_row_index")
    if row_index in (None, ""):
        return
    data = {
        str(key): "" if value is None else str(value)
        for key, value in row.items()
        if str(key) != "_row_index"
    }
    update_client(job_id, int(row_index), data)


def import_clients_from_xlsx(job_id: str | None, local_xlsx: Path) -> int:
    """Import the MO template into the ``clients`` table.

    Uses the shared ``excel_io.load_rows`` reader so the technical column names
    (row 2: ``ID``, ``SUB_RF`` ...) and the physical ``_row_index`` match the
    rest of the pipeline (generator/sender/statistics). This keeps DB keys in
    sync with ``sync_client_row``/``update_client``.
    """
    normalized = normalize_job_id(job_id)
    if not normalized:
        return 0
    from src.generator.generation.excel_io import load_rows

    _, _, rows = load_rows(Path(local_xlsx))
    now = datetime.now(timezone.utc)
    rows_written = 0
    with session_scope() as session:
        session.execute(delete(Client).where(Client.job_id == normalized))
        for source in rows:
            row_index = source.get("_row_index")
            if row_index in (None, ""):
                continue
            data = {
                str(key): "" if value is None else str(value)
                for key, value in source.items()
                if str(key) != "_row_index"
            }
            session.add(
                Client(
                    job_id=normalized,
                    row_index=int(row_index),
                    data=data,
                    updated_at=now,
                )
            )
            rows_written += 1
    return rows_written


def list_clients(job_id: str | None) -> list[dict[str, Any]]:
    normalized = normalize_job_id(job_id)
    if not normalized:
        return []
    with session_scope() as session:
        rows = session.execute(
            select(Client).where(Client.job_id == normalized).order_by(Client.row_index.asc())
        ).scalars().all()
    return [dict(row.data) if isinstance(row.data, dict) else {} for row in rows]


def _list_clients_with_index(job_id: str | None) -> list[tuple[int, dict[str, Any]]]:
    normalized = normalize_job_id(job_id)
    if not normalized:
        return []
    with session_scope() as session:
        rows = session.execute(
            select(Client).where(Client.job_id == normalized).order_by(Client.row_index.asc())
        ).scalars().all()
    return [
        (int(row.row_index), dict(row.data) if isinstance(row.data, dict) else {})
        for row in rows
    ]


def update_client(job_id: str | None, row_index: int, data: dict[str, Any]) -> None:
    normalized = normalize_job_id(job_id)
    if not normalized:
        return
    now = datetime.now(timezone.utc)
    with session_scope() as session:
        row = session.execute(
            select(Client).where(Client.job_id == normalized, Client.row_index == row_index)
        ).scalar_one_or_none()
        if row is None:
            session.add(Client(job_id=normalized, row_index=row_index, data=data, updated_at=now))
        else:
            row.data = data
            row.updated_at = now


def _atomic_save(workbook: Workbook, local_path: Path) -> None:
    fd, tmp_name = tempfile.mkstemp(dir=str(local_path.parent), suffix=".xlsxtmp")
    os.close(fd)
    tmp_path = Path(tmp_name)
    try:
        workbook.save(tmp_path)
        os.replace(tmp_path, local_path)
    except BaseException:
        tmp_path.unlink(missing_ok=True)
        raise


def materialize_xlsx(job_id: str | None, local_path: Path) -> Path:
    """Rebuild data.xlsx from the ``clients`` table.

    Writes the canonical two-row header (row 1: human labels, row 2: technical
    keys) so that ``excel_io.load_rows`` (which reads headers from row 2 and data
    from row 3) parses the result correctly. Preserves each client's stored
    ``row_index`` and keeps any extra columns that are not part of the standard
    schema.
    """
    normalized = normalize_job_id(job_id)
    local_path = Path(local_path)
    local_path.parent.mkdir(parents=True, exist_ok=True)

    # The on-disk data.xlsx is the authoritative working copy: upload, sender
    # (save_workbook + sync_client_row) and municipality verification all keep
    # the file and the DB in sync. Rebuilding it from the DB is only needed when
    # the file is missing (e.g. a fresh container that only has the DB). Never
    # overwrite an existing file, so we never clobber parser enrichment or the
    # canonical two-row header layout that excel_io.load_rows expects.
    if local_path.exists():
        return local_path

    records = _list_clients_with_index(normalized)

    from src.parser.excel_writer import COL_KEYS

    label_map = _human_label_map()
    ordered_keys: list[str] = list(COL_KEYS)
    for _, data in records:
        for key in data.keys():
            if key and key != "_row_index" and key not in ordered_keys:
                ordered_keys.append(key)

    workbook = Workbook()
    worksheet = workbook.active
    for col_idx, key in enumerate(ordered_keys, start=1):
        worksheet.cell(row=1, column=col_idx, value=label_map.get(key, key))
        worksheet.cell(row=2, column=col_idx, value=key)

    key_to_col = {key: idx for idx, key in enumerate(ordered_keys, start=1)}
    next_free_row = 3
    for row_index, data in records:
        target_row = row_index if isinstance(row_index, int) and row_index >= 3 else next_free_row
        for key, value in data.items():
            col = key_to_col.get(str(key))
            if not col:
                continue
            cell = worksheet.cell(row=target_row, column=col, value="" if value is None else str(value))
            cell.number_format = "@"
        next_free_row = max(next_free_row, target_row + 1)

    _atomic_save(workbook, local_path)
    return local_path
