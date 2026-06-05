from __future__ import annotations

from collections.abc import Iterable
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from src.generator.inflection.inflection_report import load_inflection_log
from src.jobs import resolve_job_paths
from src.jobs.state import load_agent_state


BLUE = "1F4E78"
BORDER = Side(style="thin", color="D9E2F3")
REPORT_FILENAME = "journal_corrections_report.xlsx"


def build_correction_report_xlsx(job_id: str | None = None) -> Path:
    """Build an Excel journal with inflection and philologist corrections."""

    job_paths = resolve_job_paths(job_id)
    state_dir = job_paths.root_dir / "state"
    output_path = state_dir / REPORT_FILENAME

    philologist_state = _load_philologist_state(job_id)
    inflection_rows = load_inflection_log(job_id)
    correction_rows = list(_iter_philologist_corrections(philologist_state))

    workbook = Workbook()
    inflection_sheet = workbook.active
    inflection_sheet.title = "Склонения"
    corrections_sheet = workbook.create_sheet("Правки филолога")

    _add_title(inflection_sheet, "Журнал склонений")
    _write_table(
        inflection_sheet,
        3,
        ["ID строки", "Поле", "Было", "Стало", "Падеж", "Метод", "Уверенность", "Предупреждение", "Причина", "Контекст"],
        [
            [
                _short(row.get("row_id"), 120),
                _short(row.get("field"), 120),
                _short(row.get("source_value"), 800),
                _short(row.get("result_value"), 800),
                _short(row.get("target_case"), 120),
                _short(row.get("method"), 120),
                _short(row.get("confidence"), 120),
                _short(row.get("warning"), 800),
                _short(row.get("reason"), 1000),
                _short(row.get("filled_sentence") or row.get("context_sentence"), 1200),
            ]
            for row in inflection_rows
        ],
        name="InflectionLog",
    )

    _add_title(corrections_sheet, "Правки филолога")
    _write_table(
        corrections_sheet,
        3,
        ["Документ", "Место", "Фрагмент", "Замечание", "Предложение"],
        [
            [
                _short(row.get("document"), 1200),
                _short(row.get("location"), 120),
                _short(row.get("fragment"), 900),
                _short(row.get("issue"), 1000),
                _short(row.get("suggestion"), 900),
            ]
            for row in correction_rows
        ],
        name="PhilologistCorrections",
    )

    _style_workbook(workbook)
    _autosize(
        inflection_sheet,
        {"A": 10, "B": 18, "C": 34, "D": 34, "E": 16, "F": 18, "G": 14, "H": 28, "I": 42, "J": 80},
    )
    _autosize(corrections_sheet, {"A": 70, "B": 16, "C": 38, "D": 48, "E": 38})

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def correction_report_has_data(job_id: str | None = None) -> bool:
    philologist_state = _load_philologist_state(job_id)
    return bool(load_inflection_log(job_id) or list(_iter_philologist_corrections(philologist_state)))


def _load_philologist_state(job_id: str | None) -> dict[str, Any]:
    return load_agent_state("philologist", {}, job_id, include_details=True)


def _iter_philologist_corrections(state: dict[str, Any]) -> Iterable[dict[str, Any]]:
    for document in state.get("documents") or []:
        if not isinstance(document, dict):
            continue
        document_name = document.get("path") or document.get("name") or ""
        for fix in document.get("applied_fixes") or []:
            if isinstance(fix, dict):
                yield _correction_row(document_name, fix)

    for item in (state.get("inflection_context_corrections") or {}).get("corrections") or []:
        if not isinstance(item, dict):
            continue
        yield {
            "document": item.get("document", ""),
            "location": "",
            "fragment": item.get("fragment", ""),
            "issue": item.get("comment") or "LLM-проверка контекста подстановки.",
            "suggestion": item.get("suggestion", ""),
        }


def _correction_row(document_name: Any, fix: dict[str, Any]) -> dict[str, Any]:
    return {
        "document": document_name,
        "location": fix.get("location", ""),
        "fragment": fix.get("fragment", ""),
        "issue": fix.get("issue", ""),
        "suggestion": fix.get("suggestion", ""),
    }


def _add_title(sheet, title: str) -> None:
    sheet["A1"] = title
    sheet["A1"].font = Font(bold=True, size=16, color=BLUE)


def _write_table(sheet, start_row: int, headers: list[str], rows: list[list[Any]], *, name: str) -> None:
    for column_index, header in enumerate(headers, 1):
        cell = sheet.cell(start_row, column_index, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index, row in enumerate(rows, start_row + 1):
        for column_index, value in enumerate(row, 1):
            cell = sheet.cell(row_index, column_index, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    max_row = start_row + max(len(rows), 1)
    max_column = len(headers)
    for row in sheet.iter_rows(min_row=start_row, max_row=max_row, min_col=1, max_col=max_column):
        for cell in row:
            cell.border = Border(bottom=BORDER)

    if rows:
        table = Table(displayName=name, ref=f"A{start_row}:{get_column_letter(max_column)}{max_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )
        sheet.add_table(table)


def _style_workbook(workbook: Workbook) -> None:
    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A4"
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)


def _autosize(sheet, widths: dict[str, int]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _short(value: Any, max_len: int = 500) -> str:
    text = "" if value is None else str(value)
    text = " ".join(text.split())
    return text if len(text) <= max_len else text[: max_len - 1] + "…"
