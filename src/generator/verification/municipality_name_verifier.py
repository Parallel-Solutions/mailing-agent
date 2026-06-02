from __future__ import annotations

import html
import re
from dataclasses import dataclass, replace
from pathlib import Path
from typing import Any, Callable
from urllib.parse import urlparse

import requests
from openpyxl import load_workbook

from src.generator.verification.minjust_municipality_lookup import MinjustMunicipalityLookup
from src.generator.verification.oktmo_municipality_lookup import OktmoMunicipalityLookup
from src.utils.config import settings


HEADER_ROW = 2
DATA_START_ROW = 3
MUN_NAME_COLUMN = "MUN_NAME"
MUN_R_NAME_COLUMN = "MUN_R_NAME"
ADM_NAME_COLUMN = "ADM_NAME"

ORIGINAL_COLUMN = "MUN_NAME_SOURCE_ORIGINAL"
OFFICIAL_COLUMN = "MUN_NAME_OFFICIAL"
STATUS_COLUMN = "MUN_NAME_VERIFICATION_STATUS"
CONFIDENCE_COLUMN = "MUN_NAME_VERIFICATION_CONFIDENCE"
SOURCE_COLUMN = "MUN_NAME_VERIFICATION_SOURCE"
REASON_COLUMN = "MUN_NAME_VERIFICATION_REASON"
URL_COLUMN = "MUN_NAME_SOURCE_URL"

VERIFICATION_COLUMNS = (
    ORIGINAL_COLUMN,
    OFFICIAL_COLUMN,
    STATUS_COLUMN,
    CONFIDENCE_COLUMN,
    SOURCE_COLUMN,
    REASON_COLUMN,
    URL_COLUMN,
)

ADMINISTRATION_PREFIX_RE = re.compile(
    r"^\s*(?:администрация|администрации)\s+(?:муниципального\s+образования\s+)?",
    re.IGNORECASE,
)
MUNICIPAL_FORM_RE = re.compile(
    r"\b("
    r"сельское\s+поселение|городское\s+поселение|муниципальный\s+округ|"
    r"городской\s+округ|муниципальное\s+образование|поселок|посёлок|"
    r"сельсовет|город\s+[^,]+"
    r")\b",
    re.IGNORECASE,
)
QUOTE_RE = re.compile(r"[\"«„](.*?)[\"»“]")
LOCAL_ADMINISTRATION_WRAPPER_RE = re.compile(
    r"^(?:муниципальное\s+учреждение\s+)?местная\s+администрация\s+"
    r"(?:сельского|городского)\s+поселения\s+",
    re.IGNORECASE,
)
OFFICIAL_SITE_EXCLUDED_DOMAINS = (
    "audit-it.ru",
    "checko.ru",
    "egrul.nalog.ru",
    "kartoteka.ru",
    "list-org.com",
    "nalog.ru",
    "rusprofile.ru",
    "sbis.ru",
    "spark-interfax.ru",
    "synapsenet.ru",
    "vbankcenter.ru",
    "zachestnyibiznes.ru",
    "zakupki.gov.ru",
)
OFFICIAL_SITE_DOMAIN_HINTS = (
    ".gov.ru",
    ".gosuslugi.ru",
    ".mosreg.ru",
    ".tatarstan.ru",
    ".bashkortostan.ru",
    ".lenobl.ru",
    ".ryazangov.ru",
)
OFFICIAL_SITE_URL_HINTS = (
    "adm",
    "admin",
    "ustav",
    "mo-",
    "mun",
    "municip",
    "poselen",
    "selsovet",
    "сельсовет",
)
OFFICIAL_SITE_STRONG_URL_HINTS = (
    "ustav",
    "o-munitsipalnom-obrazovanii",
    "munitsipalnoe-obrazovanie",
    "municzipalnoe-obrazovanie",
    "about",
    "obshchie-svedeniya",
    "general-information",
)
OFFICIAL_SITE_LEGAL_PHRASES = (
    "устав муниципального образования",
    "о муниципальном образовании",
    "муниципальное образование",
    "официальный сайт администрации",
    "официальный сайт муниципального образования",
)
SEARCH_ABBREVIATION_PATTERNS: tuple[tuple[re.Pattern[str], str], ...] = (
    (re.compile(r"\bр\.?\s*п\.?\b", re.IGNORECASE), "рабочий поселок"),
    (re.compile(r"\bп\.?\s*г\.?\s*т\.?\b", re.IGNORECASE), "поселок городского типа"),
    (re.compile(r"\bг\.?\s*п\.?\b", re.IGNORECASE), "городское поселение"),
    (re.compile(r"\bг\.\b", re.IGNORECASE), "город"),
)
DISPLAY_LOCALITY_ABBREVIATION_PATTERNS: tuple[tuple[re.Pattern[str], str], ...] = (
    (re.compile(r"\bр\.?\s*п\.?\b", re.IGNORECASE), "рабочий поселок"),
    (re.compile(r"\bп\.?\s*г\.?\s*т\.?\b", re.IGNORECASE), "поселок городского типа"),
    (re.compile(r"\bг\.\b", re.IGNORECASE), "город"),
    (re.compile(r"\bс\.\b", re.IGNORECASE), "село"),
    (re.compile(r"\bд\.\b", re.IGNORECASE), "деревня"),
    (re.compile(r"\bп\.\b", re.IGNORECASE), "поселок"),
)
FORBIDDEN_MUNICIPALITY_NAME_FRAGMENTS = (
    "администрация",
    "муниципальное учреждение",
    "местная администрация",
)


@dataclass(frozen=True)
class MunicipalityNameVerification:
    row_id: str
    original_name: str
    official_name: str
    status: str
    confidence: str
    source: str
    reason: str
    source_url: str = ""

    @property
    def should_replace(self) -> bool:
        return self.status == "verified" and self.confidence == "high" and bool(self.official_name)


@dataclass(frozen=True)
class OfficialSiteMatch:
    url: str
    title: str
    content: str
    score: int
    matched_name: str = ""
    evidence: str = ""


@dataclass(frozen=True)
class SearchResultSnippet:
    url: str
    title: str
    content: str


class OfficialSiteLookup:
    SEARCH_URL = "https://yandex.ru/search/?text={query}"
    MAX_FETCH_BYTES = 512 * 1024
    MAX_QUERIES_PER_ROW = 3
    MAX_SNIPPETS_PER_QUERY = 6
    MAX_PAGE_FETCHES_PER_QUERY = 2
    MIN_SNIPPET_SCORE = 3

    def __init__(
        self,
        *,
        timeout_seconds: float = 15.0,
        verify_ssl: bool = True,
        fetcher: Any | None = None,
    ) -> None:
        self.timeout_seconds = timeout_seconds
        self.verify_ssl = verify_ssl
        self.fetcher = fetcher
        self.disabled_reason = ""

    def confirm(
        self,
        row: dict[str, Any],
        candidate_names: list[str],
    ) -> OfficialSiteMatch | None:
        queries = _build_official_site_queries(row, candidate_names)
        if not queries:
            return None

        best_match: OfficialSiteMatch | None = None
        for query in queries[: self.MAX_QUERIES_PER_ROW]:
            search_url = self.SEARCH_URL.format(query=requests.utils.quote(query))
            search_html = self._fetch_text(search_url)
            if not search_html:
                continue
            snippets = _parse_yandex_search_results(search_html)
            page_fetches = 0
            for snippet in snippets[: self.MAX_SNIPPETS_PER_QUERY]:
                snippet_match = _official_site_match_from_search_result(
                    {
                        "url": snippet.url,
                        "title": snippet.title,
                        "content": snippet.content,
                    },
                    row,
                    candidate_names=candidate_names,
                )
                if not snippet_match or snippet_match.score < self.MIN_SNIPPET_SCORE:
                    continue
                page_title, page_content = self._load_page_context(snippet.url)
                page_fetches += 1
                merged_title = page_title or snippet.title
                merged_content = " ".join(part for part in (snippet.content, page_content) if part).strip()
                match = _official_site_match_from_search_result(
                    {
                        "url": snippet.url,
                        "title": merged_title,
                        "content": merged_content,
                    },
                    row,
                    candidate_names=candidate_names,
                )
                if not match:
                    if page_fetches >= self.MAX_PAGE_FETCHES_PER_QUERY:
                        break
                    continue
                if not best_match or match.score > best_match.score:
                    best_match = match
                if best_match and best_match.score >= 12:
                    break
                if page_fetches >= self.MAX_PAGE_FETCHES_PER_QUERY:
                    break
            if best_match and best_match.score >= 12:
                break
        return best_match

    def _load_page_context(self, url: str) -> tuple[str, str]:
        page_html = self._fetch_text(url, connect_timeout=3, read_timeout=min(6.0, self.timeout_seconds))
        if not page_html:
            return "", ""
        return _extract_html_title(page_html), _html_to_text(page_html, limit=12000)

    def _fetch_text(self, url: str, *, connect_timeout: float = 5, read_timeout: float | None = None) -> str:
        response = None
        try:
            if self.fetcher:
                return str(self.fetcher(url, self.timeout_seconds, self.verify_ssl) or "")
            timeout_read = self.timeout_seconds if read_timeout is None else read_timeout
            response = requests.get(
                url,
                timeout=(connect_timeout, timeout_read),
                verify=self.verify_ssl,
                headers={"User-Agent": "Mozilla/5.0 mailing-agent municipality verifier"},
                stream=True,
            )
            response.raise_for_status()
            response.encoding = response.encoding or "utf-8"
            chunks: list[bytes] = []
            remaining = self.MAX_FETCH_BYTES
            for chunk in response.iter_content(chunk_size=16384, decode_unicode=False):
                if not chunk:
                    continue
                if len(chunk) >= remaining:
                    chunks.append(chunk[:remaining])
                    break
                chunks.append(chunk)
                remaining -= len(chunk)
                if remaining <= 0:
                    break
            payload = b"".join(chunks)
            return payload.decode(response.encoding or "utf-8", errors="ignore")
        except requests.RequestException as exc:
            self.disabled_reason = str(exc) or exc.__class__.__name__
            return ""
        finally:
            try:
                response.close()  # type: ignore[name-defined]
            except Exception:
                pass


def verify_municipality_name(
    row: dict[str, Any],
    *,
    oktmo_lookup: OktmoMunicipalityLookup | None = None,
    minjust_lookup: MinjustMunicipalityLookup | None = None,
) -> MunicipalityNameVerification:
    """Choose the best local official-name candidate from row data.

    This is the deterministic first layer. It extracts candidates from the
    uploaded table, including quoted names from ADM_NAME, but only allows
    automatic replacement after stronger confirmation.
    """

    current_name = _clean(row.get(MUN_NAME_COLUMN))
    adm_name = _clean(row.get(ADM_NAME_COLUMN))

    quoted_candidate = _extract_quoted_municipality_name(adm_name, current_name)
    if quoted_candidate:
        normalized = normalize_municipality_display_name(quoted_candidate)
        if not _candidate_is_safe_for_autoreplace(normalized, current_name):
            return MunicipalityNameVerification(
                row_id=_clean(row.get("ID")),
                original_name=current_name,
                official_name=normalize_municipality_display_name(current_name) or normalized,
                status="kept",
                confidence="medium",
                source="ADM_NAME",
                reason="Из ADM_NAME извлечен кандидат, но он слишком сильно расходится с исходным MUN_NAME или выглядит как название администрации; автозамена отключена до внешнего подтверждения.",
            )
        if oktmo_lookup:
            oktmo_result = oktmo_lookup.confirm(row, normalized)
            if oktmo_result:
                return MunicipalityNameVerification(
                    row_id=_clean(row.get("ID")),
                    original_name=current_name,
                    official_name=normalize_municipality_display_name(oktmo_result.name),
                    status="verified",
                    confidence="high",
                    source="oktmo+ADM_NAME",
                    reason=(
                        "Название извлечено из кавычек в ADM_NAME и подтверждено официальным "
                        f"классификатором ОКТМО Росстата; код ОКТМО {oktmo_result.oktmo_code}."
                    ),
                    source_url=oktmo_result.source_url,
                )
        if minjust_lookup:
            minjust_result = minjust_lookup.confirm(row, normalized)
            if minjust_result:
                return MunicipalityNameVerification(
                    row_id=_clean(row.get("ID")),
                    original_name=current_name,
                    official_name=normalize_municipality_display_name(minjust_result.name),
                    status="kept",
                    confidence="medium",
                    source="minjust",
                    reason="Название выглядит корректным и даже подтверждается Минюстом, но без подтверждения через ОКТМО автозамена отключена.",
                    source_url=minjust_result.source_url,
                )
        if current_name and _normalize_for_match(current_name) == _normalize_for_match(normalized):
            return MunicipalityNameVerification(
                row_id=_clean(row.get("ID")),
                original_name=current_name,
                official_name=normalized,
                status="kept",
                confidence="medium",
                source="ADM_NAME+MUN_NAME",
                reason="Название извлечено из кавычек в полном названии администрации и совпадает с MUN_NAME, но без подтверждения через ОКТМО не считается подтвержденным.",
            )
        if _candidate_is_safe_for_autoreplace(normalized, current_name):
            return MunicipalityNameVerification(
                row_id=_clean(row.get("ID")),
                original_name=current_name,
                official_name=normalized,
                status="kept",
                confidence="medium",
                source="ADM_NAME",
                reason="Название извлечено из кавычек в полном названии администрации и выглядит безопасным, но без подтверждения через ОКТМО автозамена отключена.",
            )
        return MunicipalityNameVerification(
            row_id=_clean(row.get("ID")),
            original_name=current_name,
            official_name=normalized,
            status="kept",
            confidence="medium",
            source="ADM_NAME",
            reason="Название извлечено из кавычек в полном названии администрации, но без внешнего подтверждения исходный MUN_NAME оставлен.",
        )

    adm_candidate = extract_municipality_name_from_administration(adm_name)
    if adm_candidate:
        normalized = normalize_municipality_display_name(adm_candidate)
        if not _candidate_is_safe_for_autoreplace(normalized, current_name):
            return MunicipalityNameVerification(
                row_id=_clean(row.get("ID")),
                original_name=current_name,
                official_name=normalize_municipality_display_name(current_name) or normalized,
                status="kept",
                confidence="medium",
                source="ADM_NAME",
                reason="Из ADM_NAME извлечен кандидат, но автозамена отключена: название похоже на служебную формулировку или меняет основной топоним без внешнего подтверждения.",
            )
        if oktmo_lookup:
            oktmo_result = oktmo_lookup.confirm(row, normalized)
            if oktmo_result:
                return MunicipalityNameVerification(
                    row_id=_clean(row.get("ID")),
                    original_name=current_name,
                    official_name=normalize_municipality_display_name(oktmo_result.name),
                    status="verified",
                    confidence="high",
                    source="oktmo+ADM_NAME",
                    reason=(
                        "Название извлечено из ADM_NAME и подтверждено официальным "
                        f"классификатором ОКТМО Росстата; код ОКТМО {oktmo_result.oktmo_code}."
                    ),
                    source_url=oktmo_result.source_url,
                )
        if minjust_lookup:
            minjust_result = minjust_lookup.confirm(row, normalized)
            if minjust_result:
                return MunicipalityNameVerification(
                    row_id=_clean(row.get("ID")),
                    original_name=current_name,
                    official_name=normalize_municipality_display_name(minjust_result.name),
                    status="kept",
                    confidence="medium",
                    source="minjust",
                    reason="Название выглядит корректным и даже подтверждается Минюстом, но без подтверждения через ОКТМО автозамена отключена.",
                    source_url=minjust_result.source_url,
                )
        if current_name and _normalize_for_match(current_name) == _normalize_for_match(normalized):
            return MunicipalityNameVerification(
                row_id=_clean(row.get("ID")),
                original_name=current_name,
                official_name=normalized,
                status="kept",
                confidence="medium",
                source="ADM_NAME+MUN_NAME",
                reason="Название извлечено из полного названия администрации без кавычек и совпадает с MUN_NAME, но без подтверждения через ОКТМО не считается подтвержденным.",
            )
        return MunicipalityNameVerification(
            row_id=_clean(row.get("ID")),
            original_name=current_name,
            official_name=normalized,
            status="kept",
            confidence="medium",
            source="ADM_NAME",
            reason="Кандидат извлечен из полного названия администрации без кавычек, но без сильного внешнего подтверждения исходный MUN_NAME оставлен.",
        )

    if current_name and oktmo_lookup:
        oktmo_result = oktmo_lookup.confirm(row, current_name)
        if oktmo_result:
            return MunicipalityNameVerification(
                row_id=_clean(row.get("ID")),
                original_name=current_name,
                official_name=normalize_municipality_display_name(oktmo_result.name),
                status="verified",
                confidence="high",
                source="oktmo",
                reason=(
                    "Исходный MUN_NAME сверен с официальным классификатором ОКТМО Росстата; "
                    f"код ОКТМО {oktmo_result.oktmo_code}."
                ),
                source_url=oktmo_result.source_url,
            )

    if _adm_name_confirms_current_municipality(adm_name, current_name):
        return MunicipalityNameVerification(
            row_id=_clean(row.get("ID")),
            original_name=current_name,
            official_name=normalize_municipality_display_name(current_name),
            status="kept",
            confidence="medium",
            source="ADM_NAME+MUN_NAME",
            reason="MUN_NAME согласован с названием администрации, но без подтверждения через ОКТМО не считается подтвержденным.",
        )

    stripped_candidate = _extract_from_administration_name(adm_name)
    if stripped_candidate:
        normalized = normalize_municipality_display_name(stripped_candidate)
        return MunicipalityNameVerification(
            row_id=_clean(row.get("ID")),
            original_name=current_name,
            official_name=normalized,
            status="kept",
            confidence="medium",
            source="ADM_NAME",
            reason="Название получено после удаления слова «Администрация», но без явного официального подтверждения; исходный MUN_NAME оставлен.",
        )

    expanded_current_name = _expand_locality_abbreviations_for_display(current_name)
    if current_name and expanded_current_name and expanded_current_name != normalize_municipality_display_name(current_name):
        return MunicipalityNameVerification(
            row_id=_clean(row.get("ID")),
            original_name=current_name,
            official_name=expanded_current_name,
            status="kept",
            confidence="medium",
            source="normalization",
            reason="Сокращенная форма населенного пункта распознана, но без подтверждения через ОКТМО автоматическая замена отключена.",
        )

    if current_name:
        return MunicipalityNameVerification(
            row_id=_clean(row.get("ID")),
            original_name=current_name,
            official_name=normalize_municipality_display_name(current_name),
            status="kept",
            confidence="low",
            source="MUN_NAME",
            reason="В ADM_NAME не найдено надежного официального названия; оставлен исходный MUN_NAME.",
        )

    return MunicipalityNameVerification(
        row_id=_clean(row.get("ID")),
        original_name=current_name,
        official_name="",
        status="missing",
        confidence="low",
        source="",
        reason="Не заполнены MUN_NAME и ADM_NAME.",
    )


def verify_municipality_names_in_workbook(
    xlsx_path: Path,
    *,
    use_official_sites: bool = False,
    use_oktmo: bool = False,
    oktmo_lookup: OktmoMunicipalityLookup | None = None,
    official_site_lookup: OfficialSiteLookup | None = None,
    use_minjust: bool = True,
    minjust_lookup: MinjustMunicipalityLookup | None = None,
    progress_callback: Callable[[dict[str, Any]], None] | None = None,
) -> dict[str, Any]:
    """Update data.xlsx with municipality-name verification.

    The local deterministic layer is always used. Official-site search is an
    optional production layer because it depends on an external API and can take
    noticeable time for large tables.
    """

    workbook = load_workbook(xlsx_path)
    worksheet = workbook[workbook.sheetnames[0]]
    header_map = _ensure_headers(worksheet)
    oktmo_lookup = oktmo_lookup if use_oktmo else None
    if use_oktmo and oktmo_lookup is None:
        oktmo_lookup = OktmoMunicipalityLookup()
    official_site_lookup = official_site_lookup if use_official_sites else None
    if use_official_sites and official_site_lookup is None:
        official_site_lookup = OfficialSiteLookup(
            timeout_seconds=settings.municipality_official_sites_timeout_seconds,
            verify_ssl=settings.municipality_official_sites_verify_ssl,
        )
    minjust_lookup = minjust_lookup if use_minjust else None
    if use_minjust and minjust_lookup is None:
        minjust_lookup = MinjustMunicipalityLookup()

    mun_col = header_map.get(MUN_NAME_COLUMN)
    if not mun_col:
        district_col = header_map.get(MUN_R_NAME_COLUMN)
        if district_col:
            stats = _district_table_verification_result(worksheet, header_map)
            workbook.close()
            return stats
        workbook.close()
        return {
            "status": "skipped",
            "updated_rows": 0,
            "verified_rows": 0,
            "kept_rows": 0,
            "replacements": [],
            "replacement_samples": [],
            "reason": "В таблице нет колонки MUN_NAME.",
        }

    max_column = worksheet.max_column
    row_values_list = list(
        worksheet.iter_rows(
            min_row=DATA_START_ROW,
            max_row=worksheet.max_row,
            max_col=max_column,
            values_only=True,
        )
    )
    estimated_total_rows = sum(
        1
        for row_values in row_values_list
        if (
            any(value not in (None, "") for value in _read_row_from_values(header_map, row_values).values())
            and not _is_service_row(_read_row_from_values(header_map, row_values))
        )
    )

    stats = {
        "status": "ok",
        "total_rows": 0,
        "estimated_total_rows": estimated_total_rows,
        "unique_verification_keys": 0,
        "cached_verification_rows": 0,
        "updated_rows": 0,
        "verified_rows": 0,
        "kept_rows": 0,
        "missing_rows": 0,
        "official_site_checked_rows": 0,
        "official_site_found_rows": 0,
        "official_site_error_rows": 0,
        "official_site_lookup_enabled": bool(official_site_lookup),
        "official_site_lookup_disabled_reason": "",
        "oktmo_lookup_enabled": bool(oktmo_lookup),
        "oktmo_lookup_disabled_reason": "",
        "minjust_lookup_enabled": bool(minjust_lookup),
        "minjust_lookup_disabled_reason": "",
        "replacements": [],
        "replacement_samples": [],
        "decision_samples": [],
    }

    def report_progress(force: bool = False) -> None:
        if progress_callback is None:
            return
        estimated_total = int(stats.get("estimated_total_rows") or 0)
        processed = int(stats.get("total_rows") or 0)
        step = max(1, estimated_total // 20) if estimated_total else 1
        if force or processed <= 1 or processed % step == 0 or processed >= estimated_total:
            progress_callback(
                {
                    "processed_rows": processed,
                    "total_rows": estimated_total,
                    "verified_rows": int(stats.get("verified_rows") or 0),
                    "updated_rows": int(stats.get("updated_rows") or 0),
                    "missing_rows": int(stats.get("missing_rows") or 0),
                }
            )

    report_progress(force=True)
    verification_cache: dict[tuple[str, ...], MunicipalityNameVerification] = {}
    for offset, row_values in enumerate(row_values_list):
        row_index = DATA_START_ROW + offset
        row = _read_row_from_values(header_map, row_values)
        if not any(value not in (None, "") for value in row.values()) or _is_service_row(row):
            continue
        stats["total_rows"] += 1
        cache_key = _verification_cache_key(row)
        cached_verification = verification_cache.get(cache_key)
        if cached_verification:
            verification = replace(cached_verification, row_id=_clean(row.get("ID")))
            stats["cached_verification_rows"] += 1
        else:
            verification = verify_municipality_name(row, oktmo_lookup=oktmo_lookup, minjust_lookup=minjust_lookup)
            verification_cache[cache_key] = verification
        if official_site_lookup and _should_use_official_site_followup(row, verification):
            verification = _verify_with_official_site(row, verification, stats, official_site_lookup)
        _write_verification(worksheet, header_map, row_index, verification)

        if verification.status == "verified":
            stats["verified_rows"] += 1
        elif verification.status == "kept":
            stats["kept_rows"] += 1
        elif verification.status == "missing":
            stats["missing_rows"] += 1

        current_name = _clean(row.get(MUN_NAME_COLUMN))
        if verification.should_replace and verification.official_name != current_name:
            worksheet.cell(row=row_index, column=mun_col).value = verification.official_name
            stats["updated_rows"] += 1
            replacement = {
                "row_id": verification.row_id or str(row_index - HEADER_ROW),
                "from": current_name,
                "to": verification.official_name,
                "source": verification.source,
                "confidence": verification.confidence,
            }
            stats["replacements"].append(replacement)
            if len(stats["replacement_samples"]) < 20:
                stats["replacement_samples"].append(replacement)
        if len(stats["decision_samples"]) < 20:
            stats["decision_samples"].append(
                {
                    "row_id": verification.row_id or str(row_index - HEADER_ROW),
                    "original_name": current_name,
                    "official_name": verification.official_name,
                    "status": verification.status,
                    "confidence": verification.confidence,
                    "source": verification.source,
                    "reason": verification.reason,
                    "source_url": verification.source_url,
                    "auto_replaced": bool(verification.should_replace and verification.official_name != current_name),
                }
            )
        report_progress()

    stats["unique_verification_keys"] = len(verification_cache)
    if official_site_lookup and official_site_lookup.disabled_reason:
        stats["official_site_lookup_disabled_reason"] = official_site_lookup.disabled_reason
    if oktmo_lookup and oktmo_lookup.disabled_reason:
        stats["oktmo_lookup_disabled_reason"] = oktmo_lookup.disabled_reason
    if minjust_lookup and minjust_lookup.disabled_reason:
        stats["minjust_lookup_disabled_reason"] = minjust_lookup.disabled_reason

    _remove_verification_columns(worksheet)
    workbook.save(xlsx_path)
    workbook.close()
    report_progress(force=True)
    return stats


def _district_table_verification_result(worksheet, header_map: dict[str, int]) -> dict[str, Any]:
    max_column = worksheet.max_column
    total_rows = 0
    missing_rows = 0
    decision_samples: list[dict[str, Any]] = []
    for row_values in worksheet.iter_rows(
        min_row=DATA_START_ROW,
        max_row=worksheet.max_row,
        max_col=max_column,
        values_only=True,
    ):
        row = _read_row_from_values(header_map, row_values)
        if not any(value not in (None, "") for value in row.values()) or _is_service_row(row):
            continue
        total_rows += 1
        district_name = _clean(row.get(MUN_R_NAME_COLUMN))
        if not district_name:
            missing_rows += 1
        if len(decision_samples) < 20:
            decision_samples.append(
                {
                    "row_id": _clean(row.get("ID")) or str(total_rows),
                    "original_name": district_name,
                    "official_name": district_name,
                    "status": "kept" if district_name else "missing",
                    "confidence": "medium" if district_name else "low",
                    "source": MUN_R_NAME_COLUMN if district_name else "",
                    "reason": (
                        "Районная таблица: основная сущность взята из MUN_R_NAME, проверка MUN_NAME не требуется."
                        if district_name
                        else "В районной таблице не заполнен MUN_R_NAME."
                    ),
                    "source_url": "",
                    "auto_replaced": False,
                }
            )

    kept_rows = max(0, total_rows - missing_rows)
    return {
        "status": "ok",
        "table_mode": "district",
        "total_rows": total_rows,
        "estimated_total_rows": total_rows,
        "unique_verification_keys": kept_rows,
        "cached_verification_rows": 0,
        "updated_rows": 0,
        "verified_rows": 0,
        "kept_rows": kept_rows,
        "missing_rows": missing_rows,
        "official_site_checked_rows": 0,
        "official_site_found_rows": 0,
        "official_site_error_rows": 0,
        "official_site_lookup_enabled": False,
        "official_site_lookup_disabled_reason": "",
        "oktmo_lookup_enabled": False,
        "oktmo_lookup_disabled_reason": "Для районной таблицы без MUN_NAME проверка ОКТМО по MUN_NAME пропущена.",
        "minjust_lookup_enabled": False,
        "minjust_lookup_disabled_reason": "",
        "replacements": [],
        "replacement_samples": [],
        "decision_samples": decision_samples,
    }


def _verify_with_official_site(
    row: dict[str, Any],
    local_verification: MunicipalityNameVerification,
    stats: dict[str, Any],
    official_site_lookup: OfficialSiteLookup,
) -> MunicipalityNameVerification:
    if local_verification.should_replace:
        return local_verification
    stats["official_site_checked_rows"] += 1
    try:
        match = find_official_site_for_municipality(
            row,
            verification=local_verification,
            lookup=official_site_lookup,
        )
    except Exception as exc:
        stats["official_site_error_rows"] += 1
        return _with_reason_suffix(
            local_verification,
            f"Проверка официального сайта не выполнена: {exc}",
        )

    if not match:
        return _with_reason_suffix(local_verification, "Официальный сайт по строке не найден.")

    stats["official_site_found_rows"] += 1
    current_name = _clean(row.get(MUN_NAME_COLUMN))
    expected_name = normalize_municipality_display_name(local_verification.official_name)
    matched_name = normalize_municipality_display_name(match.matched_name)
    source_label = match.title or match.url

    if matched_name and expected_name and _normalize_for_match(matched_name) == _normalize_for_match(expected_name):
        return MunicipalityNameVerification(
            row_id=local_verification.row_id,
            original_name=local_verification.original_name,
            official_name=matched_name,
            status="verified",
            confidence="high",
            source=f"{local_verification.source}+official_site",
            reason=(
                f"{local_verification.reason} Официальный сайт подтвердил форму "
                f"«{matched_name}»: {source_label}. {match.evidence}".strip()
            ),
            source_url=match.url,
        )

    if matched_name and current_name and _normalize_for_match(matched_name) == _normalize_for_match(current_name):
        return MunicipalityNameVerification(
            row_id=local_verification.row_id,
            original_name=local_verification.original_name,
            official_name=normalize_municipality_display_name(current_name),
            status="verified",
            confidence="medium",
            source="official_site",
            reason=f"MUN_NAME найден на официальном сайте: {source_label}. Замена не требуется. {match.evidence}".strip(),
            source_url=match.url,
        )

    if (
        matched_name
        and _needs_abbreviation_resolution(current_name, local_verification.official_name)
        and match.score >= 10
        and _contains_legal_municipality_evidence(f"{match.title} {match.content}")
        and _candidate_is_safe_for_autoreplace(matched_name, current_name)
    ):
        return MunicipalityNameVerification(
            row_id=local_verification.row_id,
            original_name=local_verification.original_name,
            official_name=matched_name,
            status="verified",
            confidence="high",
            source="official_site",
            reason=(
                f"Сокращенная форма в MUN_NAME уточнена по официальному сайту: {source_label}. "
                f"{match.evidence}"
            ).strip(),
            source_url=match.url,
        )

    return MunicipalityNameVerification(
        row_id=local_verification.row_id,
        original_name=local_verification.original_name,
        official_name=local_verification.official_name,
        status=local_verification.status,
        confidence=local_verification.confidence,
        source=f"{local_verification.source}+official_site",
        reason=(
            f"{local_verification.reason} Официальный сайт найден ({source_label}), "
            "но сильного подтверждения для автозамены не получено."
        ),
        source_url=match.url,
    )


def _should_use_official_site_followup(
    row: dict[str, Any],
    verification: MunicipalityNameVerification,
) -> bool:
    if verification.should_replace:
        return False
    if verification.status == "missing":
        return False
    if _needs_abbreviation_resolution(
        _clean(row.get(MUN_NAME_COLUMN)),
        verification.official_name,
        _clean(row.get(ADM_NAME_COLUMN)),
    ):
        return True
    if verification.status == "kept":
        return True
    return verification.confidence in {"low", "medium"}


def find_official_site_for_municipality(
    row: dict[str, Any],
    *,
    verification: MunicipalityNameVerification | None = None,
    lookup: OfficialSiteLookup | None = None,
) -> OfficialSiteMatch | None:
    lookup = lookup or OfficialSiteLookup(
        timeout_seconds=settings.municipality_official_sites_timeout_seconds,
        verify_ssl=settings.municipality_official_sites_verify_ssl,
    )
    candidate_names = _build_official_site_candidates(row, verification)
    return lookup.confirm(row, candidate_names)


def _build_official_site_query(row: dict[str, Any], candidate_name: str = "") -> str:
    adm_name = _clean(row.get(ADM_NAME_COLUMN))
    mun_name = _clean(row.get(MUN_NAME_COLUMN))
    sub_rf = _clean(row.get("SUB_RF"))
    district = _clean(row.get("MUN_R_NAME"))
    base_name = _clean(candidate_name) or mun_name
    if adm_name:
        return f'"{adm_name}" официальный сайт администрация {sub_rf}'.strip()
    if base_name:
        return f'"{base_name}" официальный сайт администрация {district} {sub_rf}'.strip()
    return ""


def _build_official_site_queries(row: dict[str, Any], candidate_names: list[str]) -> list[str]:
    queries: list[str] = []
    for candidate_name in candidate_names[:4]:
        for suffix in (
            "официальный сайт администрация",
            "устав муниципального образования",
            "о муниципальном образовании",
        ):
            sub_rf = _clean(row.get("SUB_RF"))
            district = _clean(row.get("MUN_R_NAME"))
            query = f'"{candidate_name}" {suffix} {district} {sub_rf}'.strip()
            if query and query not in queries:
                queries.append(query)
    fallback_query = _build_official_site_query(row)
    if fallback_query and fallback_query not in queries:
        queries.append(fallback_query)
    return queries


def _build_official_site_candidates(
    row: dict[str, Any],
    verification: MunicipalityNameVerification | None = None,
) -> list[str]:
    candidates: list[str] = []
    for value in (
        verification.official_name if verification else "",
        extract_municipality_name_from_administration(_clean(row.get(ADM_NAME_COLUMN))),
        _extract_quoted_municipality_name(_clean(row.get(ADM_NAME_COLUMN)), _clean(row.get(MUN_NAME_COLUMN))),
        _clean(row.get(MUN_NAME_COLUMN)),
    ):
        for variant in _search_name_variants(value):
            if variant and variant not in candidates:
                candidates.append(variant)
    return candidates


def _parse_yandex_search_results(search_html: str) -> list[SearchResultSnippet]:
    snippets: list[SearchResultSnippet] = []
    for match in re.finditer(r"<a[^>]+href=\"([^\"]+)\"[^>]*>(.*?)</a>", search_html, re.IGNORECASE | re.DOTALL):
        url = html.unescape(_clean(match.group(1)))
        if not url.startswith("http"):
            continue
        title = _html_to_text(match.group(2), limit=300)
        content = ""
        snippets.append(SearchResultSnippet(url=url, title=title, content=content))
        if len(snippets) >= 20:
            break
    return snippets


def _extract_html_title(page_html: str) -> str:
    match = re.search(r"<title[^>]*>(.*?)</title>", page_html, re.IGNORECASE | re.DOTALL)
    return _html_to_text(match.group(1), limit=300) if match else ""


def _html_to_text(value: str, *, limit: int = 4000) -> str:
    text = re.sub(r"(?is)<script.*?>.*?</script>", " ", value)
    text = re.sub(r"(?is)<style.*?>.*?</style>", " ", text)
    text = re.sub(r"(?s)<[^>]+>", " ", text)
    text = html.unescape(text)
    text = re.sub(r"\s+", " ", text).strip()
    return text[:limit]


def _pick_matching_candidate_name(text: str, candidate_names: list[str]) -> str:
    normalized_text = _normalize_for_match(text)
    best_match = ""
    best_len = 0
    for candidate_name in candidate_names:
        for variant in _search_name_variants(candidate_name):
            normalized_candidate = _normalize_for_match(variant)
            if not normalized_candidate:
                continue
            if normalized_candidate in normalized_text and len(normalized_candidate) > best_len:
                best_match = normalize_municipality_display_name(variant)
                best_len = len(normalized_candidate)
    return best_match


def _contains_legal_municipality_evidence(text: str) -> bool:
    normalized_text = _normalize_for_match(text)
    return any(_normalize_for_match(phrase) in normalized_text for phrase in OFFICIAL_SITE_LEGAL_PHRASES)


def _official_site_match_from_search_result(
    result: dict[str, Any],
    row: dict[str, Any],
    *,
    candidate_names: list[str] | None = None,
) -> OfficialSiteMatch | None:
    url = _clean(result.get("url"))
    if not url:
        return None
    domain = urlparse(url).netloc.lower().removeprefix("www.")
    if not domain or any(domain == item or domain.endswith(f".{item}") for item in OFFICIAL_SITE_EXCLUDED_DOMAINS):
        return None

    title = _clean(result.get("title"))
    content = _clean(result.get("content"))
    haystack = f"{title} {content}".lower()
    score = 0
    evidence: list[str] = []
    if any(hint in domain for hint in OFFICIAL_SITE_DOMAIN_HINTS):
        score += 5
        evidence.append("домен похож на официальный")
    if any(hint in url.lower() for hint in OFFICIAL_SITE_URL_HINTS):
        score += 3
        evidence.append("url похож на сайт администрации")
    if any(hint in url.lower() for hint in OFFICIAL_SITE_STRONG_URL_HINTS):
        score += 4
        evidence.append("url похож на страницу устава или описания МО")
    if "официальный" in haystack:
        score += 2
        evidence.append("есть слово «официальный»")
    if "администрац" in haystack:
        score += 2
        evidence.append("упомянута администрация")
    if _contains_legal_municipality_evidence(f"{title} {content}"):
        score += 4
        evidence.append("есть юридически сильная формулировка про МО или устав")
    if _text_contains_municipality_name(f"{title} {content}", _clean(row.get(MUN_NAME_COLUMN))):
        score += 3
        evidence.append("есть текущее MUN_NAME")
    if _adm_name_mentions_search_result(_clean(row.get(ADM_NAME_COLUMN)), f"{title} {content}"):
        score += 3
        evidence.append("сайт совпадает с ADM_NAME")
    matched_name = _pick_matching_candidate_name(f"{title} {content}", candidate_names or [])
    if matched_name:
        score += 6
        evidence.append(f"найдено название «{matched_name}»")
    if score < 3:
        return None
    return OfficialSiteMatch(
        url=url,
        title=title,
        content=content,
        score=score,
        matched_name=matched_name,
        evidence="; ".join(evidence),
    )


def _text_contains_municipality_name(text: str, municipality_name: str) -> bool:
    normalized_text = _normalize_for_match(text)
    if not normalized_text:
        return False
    for variant in _search_name_variants(municipality_name):
        keyword = _municipality_keyword_base(_normalize_for_match(variant))
        if keyword and (keyword in normalized_text or keyword.removesuffix("ск") in normalized_text):
            return True
    return False


def _needs_abbreviation_resolution(*values: str) -> bool:
    for value in values:
        normalized = _normalize_for_match(value)
        if any(pattern.search(normalized) for pattern, _ in SEARCH_ABBREVIATION_PATTERNS):
            return True
    return False


def _adm_name_mentions_search_result(adm_name: str, text: str) -> bool:
    normalized_text = _normalize_for_match(text)
    for word in _administration_municipality_name_words(_normalize_for_match(adm_name)):
        base = _adjective_base(word)
        if base and (base in normalized_text or base.removesuffix("ск") in normalized_text):
            return True
    return False


def _with_reason_suffix(
    verification: MunicipalityNameVerification,
    suffix: str,
) -> MunicipalityNameVerification:
    return MunicipalityNameVerification(
        row_id=verification.row_id,
        original_name=verification.original_name,
        official_name=verification.official_name,
        status=verification.status,
        confidence=verification.confidence,
        source=verification.source,
        reason=f"{verification.reason} {suffix}",
        source_url=verification.source_url,
    )


def normalize_municipality_display_name(value: Any) -> str:
    text = _clean(value)
    if not text:
        return ""
    text = text.strip("\"«»“”„")
    if _is_mostly_upper(text):
        text = text.lower()
        text = _restore_municipality_capitalization(text)
    text = _normalize_quotes_spacing(text)
    return text


def _expand_locality_abbreviations_for_display(value: Any) -> str:
    text = normalize_municipality_display_name(value)
    if not text:
        return ""
    expanded = text
    for pattern, replacement in DISPLAY_LOCALITY_ABBREVIATION_PATTERNS:
        expanded = re.sub(pattern, replacement, expanded)
    expanded = re.sub(r"\s+", " ", expanded).strip()
    return normalize_municipality_display_name(expanded)


def _search_name_variants(value: Any) -> list[str]:
    normalized = normalize_municipality_display_name(value)
    if not normalized:
        return []
    variants = [normalized]
    queue = [normalized]
    while queue:
        current = queue.pop(0)
        for pattern, replacement in SEARCH_ABBREVIATION_PATTERNS:
            expanded = re.sub(pattern, replacement, current)
            expanded = normalize_municipality_display_name(expanded)
            if expanded and expanded not in variants:
                variants.append(expanded)
                queue.append(expanded)
    return variants


def _ensure_headers(worksheet) -> dict[str, int]:
    header_map = {
        str(worksheet.cell(row=HEADER_ROW, column=column_index).value).strip(): column_index
        for column_index in range(1, worksheet.max_column + 1)
        if worksheet.cell(row=HEADER_ROW, column=column_index).value
    }
    return header_map


def _read_row(worksheet, header_map: dict[str, int], row_index: int) -> dict[str, Any]:
    return {
        header: worksheet.cell(row=row_index, column=column_index).value
        for header, column_index in header_map.items()
    }


def _read_row_from_values(header_map: dict[str, int], row_values: tuple[Any, ...]) -> dict[str, Any]:
    return {
        header: row_values[column_index - 1] if column_index <= len(row_values) else None
        for header, column_index in header_map.items()
    }


def _is_service_row(row: dict[str, Any]) -> bool:
    first_value = ""
    for value in row.values():
        if value not in (None, ""):
            first_value = str(value).strip()
            break
    return first_value.lower().startswith("источники:")


def _verification_cache_key(row: dict[str, Any]) -> tuple[str, ...]:
    return (
        _normalize_for_match(row.get("SUB_RF")),
        _normalize_for_match(row.get("MUN_R_NAME")),
        _normalize_for_match(row.get(MUN_NAME_COLUMN)),
        _normalize_for_match(row.get(ADM_NAME_COLUMN)),
    )


def _write_verification(
    worksheet,
    header_map: dict[str, int],
    row_index: int,
    verification: MunicipalityNameVerification,
) -> None:
    return


def _remove_verification_columns(worksheet) -> None:
    headers_to_remove = set(VERIFICATION_COLUMNS)
    columns_to_delete = [
        column_index
        for column_index in range(1, worksheet.max_column + 1)
        if _clean(worksheet.cell(row=HEADER_ROW, column=column_index).value) in headers_to_remove
    ]
    for column_index in sorted(columns_to_delete, reverse=True):
        worksheet.delete_cols(column_index)


def _extract_quoted_municipality_name(adm_name: str, current_name: str = "") -> str:
    if not adm_name:
        return ""
    for match in QUOTE_RE.finditer(adm_name):
        candidate = _clean(match.group(1))
        unwrapped_candidate = _unwrap_local_administration_candidate(candidate)
        if unwrapped_candidate and not _looks_like_municipality(unwrapped_candidate):
            contextual_candidate = _compose_contextual_quoted_name(adm_name, unwrapped_candidate)
            if contextual_candidate and _looks_like_municipality(contextual_candidate) and not _candidate_contains_forbidden_fragments(contextual_candidate):
                return contextual_candidate
        if unwrapped_candidate and _looks_like_municipality(unwrapped_candidate) and not _candidate_contains_forbidden_fragments(unwrapped_candidate):
            return unwrapped_candidate
        if _is_bare_locality_name(candidate) and _looks_like_settlement(current_name):
            contextual_candidate = _compose_contextual_quoted_name(adm_name, candidate)
            if contextual_candidate and _looks_like_municipality(contextual_candidate) and not _candidate_contains_forbidden_fragments(contextual_candidate):
                return contextual_candidate
        if _looks_like_municipality(candidate) and not _candidate_contains_forbidden_fragments(candidate):
            return candidate
        contextual_candidate = _compose_contextual_quoted_name(adm_name, candidate)
        if contextual_candidate and _looks_like_municipality(contextual_candidate) and not _candidate_contains_forbidden_fragments(contextual_candidate):
            return contextual_candidate
    return ""


def _extract_from_administration_name(adm_name: str) -> str:
    if not adm_name:
        return ""
    without_prefix = ADMINISTRATION_PREFIX_RE.sub("", adm_name).strip()
    without_prefix = without_prefix.strip("\"«»“”„")
    if (
        without_prefix
        and without_prefix != adm_name
        and _looks_like_municipality(without_prefix)
        and not _candidate_contains_forbidden_fragments(without_prefix)
        and not _candidate_contains_region_tail(without_prefix)
    ):
        return without_prefix
    return ""


def extract_municipality_name_from_administration(adm_name: str) -> str:
    """Extract an unquoted municipality name from a full administration title."""

    text = _clean(adm_name)
    if not text:
        return ""

    prefixed_patterns = (
        (
            r"\bадминистрация\s+([а-яё-]+)\s+сельского\s+поселения\b",
            "Сельское поселение",
        ),
        (
            r"\bадминистрация\s+([а-яё-]+)\s+городского\s+поселения\b",
            "Городское поселение",
        ),
    )
    for pattern, prefix in prefixed_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if not match:
            continue
        tail = _normalize_adjective_settlement_name(match.group(1))
        candidate = _compose_settlement_name(prefix, tail)
        if _looks_like_municipality(candidate):
            return candidate

    patterns = (
        (
            r"\bгородского\s+поселения\s+((?:город|пос[её]лок|пгт)\s+.+?)"
            r"(?=\s+муниципального\s+района|\s+муниципальный\s+район|\s+республики\b|\s+области\b|\s+края\b|$)",
            "Городское поселение",
        ),
        (
            r"\bгородского\s+поселения\s+((?:[а-яё-]+(?:\s+[а-яё-]+){0,3}))"
            r"(?=\s+муниципального\s+района|\s+муниципальный\s+район|\s+республики\b|\s+области\b|\s+края\b|$)",
            "Городское поселение",
        ),
        (
            r"\bсельского\s+поселения\s+((?:село|деревня|станица|аул)\s+.+?)"
            r"(?=\s+муниципального\s+района|\s+муниципальный\s+район|\s+республики\b|\s+области\b|\s+края\b|$)",
            "Сельское поселение",
        ),
        (
            r"\bсельского\s+поселения\s+((?:[а-яё-]+(?:\s+[а-яё-]+){0,3}))"
            r"(?=\s+муниципального\s+района|\s+муниципальный\s+район|\s+республики\b|\s+области\b|\s+края\b|$)",
            "Сельское поселение",
        ),
        (
            r"\bгородского\s+округа\s+(.+?)"
            r"(?=\s+муниципального\s+района|\s+муниципальный\s+район|\s+республики\b|\s+области\b|\s+края\b|$)",
            "Городской округ",
        ),
        (
            r"\bмуниципального\s+округа\s+(.+?)"
            r"(?=\s+муниципального\s+района|\s+муниципальный\s+район|\s+республики\b|\s+области\b|\s+края\b|$)",
            "Муниципальный округ",
        ),
    )
    for pattern, prefix in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if not match:
            continue
        tail = _normalize_locality_fragment(match.group(1))
        candidate = _compose_settlement_name(prefix, tail)
        if _looks_like_municipality(candidate) and not _candidate_contains_region_tail(candidate):
            return candidate
    return ""


def _looks_like_municipality(value: str) -> bool:
    text = _clean(value)
    return bool(text and MUNICIPAL_FORM_RE.search(text))


def _looks_like_settlement(value: str) -> bool:
    lowered = _clean(value).lower()
    return "городское поселение" in lowered or "сельское поселение" in lowered


def _is_bare_locality_name(value: str) -> bool:
    lowered = _clean(value).lower()
    return bool(re.match(r"^(?:город|пос[её]лок|п\.|село|деревня)\b", lowered))


def _compose_contextual_quoted_name(adm_name: str, candidate: str) -> str:
    adm_text = _clean(adm_name).lower()
    if "городского поселения" in adm_text or "городское поселение" in adm_text:
        return _compose_settlement_name("Городское поселение", _normalize_locality_fragment(candidate))
    if "сельского поселения" in adm_text or "сельское поселение" in adm_text:
        return _compose_settlement_name("Сельское поселение", _normalize_locality_fragment(candidate))
    return ""


def _compose_settlement_name(prefix: str, tail: str) -> str:
    normalized_tail = _clean(tail)
    normalized_prefix = _clean(prefix)
    if not normalized_prefix:
        return normalized_tail
    if not normalized_tail:
        return normalized_prefix
    lowered_tail = _normalize_for_match(normalized_tail)
    locality_prefixes = (
        "город ",
        "поселок ",
        "посёлок ",
        "рабочий поселок ",
        "рабочий посёлок ",
        "село ",
        "деревня ",
        "станица ",
        "аул ",
    )
    if lowered_tail.startswith(locality_prefixes):
        return f"{normalized_prefix} {normalized_tail}".strip()
    if " " in normalized_tail:
        return f"{normalized_prefix} {normalized_tail}".strip()
    return f"{_settlement_adjective_tail(normalized_tail)} {normalized_prefix.lower()}".strip()


def _settlement_adjective_tail(value: str) -> str:
    word = _readable_capitalized_word(_clean(value))
    lowered = word.lower().replace("ё", "е")
    if lowered.endswith(("ское", "цкое", "ое", "ее")):
        return word
    if lowered.endswith(("ский", "цкий")):
        return word[:-2] + "ое"
    if lowered.endswith(("ый", "ой", "ий")):
        return word[:-2] + "ое"
    if lowered.endswith(("ово", "ево", "ино", "ыно")):
        return word[:-1] + "ское"
    if lowered.endswith("ка") and len(word) > 4:
        return word[:-2] + "ское"
    if lowered.endswith("ск"):
        return word + "ое"
    if lowered.endswith(("а", "я")):
        return word[:-1] + "ское"
    if lowered.endswith("ь"):
        return word[:-1] + "ское"
    return word + "ское"


def _unwrap_local_administration_candidate(value: str) -> str:
    text = _clean(value)
    if not text:
        return ""
    stripped = LOCAL_ADMINISTRATION_WRAPPER_RE.sub("", text).strip()
    return stripped if stripped and stripped != text else ""


def _normalize_locality_fragment(value: str) -> str:
    text = _clean(value).strip("\"«»“”„")
    text = re.sub(r"(?i)^п\.\s*", "поселок ", text)
    text = re.sub(r"(?i)^пос\.\s*", "поселок ", text)
    text = re.sub(r"(?i)^пос\.\s*([а-яё])", lambda m: "поселок " + m.group(1), text)
    text = re.sub(r"(?i)^г\.\s*", "город ", text)
    if _is_mostly_upper(text):
        text = text.lower()
    words = text.split()
    if not words:
        return ""
    if words[0].lower() in {"город", "поселок", "посёлок", "село", "деревня"}:
        words[0] = words[0].lower()
        words[1:] = [_readable_capitalized_word(word) for word in words[1:]]
        return " ".join(words)
    return " ".join(_readable_capitalized_word(word) for word in words)


def _normalize_adjective_settlement_name(value: str) -> str:
    word = _clean(value)
    if not word:
        return ""
    lower = word.lower().replace("ё", "е")
    replacements = (
        ("ского", "ское"),
        ("цкого", "цкое"),
        ("ого", "ое"),
        ("его", "ее"),
        ("ий", "ий"),
        ("ый", "ый"),
        ("ой", "ой"),
    )
    for source, target in replacements:
        if lower.endswith(source) and len(lower) > len(source) + 1:
            return _capitalize_hyphenated(lower[: -len(source)] + target)
    return _readable_capitalized_word(word)


def _candidate_contains_region_tail(value: str) -> bool:
    normalized = _normalize_for_match(value)
    blocked_tokens = (
        "района",
        "район",
        "области",
        "область",
        "республики",
        "республика",
        "края",
        "край",
    )
    return any(token in normalized.split() for token in blocked_tokens)


def _candidate_contains_forbidden_fragments(value: str) -> bool:
    normalized = _normalize_for_match(value)
    return any(fragment in normalized for fragment in FORBIDDEN_MUNICIPALITY_NAME_FRAGMENTS)


def _candidate_is_safe_for_autoreplace(candidate_name: str, current_name: str) -> bool:
    candidate_name = normalize_municipality_display_name(candidate_name)
    current_name = normalize_municipality_display_name(current_name)
    if not candidate_name:
        return False
    if _candidate_contains_forbidden_fragments(candidate_name):
        return False
    if _candidate_contains_region_tail(candidate_name):
        return False
    if _has_bad_short_form(candidate_name):
        return False
    if current_name and _drops_municipality_type(current_name, candidate_name):
        return False
    if current_name and not _same_primary_toponym(current_name, candidate_name):
        return False
    return True


def _has_bad_short_form(value: str) -> bool:
    normalized = _clean(value).lower()
    return any(token in normalized for token in ("пос.", "г.", "п.")) and "поселок " not in normalized and "город " not in normalized


def _drops_municipality_type(current_name: str, candidate_name: str) -> bool:
    current_type = _municipality_type_label(current_name)
    candidate_type = _municipality_type_label(candidate_name)
    return bool(current_type and candidate_type and current_type != candidate_type) or bool(current_type and not candidate_type)


def _municipality_type_label(value: str) -> str:
    lowered = _normalize_for_match(value)
    for label in (
        "городское поселение",
        "сельское поселение",
        "городской округ",
        "муниципальный округ",
        "муниципальное образование",
    ):
        if label in lowered:
            return label
    return ""


def _same_primary_toponym(left: str, right: str) -> bool:
    left_tokens = _primary_toponym_tokens(left)
    right_tokens = _primary_toponym_tokens(right)
    if not left_tokens or not right_tokens:
        return True
    if left_tokens == right_tokens:
        return True
    if len(left_tokens) == 1 and len(right_tokens) == 1:
        return _adjective_bases_match(_adjective_base(left_tokens[0]), _adjective_base(right_tokens[0]))
    left_joined = " ".join(left_tokens)
    right_joined = " ".join(right_tokens)
    return left_joined == right_joined


def _primary_toponym_tokens(value: str) -> list[str]:
    normalized = _normalize_for_match(value)
    words = [
        word for word in normalized.split()
        if word not in {
            "городское", "сельское", "поселение", "городской", "муниципальный",
            "округ", "муниципальное", "образование", "город", "поселок", "поселок",
            "поселок", "поселок", "поселок", "сельсовет", "поссовет", "поселок",
            "рп", "пгт", "гп", "г", "рабочий", "типа",
        }
    ]
    return words[:3]


def _adm_name_confirms_current_municipality(adm_name: str, current_name: str) -> bool:
    """Confirm common administration names against current MUN_NAME.

    These names do not contain the official municipality name in quotes, so we
    should not rewrite MUN_NAME from them. Still, they are strong enough to
    confirm that the current row is internally consistent.
    """

    adm_text = _clean(adm_name).lower()
    mun_text = _clean(current_name).lower()
    if not adm_text or not mun_text:
        return False
    mun_keyword = _municipality_keyword_base(mun_text)
    if not mun_keyword:
        return False
    for adm_word in _administration_municipality_name_words(adm_text):
        if _adjective_bases_match(_adjective_base(adm_word), mun_keyword):
            return True
    return False


def _municipality_keyword_base(mun_text: str) -> str:
    for pattern in (
        r"\bсельское\s+поселение\s+([а-яё-]+)",
        r"\bгородское\s+поселение\s+([а-яё-]+)",
    ):
        match = re.search(pattern, mun_text, re.IGNORECASE)
        if match:
            return _adjective_base(match.group(1))
    return _adjective_base(_first_word(mun_text))


def _administration_municipality_name_words(adm_text: str) -> list[str]:
    patterns = (
        r"\b([а-яё-]+)\s+сельская\s+администрация\b",
        r"\bадминистрация\s+(?:муниципального\s+образования\s+[-–—]?\s*)?([а-яё-]+)\s+сельское\s+поселение\b",
        r"\bадминистрация\s+(?:муниципального\s+образования\s+[-–—]?\s*)?([а-яё-]+)\s+городское\s+поселение\b",
        r"\bадминистрация\s+(?:муниципального\s+образования\s+[-–—]?\s*)?([а-яё-]+)\s+сельсовет\b",
        r"\bадминистрация\s+([а-яё-]+)\s+сельского\s+поселения\b",
        r"\bадминистрация\s+([а-яё-]+)\s+городского\s+поселения\b",
        r"\bадминистрация\s+([а-яё-]+)\s+сельсовета\b",
        r"\bадминистрация\s+муниципального\s+образования\s+сельского\s+поселения\s+([а-яё-]+)\b",
        r"\bадминистрация\s+муниципального\s+образования\s+городского\s+поселения\s+([а-яё-]+)\b",
        r"\bадминистрация\s+сельского\s+поселения\s+([а-яё-]+)\b",
        r"\bадминистрация\s+городского\s+поселения\s+([а-яё-]+)\b",
    )
    return [match.group(1) for pattern in patterns for match in re.finditer(pattern, adm_text, re.IGNORECASE)]


def _first_word(value: str) -> str:
    match = re.search(r"[а-яёА-ЯЁ-]+", value)
    return match.group(0) if match else ""


def _adjective_base(value: str) -> str:
    word = value.lower().replace("ё", "е").strip("-")
    for ending in (
        "ского",
        "цкого",
        "ская",
        "цкая",
        "ское",
        "цкое",
        "ого",
        "его",
        "ая",
        "ое",
        "яя",
        "ее",
        "ий",
        "ый",
        "ой",
    ):
        if word.endswith(ending) and len(word) > len(ending) + 2:
            return word[: -len(ending)]
    return word


def _adjective_bases_match(left: str, right: str) -> bool:
    if not left or not right:
        return False
    if left == right:
        return True
    return left.removesuffix("ск") == right.removesuffix("ск")


def _is_mostly_upper(value: str) -> bool:
    letters = [char for char in value if char.isalpha()]
    if not letters:
        return False
    upper_letters = [char for char in letters if char.isupper()]
    return len(upper_letters) / len(letters) > 0.8


def _title_first_word(value: str) -> str:
    words = value.split(" ")
    if not words:
        return value
    words[0] = _capitalize_hyphenated(words[0])
    return " ".join(words)


def _restore_municipality_capitalization(value: str) -> str:
    text = _title_first_word(value)
    words = text.split(" ")
    if words and words[0].lower() in {"город", "поселок", "посёлок", "село", "деревня"}:
        words[0] = words[0].lower()
        words[1:] = [_readable_capitalized_word(word) for word in words[1:]]
        return " ".join(words)
    prefixes = (
        ("Сельское", "поселение"),
        ("Городское", "поселение"),
        ("Муниципальный", "округ"),
        ("Городской", "округ"),
        ("Муниципальное", "образование"),
    )
    for prefix in prefixes:
        if len(words) > len(prefix) and tuple(words[: len(prefix)]) == prefix:
            tail_index = len(prefix)
            if words[tail_index].lower() in {"город", "село", "поселок", "посёлок", "деревня"}:
                words[tail_index] = words[tail_index].lower()
                if len(words) > tail_index + 1:
                    words[tail_index + 1] = _capitalize_hyphenated(words[tail_index + 1])
            else:
                words[tail_index] = _capitalize_hyphenated(words[tail_index])
            return " ".join(words)
    return text


def _capitalize_hyphenated(value: str) -> str:
    return "-".join(part[:1].upper() + part[1:] for part in value.split("-") if part)


def _readable_capitalized_word(value: str) -> str:
    word = value.lower() if _is_mostly_upper(value) else value
    return _capitalize_hyphenated(word)


def _normalize_quotes_spacing(value: str) -> str:
    text = re.sub(r"\s+", " ", value).strip()
    text = text.replace(" ,", ",").replace(" .", ".")
    return text


def _normalize_for_match(value: Any) -> str:
    text = _clean(value).lower().replace("ё", "е")
    text = re.sub(r"[^а-яa-z0-9-]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split()).strip()
