from __future__ import annotations

from collections import Counter
import csv
from datetime import datetime, timedelta
import io
import json
from pathlib import Path
from time import sleep
import threading
from typing import Any
from urllib.request import Request, urlopen

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from src.generator.delivery.sender_agent import (
    _build_unisender_go_url,
    _check_unisender_classic_messages,
    _load_sender_state,
    _load_sent_mail_log_items,
    _parse_emails,
    _run_unisender_request,
    _safe_text,
    _unisender_status_label,
)
from src.generator.delivery.consent_store import load_consent_records
from src.generator.delivery.mailopost_events import load_mailopost_events
from src.generator.delivery.rusender_events import load_rusender_events
from src.generator.delivery.unisender_go_events import load_unisender_go_events
from src.generator.generation.excel_io import load_rows
from src.jobs import resolve_job_paths
from src.utils.config import settings


BLUE = "1F4E78"
LIGHT_BLUE = "D9EAF7"
BORDER = Side(style="thin", color="D9E2F3")
REPORT_FILENAME = "sender_delivery_report.xlsx"
LEGACY_REPORT_FILENAME = "unisender_delivery_report.xlsx"
ANALYTICS_CACHE_FILENAME = "sender_delivery_analytics.json"
EVENT_DUMP_POLL_ATTEMPTS = 15
EVENT_DUMP_POLL_SECONDS = 2.0
_ANALYTICS_REFRESH_LOCK = threading.Lock()
_ANALYTICS_REFRESH_THREADS: dict[str, threading.Thread] = {}
UNISENDER_GO_SOFT_ERROR_STATUSES = {
    "soft_bounced",
    "err_will_retry",
    "skip_dup_temp_unreachable",
}
UNISENDER_GO_HARD_ERROR_STATUSES = {
    "hard_bounced",
    "err_user_unknown",
    "err_user_inactive",
    "err_mailbox_full",
    "err_delivery_failed",
}
UNISENDER_GO_TECHNICAL_ERROR_STATUSES = {
    "err_lost",
    "err_spam_rejected",
    "err_spam_skipped",
    "skip_dup_unreachable",
}
PROVIDER_LABELS = {
    "rusender": "RuSender",
    "unisender": "UniSender",
    "unisender_go": "UniSender Go",
    "unisender_classic": "UniSender",
    "smtp": "SMTP",
    "mailopost": "MailoPost",
}


def build_sender_delivery_report_xlsx(job_id: str | None = None, *, refresh: bool = True) -> Path:
    """Build an Excel journal for delivery attempts."""

    job_paths = resolve_job_paths(job_id)
    output_path = job_paths.root_dir / "state" / REPORT_FILENAME
    rows, refresh_error = _build_delivery_rows(job_id, refresh=refresh)
    consent_rows = _build_consent_rows(job_id)

    workbook = Workbook()
    stats_sheet = workbook.active
    stats_sheet.title = "Статистика"
    journal_sheet = workbook.create_sheet("Журнал отправки")
    consent_sheet = workbook.create_sheet("Согласия")

    _write_statistics_sheet(stats_sheet, rows, job_id=job_id, refresh_error=refresh_error, consent_rows=consent_rows)
    _write_journal_sheet(journal_sheet, rows)
    _write_consent_sheet(consent_sheet, consent_rows)
    _style_workbook(workbook)
    _autosize(
        stats_sheet,
        {"A": 34, "B": 22, "D": 28, "E": 16},
    )
    _autosize(
        journal_sheet,
        {
            "A": 10,
            "B": 34,
            "C": 32,
            "D": 22,
            "E": 42,
            "F": 20,
            "G": 24,
            "H": 34,
            "I": 16,
            "J": 18,
            "K": 18,
            "L": 24,
            "M": 46,
        },
    )
    _autosize(
        consent_sheet,
        {
            "A": 10,
            "B": 34,
            "C": 32,
            "D": 18,
            "E": 22,
            "F": 22,
            "G": 24,
            "H": 22,
            "I": 22,
            "J": 28,
            "K": 18,
            "L": 18,
            "M": 38,
            "N": 48,
        },
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def build_sender_delivery_analytics(
    job_id: str | None = None,
    *,
    refresh: bool = False,
    refresh_wait: bool = False,
) -> dict[str, Any]:
    """Return lightweight dashboard metrics for the current sending job."""

    refresh_started = False
    if refresh and refresh_wait:
        rows, refresh_error = _build_delivery_rows(job_id, refresh=True)
    else:
        if refresh:
            refresh_started = _start_sender_delivery_refresh(job_id)
        rows, refresh_error = _build_delivery_rows(job_id, refresh=False)
    consent_rows = _build_consent_rows(job_id)
    consent_summary = _consent_summary(consent_rows)
    statuses = Counter(_normalize_provider_status(row["provider_status"] or "unknown") for row in rows)
    providers = Counter(row.get("provider") or "unisender" for row in rows)
    provider_key, provider_label = _primary_provider(providers)

    total = len(rows)
    accepted_statuses = {"accepted", "success", "ok_sent", "sent", "queued", "not_sent", "processing"}
    delivered_statuses = {
        "delivered",
        "opened",
        "clicked",
        "subscribed",
        "unsubscribed",
        "spam",
        "ok_delivered",
        "ok_read",
        "ok_link_visited",
        "ok_unsubscribed",
        "ok_spam_folder",
    }
    read_statuses = {"opened", "clicked", "subscribed", "unsubscribed", "spam", "ok_read", "ok_link_visited"}
    clicked_statuses = {"clicked", "ok_link_visited"}
    unsubscribed_statuses = {"unsubscribed", "ok_unsubscribed"}
    spam_statuses = {"spam", "ok_spam_folder"}
    soft_bounce_statuses = UNISENDER_GO_SOFT_ERROR_STATUSES
    hard_bounce_statuses = UNISENDER_GO_HARD_ERROR_STATUSES
    pending_statuses = accepted_statuses | {"unknown"}

    delivered = sum(statuses.get(status, 0) for status in delivered_statuses)
    read = sum(statuses.get(status, 0) for status in read_statuses)
    clicked = sum(statuses.get(status, 0) for status in clicked_statuses)
    unsubscribed = sum(statuses.get(status, 0) for status in unsubscribed_statuses)
    spam = sum(statuses.get(status, 0) for status in spam_statuses)
    soft_bounced = sum(statuses.get(status, 0) for status in soft_bounce_statuses)
    hard_bounced = sum(statuses.get(status, 0) for status in hard_bounce_statuses)
    warnings = unsubscribed + spam
    generic_errors = sum(
        count
        for status, count in statuses.items()
        if (status.startswith("err_") or status.startswith("skip_"))
        and status not in soft_bounce_statuses
        and status not in hard_bounce_statuses
    )
    errors = hard_bounced + soft_bounced + generic_errors
    pending = sum(statuses.get(status, 0) for status in pending_statuses)
    accepted = total
    checked = sum(1 for row in rows if row.get("checked_at"))
    provider_events_count = sum(
        count
        for status, count in statuses.items()
        if status not in accepted_statuses and status not in {"unknown"}
    )
    awaiting_provider_events = bool(
        total > 0 and provider_events_count <= 0 and pending >= accepted
    )

    def pct(value: int, base: int | None = None) -> float:
        denominator = total if base is None else base
        if denominator <= 0:
            return 0.0
        return round((value / denominator) * 100, 1)

    def consent_pct(value: int) -> float:
        denominator = consent_summary["total"]
        if denominator <= 0:
            return 0.0
        return round((value / denominator) * 100, 1)

    cards = [
        {
            "id": "accepted",
            "title": f"Передано в {provider_label}" if provider_label else "Передано провайдеру",
            "value": accepted,
            "percent": 100.0 if total else 0.0,
            "hint": f"Email-получатели, которых наш сервис успешно передал в {provider_label}." if provider_label else "Email-получатели, которых наш сервис успешно передал провайдеру.",
            "tone": "good" if accepted and not errors else "neutral",
        },
        {
            "id": "confirmed_consents",
            "title": "Дали согласие",
            "value": consent_summary["confirmed"],
            "percent": consent_pct(consent_summary["confirmed"]),
            "hint": "Получатели, которые перешли по consent-ссылке и подтвердили получение материалов.",
            "tone": "good" if consent_summary["confirmed"] else "neutral",
        },
        {
            "id": "materials_sent",
            "title": "Материалы отправлены",
            "value": consent_summary["materials_sent"],
            "percent": consent_pct(consent_summary["materials_sent"]),
            "hint": "Сколько подтверждённых получателей уже получили КП/договор после согласия.",
            "tone": "good" if consent_summary["materials_sent"] else "neutral",
        },
        {
            "id": "delivered",
            "title": "Доставлено",
            "value": delivered,
            "percent": pct(delivered),
            "hint": "Подтверждённая доставка по событиям провайдера, если они доступны.",
            "tone": "neutral" if awaiting_provider_events else "good",
        },
        {
            "id": "opened",
            "title": "Открытия",
            "value": read,
            "percent": pct(read),
            "hint": "Письма, где провайдер зафиксировал открытие или более позднее действие.",
            "tone": "neutral" if awaiting_provider_events else "good",
        },
        {
            "id": "clicked",
            "title": "Переходы",
            "value": clicked,
            "percent": pct(clicked),
            "hint": "Письма, где провайдер зафиксировал переход по ссылке.",
            "tone": "neutral" if awaiting_provider_events else "good",
        },
        {
            "id": "hard_bounced",
            "title": "Недоставлено",
            "value": hard_bounced,
            "percent": pct(hard_bounced),
            "hint": "Финальные ошибки доставки: письмо больше не будет доставляться.",
            "tone": "bad" if hard_bounced else "neutral",
        },
        {
            "id": "soft_bounced",
            "title": "Временные ошибки",
            "value": soft_bounced,
            "percent": pct(soft_bounced),
            "hint": "Временные недоставки: провайдер ещё может повторить попытку.",
            "tone": "warn" if soft_bounced else "neutral",
        },
        {
            "id": "complaints",
            "title": "Отписки/спам",
            "value": unsubscribed + spam,
            "percent": pct(unsubscribed + spam),
            "hint": "Отписки и жалобы на спам по событиям провайдера.",
            "tone": "warn" if (unsubscribed + spam) else "neutral",
        },
        {
            "id": "pending",
            "title": "В обработке",
            "value": pending,
            "percent": pct(pending),
            "hint": "Письма без финального события доставки, открытия, клика или ошибки.",
            "tone": "warn" if pending else "neutral",
        },
    ]
    rates = {
        "delivery_rate": pct(delivered),
        "open_rate": pct(read, delivered or total),
        "ctr": pct(clicked),
        "error_rate": pct(errors),
        "pending_rate": pct(pending),
    }

    return {
        "status": "ok",
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "job_id": job_id or "",
        "provider": provider_key,
        "provider_label": provider_label,
        "provider_labels": {
            provider: _provider_label(provider)
            for provider, count in providers.items()
            if count
        },
        "total": total,
        "checked": checked,
        "refresh_error": refresh_error,
        "refresh_started": refresh_started,
        "refresh_in_progress": _is_sender_delivery_refresh_running(job_id),
        "awaiting_provider_events": awaiting_provider_events,
        "provider_events_count": provider_events_count,
        "summary": {
            "accepted": accepted,
            "delivered": delivered,
            "opened": read,
            "clicked": clicked,
            "unsubscribed": unsubscribed,
            "spam": spam,
            "soft_bounced": soft_bounced,
            "hard_bounced": hard_bounced,
            "errors": errors,
            "pending": pending,
            "providers": dict(providers),
            "rates": rates,
            "consents": consent_summary,
        },
        "consents": consent_summary,
        "cards": cards,
        "statuses": [
            {"status": status, "label": _report_status_label(status), "count": count}
            for status, count in sorted(statuses.items(), key=lambda item: (-item[1], item[0]))
        ],
        "note": _analytics_note(total=total, checked=checked, providers=providers, refresh_error=refresh_error),
    }


def build_unisender_delivery_report_xlsx(job_id: str | None = None, *, refresh: bool = True) -> Path:
    return build_sender_delivery_report_xlsx(job_id, refresh=refresh)


def build_unisender_delivery_analytics(
    job_id: str | None = None,
    *,
    refresh: bool = False,
    refresh_wait: bool = False,
) -> dict[str, Any]:
    return build_sender_delivery_analytics(job_id, refresh=refresh, refresh_wait=refresh_wait)


def _start_sender_delivery_refresh(job_id: str | None) -> bool:
    key = str(job_id or "__legacy__")
    with _ANALYTICS_REFRESH_LOCK:
        current = _ANALYTICS_REFRESH_THREADS.get(key)
        if current and current.is_alive():
            return False
        thread = threading.Thread(
            target=_run_sender_delivery_refresh_background,
            kwargs={"job_id": job_id},
            daemon=True,
            name=f"sender-analytics-{key}",
        )
        _ANALYTICS_REFRESH_THREADS[key] = thread
        thread.start()
        return True


def _is_sender_delivery_refresh_running(job_id: str | None) -> bool:
    key = str(job_id or "__legacy__")
    with _ANALYTICS_REFRESH_LOCK:
        current = _ANALYTICS_REFRESH_THREADS.get(key)
        return bool(current and current.is_alive())


def _run_sender_delivery_refresh_background(job_id: str | None) -> None:
    key = str(job_id or "__legacy__")
    try:
        _build_delivery_rows(job_id, refresh=True)
    finally:
        with _ANALYTICS_REFRESH_LOCK:
            current = _ANALYTICS_REFRESH_THREADS.get(key)
            if current is threading.current_thread():
                _ANALYTICS_REFRESH_THREADS.pop(key, None)


def sender_delivery_report_has_data(job_id: str | None = None) -> bool:
    return bool(_load_delivery_log_items(job_id) or load_consent_records(job_id))


def unisender_delivery_report_has_data(job_id: str | None = None) -> bool:
    return sender_delivery_report_has_data(job_id)


def _build_delivery_rows(job_id: str | None, *, refresh: bool) -> tuple[list[dict[str, Any]], str]:
    items = _load_delivery_log_items(job_id)
    refresh_messages: list[str] = []
    if refresh:
        go_refresh_error = _refresh_unisender_go_event_dump(job_id, items)
        if go_refresh_error:
            refresh_messages.append(go_refresh_error)
    go_events = _latest_unisender_go_events(job_id)
    rusender_events = _latest_rusender_events(job_id)
    mailopost_events = _latest_mailopost_events(job_id)
    cached_statuses = _load_delivery_status_cache(job_id)
    provider_statuses: dict[str, dict[str, Any]] = {}
    refresh_error = ""

    classic_ids = [_message_id(item) for item in items if _provider_name(item) == "unisender_classic" and _message_id(item)]
    if refresh and classic_ids:
        try:
            provider_statuses = _check_classic_statuses(classic_ids)
            if provider_statuses:
                cached_statuses.update(provider_statuses)
                _save_delivery_status_cache(job_id, cached_statuses)
        except Exception as exc:
            refresh_error = _safe_text(exc) or "не удалось обновить статусы UniSender"
    if refresh_error:
        refresh_messages.append(refresh_error)
    refresh_error = " ".join(message for message in refresh_messages if message).strip()

    rows: list[dict[str, Any]] = []
    for item in items:
        provider = item.get("provider") if isinstance(item.get("provider"), dict) else {}
        message_id = _message_id(item)
        accepted_status = _safe_text(provider.get("status")) or "accepted"
        provider_status = accepted_status
        checked_at = _safe_text(provider.get("checked_at"))
        if message_id and message_id in provider_statuses:
            provider_status = _safe_text(provider_statuses[message_id].get("provider_status")) or provider_status
            checked_at = _safe_text(provider_statuses[message_id].get("checked_at")) or checked_at
        elif message_id and message_id in cached_statuses:
            provider_status = _safe_text(cached_statuses[message_id].get("provider_status")) or provider_status
            checked_at = _safe_text(cached_statuses[message_id].get("checked_at")) or checked_at
        go_event = _match_unisender_go_event(item, go_events)
        if go_event:
            provider_status = _safe_text(go_event.get("provider_status")) or provider_status
            checked_at = _safe_text(go_event.get("checked_at")) or checked_at
        rusender_event = _match_rusender_event(item, rusender_events)
        if rusender_event:
            provider_status = _safe_text(rusender_event.get("provider_status")) or provider_status
            checked_at = _safe_text(rusender_event.get("checked_at")) or checked_at
        mailopost_event = _match_mailopost_event(item, mailopost_events)
        if mailopost_event:
            provider_status = _safe_text(mailopost_event.get("provider_status")) or provider_status
            checked_at = _safe_text(mailopost_event.get("checked_at")) or checked_at

        label = _report_status_label(provider_status)
        delivery_response = _delivery_response_text(go_event, rusender_event, mailopost_event)
        rows.append(
            {
                "row_id": _safe_text(item.get("row_id")),
                "mun_name": _safe_text(item.get("mun_name")),
                "recipient": _safe_text(item.get("recipient")),
                "sent_at": _safe_text(item.get("sent_at")),
                "subject": _safe_text(item.get("subject")),
                "accepted_status": accepted_status,
                "provider": _provider_name(item),
                "provider_status": _normalize_provider_status(provider_status),
                "provider_status_label": label,
                "delivery_response": delivery_response,
                "outcome": _delivery_outcome(provider_status),
                "email_id": message_id,
                "message_id": _safe_text(item.get("provider_job_id") or provider.get("job_id")),
                "checked_at": checked_at,
                "comment": _comment_text(item, refresh_error),
            }
        )
    return rows, refresh_error


def _refresh_unisender_go_event_dump(job_id: str | None, items: list[dict[str, Any]]) -> str:
    if not job_id:
        return ""
    go_items = [item for item in items if _provider_name(item) == "unisender_go"]
    if not go_items:
        return ""
    api_key = _safe_text(settings.unisender_api_key)
    if not api_key:
        return "Не указан API-ключ UniSender Go, поэтому не удалось добрать события доставки."

    start_time, end_time = _event_dump_time_window(go_items)
    request_body = {
        "start_time": start_time,
        "end_time": end_time,
        "limit": 100000,
        "dump_fields": [
            "event_time",
            "job_id",
            "email",
            "status",
            "delivery_status",
            "metadata",
            "url",
        ],
        "format": "csv",
        "delimiter": ",",
    }
    email_from = _safe_text(settings.unisender_sender_email or settings.smtp_sender_email)
    if email_from:
        request_body["filter"] = {"email_from": email_from}
    try:
        create_data = _unisender_go_json_request("event-dump/create.json", request_body, api_key=api_key)
    except RuntimeError as exc:
        return _safe_text(exc)

    dump_id = _safe_text(create_data.get("dump_id"))
    if not dump_id:
        return "UniSender Go не вернул идентификатор event-dump."

    try:
        files = _wait_unisender_go_dump_files(dump_id, api_key=api_key)
        imported = _import_unisender_go_dump_files(job_id, files, go_items)
    except RuntimeError as exc:
        return _safe_text(exc)
    finally:
        _delete_unisender_go_dump_safely(dump_id, api_key=api_key)

    if imported <= 0:
        return "Новые события UniSender Go пока не найдены. Возможно, провайдер ещё не сформировал доставку/открытия или события уже недоступны."
    return f"Добрали из UniSender Go {imported} событий доставки."


def _event_dump_time_window(items: list[dict[str, Any]]) -> tuple[str, str]:
    sent_times: list[datetime] = []
    for item in items:
        text = _safe_text(item.get("sent_at"))
        if not text:
            continue
        try:
            sent_times.append(datetime.fromisoformat(text))
        except ValueError:
            continue
    now = datetime.now()
    if sent_times:
        start = min(sent_times) - timedelta(days=1)
    else:
        start = now - timedelta(days=7)
    # UniSender хранит историю ограниченно, а серверные часы/таймзона могут плавать.
    floor = now - timedelta(days=45)
    if start < floor:
        start = floor
    end = now + timedelta(minutes=5)
    return start.strftime("%Y-%m-%d %H:%M:%S"), end.strftime("%Y-%m-%d %H:%M:%S")


def _unisender_go_json_request(path: str, payload: dict[str, Any], *, api_key: str) -> dict[str, Any]:
    request = Request(
        _build_unisender_go_url(path),
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        method="POST",
        headers={
            "Content-Type": "application/json",
            "Accept": "application/json",
            "X-API-KEY": api_key,
        },
    )
    raw = _run_unisender_request(request, timeout=60, request_label="UniSender Go")
    try:
        data = json.loads(raw)
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"UniSender Go вернул непонятный ответ event-dump: {raw[:300]}") from exc
    if _safe_text(data.get("status")).lower() != "success":
        raise RuntimeError(_safe_text(data.get("message")) or "UniSender Go не подтвердил event-dump.")
    return data


def _wait_unisender_go_dump_files(dump_id: str, *, api_key: str) -> list[dict[str, Any]]:
    last_status = ""
    files: list[dict[str, Any]] = []
    for _ in range(EVENT_DUMP_POLL_ATTEMPTS):
        data = _unisender_go_json_request("event-dump/get.json", {"dump_id": dump_id}, api_key=api_key)
        event_dump = data.get("event_dump") if isinstance(data.get("event_dump"), dict) else {}
        last_status = _safe_text(event_dump.get("dump_status")).lower()
        files = event_dump.get("files") if isinstance(event_dump.get("files"), list) else []
        if last_status == "ready" and files:
            return [item for item in files if isinstance(item, dict) and _safe_text(item.get("url"))]
        if last_status == "failed":
            raise RuntimeError("UniSender Go не смог подготовить event-dump для аналитики.")
        sleep(EVENT_DUMP_POLL_SECONDS)
    raise RuntimeError(
        f"UniSender Go ещё не подготовил event-dump (текущий статус: {last_status or 'unknown'}). Попробуйте обновить аналитику через минуту."
    )


def _import_unisender_go_dump_files(job_id: str, files: list[dict[str, Any]], items: list[dict[str, Any]]) -> int:
    path = resolve_job_paths(job_id).root_dir / "state" / "unisender_go_events.jsonl"
    path.parent.mkdir(parents=True, exist_ok=True)
    existing_events = load_unisender_go_events(job_id)
    existing_keys = {_unisender_go_event_record_key(item) for item in existing_events if isinstance(item, dict)}
    provider_job_ids = {
        _safe_text(item.get("provider_job_id") or item.get("message_id"))
        for item in items
        if _safe_text(item.get("provider_job_id") or item.get("message_id"))
    }
    recipient_row_pairs = {
        (_safe_text(item.get("recipient")).lower(), _safe_text(item.get("row_id")))
        for item in items
        if _safe_text(item.get("recipient"))
    }
    imported = 0
    with path.open("a", encoding="utf-8") as handle:
        for file_info in files:
            url = _safe_text(file_info.get("url"))
            if not url:
                continue
            for record in _load_unisender_go_dump_records(
                url,
                job_id=job_id,
                provider_job_ids=provider_job_ids,
                recipient_row_pairs=recipient_row_pairs,
            ):
                key = _unisender_go_event_record_key(record)
                if key in existing_keys:
                    continue
                handle.write(json.dumps(record, ensure_ascii=False) + "\n")
                existing_keys.add(key)
                imported += 1
    return imported


def _load_unisender_go_dump_records(
    url: str,
    *,
    job_id: str,
    provider_job_ids: set[str],
    recipient_row_pairs: set[tuple[str, str]],
) -> list[dict[str, Any]]:
    with urlopen(url, timeout=60) as response:
        raw = response.read().decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(raw))
    records: list[dict[str, Any]] = []
    for row in reader:
        metadata = _parse_event_dump_metadata(row.get("metadata"))
        recipient = _safe_text(row.get("email")).lower()
        row_id = _safe_text(metadata.get("app_row_id") or metadata.get("row_id"))
        event_job_id = _safe_text(metadata.get("app_job_id") or metadata.get("job_id"))
        provider_job_id = _safe_text(row.get("job_id"))
        if event_job_id:
            if event_job_id != job_id:
                continue
        elif provider_job_id:
            if provider_job_id not in provider_job_ids:
                continue
        elif (recipient, row_id) not in recipient_row_pairs:
            continue
        event_type = _resolve_event_dump_status(row)
        if not event_type:
            continue
        records.append(
            {
                "received_at": _safe_text(row.get("event_time")) or datetime.now().isoformat(timespec="seconds"),
                "event": {
                    "job_id": _safe_text(row.get("job_id")),
                    "email": recipient,
                    "status": _safe_text(row.get("status")),
                    "delivery_status": _safe_text(row.get("delivery_status")),
                    "metadata": metadata,
                    "url": _safe_text(row.get("url")),
                },
                "event_type": event_type,
                "recipient": recipient,
                "provider_job_id": provider_job_id,
                "row_id": row_id,
                "mun_name": _safe_text(metadata.get("app_mun_name") or metadata.get("mun_name")),
            }
        )
    return records


def _parse_event_dump_metadata(value: Any) -> dict[str, Any]:
    if isinstance(value, dict):
        return value
    text = _safe_text(value)
    if not text:
        return {}
    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        return {}
    return data if isinstance(data, dict) else {}


def _resolve_event_dump_status(row: dict[str, Any]) -> str:
    status = _normalize_provider_status(_safe_text(row.get("status")))
    delivery_status = _normalize_provider_status(_safe_text(row.get("delivery_status")))
    if status in {"soft_bounced", "hard_bounced"} and delivery_status and delivery_status != "unknown":
        return delivery_status
    return status if status != "unknown" else delivery_status


def _unisender_go_event_record_key(record: dict[str, Any]) -> tuple[str, str, str, str, str]:
    event = record.get("event") if isinstance(record.get("event"), dict) else {}
    return (
        _safe_text(record.get("recipient")).lower(),
        _safe_text(record.get("row_id")),
        _normalize_provider_status(_safe_text(record.get("event_type"))),
        _safe_text(record.get("received_at")),
        _safe_text(event.get("url")),
    )


def _delete_unisender_go_dump_safely(dump_id: str, *, api_key: str) -> None:
    if not dump_id:
        return
    try:
        _unisender_go_json_request("event-dump/delete.json", {"dump_id": dump_id}, api_key=api_key)
    except Exception:
        return


def _delivery_status_cache_path(job_id: str | None) -> Path:
    job_paths = resolve_job_paths(job_id)
    return job_paths.root_dir / "state" / ANALYTICS_CACHE_FILENAME


def _load_delivery_status_cache(job_id: str | None) -> dict[str, dict[str, Any]]:
    path = _delivery_status_cache_path(job_id)
    if not path.exists():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError):
        return {}
    statuses = data.get("statuses") if isinstance(data, dict) else {}
    if not isinstance(statuses, dict):
        return {}
    return {
        _safe_text(message_id): payload
        for message_id, payload in statuses.items()
        if _safe_text(message_id) and isinstance(payload, dict)
    }


def _save_delivery_status_cache(job_id: str | None, statuses: dict[str, dict[str, Any]]) -> None:
    path = _delivery_status_cache_path(job_id)
    payload = {
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "statuses": statuses,
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _load_delivery_log_items(job_id: str | None) -> list[dict[str, Any]]:
    job_paths = resolve_job_paths(job_id)
    sent_mail_log_path = None if job_paths.uses_legacy_layout else job_paths.sent_mail_log_path
    items = [
        item
        for item in _load_sent_mail_log_items(sent_mail_log_path)
        if _safe_text(item.get("transport")) in {"unisender", "rusender", "smtp", "mailopost"}
    ]
    items = _filter_items_by_current_sender_state(job_id, items)
    items = _filter_items_by_current_data(job_id, items)
    send_run_id, send_run_started_at = _current_send_scope(job_id)
    items = _filter_items_by_send_scope(
        items,
        send_run_id=send_run_id,
        send_run_started_at=send_run_started_at,
    )
    return _dedupe_latest_log_items(items)


def _filter_items_by_current_sender_state(job_id: str | None, items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not job_id or not items:
        return items
    state = _load_sender_state(job_id)
    if _sender_state_is_materials_followup(job_id, state):
        return items
    if bool(state.get("selection_scoped")):
        return items
    state_rows = state.get("rows")
    if not isinstance(state_rows, list) or not state_rows:
        return items
    total_rows = _safe_int(state.get("total_rows"))
    if total_rows and len(state_rows) < total_rows:
        return items

    sent_pairs = _sender_state_sent_pairs(state_rows)
    if not sent_pairs:
        return items
    return [
        item
        for item in items
        if (_safe_text(item.get("row_id")), _safe_text(item.get("recipient")).lower()) in sent_pairs
    ]


def _sender_state_sent_pairs(state_rows: list[Any]) -> set[tuple[str, str]]:
    pairs: set[tuple[str, str]] = set()
    for entry in state_rows:
        if not isinstance(entry, dict):
            continue
        row_id = _safe_text(entry.get("id") or entry.get("row_id"))
        if not row_id:
            continue
        recipients = entry.get("sent_recipients")
        if not isinstance(recipients, list) or not recipients:
            recipient = _safe_text(entry.get("recipient"))
            recipients = [recipient] if recipient else []
        for recipient in recipients:
            recipient_key = _safe_text(recipient).lower()
            if recipient_key:
                pairs.add((row_id, recipient_key))
    return pairs


def _safe_int(value: Any) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def _filter_items_by_current_data(job_id: str | None, items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not job_id or not items:
        return items
    data_path = resolve_job_paths(job_id).data_xlsx
    if not data_path.exists():
        return items
    try:
        _, _, rows = load_rows(data_path)
    except Exception:
        return items

    current_row_ids: set[str] = set()
    current_row_email_pairs: set[tuple[str, str]] = set()
    for row in rows:
        row_id = _safe_text(row.get("ID"))
        if not row_id:
            continue
        current_row_ids.add(row_id)
        for email in _parse_emails(row.get("EMAIL_OSN")) + _parse_emails(row.get("EMAIL_DOP")):
            if _safe_text(email):
                current_row_email_pairs.add((row_id, _safe_text(email).lower()))
    if not current_row_ids:
        return items
    return [
        item
        for item in items
        if _sent_log_item_matches_current_data(
            item,
            current_row_ids=current_row_ids,
            current_row_email_pairs=current_row_email_pairs,
        )
    ]


def _sent_log_item_matches_current_data(
    item: dict[str, Any],
    *,
    current_row_ids: set[str],
    current_row_email_pairs: set[tuple[str, str]],
) -> bool:
    row_id = _safe_text(item.get("row_id"))
    if not row_id:
        return True
    if row_id not in current_row_ids:
        return False
    if not current_row_email_pairs:
        return True
    recipient = _safe_text(item.get("recipient")).lower()
    if not recipient:
        return True
    return (row_id, recipient) in current_row_email_pairs


def _dedupe_latest_log_items(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    latest: dict[str, dict[str, Any]] = {}
    passthrough: list[dict[str, Any]] = []
    for item in items:
        key = _sent_log_dedupe_key(item)
        if not key:
            passthrough.append(item)
            continue
        latest[key] = item
    return passthrough + list(latest.values())


def _sent_log_dedupe_key(item: dict[str, Any]) -> str:
    row_id = _safe_text(item.get("row_id"))
    recipient = _safe_text(item.get("recipient")).lower()
    if row_id and recipient:
        return f"row_email:{row_id}:{recipient}"
    provider = item.get("provider") if isinstance(item.get("provider"), dict) else {}
    provider_job_id = _safe_text(item.get("provider_job_id") or item.get("message_id") or provider.get("job_id"))
    if provider_job_id:
        return f"provider_job:{provider_job_id}"
    return ""


def _sender_state_is_materials_followup(job_id: str | None, state: dict[str, Any]) -> bool:
    if not job_id or _safe_text(state.get("send_mode")) != "materials":
        return False
    state_rows = state.get("rows")
    state_row_count = len(state_rows) if isinstance(state_rows, list) else 0
    return len(load_consent_records(job_id)) > state_row_count


def _current_send_scope(job_id: str | None) -> tuple[str, str]:
    state = _load_sender_state(job_id)
    if _sender_state_is_materials_followup(job_id, state):
        return "", ""
    send_run_id = _safe_text(state.get("send_run_id"))
    send_run_started_at = _safe_text(state.get("send_run_started_at"))
    if not send_run_started_at and _safe_text(state.get("mode")) == "send":
        send_run_started_at = _safe_text(state.get("started_at"))
    return send_run_id, send_run_started_at


def _filter_items_by_send_scope(
    items: list[dict[str, Any]],
    *,
    send_run_id: str,
    send_run_started_at: str,
) -> list[dict[str, Any]]:
    expected_run_id = _safe_text(send_run_id)
    if expected_run_id and any(_safe_text(item.get("send_run_id")) for item in items):
        return [item for item in items if _safe_text(item.get("send_run_id")) == expected_run_id]

    started_at = _safe_text(send_run_started_at)
    if started_at:
        scoped_items = [
            item
            for item in items
            if not _safe_text(item.get("sent_at")) or _safe_text(item.get("sent_at")) >= started_at
        ]
        if scoped_items:
            return scoped_items
    return items


def _latest_unisender_go_events(job_id: str | None) -> dict[str, dict[str, Any]]:
    latest: dict[str, dict[str, Any]] = {}
    priority = {
        "accepted": 10,
        "sent": 20,
        "delivered": 30,
        "opened": 40,
        "clicked": 50,
        "subscribed": 55,
        "unsubscribed": 60,
        "soft_bounced": 70,
        "spam": 80,
        "hard_bounced": 90,
    }
    for event in load_unisender_go_events(job_id):
        status = _normalize_provider_status(_safe_text(event.get("event_type")))
        checked_at = _safe_text(event.get("received_at"))
        recipient = _safe_text(event.get("recipient")).lower()
        row_id = _safe_text(event.get("row_id"))
        provider_job_id = _safe_text(event.get("provider_job_id"))
        keys = []
        if provider_job_id and recipient:
            keys.append(f"provider_job_email:{provider_job_id}:{recipient}")
        if provider_job_id and not recipient:
            keys.append(f"provider_job:{provider_job_id}")
        if row_id and recipient:
            keys.append(f"row_email:{row_id}:{recipient}")
        if recipient:
            keys.append(f"email:{recipient}")
        for key in keys:
            previous = latest.get(key)
            previous_rank = _unisender_go_status_priority(
                _safe_text(previous.get("provider_status")) if previous else "",
                priority,
            )
            if not previous or _unisender_go_status_priority(status, priority) >= previous_rank:
                latest_event = dict(event)
                latest_event["provider_status"] = status
                latest_event["checked_at"] = checked_at
                latest[key] = latest_event
    return latest


def _latest_mailopost_events(job_id: str | None) -> dict[str, dict[str, Any]]:
    latest: dict[str, dict[str, Any]] = {}
    priority = {
        "queued": 10,
        "sent": 20,
        "delivered": 30,
        "opened": 40,
        "clicked": 50,
        "unsubscribed": 60,
        "soft_bounced": 70,
        "skipped": 75,
        "spam": 80,
        "complained": 80,
        "hard_bounced": 90,
    }
    for event in load_mailopost_events(job_id):
        status = _normalize_provider_status(_safe_text(event.get("provider_status") or event.get("event_type")))
        checked_at = _safe_text(event.get("occurred_at") or event.get("received_at"))
        recipient = _safe_text(event.get("recipient")).lower()
        message_id = _safe_text(event.get("message_id"))
        row_id = _safe_text(event.get("row_id"))
        keys = []
        if message_id and recipient:
            keys.append(f"message_email:{message_id}:{recipient}")
        if message_id:
            keys.append(f"message:{message_id}")
        if row_id and recipient:
            keys.append(f"row_email:{row_id}:{recipient}")
        if recipient:
            keys.append(f"email:{recipient}")
        for key in keys:
            previous = latest.get(key)
            previous_rank = priority.get(_normalize_provider_status(_safe_text(previous.get("provider_status")) if previous else ""), 0)
            if not previous or priority.get(status, 0) >= previous_rank:
                latest_event = dict(event)
                latest_event["provider_status"] = status
                latest_event["checked_at"] = checked_at
                latest[key] = latest_event
    return latest


def _latest_rusender_events(job_id: str | None) -> dict[str, dict[str, Any]]:
    latest: dict[str, dict[str, Any]] = {}
    priority = {
        "accepted": 10,
        "delivered": 30,
        "opened": 40,
        "clicked": 50,
        "unsubscribed": 60,
        "soft_bounced": 70,
        "spam": 80,
        "hard_bounced": 90,
        "err_delivery_failed": 95,
    }
    for event in load_rusender_events(job_id):
        status = _normalize_provider_status(_safe_text(event.get("provider_status") or event.get("event_type")))
        checked_at = _safe_text(event.get("occurred_at") or event.get("received_at"))
        recipient = _safe_text(event.get("recipient")).lower()
        task_id = _safe_text(event.get("task_id"))
        keys = []
        if task_id and recipient:
            keys.append(f"task_email:{task_id}:{recipient}")
        if task_id:
            keys.append(f"task:{task_id}")
        if recipient:
            keys.append(f"email:{recipient}")
        for key in keys:
            previous = latest.get(key)
            previous_rank = priority.get(_normalize_provider_status(_safe_text(previous.get("provider_status")) if previous else ""), 0)
            if not previous or priority.get(status, 0) >= previous_rank:
                latest_event = dict(event)
                latest_event["provider_status"] = status
                latest_event["checked_at"] = checked_at
                latest[key] = latest_event
    return latest


def _unisender_go_status_priority(status: str, priority: dict[str, int]) -> int:
    normalized = _normalize_provider_status(status)
    if normalized in UNISENDER_GO_TECHNICAL_ERROR_STATUSES:
        return 22
    if normalized in UNISENDER_GO_HARD_ERROR_STATUSES or normalized.startswith("err_"):
        return 28
    if normalized in UNISENDER_GO_SOFT_ERROR_STATUSES:
        return 26
    if normalized.startswith("skip_"):
        return 22
    return priority.get(normalized, 0)


def _match_unisender_go_event(item: dict[str, Any], events: dict[str, dict[str, Any]]) -> dict[str, Any] | None:
    provider = item.get("provider") if isinstance(item.get("provider"), dict) else {}
    provider_job_id = _safe_text(item.get("provider_job_id") or item.get("message_id") or provider.get("job_id"))
    recipient = _safe_text(item.get("recipient")).lower()
    row_id = _safe_text(item.get("row_id"))
    if provider_job_id:
        precise_provider_match = events.get(f"provider_job_email:{provider_job_id}:{recipient}") if recipient else None
        if precise_provider_match:
            return precise_provider_match
        provider_match = events.get(f"provider_job:{provider_job_id}") if not recipient else None
        if provider_match:
            return provider_match

    keys = [
        f"row_email:{row_id}:{recipient}",
        f"email:{recipient}",
    ]
    for key in keys:
        if key and key in events:
            return events[key]
    return None


def _match_mailopost_event(item: dict[str, Any], events: dict[str, dict[str, Any]]) -> dict[str, Any] | None:
    if _provider_name(item) != "mailopost":
        return None
    provider = item.get("provider") if isinstance(item.get("provider"), dict) else {}
    message_id = _safe_text(
        item.get("provider_message_id")
        or item.get("message_id")
        or provider.get("message_id")
        or provider.get("id")
    )
    recipient = _safe_text(item.get("recipient")).lower()
    row_id = _safe_text(item.get("row_id"))
    keys = []
    if message_id and recipient:
        keys.append(f"message_email:{message_id}:{recipient}")
    if message_id:
        keys.append(f"message:{message_id}")
    if row_id and recipient:
        keys.append(f"row_email:{row_id}:{recipient}")
    if recipient:
        keys.append(f"email:{recipient}")
    for key in keys:
        if key in events:
            return events[key]
    return None

def _match_rusender_event(item: dict[str, Any], events: dict[str, dict[str, Any]]) -> dict[str, Any] | None:
    if _provider_name(item) != "rusender":
        return None
    provider = item.get("provider") if isinstance(item.get("provider"), dict) else {}
    task_id = _safe_text(
        item.get("provider_message_id")
        or item.get("message_id")
        or item.get("provider_job_id")
        or provider.get("message_id")
        or provider.get("uuid")
    )
    recipient = _safe_text(item.get("recipient")).lower()
    keys = []
    if task_id and recipient:
        keys.append(f"task_email:{task_id}:{recipient}")
    if task_id:
        keys.append(f"task:{task_id}")
    if recipient:
        keys.append(f"email:{recipient}")
    for key in keys:
        if key in events:
            return events[key]
    return None


def _delivery_response_text(*events: dict[str, Any] | None) -> str:
    for event in events:
        response = _extract_delivery_response(event)
        if response:
            return response
    return ""


def _extract_delivery_response(value: Any) -> str:
    if not isinstance(value, dict):
        return ""

    direct_keys = (
        "delivery_response",
        "deliveryResponse",
        "smtp_response",
        "smtpResponse",
        "diagnostic_code",
        "diagnosticCode",
        "diagnostic",
        "reason",
        "error_message",
        "errorMessage",
        "error",
    )
    for key in direct_keys:
        text = _safe_text(value.get(key))
        if text:
            return text

    for key in ("event", "payload", "message", "data", "raw", "response"):
        nested = value.get(key)
        if isinstance(nested, dict):
            text = _extract_delivery_response(nested)
            if text:
                return text

    return ""


def _check_classic_statuses(email_ids: list[str]) -> dict[str, dict[str, Any]]:
    statuses: dict[str, dict[str, Any]] = {}
    unique_ids = [_safe_text(item) for item in dict.fromkeys(email_ids) if _safe_text(item)]
    for index in range(0, len(unique_ids), 500):
        statuses.update(_check_unisender_classic_messages(unique_ids[index : index + 500]))
    return statuses


def _build_consent_rows(job_id: str | None) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for record in load_consent_records(job_id):
        status = _safe_text(record.get("status")) or "pending"
        materials_status = _safe_text(record.get("materials_status"))
        materials_error = _safe_text(record.get("materials_error"))
        rows.append(
            {
                "row_id": _safe_text(record.get("row_id")),
                "mun_name": _safe_text(record.get("mun_name")),
                "recipient": _safe_text(record.get("recipient")),
                "status": status,
                "status_label": _consent_status_label(status),
                "request_sent_at": _safe_text(record.get("request_sent_at")),
                "created_at": _safe_text(record.get("created_at")),
                "confirmed_at": _safe_text(record.get("confirmed_at")),
                "expires_at": _safe_text(record.get("expires_at")),
                "materials_status": materials_status,
                "materials_status_label": _materials_status_label(
                    materials_status,
                    sent_at=_safe_text(record.get("materials_sent_at")),
                    error=materials_error,
                ),
                "materials_sent_at": _safe_text(record.get("materials_sent_at")),
                "materials_error": materials_error,
                "materials_dispatch_summary": _safe_text(record.get("materials_dispatch_summary")),
                "transport": _safe_text(record.get("transport")),
                "attachment_mode": _safe_text(record.get("attachment_mode")),
                "work_type": _safe_text(record.get("work_type")),
                "confirmed_ip": _safe_text(record.get("confirmed_ip")),
                "confirmed_user_agent": _safe_text(record.get("confirmed_user_agent")),
                "consent_document_path": _safe_text(record.get("consent_document_path")),
            }
        )
    return sorted(rows, key=lambda row: (_sort_key(row.get("row_id")), row.get("recipient") or ""))


def _sort_key(value: Any) -> tuple[int, str]:
    text = _safe_text(value)
    try:
        return (0, f"{int(text):012d}")
    except ValueError:
        return (1, text)


def _consent_summary(consent_rows: list[dict[str, Any]]) -> dict[str, int]:
    total = len(consent_rows)
    requested = sum(
        1
        for row in consent_rows
        if row.get("request_sent_at") or row.get("status") in {"request_sent", "confirmed"}
    )
    confirmed = sum(1 for row in consent_rows if row.get("status") == "confirmed")
    expired = sum(1 for row in consent_rows if row.get("status") == "expired")
    materials_sent = sum(
        1
        for row in consent_rows
        if row.get("materials_status") == "sent" or row.get("materials_sent_at")
    )
    materials_error = sum(
        1
        for row in consent_rows
        if row.get("materials_status") == "error" or row.get("materials_error")
    )
    pending = max(0, total - confirmed - expired)
    return {
        "total": total,
        "requested": requested,
        "confirmed": confirmed,
        "pending": pending,
        "expired": expired,
        "materials_sent": materials_sent,
        "materials_error": materials_error,
    }


def _consent_status_label(status: str) -> str:
    labels = {
        "pending": "Ожидает отправки запроса",
        "request_sent": "Запрос согласия отправлен",
        "confirmed": "Согласие получено",
        "expired": "Ссылка истекла",
    }
    return labels.get(_safe_text(status), _safe_text(status) or "Статус неизвестен")


def _materials_status_label(status: str, *, sent_at: str, error: str) -> str:
    normalized = _safe_text(status)
    if normalized == "sent" or sent_at:
        return "Материалы отправлены"
    if normalized == "error" or error:
        return "Ошибка отправки материалов"
    return "Материалы ещё не отправлялись"


def _write_statistics_sheet(
    sheet,
    rows: list[dict[str, Any]],
    *,
    job_id: str | None,
    refresh_error: str,
    consent_rows: list[dict[str, Any]] | None = None,
) -> None:
    _add_title(sheet, "Статистика отправки")
    outcomes = Counter(row["outcome"] for row in rows)
    statuses = Counter(row["provider_status"] or "статус неизвестен" for row in rows)
    checked_count = sum(1 for row in rows if row.get("checked_at"))
    unique_recipients = len({row["recipient"].lower() for row in rows if row["recipient"]})
    unique_municipalities = len({row["mun_name"].lower() for row in rows if row["mun_name"]})

    summary_rows = [
        ["Дата формирования", datetime.now().isoformat(timespec="seconds")],
        ["Job ID", job_id or "текущий/legacy"],
        ["Всего писем", len(rows)],
        ["Уникальных получателей", unique_recipients],
        ["Уникальных МО", unique_municipalities],
        ["Успешно", outcomes.get("Успешно", 0)],
        ["В обработке", outcomes.get("В обработке", 0)],
        ["Предупреждение", outcomes.get("Предупреждение", 0)],
        ["Ошибка", outcomes.get("Ошибка", 0)],
        ["Статус проверен", checked_count],
        ["Комментарий обновления", refresh_error or "Статусы обновлены или обновление не требовалось."],
    ]
    _write_table(sheet, 3, ["Показатель", "Значение"], summary_rows, name="DeliveryStats")

    status_rows = [
        [status, _report_status_label(status), count]
        for status, count in sorted(statuses.items(), key=lambda item: (-item[1], item[0]))
    ]
    _write_table(sheet, 3, ["Код статуса", "Расшифровка", "Количество"], status_rows, name="DeliveryStatusStats", start_column=4)

    consent_summary = _consent_summary(consent_rows or [])
    consent_stats_rows = [
        ["Запросов согласия всего", consent_summary["total"]],
        ["Запросов согласия отправлено", consent_summary["requested"]],
        ["Согласие дали", consent_summary["confirmed"]],
        ["Ожидаем согласие", consent_summary["pending"]],
        ["Ссылок истекло", consent_summary["expired"]],
        ["Материалы отправлены после согласия", consent_summary["materials_sent"]],
        ["Ошибки отправки материалов", consent_summary["materials_error"]],
    ]
    _write_table(sheet, 17, ["Показатель", "Значение"], consent_stats_rows, name="ConsentStats")


def _write_journal_sheet(sheet, rows: list[dict[str, Any]]) -> None:
    _add_title(sheet, "Журнал отправки")
    _write_table(
        sheet,
        3,
        [
            "№ строки",
            "Муниципальное образование",
            "Получатель",
            "Время отправки",
            "Тема письма",
            "Принято провайдером",
            "Код статуса провайдера",
            "Расшифровка статуса",
            "Причина недоставки",
            "Итог",
            "Email ID",
            "Message ID",
            "Время проверки статуса",
            "Комментарий",
        ],
        [
            [
                row["row_id"],
                row["mun_name"],
                row["recipient"],
                row["sent_at"],
                row["subject"],
                row["accepted_status"],
                row["provider_status"],
                row["provider_status_label"],
                row.get("delivery_response", ""),
                row["outcome"],
                row["email_id"],
                row["message_id"],
                row["checked_at"],
                row["comment"],
            ]
            for row in rows
        ],
        name="UnisenderDeliveryLog",
    )


def _write_consent_sheet(sheet, rows: list[dict[str, Any]]) -> None:
    _add_title(sheet, "Согласия")
    _write_table(
        sheet,
        3,
        [
            "№ строки",
            "Муниципальное образование",
            "Получатель",
            "Статус согласия",
            "Запрос отправлен",
            "Согласие получено",
            "Истекает",
            "Статус материалов",
            "Материалы отправлены",
            "Ошибка материалов",
            "Транспорт",
            "Что отправляем",
            "IP подтверждения",
            "User-Agent подтверждения",
        ],
        [
            [
                row["row_id"],
                row["mun_name"],
                row["recipient"],
                row["status_label"],
                row["request_sent_at"],
                row["confirmed_at"],
                row["expires_at"],
                row["materials_status_label"],
                row["materials_sent_at"],
                row["materials_error"],
                row["transport"],
                row["attachment_mode"],
                row["confirmed_ip"],
                row["confirmed_user_agent"],
            ]
            for row in rows
        ],
        name="ConsentLog",
    )


def _provider_name(item: dict[str, Any]) -> str:
    provider = item.get("provider") if isinstance(item.get("provider"), dict) else {}
    return _safe_text(provider.get("provider")) or _safe_text(item.get("transport")) or "unknown"


def _message_id(item: dict[str, Any]) -> str:
    provider = item.get("provider") if isinstance(item.get("provider"), dict) else {}
    return _safe_text(item.get("provider_message_id") or provider.get("message_id") or provider.get("uuid"))


def _comment_text(item: dict[str, Any], refresh_error: str) -> str:
    warning = _safe_text(item.get("warning"))
    if warning and refresh_error:
        return f"{warning} Статус не обновлен: {refresh_error}"
    if warning:
        return warning
    if refresh_error:
        return f"Статус не обновлен: {refresh_error}"
    return ""


def _delivery_outcome(status: str) -> str:
    normalized = _normalize_provider_status(status)
    if normalized in {
        "delivered",
        "opened",
        "clicked",
        "subscribed",
        "ok_delivered",
        "ok_read",
        "ok_link_visited",
    }:
        return "Успешно"
    if normalized in {"unsubscribed", "spam", "ok_unsubscribed", "ok_spam_folder"}:
        return "Предупреждение"
    if normalized in UNISENDER_GO_SOFT_ERROR_STATUSES:
        return "Временная ошибка"
    if normalized in UNISENDER_GO_HARD_ERROR_STATUSES or normalized.startswith("err_") or normalized.startswith("skip_"):
        return "Ошибка"
    return "В обработке"


def _report_status_label(status: str) -> str:
    normalized = _normalize_provider_status(status)
    overrides = {
        "accepted": "Принято, ожидает отправки",
        "sent": "Отправлено, ждём доставку",
        "success": "Принято UniSender Go",
        "delivered": "Доставлено",
        "opened": "Открыто",
        "clicked": "Переход по ссылке",
        "unsubscribed": "Получатель отписался",
        "subscribed": "Получатель снова подписался",
        "soft_bounced": "Временная недоставка",
        "hard_bounced": "Не доставлено",
        "spam": "Жалоба на спам",
        "err_user_unknown": "Ошибка: адрес не существует",
        "err_user_inactive": "Ошибка: ящик неактивен",
        "err_mailbox_full": "Ошибка: ящик переполнен",
        "err_spam_rejected": "Ошибка: отклонено как спам",
        "err_spam_skipped": "Ошибка: пропущено как спам",
        "err_delivery_failed": "Ошибка доставки",
        "err_will_retry": "Временная ошибка, UniSender повторит",
        "err_lost": "Статус потерян, нужна проверка вручную",
        "skip_dup_unreachable": "Ошибка: адрес уже недоступен",
        "skip_dup_temp_unreachable": "Временная ошибка: адрес временно недоступен",
    }
    label = overrides.get(normalized) or _unisender_status_label(normalized)
    return label[:1].upper() + label[1:] if label else "Статус неизвестен"


def _normalize_provider_status(status: str) -> str:
    normalized = _safe_text(status).strip().lower().replace("-", "_").replace(" ", "_")
    return normalized or "unknown"


def _provider_label(provider: str) -> str:
    normalized = _safe_text(provider).strip().lower()
    return PROVIDER_LABELS.get(normalized) or (_safe_text(provider) or "провайдер")


def _primary_provider(providers: Counter[str]) -> tuple[str, str]:
    if not providers:
        return "", ""
    provider, _ = providers.most_common(1)[0]
    return provider, _provider_label(provider)


def _analytics_note(*, total: int, checked: int, providers: Counter[str], refresh_error: str) -> str:
    if total <= 0:
        return "Статистика появится после реальной отправки писем."
    provider_names = ", ".join(_provider_label(name) for name, count in providers.most_common() if count)
    has_unisender = any("unisender" in name for name in providers)
    has_rusender = any("rusender" in name for name in providers)
    has_mailopost = any("mailopost" in name for name in providers)
    if refresh_error:
        return f"Показаны локальные данные отправки. Не удалось обновить часть статусов: {refresh_error}"
    if has_rusender and not has_unisender:
        return (
            "Для RuSender доставка, открытия и клики приходят через webhook. "
            "Если в кабинете RuSender события уже есть, а здесь пока нули, проверьте URL webhook и токен."
        )
    if has_unisender or has_mailopost:
        return (
            "Доставка, открытия и клики появляются после событий провайдера. "
            "Если письма уже переданы, а здесь пока нули, значит сервис ещё ждёт webhook или добор событий."
        )
    if checked:
        return f"Статусы обновлены для {checked} писем. Провайдеры: {provider_names or 'неизвестно'}."
    return (
        "Показаны локальные данные отправки. Для этого провайдера пока нет подключённых событий "
        f"доставки, открытий и переходов. Провайдеры: {provider_names or 'неизвестно'}."
    )


def _add_title(sheet, title: str) -> None:
    sheet["A1"] = title
    sheet["A1"].font = Font(bold=True, size=16, color=BLUE)


def _write_table(
    sheet,
    start_row: int,
    headers: list[str],
    rows: list[list[Any]],
    *,
    name: str,
    start_column: int = 1,
) -> None:
    for column_offset, header in enumerate(headers):
        cell = sheet.cell(start_row, start_column + column_offset, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index, row in enumerate(rows, start_row + 1):
        for column_offset, value in enumerate(row):
            cell = sheet.cell(row_index, start_column + column_offset, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    max_row = start_row + max(len(rows), 1)
    max_column = start_column + len(headers) - 1
    for row in sheet.iter_rows(min_row=start_row, max_row=max_row, min_col=start_column, max_col=max_column):
        for cell in row:
            cell.border = Border(bottom=BORDER)

    if rows:
        ref = f"{get_column_letter(start_column)}{start_row}:{get_column_letter(max_column)}{max_row}"
        table = Table(displayName=name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )
        sheet.add_table(table)


def _style_workbook(workbook: Workbook) -> None:
    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A4"
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for cell in sheet[3]:
            if cell.value:
                cell.fill = PatternFill("solid", fgColor=BLUE)
                cell.font = Font(bold=True, color="FFFFFF")
    workbook["Статистика"]["A1"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)


def _autosize(sheet, widths: dict[str, int]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
