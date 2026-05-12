from pathlib import Path
from typing import Optional

from openpyxl import load_workbook


HEADER_ROW = 2
STATUS_HEADER_ALIASES = ("STATUS", "Статус отправки")


def load_rows(xlsx_path: Path, sheet_name: Optional[str] = None) -> tuple[object, object, list[dict]]:
    workbook = load_workbook(xlsx_path)
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]

    headers = [
        worksheet.cell(row=HEADER_ROW, column=column_index).value
        for column_index in range(1, worksheet.max_column + 1)
    ]

    rows: list[dict] = []
    for row_index in range(HEADER_ROW + 1, worksheet.max_row + 1):
        row = {}
        is_empty = True
        for column_index, header in enumerate(headers, start=1):
            if not header:
                continue
            value = worksheet.cell(row=row_index, column=column_index).value
            row[header] = value
            if value not in (None, ""):
                is_empty = False
        if not is_empty:
            row["_row_index"] = row_index
            rows.append(row)

    return workbook, worksheet, rows


def update_status(worksheet, row_index: int, status_value: str) -> None:
    header_map = {
        worksheet.cell(row=HEADER_ROW, column=column_index).value: column_index
        for column_index in range(1, worksheet.max_column + 1)
    }
    status_column = None
    for header in STATUS_HEADER_ALIASES:
        if header in header_map:
            status_column = header_map[header]
            break
    if status_column is None:
        raise KeyError("STATUS")
    worksheet.cell(row=row_index, column=status_column).value = status_value


def save_workbook(workbook, xlsx_path: Path) -> None:
    workbook.save(xlsx_path)
