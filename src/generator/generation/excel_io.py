import os
import tempfile
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook


HEADER_ROW = 2
STATUS_HEADER_ALIASES = ("STATUS", "Статус отправки")


def _is_service_row(row: dict) -> bool:
    first_value = ""
    for value in row.values():
        if value not in (None, ""):
            first_value = str(value).strip()
            break
    return first_value.lower().startswith("источники:")


def _apply_header_aliases(row: dict) -> None:
    if "REQUISITES_OKTNO" not in row and "REQUISITES_OKTMO" in row:
        row["REQUISITES_OKTNO"] = row.get("REQUISITES_OKTMO")
    if "REQUISITES_OKTMO" not in row and "REQUISITES_OKTNO" in row:
        row["REQUISITES_OKTMO"] = row.get("REQUISITES_OKTNO")


def load_rows(xlsx_path: Path, sheet_name: Optional[str] = None) -> tuple[object, object, list[dict]]:
    workbook = load_workbook(xlsx_path)
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]

    headers = [
        worksheet.cell(row=HEADER_ROW, column=column_index).value
        for column_index in range(1, worksheet.max_column + 1)
    ]

    rows: list[dict] = []
    for row_index in range(HEADER_ROW + 1, worksheet.max_row + 1):
        row = {}
        is_empty = True
        for column_index, header in enumerate(headers, start=1):
            if not header:
                continue
            value = worksheet.cell(row=row_index, column=column_index).value
            row[header] = value
            if value not in (None, ""):
                is_empty = False
        if not is_empty and not _is_service_row(row):
            _apply_header_aliases(row)
            row["_row_index"] = row_index
            rows.append(row)

    return workbook, worksheet, rows


def update_status(worksheet, row_index: int, status_value: str) -> None:
    header_map = {
        worksheet.cell(row=HEADER_ROW, column=column_index).value: column_index
        for column_index in range(1, worksheet.max_column + 1)
    }
    status_column = None
    for header in STATUS_HEADER_ALIASES:
        if header in header_map:
            status_column = header_map[header]
            break
    if status_column is None:
        status_column = worksheet.max_column + 1
        worksheet.cell(row=HEADER_ROW, column=status_column).value = STATUS_HEADER_ALIASES[0]
    worksheet.cell(row=row_index, column=status_column).value = status_value


def save_workbook(workbook, xlsx_path: Path) -> None:
    xlsx_path = Path(xlsx_path)
    fd, tmp_name = tempfile.mkstemp(dir=str(xlsx_path.parent), suffix=".xlsxtmp")
    os.close(fd)
    tmp_path = Path(tmp_name)
    try:
        workbook.save(tmp_path)
        os.replace(tmp_path, xlsx_path)
    except BaseException:
        tmp_path.unlink(missing_ok=True)
        raise
