"""
batch_processor.py — детерминированный скрипт для заполнения данных по МО.

Работает БЕЗ LLM в основном цикле:
  1. Читает Excel — находит незаполненные строки
  2. Для каждой строки: Yandex Search → rusprofile → ИНН → Checko → данные
  3. Дополняет только пустые поля
  4. Сохраняет прогресс каждые N строк

Запуск из консоли:
  python batch_processor.py --file путь.xlsx
  python batch_processor.py --file путь.xlsx --max-cycles 3

Запуск из FastAPI:
  from batch_processor import run
  result = run(file_path, output_dir=str(job_output_dir))
"""
from __future__ import annotations

import argparse
import re
import sys
import time
import random
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import httpx
from bs4 import BeautifulSoup
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))

# Импорты адаптируются к контексту запуска
try:
    from src.parser_new import config
    from src.parser_new.logger import logger
    from src.parser_new.tools.regions import get_region_code
except ImportError:
    import config
    from logger import logger
    from tools.regions import get_region_code

try:
    from src.parser_new.progress import emit as _emit
except Exception:
    def _emit(*a, **k):
        pass

def _label() -> str:
    """Как называть обрабатываемые строки в сообщениях пользователю.

    Тип («городские округа», «районы», «сельские поселения») разбирается из
    исходной фразы в oktmo_tool. Здесь он только подставляется в текст, чтобы
    интерфейс говорил словами пользователя, а не обобщённым «МО».
    """
    try:
        try:
            from src.parser_new.tools.oktmo_tool import mo_label
        except ImportError:
            from tools.oktmo_tool import mo_label
        return mo_label()
    except Exception:
        return "МО"


# ==============================
# КОНСТАНТЫ
# ==============================

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "ru-RU,ru;q=0.9",
}

COL = {
    "ID": 1, "SUB_RF": 2, "MUN_R_NAME": 3, "MUN_NAME": 4,
    "ADM_NAME": 5, "ADRES": 6, "HEAD_FIO": 7, "POPULATION": 8,
    "EMAIL_OSN": 9, "EMAIL_DOP": 10, "TEL_OSN": 11, "TEL_DOP": 12,
    "REQUISITES_INN": 13, "REQUISITES_KPP": 14, "REQUISITES_OGRN": 15,
    "REQUISITES_OKPO": 16, "REQUISITES_OKTMO": 17, "STATUS": 18, "NOTE": 19,
}

DATA_START_ROW = 3


# ==============================
# СТРУКТУРА ДАННЫХ
# ==============================

@dataclass
class MORecord:
    excel_row: int
    sub_rf: str = ""
    mun_r_name: str = ""
    mun_name: str = ""
    adm_name: str = ""
    adres: str = ""
    head_fio: str = ""
    population: str = ""
    email_osn: str = ""
    email_dop: str = ""
    tel_osn: str = ""
    tel_dop: str = ""
    inn: str = ""
    kpp: str = ""
    ogrn: str = ""
    okpo: str = ""
    oktmo: str = ""
    status: str = ""
    note: str = ""

    def missing_fields(self) -> list[str]:
        checks = {
            "ADM_NAME": self.adm_name, "ADRES": self.adres,
            "HEAD_FIO": self.head_fio, "EMAIL_OSN": self.email_osn,
            "TEL_OSN": self.tel_osn, "REQUISITES_INN": self.inn,
            "REQUISITES_KPP": self.kpp, "REQUISITES_OGRN": self.ogrn,
            "REQUISITES_OKPO": self.okpo, "REQUISITES_OKTMO": self.oktmo,
        }
        return [k for k, v in checks.items() if not v]

    def is_complete(self) -> bool:
        return len(self.missing_fields()) == 0

    def needs_processing(self) -> bool:
        has_anchor = bool(
            self.mun_name or self.mun_r_name or self.adm_name
            or self.inn or self.email_osn
        )
        return has_anchor and not self.is_complete()
    
    @property
    def search_name(self) -> str:
        """Название для поиска: МО (D) если есть, иначе район (C).
        Для МО возвращает прежнее значение — старое поведение не меняется."""
        return self.mun_name or self.mun_r_name

    @property
    def is_district(self) -> bool:
        return not self.mun_name and bool(self.mun_r_name)


# ==============================
# ЧТЕНИЕ / ЗАПИСЬ EXCEL
# ==============================

def read_excel(file_path: str) -> tuple[object, list[MORecord]]:
    wb = load_workbook(file_path)
    ws = wb.active
    records = []
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        def v(col_name: str) -> str:
            val = ws.cell(row_idx, COL[col_name]).value
            return str(val).strip() if val else ""
        rec = MORecord(
            excel_row=row_idx, sub_rf=v("SUB_RF"), mun_r_name=v("MUN_R_NAME"),
            mun_name=v("MUN_NAME"), adm_name=v("ADM_NAME"), adres=v("ADRES"),
            head_fio=v("HEAD_FIO"), population=v("POPULATION"),
            email_osn=v("EMAIL_OSN"), email_dop=v("EMAIL_DOP"),
            tel_osn=v("TEL_OSN"), tel_dop=v("TEL_DOP"),
            inn=v("REQUISITES_INN"), kpp=v("REQUISITES_KPP"),
            ogrn=v("REQUISITES_OGRN"), okpo=v("REQUISITES_OKPO"),
            oktmo=v("REQUISITES_OKTMO"), status=v("STATUS"), note=v("NOTE"),
        )
        if rec.needs_processing():
            records.append(rec)
    logger.info(f"Найдено строк для обработки: {len(records)}")
    return wb, records


def write_record(ws, rec: MORecord) -> None:
    updates = {
        "ADM_NAME": rec.adm_name, "ADRES": rec.adres, "HEAD_FIO": rec.head_fio,
        "EMAIL_OSN": rec.email_osn, "EMAIL_DOP": rec.email_dop,
        "TEL_OSN": rec.tel_osn, "TEL_DOP": rec.tel_dop,
        "REQUISITES_INN": rec.inn, "REQUISITES_KPP": rec.kpp,
        "REQUISITES_OGRN": rec.ogrn, "REQUISITES_OKPO": rec.okpo,
        "REQUISITES_OKTMO": rec.oktmo, "NOTE": rec.note,
    }
    for col_name, value in updates.items():
        if value:
            existing = ws.cell(rec.excel_row, COL[col_name]).value
            if not existing:
                ws.cell(rec.excel_row, COL[col_name], value)


# ==============================
# УТИЛИТЫ ПОИСКА
# ==============================

def _is_valid_admin(title: str) -> bool:
    t = title.lower()
    bad = [
        "потребительск", "общество", "колхоз", "совхоз", "культур", "досуг",
        "библиотек", "музей", "школ", "больниц", "спорт", "казначейств",
        "налогов", "пенсион", "соцзащит", "мфц", "водоканал", "жкх",
    ]
    if any(k in t for k in bad):
        return False
    good = [
        "администрац", "сельсовет", "поселени", "мэри", "управ",
        "исполком", "исполнител", "комитет", "округ", "сельского",
    ]
    return any(k in t for k in good)


def _extract_key_words(text: str) -> list[str]:
    text = text.lower().replace("ё", "е").replace("-", " ")
    stop = [
        "сельское поселение", "городское поселение", "муниципальный район",
        "муниципальное образование", "сельсовет", "посёлок", "поселок",
        "село", "город", "район", "поселение", "администрация", "мо",
        "республика", "край", "область", "округ",
    ]
    for sw in stop:
        # \b — граница слова: вырезаем "село" как слово, но не из "веселовский"
        text = re.sub(rf"\b{sw}\b", " ", text)
    return [w.strip() for w in text.split() if len(w.strip()) > 3]

def _check_mun_match(combined: str, mun_keywords: list[str]) -> bool:
    if not mun_keywords:
        return True
    combined = combined.replace("ё", "е")
    for kw in mun_keywords:
        kw = kw.replace("ё", "е")
        root = kw[:max(4, len(kw) - 4)]
        if root in combined:
            return True
    return False


def _extract_inn_from_text(text: str) -> str | None:
    """Извлекает ИНН организации (10 цифр), игнорируя ИНН людей (12 цифр)."""
    for pat in [
        r"[Оо]рганизации присвоен ИНН\s*(\d{10})\b",
        r"[Оо]рганизации[^.]*?ИНН[\s/:]+(\d{10})\b",
        r"присвоен ИНН\s*(\d{10})\b",
        r"ИНН/КПП[\s\n:]+(\d{10})\b",
        r"ИНН:\s*(\d{10})\b",
    ]:
        m = re.search(pat, text)
        if m:
            candidate = m.group(1)
            pos = m.end()
            if pos < len(text) and text[pos:pos + 2].isdigit():
                continue
            return candidate
    clean_text = re.sub(r"\b\d{12}\b", "XXXXXXXXXXXX", text)
    m = re.search(r"ИНН[\s/:]+(\d{10})\b", clean_text)
    if m:
        candidate = m.group(1)
        pos = m.end()
        if pos < len(clean_text) and clean_text[pos:pos + 2].isdigit():
            return None
        return candidate
    return None

def _is_liquidated_status(status: str) -> bool:
    """Ликвидирована / недействующая — по статусу из Checko."""
    s = (status or "").lower()
    return "ликвид" in s or "не действ" in s or "прекращ" in s

def _check_liquidated(text: str) -> bool:
    t = text.lower()
    return bool(
        re.search(r"статус[:\s]+ликвидирован", t)
        or re.search(r"организация ликвидирована", t)
        or re.search(r"ликвидирована с \d{2}\.\d{2}\.\d{4}", t)
    )


def validate_org_matches(org_name: str, rec: MORecord) -> bool:
    if not org_name:
        return False
    mun_keywords = _extract_key_words(rec.search_name)
    return _check_mun_match(org_name.lower(), mun_keywords)


def _parse_rusprofile_page(soup) -> dict:
    """Извлекает название, статус и ИНН со страницы rusprofile."""
    result = {"name": "", "liquidated": False, "inn": None}

    name_el = soup.select_one("[itemprop='legalName'] .copy_target")
    result["name"] = name_el.get_text(strip=True) if name_el else ""

    status_el = soup.select_one(".warning-text")
    if status_el:
        st = status_el.get_text().lower()
        if "ликвид" in st or "не действ" in st:
            result["liquidated"] = True
            return result

    page_text = soup.get_text()
    result["inn"] = _extract_inn_from_text(page_text)
    return result

def _rusprofile_clean_name(url: str) -> str:
    """Открывает карточку rusprofile и возвращает чистое название организации."""
    try:
        time.sleep(2)
        resp = httpx.get(url, headers=HEADERS, timeout=10, follow_redirects=True)
        return _parse_rusprofile_page(BeautifulSoup(resp.text, "lxml")).get("name", "")
    except Exception as e:
        logger.debug(f"[yandex] не удалось снять имя с {url[:50]}: {e}")
        return ""


# ==============================
# YANDEX SEARCH API
# ==============================

_yandex_sdk = None
_yandex_search = None


def _get_yandex_search():
    global _yandex_sdk, _yandex_search
    if _yandex_search is None:
        from yandex_ai_studio_sdk import AIStudio
        _yandex_sdk = AIStudio(
            folder_id=config.YANDEX_FOLDER_ID,
            auth=config.YANDEX_API_KEY,
        )
        _yandex_search = _yandex_sdk.search_api.web("RU", groups_on_page=5)
    return _yandex_search


def _parse_yandex_xml(xml_bytes: bytes) -> list[dict]:
    """Парсит XML-ответ Яндекса. Возвращает [{url, domain, title, snippet}].
    Берёт именно <url> (не <saved-copy-url>), устойчиво к порядку тегов."""
    if isinstance(xml_bytes, bytes):
        xml_text = xml_bytes.decode("utf-8", errors="ignore")
    else:
        xml_text = str(xml_bytes or "")

    results = []
    for doc in re.findall(r"<doc\b[^>]*>(.*?)</doc>", xml_text, re.DOTALL):
        # настоящий URL: тег <url>...</url>, НО не часть <saved-copy-url>
        url_m = re.search(r"(?<![\w-])<url>\s*(.*?)\s*</url>", doc, re.DOTALL)
        if not url_m:
            continue
        url = re.sub(r"<[^>]+>", "", url_m.group(1)).strip()

        domain_m = re.search(r"<domain>\s*(.*?)\s*</domain>", doc, re.DOTALL)
        title_m = re.search(r"<title>(.*?)</title>", doc, re.DOTALL)
        pass_m = re.search(r"<passages>(.*?)</passages>", doc, re.DOTALL)

        def _clean(s: str) -> str:
            s = re.sub(r"<[^>]+>", "", s)          # снимаем <hlword> и пр.
            s = (s.replace("&quot;", '"').replace("&amp;", "&")
                   .replace("&lt;", "<").replace("&gt;", ">").replace("&nbsp;", " "))
            return s.strip()

        results.append({
            "url": url,
            "domain": _clean(domain_m.group(1)) if domain_m else "",
            "title": _clean(title_m.group(1)) if title_m else "",
            "snippet": _clean(pass_m.group(1)) if pass_m else "",
        })
    return results

def _search_yandex(rec: MORecord) -> tuple[str | None, bool, str]:
    """Ищет администрацию МО через Yandex Search API → rusprofile.
    Возвращает (inn, liquidated, name). name заполняется прежде всего для
    ликвидированных — чтобы записать настоящее название, а не слово."""
    if not config.YANDEX_API_KEY or not config.YANDEX_FOLDER_ID:
        return None, False, ""

    try:
        search = _get_yandex_search()
        query = f"site:rusprofile.ru {rec.sub_rf} {rec.mun_r_name} {rec.mun_name} администрация"
        logger.debug(f"[yandex] Запрос: {query}")

        time.sleep(0.5)
        xml_result = search.run(query, format="xml", page=0)
        parsed = _parse_yandex_xml(xml_result)

        logger.debug(f"[yandex] Результатов из XML: {len(parsed)}")
        for i, r in enumerate(parsed):
            logger.debug(f"[yandex]   {i+1}. {r['url'][:60]}  title={r['title'][:50]}")

        mun_keywords = _extract_key_words(rec.search_name)

        # Способ 1: ищем по title/snippet из XML
        candidates = []
        for r in parsed:
            if "rusprofile.ru" not in r["url"] or "/id/" not in r["url"]:
                continue
            combined = (r["title"] + " " + r["snippet"]).lower()
            if not _is_valid_admin(r["title"]) and not _is_valid_admin(r["snippet"]):
                continue
            if not _check_mun_match(combined, mun_keywords):
                continue
            candidates.append(r)

        if candidates:
            prio = [c for c in candidates if any(k in c["title"].lower() for k in ["администрац", "исполнител", "комитет"])]
            best = (prio + [c for c in candidates if c not in prio])[0]
            full_text = best["title"] + " " + best["snippet"]

            if _check_liquidated(full_text):
                # имя из сниппета грязное — открываем страницу за чистым названием
                name = _rusprofile_clean_name(best["url"])
                logger.info(f"[yandex] Ликвидирована: {name or best['title'][:60]}")
                return None, True, name

            inn = _extract_inn_from_text(full_text)
            if inn:
                logger.info(f"[yandex] ИНН из XML: {inn} ({best['title'][:50]})")
                return inn, False, ""

            logger.debug(f"[yandex] ИНН не в XML, открываю: {best['url']}")
            time.sleep(2)
            resp = httpx.get(best["url"], headers=HEADERS, timeout=10, follow_redirects=True)
            page = _parse_rusprofile_page(BeautifulSoup(resp.text, "lxml"))
            if page["liquidated"]:
                logger.info(f"[yandex] Ликвидирована: {page['name'][:60]}")
                return None, True, page["name"]
            if page["inn"]:
                logger.info(f"[yandex] ИНН со страницы: {page['inn']} ({page['name'][:50]})")
                return page["inn"], False, page["name"]
            return None, False, ""

        # Способ 2: XML не дал title — заходим на первую /id/ ссылку
        rusprofile_urls = [r["url"] for r in parsed if "rusprofile.ru" in r["url"] and "/id/" in r["url"]]
        if not rusprofile_urls:
            raw_urls = re.findall(r"<url>(.*?)</url>", xml_result.decode("utf-8", errors="ignore"))
            rusprofile_urls = [u for u in raw_urls if "rusprofile.ru" in u and "/id/" in u]

        if rusprofile_urls:
            url = rusprofile_urls[0]
            logger.debug(f"[yandex] Пробуем первую ссылку: {url}")
            time.sleep(2)
            try:
                resp = httpx.get(url, headers=HEADERS, timeout=10, follow_redirects=True)
                page = _parse_rusprofile_page(BeautifulSoup(resp.text, "lxml"))
                logger.debug(f"[yandex] Со страницы: name={page['name'][:60]}, liq={page['liquidated']}, inn={page['inn']}")

                if page["name"] and _is_valid_admin(page["name"]) and _check_mun_match(page["name"].lower(), mun_keywords):
                    if page["liquidated"]:
                        logger.info(f"[yandex] Ликвидирована: {page['name'][:60]}")
                        return None, True, page["name"]
                    if page["inn"]:
                        logger.info(f"[yandex] ИНН найден: {page['inn']} ({page['name'][:50]})")
                        return page["inn"], False, page["name"]
                    logger.warning(f"[yandex] Организация найдена но ИНН не извлечён: {page['name'][:60]}")
            except Exception as e:
                logger.debug(f"[yandex] Ошибка при заходе на страницу: {e}")

        logger.debug(f"[yandex] Подходящих результатов не найдено для: {rec.mun_name}")
        return None, False, ""

    except Exception as e:
        logger.warning(f"[yandex] Ошибка для {rec.mun_name}: {e}")
        return None, False, ""


def search_rusprofile(rec: MORecord) -> tuple[str | None, bool, str]:
    """Ищет администрацию МО через Yandex Search API. Возвращает (inn, liquidated, name)."""
    return _search_yandex(rec)


# ==============================
# CHECKO API
# ==============================

_checko_cache: dict[str, dict] = {}


def fetch_checko_by_inn(inn: str) -> dict | None:
    if inn in _checko_cache:
        return _checko_cache[inn]
    if not config.CHECKO_API_KEY:
        return None
    try:
        resp = httpx.get(
            "https://api.checko.ru/v2/company",
            params={"key": config.CHECKO_API_KEY, "inn": inn},
            timeout=15,
        )
        data = resp.json()
        if data.get("meta", {}).get("status") == "error":
            return None
        org = data.get("data", {})
        if not org:
            return None

        contacts = org.get("Контакты", {}) or {}
        phones = contacts.get("Тел", []) or []
        emails = contacts.get("Емэйл", []) or []
        oktmo = org.get("ОКТМО", {}) or {}
        adres = org.get("ЮрАдрес", "")
        if isinstance(adres, dict):
            adres = adres.get("АдресРФ", "") or str(adres)

        # статус: в company-ответе Checko это бывает и строкой, и объектом
        status_raw = org.get("Статус", "")
        if isinstance(status_raw, dict):
            status_str = status_raw.get("Наим") or status_raw.get("Код") or ""
        else:
            status_str = status_raw or ""
        
        # блок "Ликвид" — определённый признак ликвидации, даже если статус странный
        if isinstance(org.get("Ликвид"), dict) and org["Ликвид"]:
            status_str = (status_str + " ликвидирована").strip()

        head_fio = ""
        for r in (org.get("Руковод", []) or []):
            if not r.get("Недост") and not r.get("ДисквЛицо"):
                head_fio = r.get("ФИО", "")
                break

        result = {
            "adm_name": org.get("НаимПолн", ""),
            "adres": adres,
            "head_fio": head_fio,
            "email_osn": emails[0] if emails else "",
            "email_dop": ", ".join(emails[1:]) if len(emails) > 1 else "",
            "tel_osn": phones[0] if phones else "",
            "tel_dop": ", ".join(phones[1:]) if len(phones) > 1 else "",
            "inn": org.get("ИНН", inn),
            "kpp": org.get("КПП", ""),
            "ogrn": org.get("ОГРН", ""),
            "okpo": org.get("ОКПО", ""),
            "oktmo": oktmo.get("Код", ""),
            "status": status_str,
        }
        logger.info(f"[checko] ИНН {inn} → {result['adm_name'][:60]} | статус: {status_str or '—'}")
        _checko_cache[inn] = result
        return result
    except Exception as e:
        logger.warning(f"[checko] Ошибка для ИНН {inn}: {e}")
        return None


def search_checko_by_name(rec: MORecord) -> str | None:
    region_code = get_region_code(rec.sub_rf)
    if not region_code:
        return None
    try:
        resp = httpx.get(
            "https://api.checko.ru/v2/search",
            params={
                "key": config.CHECKO_API_KEY, "by": "name",
                "obj": "org", "query": rec.search_name, "region": region_code,
            },
            timeout=15,
        )
        data = resp.json()
        records = data.get("data", {}).get("Записи", []) or []
        admin_kw = ["администрац", "сельсовет", "поселени", "исполнител", "комитет"]
        district_key = rec.mun_r_name.lower().replace("муниципальный район", "").strip()
        district_words = [w for w in district_key.split() if len(w) > 3]

        for r in records:
            status = (r.get("Статус", "") or "").lower()
            if "не действует" in status or "ликвид" in status:
                continue
            name = (r.get("НаимПолн", "") or r.get("НаимСокр", "") or "").lower()
            if not any(k in name for k in admin_kw):
                continue
            if district_words and not any(w in name for w in district_words):
                continue
            return r.get("ИНН")
        return None
    except Exception as e:
        logger.warning(f"[checko/search] Ошибка для {rec.mun_name}: {e}")
        return None


# ==============================
# ПОИСК КОНТАКТОВ (через Yandex вместо Tavily)
# ==============================

def search_contacts_online(adm_name: str) -> dict:
    """Ищет контакты администрации на официальном сайте через Yandex."""
    if not config.YANDEX_API_KEY or not config.YANDEX_FOLDER_ID or not adm_name:
        return {}
    try:
        search = _get_yandex_search()
        query = f"{adm_name} официальный сайт контакты"
        xml_result = search.run(query, format="xml", page=0)
        parsed = _parse_yandex_xml(xml_result)

        # Фильтруем агрегаторы
        skip = ["rusprofile.ru", "egrul.ru", "checko.ru", "sbis.ru", "list-org.com"]
        official = [r for r in parsed if not any(d in r["url"] for d in skip)]
        if not official:
            return {}

        url = official[0]["url"]
        time.sleep(1)
        resp = httpx.get(url, headers=HEADERS, timeout=10, follow_redirects=True)
        text = BeautifulSoup(resp.text, "lxml").get_text()

        phones = re.findall(
            r"(?:\+7|8)[\s\-]?\(?\d{3}\)?[\s\-]?\d{3}[\s\-]?\d{2}[\s\-]?\d{2}", text
        )
        emails = re.findall(
            r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", text
        )
        return {
            "tel_osn": phones[0] if phones else "",
            "tel_dop": phones[1] if len(phones) > 1 else "",
            "email_osn": emails[0] if emails else "",
            "email_dop": ", ".join(emails[1:3]) if len(emails) > 1 else "",
        }
    except Exception:
        return {}


# ==============================
# ОБРАБОТКА ОДНОЙ СТРОКИ
# ==============================

def process_record(rec: MORecord) -> MORecord:
    inn = rec.inn
    source = "inn" if inn else None

    # 1) ИНН нет, но есть почта — ищем организацию по почте
    if not inn and rec.email_osn:
        try:
            from src.parser_new.email_lookup import search_inn_by_email
        except ImportError:
            from email_lookup import search_inn_by_email
        found_inn, liquidated, found_name = search_inn_by_email(rec.email_osn)
        if found_inn:
            inn, source = found_inn, "email"
            # ликвидация видна прямо на rusprofile — в Checko НЕ идём
            if liquidated:
                rec.inn = inn
                if not rec.adm_name:
                    rec.adm_name = found_name
                rec.note = "ликвидирована"
                logger.info(f"[email] {rec.email_osn} → {inn} ликвидирована, Checko пропущен")
                return rec
            logger.info(f"[email] {rec.email_osn} → ИНН {inn}")

    # 2) ИНН всё ещё нет — старый путь по названию МО.
    #    ВАЖНО: только если в строке есть название/район/администрация.
    #    Иначе (строка «только с почтой») search_rusprofile уйдёт в Яндекс
    #    с пустым запросом и притащит случайную организацию.
    if not inn:
        if rec.mun_name or rec.mun_r_name or rec.adm_name:
            found_inn, liquidated, org_name = search_rusprofile(rec)
            if liquidated:
                if not rec.adm_name and org_name:
                    rec.adm_name = org_name        # настоящее название, а не слово
                rec.note = "ликвидирована"
                return rec
            if found_inn:
                inn, source = found_inn, "name"
            else:
                inn = search_checko_by_name(rec)
                if inn:
                    source = "name"
        if not inn:
            rec.note = "не найдено"
            return rec

    checko_data = fetch_checko_by_inn(inn)
    if not checko_data:
        rec.note = "ИНН найден но данные Checko недоступны"
        rec.inn = inn
        return rec

    # Ликвидированная организация — помечаем и НЕ дозаполняем (любой источник)
    if _is_liquidated_status(checko_data.get("status", "")):
        rec.inn = inn
        if not rec.adm_name:
            rec.adm_name = checko_data.get("adm_name", "")
        rec.note = "ликвидирована"
        logger.info(f"[process] ИНН {inn} ликвидирована — дозаполнение пропущено")
        return rec

    # Проверка соответствия — ТОЛЬКО для поиска по названию.
    if source == "name":
        found_name = checko_data.get("adm_name", "")
        if not validate_org_matches(found_name, rec):
            logger.warning(f"Несоответствие: искали '{rec.search_name}', нашли '{found_name}'")
            rec.note = f"проверить вручную: нашли {found_name[:80]}"
            return rec

    if not rec.inn:       rec.inn = checko_data["inn"]
    if not rec.adm_name:  rec.adm_name = checko_data["adm_name"]
    if not rec.adres:     rec.adres = checko_data["adres"]
    if not rec.head_fio:  rec.head_fio = checko_data["head_fio"]
    if not rec.email_osn: rec.email_osn = checko_data["email_osn"]
    if not rec.email_dop: rec.email_dop = checko_data["email_dop"]
    if not rec.tel_osn:   rec.tel_osn = checko_data["tel_osn"]
    if not rec.tel_dop:   rec.tel_dop = checko_data["tel_dop"]
    if not rec.kpp:       rec.kpp = checko_data["kpp"]
    if not rec.ogrn:      rec.ogrn = checko_data["ogrn"]
    if not rec.okpo:      rec.okpo = checko_data["okpo"]
    if not rec.oktmo:     rec.oktmo = checko_data["oktmo"]

    if not rec.email_osn or not rec.tel_osn:
        name = rec.adm_name or checko_data.get("adm_name", "")
        if name:
            contacts = search_contacts_online(name)
            if not rec.email_osn and contacts.get("email_osn"):
                rec.email_osn = contacts["email_osn"]
            if not rec.email_dop and contacts.get("email_dop"):
                rec.email_dop = contacts["email_dop"]
            if not rec.tel_osn and contacts.get("tel_osn"):
                rec.tel_osn = contacts["tel_osn"]
            if not rec.tel_dop and contacts.get("tel_dop"):
                rec.tel_dop = contacts["tel_dop"]

    return rec


# ==============================
# УТИЛИТЫ
# ==============================

def _copy_row_to_failed(ws_src, ws_dst, src_row: int, dst_row: int) -> None:
    for col_idx in range(1, len(COL) + 1):
        value = ws_src.cell(src_row, col_idx).value
        if value is not None:
            ws_dst.cell(dst_row, col_idx, value)


# ==============================
# ГЛАВНАЯ ФУНКЦИЯ
# ==============================

def run(
    file_path: str,
    save_every: int = 10,
    output_dir: str | None = None,
    cycle: int = 1,
) -> dict:
    import shutil

    path = Path(file_path)
    if not path.exists():
        logger.error(f"Файл не найден: {file_path}")
        return {"error": f"Файл не найден: {file_path}"}

    # Определяем папку для результатов
    if output_dir:
        out_dir = Path(output_dir)
    else:
        out_dir = Path(config.OUTPUT_DIR) / "latest"
    out_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    out_path = out_dir / f"batch_{timestamp}.xlsx"
    failed_path = out_dir / f"batch_FAILED_{timestamp}.xlsx"

    shutil.copy2(str(path), str(out_path))
    shutil.copy2(str(path), str(failed_path))
    logger.info(f"Рабочая копия: {out_path}")
    logger.info(f"Файл для непроверенных: {failed_path}")

    wb, records = read_excel(str(out_path))
    ws = wb.active

    wb_failed = load_workbook(str(failed_path))
    ws_failed = wb_failed.active
    if ws_failed.max_row >= DATA_START_ROW:
        ws_failed.delete_rows(DATA_START_ROW, ws_failed.max_row)
    failed_row_idx = DATA_START_ROW

    total = len(records)
    processed = found = not_found = liquidated_count = 0
    progress_step = max(1, min(save_every, total // 5)) if total < 50 else max(save_every, total // 12)   # ~10–12 апдейтов, без спама

    logger.info(f"Начинаю обработку {total} строк...")
    print(f"\n{'='*50}")
    print(f"Найдено строк для обработки: {total}")
    print(f"Файл результата: {out_path}")
    print(f"{'='*50}\n")

    if total:
        if cycle > 1:
            _emit(f"Прохожу ещё раз по {total} {_label()} с неполными данными…")
        else:
            _emit(f"Начинаю сбор данных: {total} {_label()}.")

    for i, rec in enumerate(records, 1):
        try:
            print(f"[{i}/{total}] {rec.sub_rf} | {rec.mun_r_name} | {rec.mun_name[:40]}...", end=" ", flush=True)

            updated = process_record(rec)
            write_record(ws, updated)

            if updated.note == "ликвидирована":
                liquidated_count += 1
                print("⚠️  ликвидирована")
            elif updated.note == "не найдено":
                not_found += 1
                print("❌ не найдено")
                _copy_row_to_failed(ws, ws_failed, updated.excel_row, failed_row_idx)
                failed_row_idx += 1
            elif "проверить вручную" in updated.note:
                not_found += 1
                print(f"⚠️  {updated.note[:60]}")
                _copy_row_to_failed(ws, ws_failed, updated.excel_row, failed_row_idx)
                failed_row_idx += 1
            elif "недоступны" in updated.note:
                not_found += 1
                print(f"⚠️  {updated.note[:60]}")
                _copy_row_to_failed(ws, ws_failed, updated.excel_row, failed_row_idx)
                failed_row_idx += 1
            else:
                found += 1
                print(f"✅ ИНН: {updated.inn}")

            processed += 1

            if total and i % progress_step == 0 and i < total:
                _emit(f"Собрал данные: {i} из {total} {_label()}…")

            if i % save_every == 0:
                wb.save(str(out_path))
                wb_failed.save(str(failed_path))
                logger.info(f"Прогресс сохранён: {i}/{total}")

            time.sleep(random.uniform(6.0, 16.0))

        except KeyboardInterrupt:
            logger.info("Прерывание пользователем — сохраняю прогресс...")
            wb.save(str(out_path))
            wb_failed.save(str(failed_path))
            print(f"\n\nПрервано. Обработано {processed}/{total}. Файл сохранён.")
            return {
                "processed": processed, "found": found, "not_found": not_found,
                "liquidated": liquidated_count,
                "output_path": str(out_path), "failed_path": str(failed_path),
                "interrupted": True,
            }

        except Exception as e:
            logger.error(f"Ошибка для строки {rec.excel_row}: {e}")
            ws.cell(rec.excel_row, COL["NOTE"], f"ошибка: {str(e)[:50]}")
            not_found += 1
            print(f"❌ ошибка: {e}")

    # Финальное сохранение
    # Финальное сохранение
    wb.save(str(out_path))
    wb_failed.save(str(failed_path))

    # Дозаполнение B/C/D по названию из E — ДО постпроверки названий МО,
    # чтобы верификатор проверил уже заполненный столбец D
    try:
        from src.parser_new.admin_levels import fill_admin_levels
        lvl = fill_admin_levels(str(out_path))
        logger.info(f"Уровни B/C/D: {lvl}")
        print(f"   Уровни B/C/D: B={lvl.get('filled_b',0)} C={lvl.get('filled_c',0)} "
              f"D={lvl.get('filled_d',0)} (ликв. пропущено: {lvl.get('skipped_liquidated',0)})")
    except Exception as e:
        logger.warning(f"Не удалось заполнить B/C/D: {e}")

    # Постпроверка
    try:
        from src.generator.verification.municipality_name_verifier import (
            verify_municipality_names_in_workbook,
        )
        logger.info("Запуск постпроверки названий МО...")
        verify_result = verify_municipality_names_in_workbook(out_path)
        verified = verify_result.get("verified_rows", 0)
        replaced = verify_result.get("updated_rows", 0)
        logger.info(f"Постпроверка: проверено {verified}, исправлено {replaced}")
        print(f"   Постпроверка: проверено {verified}, исправлено {replaced}")
    except ImportError:
        logger.warning("Модуль municipality_name_verifier не найден — постпроверка пропущена")
    except Exception as e:
        logger.warning(f"Ошибка постпроверки: {e}")
    
    try:
        from src.parser_new.admin_levels import fill_admin_levels
        lvl = fill_admin_levels(str(out_path))
        logger.info(f"Уровни B/C/D: {lvl}")
        print(f"   Уровни B/C/D: B={lvl.get('filled_b',0)} C={lvl.get('filled_c',0)} D={lvl.get('filled_d',0)}")
    except Exception as e:
        logger.warning(f"Не удалось заполнить B/C/D: {e}")

    # Архив (только если не job-папка)
    if not output_dir:
        arc_dir = Path(config.OUTPUT_DIR) / "archive" / datetime.now().strftime("%Y-%m-%d")
        arc_dir.mkdir(parents=True, exist_ok=True)
        shutil.copy2(str(out_path), str(arc_dir / out_path.name))
        shutil.copy2(str(failed_path), str(arc_dir / failed_path.name))

    print(f"\n{'='*50}")
    print(f"✅ Готово!")
    print(f"   Обработано:    {processed}")
    print(f"   Найдено:       {found}")
    print(f"   Ликвидировано: {liquidated_count}")
    print(f"   Не найдено:    {not_found}")
    print(f"   Основной файл: {out_path}")
    print(f"   Непроверенные: {failed_path}")
    print(f"{'='*50}\n")

    return {
        "processed": processed, "found": found, "not_found": not_found,
        "liquidated": liquidated_count,
        "output_path": str(out_path), "failed_path": str(failed_path),
    }

def _count_final_stats(file_path: str) -> dict:
    """Итог по всему файлу, а не по последнему циклу.
    Считаем по наличию ИНН: notes между циклами могут устаревать
    (строка, провалившаяся в 1-м проходе, могла дозаполниться во 2-м,
    но пометка 'не найдено' в NOTE уже не перезаписывается)."""
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        total = found = not_found = liquidated = 0
        for row_idx in range(DATA_START_ROW, ws.max_row + 1):
            mun = ws.cell(row_idx, COL["MUN_NAME"]).value or ws.cell(row_idx, COL["MUN_R_NAME"]).value
            if not (mun and str(mun).strip()):
                continue
            total += 1
            note = str(ws.cell(row_idx, COL["NOTE"]).value or "").lower()
            inn = str(ws.cell(row_idx, COL["REQUISITES_INN"]).value or "").strip()
            if "ликвид" in note:
                liquidated += 1
            elif inn:
                found += 1
            else:
                not_found += 1
        wb.close()
        return {"processed": total, "found": found,
                "not_found": not_found, "liquidated": liquidated}
    except Exception as e:
        logger.warning(f"[batch] Не удалось пересчитать итог по файлу: {e}")
        return {}

def run_with_retries(
    file_path: str,
    save_every: int = 10,
    output_dir: str | None = None,
    max_cycles: int = 2,
) -> dict:
    max_cycles = max(1, min(max_cycles, 3))
    current_file = file_path
    result: dict = {}

    for cycle in range(1, max_cycles + 1):
        logger.info(f"[batch] Цикл {cycle}/{max_cycles} по файлу {current_file}")
        result = run(current_file, save_every, output_dir, cycle=cycle)   # + cycle

        if not result or result.get("error"):
            break

        not_found = result.get("not_found", 0)
        if not_found == 0:
            logger.info(f"[batch] Все строки обработаны за {cycle} цикл(ов)")
            break

        if cycle < max_cycles:
            current_file = result.get("output_path", current_file)
            logger.info(f"[batch] Осталось ненайденных: {not_found}, повтор...")
            # _emit(...) убрали — сообщение о повторе теперь даёт сам run() с верным числом
            time.sleep(2)
        else:
            logger.info(f"[batch] Лимит циклов достигнут, осталось: {not_found}")

    if result and not result.get("error"):
        _emit("Завершил сбор, сохраняю результат в файл…")
        final = _count_final_stats(result.get("output_path", current_file))
        if final:
            result.update(final)   # found/not_found/processed теперь по всему региону

    return result
# ==============================
# ТОЧКА ВХОДА
# ==============================

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Пакетная обработка МО")
    parser.add_argument("--file", required=True, help="Путь к Excel файлу")
    parser.add_argument("--save-every", type=int, default=10,
                        help="Сохранять каждые N строк (по умолчанию 10)")
    parser.add_argument("--max-cycles", type=int, default=1,
                        help="Максимум циклов обработки (по умолчанию 1, макс 3)")
    args = parser.parse_args()

    max_cycles = min(args.max_cycles, 3)
    current_file = args.file

    for cycle in range(1, max_cycles + 1):
        print(f"\n{'='*50}")
        print(f"ЦИКЛ {cycle}/{max_cycles}")
        print(f"{'='*50}")

        result = run(current_file, args.save_every)
        if not result:
            break

        not_found_count = result.get("not_found", 0)
        if not_found_count == 0:
            print(f"\n✅ Все строки обработаны, повторный цикл не нужен")
            break

        if cycle < max_cycles:
            current_file = result.get("output_path", current_file)
            print(f"\n🔄 Осталось необработанных: {not_found_count}, запускаю цикл {cycle + 1}...")
            time.sleep(2)
        else:
            print(f"\n⚠️  Осталось необработанных: {not_found_count} (достигнут лимит циклов)")