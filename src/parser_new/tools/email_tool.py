"""
tools/email_tool.py — обновление e-mail администраций МО с их официальных сайтов.

Сценарий: в файле есть строки МО с заполненными реквизитами, но почты неверные
(письма не дошли). Нужно найти актуальную почту НА ОФИЦИАЛЬНОМ САЙТЕ каждого МО.

Логика по каждой строке (детерминированная, без LLM в цикле):
  1. Чистим формат обоих адресов (гомоглифы кириллицы -> латиница, опечатки
     доменов, пробелы). Запоминаем, был ли исходный EMAIL_OSN битым.
  2. Ищем официальный сайт через Яндекс Search API (приоритет gosweb.gosuslugi.ru
     и собственным сайтам администрации; vk/ok/2gis/rusprofile/zakupki — не сайт).
  3. Достаём почту со страницы (mailto + текст; если на главной пусто — страница
     "Контакты" того же домена).
  4. Выбираем приоритетную почту (госдомены .gov.ru / региональные выше общих
     mail.ru / yandex.ru).
  5. Раскладываем результат (см. _apply_result) и помечаем STATUS.

Прогресс идёт в чат через progress.emit (тот же механизм, что у batch).
"""
from __future__ import annotations

import re
import time
from datetime import datetime
from pathlib import Path
from urllib.parse import urljoin

import httpx
from bs4 import BeautifulSoup
from openpyxl import load_workbook

try:
    from langchain.tools import tool
except ImportError:
    def tool(fn):
        return fn

# инфраструктура проекта (с fallback на случай запуска вне пакета)
try:
    from src.parser_new import config
    from src.parser_new.logger import logger
    from src.parser_new.batch_processor import (
        _get_yandex_search, _parse_yandex_xml, HEADERS, COL, DATA_START_ROW,
    )
    from src.parser_new.progress import emit as _emit
except ImportError:  # запуск из каталога parser_new
    import config
    from logger import logger
    from batch_processor import (
        _get_yandex_search, _parse_yandex_xml, HEADERS, COL, DATA_START_ROW,
    )
    try:
        from progress import emit as _emit
    except Exception:
        def _emit(*a, **k):
            pass


# ==============================
# ЧИСТКА ФОРМАТА E-MAIL
# ==============================

# кириллические буквы, визуально совпадающие с латиницей (гомоглифы)
_HOMOGLYPHS = {
    "а": "a", "е": "e", "о": "o", "р": "p", "с": "c", "у": "y", "х": "x",
    "А": "A", "Е": "E", "О": "O", "Р": "P", "С": "C", "У": "Y", "Х": "X",
    "к": "k", "К": "K", "м": "M", "М": "M", "т": "t", "Т": "T",
    "в": "B", "В": "B", "н": "H", "Н": "H",
}

# частые опечатки доменов верхнего уровня
_TLD_FIXES = [
    (re.compile(r"\.r[uy]s\b"), ".ru"),   # .rus
    (re.compile(r"\.ry\b"), ".ru"),
    (re.compile(r"\.ri\b"), ".ru"),
    (re.compile(r"\.tu\b"), ".ru"),
    (re.compile(r"\.r\b"), ".ru"),        # оборванный .r
]

_EMAIL_RE = re.compile(r"^[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}$")
_EMAIL_FINDALL = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")

# мусорные адреса — не считаем их почтой администрации
_JUNK = ["@gosuslugi.ru", "example.", "@sentry", "noreply@", "no-reply@",
         "@2x.", "@sentry.io", "your@", "mail@mail."]


def clean_email(raw: str) -> str:
    """Чистит адрес: гомоглифы -> латиница, опечатки доменов, пробелы."""
    if not raw:
        return ""
    s = str(raw).strip().replace(" ", "")
    s = "".join(_HOMOGLYPHS.get(ch, ch) for ch in s)
    low = s.lower()
    for pat, repl in _TLD_FIXES:
        low2 = pat.sub(repl, low)
        if low2 != low:
            # применяем фикс с сохранением регистра локальной части
            at = s.find("@")
            if at != -1:
                s = s[:at] + "@" + s[at + 1:].lower()
                s = pat.sub(repl, s)
            low = low2
    return s


def is_valid_email(value: str) -> bool:
    return bool(_EMAIL_RE.match((value or "").strip()))


def _clean_list(raw: str) -> list[str]:
    """Разбивает ячейку с несколькими адресами (через запятую) и чистит каждый."""
    if not raw:
        return []
    parts = re.split(r"[,;\s]+", str(raw))
    out, seen = [], set()
    for p in parts:
        c = clean_email(p)
        if c and c.lower() not in seen:
            seen.add(c.lower())
            out.append(c)
    return out


# ==============================
# ПОИСК ОФИЦИАЛЬНОГО САЙТА
# ==============================

# не считаем официальным сайтом
_NOT_OFFICIAL = [
    "vk.com", "ok.ru", "rusprofile", "checko.ru", "egrul", "list-org",
    "sbis.ru", "2gis.", "zakupki.gov.ru", "otc.ru", "vsem-podryad",
    "yandex.ru/maps", "wikipedia", "web.archive.org", "audit-it",
]


def _pick_site(results: list[dict]) -> str | None:
    """Выбирает лучшую ссылку: gosweb.gosuslugi.ru -> прочие неагрегаторные."""
    def ok(url: str) -> bool:
        return bool(url) and not any(d in url for d in _NOT_OFFICIAL)

    # 1) приоритет: движок госсайтов
    for r in results:
        if "gosweb.gosuslugi.ru" in r.get("url", ""):
            return r["url"]
    # 2) любой другой неагрегаторный
    for r in results:
        if ok(r.get("url", "")):
            return r["url"]
    return None


def _find_official_site(adm_name: str, region: str) -> str | None:
    try:
        search = _get_yandex_search()
        query = f"{adm_name} {region} официальный сайт"
        time.sleep(0.6)
        xml = search.run(query, format="xml", page=0)
        results = _parse_yandex_xml(xml)
        return _pick_site(results)
    except Exception as e:
        logger.warning(f"[email] поиск сайта не удался для {adm_name[:40]}: {e}")
        return None


# ==============================
# ИЗВЛЕЧЕНИЕ ПОЧТЫ СО СТРАНИЦЫ
# ==============================

def _emails_from_page(url: str) -> list[str]:
    try:
        time.sleep(1.0)
        resp = httpx.get(url, headers=HEADERS, timeout=12, follow_redirects=True)
        soup = BeautifulSoup(resp.text, "lxml")
    except Exception as e:
        logger.debug(f"[email] заход на {url[:50]} не удался: {e}")
        return []

    found: list[str] = []
    for a in soup.select("a[href^='mailto:']"):
        addr = a.get("href", "")[len("mailto:"):].split("?")[0].strip()
        if addr:
            found.append(addr)
    found.extend(_EMAIL_FINDALL.findall(soup.get_text(" ", strip=True)))

    clean, seen = [], set()
    for e in found:
        c = clean_email(e)
        cl = c.lower()
        if not c or cl in seen:
            continue
        if any(j in cl for j in _JUNK):
            continue
        seen.add(cl)
        clean.append(c)
    return clean


def _contacts_link(url: str) -> str | None:
    """Ищет на странице внутреннюю ссылку 'Контакты' (того же домена)."""
    try:
        time.sleep(0.4)
        resp = httpx.get(url, headers=HEADERS, timeout=12, follow_redirects=True)
        soup = BeautifulSoup(resp.text, "lxml")
        from urllib.parse import urlparse
        base_host = urlparse(str(resp.url)).netloc
        for a in soup.find_all("a", href=True):
            if "контакт" in a.get_text().lower():
                href = urljoin(str(resp.url), a["href"])
                if urlparse(href).netloc == base_host:   # только тот же сайт
                    return href
    except Exception:
        pass
    return None


def _emails_from_site(site_url: str) -> list[str]:
    """Главная + (если пусто) страница 'Контакты' того же домена."""
    emails = _emails_from_page(site_url)
    if emails:
        return emails
    link = _contacts_link(site_url)
    if link and link != site_url:
        return _emails_from_page(link)
    return []


def _prioritize(emails: list[str]) -> list[str]:
    """Госдомены (.gov.ru и региональные) — вперёд, общие mail/yandex — назад."""
    def score(e: str) -> int:
        d = e.split("@")[-1].lower()
        if d.endswith(".gov.ru"):
            return 0
        if d.endswith(".ru") and d not in ("mail.ru", "yandex.ru", "list.ru",
                                            "rambler.ru", "bk.ru", "inbox.ru",
                                            "gmail.com", "ngs.ru"):
            return 1   # ведомственный/региональный домен
        return 2       # общий публичный ящик
    return sorted(emails, key=score)


# ==============================
# РАСКЛАДКА РЕЗУЛЬТАТА ПО ОДНОЙ СТРОКЕ
# ==============================

def _apply_result(
    ws, row: int,
    old_osn_raw: str, old_dop_raw: str,
    site_emails: list[str],
) -> str:
    """
    Записывает результат в строку по согласованным правилам.
    Возвращает короткий статус для STATUS.
    """
    old_osn_clean = clean_email(old_osn_raw)
    old_dop_list = _clean_list(old_dop_raw)
    was_bity = bool(old_osn_raw) and not is_valid_email(old_osn_raw)

    # почту на сайте не нашли
    if not site_emails:
        # оставляем почищенный старый OSN, доп — почищенный
        if old_osn_clean:
            ws.cell(row, COL["EMAIL_OSN"], old_osn_clean)
        if old_dop_list:
            ws.cell(row, COL["EMAIL_DOP"], ", ".join(old_dop_list))
        return "сайт: почта не найдена"

    prioritized = _prioritize(site_emails)
    picked = prioritized[0]
    others = prioritized[1:]
    same_as_old = old_osn_clean and picked.lower() == old_osn_clean.lower()

    # совпало с тем, что в файле
    if same_as_old:
        if was_bity:
            # адрес был верный, просто кривой в файле -> сайт подтвердил, пишем
            ws.cell(row, COL["EMAIL_OSN"], picked)
            dop = _merge_dop(old_dop_list, others, exclude=picked)
            if dop:
                ws.cell(row, COL["EMAIL_DOP"], ", ".join(dop))
            return "формат исправлен, подтверждён сайтом"
        else:
            # валидный адрес, но отправитель его забраковал, а сайт даёт тот же
            if old_dop_list:
                ws.cell(row, COL["EMAIL_DOP"], ", ".join(old_dop_list))
            return "на сайте та же почта — проверить вручную"

    # сайт дал НОВЫЙ адрес
    ws.cell(row, COL["EMAIL_OSN"], picked)

    dop = list(old_dop_list)
    # прочие найденные на сайте
    for e in others:
        if e.lower() != picked.lower():
            dop.append(e)
    # судьба старого OSN
    if was_bity and old_osn_clean and old_osn_clean.lower() != picked.lower():
        dop.append(old_osn_clean)   # битый старый -> в доп почищенным
    # (нормальный старый OSN просто удаляется — мы его в OSN перезаписали)

    # убрать дубли и сам picked из доп
    dop_final, seen = [], {picked.lower()}
    for e in dop:
        if e.lower() not in seen:
            seen.add(e.lower())
            dop_final.append(e)
    if dop_final:
        ws.cell(row, COL["EMAIL_DOP"], ", ".join(dop_final))

    return "обновлено с сайта"


def _merge_dop(old_dop: list[str], others: list[str], exclude: str) -> list[str]:
    out, seen = [], {exclude.lower()}
    for e in list(old_dop) + list(others):
        if e.lower() not in seen:
            seen.add(e.lower())
            out.append(e)
    return out


# ==============================
# ГЛАВНАЯ ФУНКЦИЯ
# ==============================

def run_email_refresh(file_path: str, output_dir: str | None = None, limit: int | None = None) -> dict:
    import shutil

    path = Path(file_path)
    if not path.exists():
        return {"error": f"Файл не найден: {file_path}"}

    out_dir = Path(output_dir) if output_dir else (Path(config.OUTPUT_DIR) / "latest")
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M")
    # префикс batch_ — чтобы существующая логика chat() подхватила файл на скачивание
    out_path = out_dir / f"batch_emails_{ts}.xlsx"
    shutil.copy2(str(path), str(out_path))

    wb = load_workbook(str(out_path))
    ws = wb.active

    # собираем строки с МО
    rows = []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        adm = str(ws.cell(r, COL["ADM_NAME"]).value or "").strip()
        mun = str(ws.cell(r, COL["MUN_NAME"]).value or "").strip()
        if adm or mun:
            rows.append(r)

    if limit:
        rows = rows[:limit]
    
    total = len(rows)
    if not total:
        return {"error": "В файле нет строк с администрациями/МО."}

    updated = same = not_found = 0
    progress_step = max(5, total // 12)
    _emit(f"Начинаю обновление почт по {total} МО.")

    for i, r in enumerate(rows, 1):
        adm = str(ws.cell(r, COL["ADM_NAME"]).value or "").strip()
        region = str(ws.cell(r, COL["SUB_RF"]).value or "").strip()
        old_osn = str(ws.cell(r, COL["EMAIL_OSN"]).value or "")
        old_dop = str(ws.cell(r, COL["EMAIL_DOP"]).value or "")

        try:
            site = _find_official_site(adm, region)
            emails = _emails_from_site(site) if site else []
            status = _apply_result(ws, r, old_osn, old_dop, emails)

            if status == "обновлено с сайта" or status == "формат исправлен, подтверждён сайтом":
                updated += 1
            elif "проверить вручную" in status:
                same += 1
            else:
                not_found += 1

            ws.cell(r, COL["STATUS"], status)
        except Exception as e:
            logger.warning(f"[email] строка {r} ({adm[:40]}): {e}")
            ws.cell(r, COL["STATUS"], f"ошибка обработки: {str(e)[:50]}")
            not_found += 1

        if i % progress_step == 0 and i < total:
            _emit(f"Обновил почты для {i} из {total} МО…")

        if i % 10 == 0:
            wb.save(str(out_path))

    _emit("Завершаю, сохраняю результат в файл…")
    wb.save(str(out_path))

    return {
        "processed": total,
        "updated": updated,
        "manual_check": same,
        "not_found": not_found,
        "output_path": str(out_path),
    }


# ==============================
# ИНСТРУМЕНТ ДЛЯ АГЕНТА
# ==============================

@tool
def fix_emails_tool(file_path: str, limit: int = 0) -> str:
    """
    Обновляет e-mail администраций МО, беря актуальные адреса с их ОФИЦИАЛЬНЫХ
    САЙТОВ. Используй когда пользователь говорит, что почты в файле неверные,
    устарели, или письма по ним НЕ ДОШЛИ, и просит найти/обновить/перепроверить
    адреса электронной почты.

    НЕ используй batch_search_tool для этой задачи — он берёт почту из Checko,
    а здесь источник принципиально другой: официальный сайт МО.

    Что делает автоматически по каждой строке:
      - чистит битый формат почты (кириллические буквы-двойники, опечатки);
      - находит официальный сайт МО и достаёт с него актуальную почту;
      - ставит её основной (EMAIL_OSN), прочие найденные — в дополнительные;
      - если на сайте та же почта, что уже была — помечает на ручную проверку;
      - если сайт/почту найти не удалось — помечает строку.

    Параметр:
      file_path: путь к Excel-файлу с МО (тот, что прислал пользователь).

    Возвращает краткий отчёт.
    """
    try:
        logger.info(f"[email] Запуск обновления почт: {file_path}")
        result = run_email_refresh(file_path=file_path, limit=limit or None)

        if not result or result.get("error"):
            return f"Ошибка: {result.get('error', 'не удалось обработать файл')}"

        return (
            f"Обновление почт завершено:\n"
            f"  Обработано: {result['processed']}\n"
            f"  Обновлено с сайта: {result['updated']}\n"
            f"  На ручную проверку: {result['manual_check']}\n"
            f"  Не найдено: {result['not_found']}"
        )
    except Exception as e:
        logger.error(f"[email] Ошибка: {e}")
        return f"Ошибка обновления почт: {e}"