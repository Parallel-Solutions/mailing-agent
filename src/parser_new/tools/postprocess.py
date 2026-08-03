"""
tools/postprocess.py — детерминированная доводка собранной таблицы.

ЗАЧЕМ: агент (LLM) систематически пропускает шаги «сходи на официальный сайт за
почтой» и «допиши город в название», сколько бы ни было инструкций в промпте.
Поэтому эти два шага вынесены в КОД и выполняются автоматически после того, как
файл создан — неважно, кем: коллектором, discovery или самим агентом.

Что делает:
  1. ПОЧТА. Для строк с пустой или подозрительной (личной) почтой ищет
     официальный сайт организации и забирает оттуда ведомственный адрес.
     Вся тяжёлая логика переиспользуется из email_tool.py.
  2. НАЗВАНИЕ. Если в названии есть «АДМИНИСТРАЦИИ ГОРОДА» и т.п. без топонима —
     дописывает город из адреса («…АДМИНИСТРАЦИИ ГОРОДА» -> «…ГОРОДА СУРГУТА»).

LLM здесь не участвует вообще.
"""
from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

try:
    from src.parser_new.logger import logger
    from src.parser_new.batch_processor import COL, DATA_START_ROW
    from src.parser_new.tools.email_tool import (
        clean_email, _find_official_site, _emails_from_site, _prioritize,
    )
    from src.parser_new.progress import emit as _emit
except ImportError:                      # запуск из каталога parser_new
    from logger import logger
    from batch_processor import COL, DATA_START_ROW
    from tools.email_tool import (
        clean_email, _find_official_site, _emails_from_site, _prioritize,
    )
    try:
        from progress import emit as _emit
    except Exception:
        def _emit(*a, **k):
            pass


# ==============================
# ЛИЧНАЯ ПОЧТА ИЛИ ВЕДОМСТВЕННАЯ
# ==============================

# Ролевые ящики организации — такие оставляем как есть.
_ROLE_WORDS = {
    "info", "mail", "email", "office", "adm", "admin", "administration",
    "priemnaya", "priem", "priemn", "kanc", "kanclr", "kancelyariya", "secretar",
    "sekretar", "general", "common", "contact", "contacts", "kontakt", "org",
    "dep", "depart", "reception", "post", "pochta", "official", "site", "web",
    "glava", "head", "public", "press", "zakupki", "buh", "buhgalter",
}

# Публичные почтовые провайдеры: для организации это уже слабый признак.
_PUBLIC_DOMAINS = {
    "mail.ru", "yandex.ru", "ya.ru", "list.ru", "bk.ru", "inbox.ru",
    "rambler.ru", "gmail.com", "outlook.com", "icloud.com", "mail.com",
}


def _local_part(email: str) -> str:
    return (email or "").split("@", 1)[0].lower().strip()


def _looks_personal(email: str) -> bool:
    """Похож ли адрес на ящик КОНКРЕТНОГО СОТРУДНИКА, а не организации.

    Checko берёт контакты в том числе из карточек госзакупок, где указан ящик
    специалиста, который вёл процедуру (kuklina@admsurgut.ru). Для рассылки такой
    адрес почти бесполезен — надо перепроверить по официальному сайту.

    Эвристика (намеренно осторожная — при сомнении считаем НЕ личной):
      - ролевое слово в локальной части        -> не личная
      - короткая аббревиатура (<= 4 символа)   -> не личная (ds@, dga@, uks@)
      - вид «и.фамилия» / «имя_фамилия»        -> личная
      - одно слово из 5+ букв без цифр и дефиса -> личная (похоже на фамилию)
    """
    local = _local_part(email)
    if not local:
        return False

    # ролевое слово целиком или как часть (info-adm, adm-priem)
    parts = re.split(r"[.\-_]+", local)
    if any(p in _ROLE_WORDS for p in parts if p):
        return False
    if local in _ROLE_WORDS:
        return False

    # короткая аббревиатура подразделения
    if len(local) <= 4:
        return False

    # «i.ivanov», «ivan_petrov» — имя + фамилия
    if len(parts) == 2 and all(p.isalpha() for p in parts):
        if len(parts[0]) <= 2 or len(parts[1]) >= 5:
            return True

    # одно слово из букв, 5+ символов — скорее всего фамилия
    if local.isalpha() and len(local) >= 5:
        return True

    return False


def _better_than(candidate: str, current: str) -> bool:
    """Стоит ли заменить текущий адрес найденным на сайте."""
    if not candidate:
        return False
    if not current:
        return True
    if candidate.lower() == current.lower():
        return False
    # ведомственный вместо личного — да
    if _looks_personal(current) and not _looks_personal(candidate):
        return True
    # собственный домен вместо публичного — да
    cur_dom = current.split("@")[-1].lower()
    cand_dom = candidate.split("@")[-1].lower()
    if cur_dom in _PUBLIC_DOMAINS and cand_dom not in _PUBLIC_DOMAINS:
        return True
    return False


def _pick_from_site(emails: list[str]) -> str:
    """Лучший адрес со страницы: сперва по домену, затем ролевые вперёд личных."""
    if not emails:
        return ""
    ranked = _prioritize(emails)
    for e in ranked:
        if not _looks_personal(e):
            return e
    return ranked[0]


# ==============================
# НАЗВАНИЕ: ДОПИСАТЬ ГОРОД
# ==============================

# «...АДМИНИСТРАЦИИ ГОРОДА» в конце и никакого топонима после
_HEADLESS_TAIL = re.compile(
    r"(АДМИНИСТРАЦИИ\s+(?:ГОРОДА|РАЙОНА|ОКРУГА|ПОСЕЛЕНИЯ))\s*$",
    re.IGNORECASE,
)

_CITY_IN_ADDR = re.compile(
    r"\bг\.?\s*([А-ЯЁ][А-Яа-яЁё\-]{2,})", re.UNICODE
)

# города, которые в адресе идут как субъект РФ, а не как место организации
_FEDERAL_CITIES = {"москва", "санкт-петербург", "севастополь"}


def _genitive_city(city: str) -> str:
    """Именительный -> родительный, грубо, но для городов работает.
    Сургут -> Сургута, Ханты-Мансийск -> Ханты-Мансийска, Тюмень -> Тюмени."""
    c = city.strip()
    low = c.lower()
    if low.endswith(("ь",)):
        return c[:-1] + "и"
    if low.endswith(("а", "я")):
        return c
    if low.endswith(("ы", "и", "о", "е")):
        return c
    return c + "а"


def _fix_name(name: str, address: str) -> str:
    """Дописывает город из адреса, если в названии его нет."""
    if not name or not _HEADLESS_TAIL.search(name.strip()):
        return name

    cities = _CITY_IN_ADDR.findall(address or "")
    city = ""
    for c in cities:
        if c.lower().strip(".,") not in _FEDERAL_CITIES:
            city = c.strip(".,")
            break
    if not city:
        return name

    return f"{name.strip()} {_genitive_city(city).upper()}"


# ==============================
# ГЛАВНАЯ ФУНКЦИЯ
# ==============================

def _site_candidates(adm: str, region: str) -> list[tuple[str, str]]:
    """Варианты запроса для поиска официального сайта, от точного к общему.

    Одной формулировки мало: у «ДЕПАРТАМЕНТ … АДМИНИСТРАЦИИ ГОРОДА СУРГУТА»
    добавка региона («Ханты-Мансийский автономный округ») уводит выдачу на
    окружной сайт вместо городского. Поэтому пробуем несколько заходов.
    """
    adm = (adm or "").strip()
    out: list[tuple[str, str]] = []
    if not adm:
        return out

    # если в названии есть город — он точнее региона
    m = re.search(r"ГОРОДА\s+([А-ЯЁ\-]{3,})", adm.upper())
    city = m.group(1).capitalize() if m else ""

    if city:
        out.append((adm, city))          # название + город из названия
    if region:
        out.append((adm, region))        # название + регион
    out.append((adm, ""))                # только название
    return out


def _lookup_site_emails(adm: str, region: str) -> tuple[str, list[str]]:
    """Ищет официальный сайт и почты на нём. Возвращает (сайт, список почт)."""
    tried: list[str] = []
    for name, place in _site_candidates(adm, region):
        try:
            site = _find_official_site(name, place)
        except Exception as e:
            logger.debug(f"[postprocess] поиск сайта ({place or 'без региона'}): {e}")
            site = None
        if not site:
            tried.append(f"{place or 'без региона'}: сайт не найден")
            continue
        emails = _emails_from_site(site)
        if emails:
            return site, emails
        tried.append(f"{site}: почт на странице нет")

    if tried:
        logger.info(f"[postprocess] {adm[:45]}: не удалось — {'; '.join(tried[:3])}")
    return "", []


def postprocess_file(file_path: str,
                     max_email_lookups: int = 40,
                     enrich_suspicious: bool = True,
                     check_all: bool = False) -> dict:
    """Доводит уже созданный файл: почта с сайтов + город в названии.

    Работает НА МЕСТЕ (файл перезаписывается) — он и так лежит в output/latest.

    max_email_lookups: предохранитель по времени. Каждый добор — это поиск в
        Яндексе плюс 1-2 захода на сайт, примерно 3-5 секунд. При большом файле
        обогащаем только первые N строк, чтобы не упереться в таймаут интерфейса.
    enrich_suspicious: добирать ли строки, где почта есть, но выглядит личной.
        Пустые почты добираются всегда.
    """
    path = Path(file_path)
    if not path.exists():
        return {"error": f"Файл не найден: {file_path}"}

    try:
        wb = load_workbook(str(path))
        ws = wb.active
    except Exception as e:
        return {"error": f"Не удалось открыть файл: {e}"}

    col_adm = COL.get("ADM_NAME")
    col_addr = COL.get("ADRES")
    col_osn = COL.get("EMAIL_OSN")
    col_dop = COL.get("EMAIL_DOP")
    col_sub = COL.get("SUB_RF")
    col_site = COL.get("WEBSITE")
    if not all([col_adm, col_osn]):
        return {"error": "В шаблоне файла нет колонок ADM_NAME/EMAIL_OSN"}

    stats = {"rows": 0, "names_fixed": 0, "email_found": 0,
             "email_replaced": 0, "email_still_empty": 0, "looked_up": 0}

    # --- 1. Названия: быстро, без сети ---
    targets: list[int] = []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        adm = str(ws.cell(r, col_adm).value or "").strip()
        if not adm:
            continue
        stats["rows"] += 1

        addr = str(ws.cell(r, col_addr).value or "") if col_addr else ""
        fixed = _fix_name(adm, addr)
        if fixed != adm:
            ws.cell(r, col_adm, fixed)
            stats["names_fixed"] += 1
            logger.info(f"[postprocess] название дополнено: …{fixed[-40:]}")

        # нужен ли добор/проверка почты
        cur = clean_email(str(ws.cell(r, col_osn).value or ""))
        has_site = bool(str(ws.cell(r, col_site).value or "").strip()) if col_site else False
        if check_all:
            targets.append(r)
        elif not cur:
            targets.append(r)
        elif has_site:
            targets.append(r)
        elif enrich_suspicious and _looks_personal(cur):
            targets.append(r)

    if not targets:
        wb.save(str(path))
        logger.info(f"[postprocess] почта в порядке, названий исправлено: "
                    f"{stats['names_fixed']}")
        return stats

    # --- 2. Почта: с сетью, поэтому с предохранителем ---
    targets = targets[:max_email_lookups]
    logger.info(f"[postprocess] добираю почту по {len(targets)} строкам")
    _emit(f"Уточняю почту на официальных сайтах: {len(targets)} организаций…")

    for i, r in enumerate(targets, 1):
        adm = str(ws.cell(r, col_adm).value or "").strip()
        region = str(ws.cell(r, col_sub).value or "").strip() if col_sub else ""
        cur = clean_email(str(ws.cell(r, col_osn).value or ""))
        known_site = str(ws.cell(r, col_site).value or "").strip() if col_site else ""

        try:
            stats["looked_up"] += 1
            if known_site and re.search(r"https?://|[A-Za-zА-Яа-я0-9-]+\.[a-z]{2,}", known_site):
                site = known_site if known_site.lower().startswith("http") else "http://" + known_site
                emails = _emails_from_site(site)
            else:
                site, emails = _lookup_site_emails(adm, region)
            if not emails:
                if not cur:
                    stats["email_still_empty"] += 1
                continue

            found = _pick_from_site(emails)
            if not found:
                logger.info(f"[postprocess] {adm[:45]}: на {site} только личные адреса")
                if not cur:
                    stats["email_still_empty"] += 1
                continue

            if not cur:
                ws.cell(r, col_osn, found)
                stats["email_found"] += 1
                logger.info(f"[postprocess] {adm[:45]}: почта найдена {found}")
            elif found.strip().lower() != cur.strip().lower():
                # почта на сайте отличается — делаем её основной,
                # прежнюю НЕ теряем: уводим в дополнительные (без дублей)
                ws.cell(r, col_osn, found)
                if col_dop:
                    old_dop = str(ws.cell(r, col_dop).value or "").strip()
                    parts, seen = [], set()
                    for p in [x.strip() for x in f"{cur}, {old_dop}".split(",") if x.strip()]:
                        if p.lower() not in seen:
                            seen.add(p.lower())
                            parts.append(p)
                    ws.cell(r, col_dop, ", ".join(parts))
                stats["email_replaced"] += 1
                logger.info(f"[postprocess] {adm[:45]}: {cur} -> {found} (прежняя в доп.)")
            else:
                logger.info(f"[postprocess] {adm[:45]}: почта совпала ({cur})")

        except Exception as e:
            logger.warning(f"[postprocess] строка {r} ({adm[:35]}): {e}")
            if not cur:
                stats["email_still_empty"] += 1

        if i % 10 == 0:
            wb.save(str(path))
            _emit(f"Проверил почту у {i} из {len(targets)}…")

    wb.save(str(path))
    logger.info(
        f"[postprocess] готово: строк {stats['rows']}, названий исправлено "
        f"{stats['names_fixed']}, почт найдено {stats['email_found']}, "
        f"заменено {stats['email_replaced']}, осталось пустых "
        f"{stats['email_still_empty']}"
    )
    return stats