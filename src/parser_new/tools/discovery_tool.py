"""
tools/discovery_tool.py — сбор МНОГИХ контрагентов по параметрам поиска.

ЯДРО ТЗ: пользователь задаёт запрос ("строительные компании в Московской
области"), модуль сам определяет ОКВЭД + регион, тянет список организаций из
Checko, обогащает каждую полными реквизитами и ЗАПИСЫВАЕТ готовую таблицу
batch_<ts>.xlsx в output/latest — в формате, который понимает импортёр
получателей (parse_recipients_xlsx). Дубликаты по ИНН убираются.

Три точки входа:
  1. discover_companies(...)      -> dict с готовыми строками (чистые данные)
  2. discover_and_write(...)      -> собрать + записать xlsx, вернуть путь
                                     (используется chat() напрямую — путь B)
  3. discover_companies_tool      -> @tool для агента (тоже пишет файл — путь A)

Формат xlsx (одна строка техзаголовков; остальное импортёр кладёт в extra{}):
  ADM_NAME -> company, HEAD_FIO -> contact_name, EMAIL_OSN -> email,
  EMAIL_DOP -> email_fallback, SUB_RF -> region; ИНН/КПП/ОГРН/Телефон/Адрес/
  Отрасль/Статус/Сайт -> extra{} (видны и правятся в списке получателей).
"""
from __future__ import annotations

import time
from datetime import datetime

from langchain.tools import tool

from src.parser_new import config
from src.parser_new.logger import logger
from src.parser_new.tools.regions import get_region_code
from src.parser_new.memory.cache_memory import org_cache_get, org_cache_set

# Переиспользуем готовые части Checko — не дублируем ключ и парсер
from src.parser_new.tools.checko_tool import _checko_get, _parse_company, CHECKO_BASE


# ==============================
# КАРТА "ТЕМА ЗАПРОСА -> ОКВЭД"
# ==============================
# Стартовый набор частых секторов. Стемы подобраны под падежи/прилагательные
# ("строит" ловит "строительные/строительство"). Явный okved=... обходит карту.
SECTOR_OKVED: dict[str, list[str]] = {
    # строительство
    "строит":     ["41.20", "43.31", "43.32", "43.33", "43.39", "43.29", "43.21", "42.11"],
    "строй":      ["41.20", "43.31", "43.39", "43.29"],
    "ремонт":     ["43.31", "43.32", "43.33", "43.34", "43.39"],
    # медицина
    "медицин":    ["86.10", "86.21", "86.22", "86.23", "86.90"],
    "клиник":     ["86.10", "86.21", "86.22", "86.90"],
    "больниц":    ["86.10"],
    "поликлин":   ["86.21"],
    "стоматолог": ["86.23"],
    "врач":       ["86.21", "86.22", "86.10"],
    "аптек":      ["47.73"],
    "фармац":     ["21.20", "21.10"],
    # образование
    "образован":  ["85.14", "85.13", "85.41", "85.42"],
    "школ":       ["85.14", "85.13"],
    "детск":      ["85.11"],
    "университет":["85.22", "85.23"],
    # общепит
    "ресторан":   ["56.10"],
    "кафе":       ["56.10"],
    "общепит":    ["56.10", "56.29", "56.30"],
    "столов":     ["56.29"],
    # гостеприимство
    "гостиниц":   ["55.10"],
    "отель":      ["55.10"],
    "хостел":     ["55.20"],
    # торговля
    "магазин":    ["47.11", "47.19", "47.71", "47.99"],
    "рознич":     ["47.11", "47.19", "47.99"],
    "торгов":     ["47.19", "46.90"],
    "оптов":      ["46.90", "46.49", "46.69"],
    # IT
    "программн":  ["62.01", "62.02", "62.09"],
    "разработк":  ["62.01", "62.02"],
    "айти":       ["62.01", "62.02", "62.09", "63.11"],
    "софт":       ["62.01"],
    # транспорт/логистика
    "транспорт":  ["49.41", "49.31", "52.29"],
    "логистик":   ["52.29", "52.10"],
    "перевозк":   ["49.41", "49.39"],
    "грузопер":   ["49.41"],
    # услуги
    "юридическ":  ["69.10"],
    "юрист":      ["69.10"],
    "адвокат":    ["69.10"],
    "бухгалтер":  ["69.20"],
    "аудит":      ["69.20"],
    "недвижим":   ["68.31", "68.20", "68.10"],
    "риелтор":    ["68.31"],
    "риэлтор":    ["68.31"],
    # сельское хозяйство / производство
    "сельск":     ["01.11", "01.13", "01.41", "01.50"],
    "фермер":     ["01.11", "01.41"],
    "агро":       ["01.11", "01.13"],
    "производств":["25.11", "10.71", "13.20", "22.29"],
    "завод":      ["25.11", "24.10"],
    "фабрик":     ["13.20", "14.13"],
    # финансы
    "страхов":    ["65.11", "65.12"],
    "банк":       ["64.19"],
    "финанс":     ["64.19", "64.99"],
    # прочее
    "красот":     ["96.02"],
    "парикмахер": ["96.02"],
    "автосервис": ["45.20"],
    "автосалон":  ["45.11"],
    # авто-ремонт (частые формулировки)
    "авторемонт": ["45.20"],
    "автотехцентр": ["45.20"],
    "автомастерск": ["45.20"],
    "шиномонтаж": ["45.20"],
    "автомойк":   ["45.20"],
    "автозапчаст": ["45.31", "45.32"],
    # двери и дверная фурнитура
    "фурнитур":   ["25.72", "46.74", "47.52", "22.23"],
    "скобян":     ["25.72", "46.74", "47.52"],
    "замк":       ["25.72", "46.74"],
    "петл":       ["25.72"],
    "двер":       ["16.23", "25.12", "22.23", "46.73", "47.52"],
    "окн":        ["16.23", "22.23", "25.12", "43.32"],
}


def resolve_okved(query: str, explicit: str = "") -> list[str] | None:
    """Список листовых кодов ОКВЭД: явный приоритетнее, иначе по словам запроса.
    Явный может быть с запятыми ("41.20,43.31")."""
    if explicit and explicit.strip():
        return [c.strip() for c in explicit.split(",") if c.strip()]
    q = (query or "").lower()
    for kw, codes in SECTOR_OKVED.items():
        if kw in q:
            return list(codes)
    return None


# ==============================
# КАРТА "ГОРОД/РАЙОН -> СУБЪЕКТ РФ"
# ==============================
# Города и районы, которых НЕТ в карте субъектов (get_region_code): они входят
# в состав субъекта. Ключ — корень названия (нижний регистр), как пишет
# пользователь; значение — (код субъекта РФ, подстрока для фильтра по адресу).
# Расширяется одной строкой. ВНИМАНИЕ: фильтр по адресу идёт по подстроке,
# поэтому для «размытых» названий (напр. "Пушкин" поймает "ул. Пушкина")
# возможны ложные совпадения — добавляй такие с осторожностью.
CITY_TO_SUBJECT: dict[str, tuple[str, str]] = {
    "колпин":     ("78", "колпино"),     # Санкт-Петербург
    "кронштадт":  ("78", "кронштадт"),   # Санкт-Петербург
    "сестрорецк": ("78", "сестрорецк"),  # Санкт-Петербург
    "зеленоград": ("77", "зеленоград"),  # Москва
}


def resolve_region(region: str) -> tuple[str | None, str]:
    """Определяет (код субъекта РФ, подстрока-фильтр по адресу).
    - субъект РФ ("Санкт-Петербург")     -> (код, "")      без фильтра
    - город/район из карты ("Колпино")   -> (код субъекта, подстрока) с фильтром
    - иначе                              -> (None, "")
    """
    text = (region or "").strip()
    # сначала пробуем как субъект РФ
    code = get_region_code(text)
    if code:
        return code, ""
    # затем как город/район внутри субъекта
    key = text.lower()
    for city_root, (subj_code, addr_sub) in CITY_TO_SUBJECT.items():
        if city_root in key:
            return subj_code, addr_sub
    return None, ""


# ==============================
# ПОИСК СПИСКА ОРГАНИЗАЦИЙ
# ==============================

def _search_page(okved: str, region_code: str, page: int) -> list[dict]:
    """Одна страница выдачи Checko /search по ОКВЭД + регион."""
    params = {
        "by":     "okved",
        "obj":    "org",
        "query":  okved,     # существующий tool ищет по query при by=okved
        "okved":  okved,     # дублируем в отдельный параметр — если API его ждёт
        "region": region_code,
        "active": "true",    # только действующие (Checko ждёт true, не 1)
        "page":   page,
    }
    # DEBUG: что реально уходит в Checko (ключ не печатаем)
    resp = _checko_get("search", params)
    meta = resp.get("meta", {})
    if meta.get("status") == "error":
        logger.warning(f"[discovery] Checko search error: {meta.get('message')}")
        return []
    data = resp.get("data", {}) or {}
    records = data.get("Записи", []) or []
    return records


def _collect_inns(okved_codes: list[str], region_code: str, limit: int, oversample: int = 3) -> list[str]:
    """Собирает уникальные ИНН по ОКВЭД+региону, листая страницы до limit."""
    inns: list[str] = []
    seen: set[str] = set()
    target = max(limit * oversample, limit)   # берём с запасом под отсев недействующих
    max_pages = 6

    for code in okved_codes:                  # перебираем листовые коды ОКВЭД
        if len(inns) >= target:
            break
        for page in range(max_pages):
            try:
                records = _search_page(code, region_code, page)
            except Exception as e:
                logger.error(f"[discovery] Ошибка (код {code}, стр {page}): {e}")
                break
            if not records:
                break
            before = len(inns)
            for rec in records:
                inn = (rec.get("ИНН") or "").strip()
                if inn and inn not in seen:
                    seen.add(inn)
                    inns.append(inn)
            added = len(inns) - before
            if added == 0:                    # страница не дала нового — к следующему коду
                break
            if len(inns) >= target:
                break
            time.sleep(0.3)

    logger.info(f"[discovery] Кандидатов ИНН: {len(inns)} по {len(okved_codes)} кодам ОКВЭД")
    return inns[:target]


# ==============================
# ОБОГАЩЕНИЕ ОДНОЙ ОРГАНИЗАЦИИ
# ==============================

def _flat(value, *keys) -> str:
    """Checko часто отдаёт поле объектом ({АдресРФ:...}/{Наим:...}). Достаём строку.
    Если value уже строка — возвращаем как есть. keys — приоритет ключей."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, dict):
        for k in keys:
            v = value.get(k)
            if v:
                return str(v).strip()
        return ""
    return str(value).strip()


def _status_active(value) -> bool:
    """True, если организация действует (по тексту статуса)."""
    text = _flat(value, "Наим", "Код").lower()
    if not text:
        return True  # статус неизвестен — не отсеиваем
    bad = ["не действ", "ликвид", "прекращ", "реорганизац", "исключен"]
    return not any(b in text for b in bad)


def _enrich_one(inn: str) -> dict | None:
    """Тянет полные данные по ИНН и раскладывает в строку базы по ТЗ."""
    cached = org_cache_get(inn)
    if cached:
        parsed = cached
    else:
        try:
            resp = _checko_get("company", {"inn": inn})
        except Exception as e:
            logger.error(f"[discovery] /company {inn}: {e}")
            return None
        data = resp.get("data", {})
        if not data:
            return None
        parsed = _parse_company(data)
        org_cache_set(inn, parsed)

    phones = parsed.get("Телефоны") or []
    emails = parsed.get("Емэйл") or []

    return {
        "company":        parsed.get("НаимПолн") or parsed.get("НаимСокр", ""),
        "contact_name":   parsed.get("ГлаваФИО", ""),
        "email":          emails[0] if emails else "",
        "email_fallback": ", ".join(emails[1:]) if len(emails) > 1 else "",
        "inn":            parsed.get("ИНН", ""),
        "kpp":            parsed.get("КПП", ""),
        "ogrn":           parsed.get("ОГРН", ""),
        "phone":          phones[0] if phones else "",
        "phone2":         ", ".join(phones[1:]) if len(phones) > 1 else "",
        "address":        _flat(parsed.get("ЮрАдрес", ""), "АдресРФ", "Адрес"),
        "industry":       _flat(parsed.get("ОКВЭД", ""), "Наим", "Код"),   # <- раньше терялось
        "status":         _flat(parsed.get("Статус", ""), "Наим", "Код"),  # <- раньше глушилось промптом
        "_status_raw":    parsed.get("Статус", ""),
        "website":        parsed.get("ВебСайт", ""),
        "head_post":      parsed.get("ГлаваДолжность", ""),
        "source":         "Checko",
    }


# ==============================
# СБОР (чистые данные)
# ==============================

def discover_companies(query: str, region: str, okved: str = "", limit: int = 25) -> dict:
    """
    Собирает список контрагентов по параметрам поиска. Возвращает
    {"success", "rows", "okved", "region_code", "error"}.
    """
    if not config.CHECKO_API_KEY:
        return {"success": False, "rows": [], "error": "CHECKO_API_KEY не задан"}

    region_code, address_filter = resolve_region(region)
    if not region_code:
        return {"success": False, "rows": [],
                "error": f"Не удалось определить код региона для '{region}'"}

    okved_codes = resolve_okved(query, okved)
    if not okved_codes:
        return {"success": False, "rows": [],
                "error": "Не удалось определить отрасль (ОКВЭД) по запросу"}

    okved_str = ",".join(okved_codes)
    logger.info(f"[discovery] query={query!r} region={region_code} okved=[{okved_str}] limit={limit}")

    # email-фильтр (только с почтой) и фильтр по городу снижают выход,
    # поэтому берём кандидатов с запасом: сильнее, если фильтруем по городу.
    oversample = 8 if address_filter else 4
    inns = _collect_inns(okved_codes, region_code, limit, oversample=oversample)
    if not inns:
        return {"success": False, "rows": [], "okved": okved_str,
                "region_code": region_code,
                "error": "По заданным параметрам организации не найдены в Checko"}

    rows: list[dict] = []
    seen_inn: set[str] = set()
    for inn in inns:
        if len(rows) >= limit:       # набрали нужное число — стоп
            break
        row = _enrich_one(inn)
        if not row:
            continue
        raw = row.pop("_status_raw", "")
        if not _status_active(raw):          # только действующие
            continue
        if not (row.get("email") or "").strip():   # только компании с почтой
            continue
        # фильтр по городу/району внутри субъекта (напр. "Колпино")
        if address_filter and address_filter not in (row.get("address") or "").lower():
            continue
        key = (row.get("inn") or inn).strip()
        if key in seen_inn:
            continue
        seen_inn.add(key)
        row["region"] = region
        rows.append(row)
        time.sleep(0.2)

    logger.info(f"[discovery] Собрано {len(rows)} организаций из {len(inns)} ИНН")
    if address_filter and not rows:
        return {"success": False, "rows": [], "okved": okved_str,
                "region_code": region_code,
                "error": f"В '{region}' не найдено ни одной компании с почтой."}
    return {"success": True, "rows": rows,
            "okved": okved_str, "region_code": region_code}


# ==============================
# ЗАПИСЬ XLSX В ФОРМАТЕ ИМПОРТЁРА
# ==============================

# (заголовок, ключ строки). Первые пять — техключи, которые ловит parse_recipients_xlsx;
# остальные уходят в extra{} и показываются колонками в списке получателей.
_COLUMNS: list[tuple[str, str]] = [
    ("ADM_NAME",     "company"),
    ("HEAD_FIO",     "contact_name"),
    ("EMAIL_OSN",    "email"),
    ("EMAIL_DOP",    "email_fallback"),
    ("SUB_RF",       "region"),
    ("ИНН",          "inn"),
    ("КПП",          "kpp"),
    ("ОГРН",         "ogrn"),
    ("Телефон",      "phone"),
    ("Телефон_доп",  "phone2"),
    ("Адрес",        "address"),
    ("Отрасль",      "industry"),
    ("Статус",       "status"),
    ("Сайт",         "website"),
    ("Должность",    "head_post"),
    ("Источник",     "source"),
]


def _write_batch_xlsx(rows: list[dict]) -> str:
    """Пишет строки в output/latest/batch_<ts>.xlsx. Возвращает путь."""
    from openpyxl import Workbook

    latest_dir = config.OUTPUT_DIR / "latest"
    latest_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    path = latest_dir / f"batch_{ts}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.append([header for header, _ in _COLUMNS])          # строка техзаголовков
    for r in rows:
        ws.append([str(r.get(key, "") or "") for _, key in _COLUMNS])
    wb.save(path)

    logger.info(f"[discovery] Таблица записана: {path} ({len(rows)} строк)")
    return str(path)

def read_batch_rows(path: str) -> list[dict]:
    """Читает batch_*.xlsx обратно в строки {ключ: значение} по _COLUMNS.
    Нужно, чтобы построить ключи уже имеющихся строк для дедупа."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    data = list(ws.iter_rows(values_only=True))
    wb.close()
    if not data:
        return []
    headers = [str(h) if h is not None else "" for h in data[0]]
    key_by_col = {}
    for i, h in enumerate(headers):
        for col_header, key in _COLUMNS:
            if h == col_header:
                key_by_col[i] = key
                break
    rows = []
    for r in data[1:]:
        row = {key_by_col[i]: (r[i] if r[i] is not None else "")
               for i in key_by_col if i < len(r)}
        rows.append(row)
    return rows


def append_batch_xlsx(path: str, rows: list[dict]) -> int:
    """Дописывает строки в существующий batch-файл. Дубли по ИНН пропускает.
    Пишет во временную копию и атомарно заменяет оригинал."""
    import os, tempfile
    from pathlib import Path
    from openpyxl import load_workbook

    if not rows:
        return 0

    wb = load_workbook(path)
    ws = wb.worksheets[0]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    try:
        inn_col = headers.index("ИНН") + 1
    except ValueError:
        inn_col = None

    existing = set()
    if inn_col:
        for row in ws.iter_rows(min_row=2, max_col=inn_col, values_only=True):
            v = row[inn_col - 1]
            if v:
                existing.add(str(v).strip())

    added = 0
    for r in rows:
        inn = str(r.get("inn", "") or "").strip()
        if inn and inn in existing:
            continue
        ws.append([str(r.get(key, "") or "") for _, key in _COLUMNS])
        if inn:
            existing.add(inn)
        added += 1

    fd, tmp = tempfile.mkstemp(suffix=".xlsx", dir=str(Path(path).parent))
    os.close(fd)
    try:
        wb.save(tmp)
        os.replace(tmp, path)      # атомарная замена — оригинал не бьётся при обрыве
    finally:
        if os.path.exists(tmp):
            os.remove(tmp)
    return added


def discover_and_write(query: str, region: str, okved: str = "", limit: int = 25) -> dict:
    """
    Собирает список и СРАЗУ пишет batch_*.xlsx. Возвращает
    {"success", "path", "count", "error"}.  Используется chat() (путь B) и агентом.
    """
    result = discover_companies(query, region, okved, limit)
    if not result["success"]:
        return {"success": False, "path": None, "count": 0, "error": result["error"]}

    rows = result["rows"]
    if not rows:
        return {"success": False, "path": None, "count": 0,
                "error": "Ничего не собрано"}

    path = _write_batch_xlsx(rows)
    return {"success": True, "path": path, "count": len(rows),
            "okved": result.get("okved", ""), "error": None}


# ==============================
# ИНСТРУМЕНТ ДЛЯ АГЕНТА (путь A)
# ==============================

@tool
def discover_companies_tool(query: str, region: str, okved: str = "", limit: int = 25) -> str:
    """
    Собирает СПИСОК коммерческих организаций по параметрам поиска (не по файлу!)
    и записывает готовую таблицу для скачивания.
    Используй, когда пользователь просит найти КЛАСС контрагентов по сфере и региону,
    например: "строительные компании в Московской области", "частные клиники в Москве",
    "аптеки в Татарстане". НЕ используй для администраций МО.

    Сам определяет отрасль (ОКВЭД) по запросу, тянет список из Checko, обогащает
    каждую организацию (наименование, адрес, ИНН/КПП/ОГРН, телефоны, email осн.+доп.,
    руководитель, отрасль, статус), убирает дубликаты по ИНН и пишет таблицу.

    Параметры:
      query:  формулировка пользователя
      region: регион РФ ("Московская область", "Татарстан", "Москва")
      okved:  код ОКВЭД явно, если известен (иначе определяется автоматически)
      limit:  сколько организаций собрать (по умолчанию 25)
    """
    res = discover_and_write(query, region, okved, limit)
    if not res["success"]:
        return f"Не удалось собрать список: {res['error']}"
    return (f"Собрано организаций: {res['count']} (ОКВЭД {res.get('okved','')}). "
            f"Таблица готова для скачивания и импорта.")