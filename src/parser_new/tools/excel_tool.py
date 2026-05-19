"""
tools/excel_tool.py — работа с Excel файлами.

Поддерживает точную структуру файла для МО:
  Строка 1: заголовки (человекочитаемые)
  Строка 2: технические имена колонок
  Строка 3+: данные
"""
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_TEXT
from langchain.tools import tool

from src.parser_new import config
from src.parser_new.logger import logger


# ==============================
# СТРУКТУРА ФАЙЛА МО
# ==============================

# (col_index, header_row1, header_row2, width)
MO_COLUMNS = [
    (1,  "№",                          "ID",               6),
    (2,  "Субъект РФ",                 "SUB_RF",           20),
    (3,  "Муниципальный район",        "MUN_R_NAME",       25),
    (4,  "Муниципальное образование",  "MUN_NAME",         25),
    (5,  "Полное название администрации", "ADM_NAME",      35),
    (6,  "Адрес",                      "ADRES",            35),
    (7,  "Глава МО",                   "HEAD_FIO",         25),
    (8,  "Численность",                "POPULATION",       12),
    (9,  "Эл. Адрес (основной)",       "EMAIL_OSN",        25),
    (10, "Эл. Адрес (доп)",            "EMAIL_DOP",        25),
    (11, "Телефон",                    "TEL_OSN",          18),
    (12, "Телефон (доп)",              "TEL_DOP",          18),
    (13, "ИНН",                        "REQUISITES_INN",   15),
    (14, "КПП",                        "REQUISITES_KPP",   12),
    (15, "ОГРН",                       "REQUISITES_OGRN",  16),
    (16, "ОКПО",                       "REQUISITES_OKPO",  12),
    (17, "ОКТМО",                      "REQUISITES_OKTMO", 12),
    (18, "Статус",                     "STATUS",           15),
    (19, "Примечание",                 "NOTE",             30),
]

# Колонки реквизитов для объединения заголовка
REQUISITES_START = 13
REQUISITES_END   = 17

# Цвета
COLOR_HEADER1_BG  = "1F3864"
COLOR_HEADER1_FG  = "FFFFFF"
COLOR_HEADER2_BG  = "2E5FA3"
COLOR_HEADER2_FG  = "FFFFFF"
COLOR_REQUIS_BG   = "2E75B6"
COLOR_ALT_ROW     = "EBF3FB"


def _thin_border() -> Border:
    s = Side(style="thin", color="B0C4DE")
    return Border(left=s, right=s, top=s, bottom=s)


def _create_mo_header(ws) -> None:
    """Создаёт двухстрочную шапку файла МО с объединённой ячейкой Реквизиты."""

    b = _thin_border()
    font1  = Font(name="Arial", bold=True, color=COLOR_HEADER1_FG, size=10)
    fill1  = PatternFill("solid", start_color=COLOR_HEADER1_BG)
    font2  = Font(name="Arial", bold=True, color=COLOR_HEADER2_FG, size=9)
    fill2  = PatternFill("solid", start_color=COLOR_HEADER2_BG)
    align_c = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Строка 1
    for col_idx, h1, h2, width in MO_COLUMNS:
        cell = ws.cell(1, col_idx, h1)
        cell.font      = font1
        cell.fill      = fill1
        cell.alignment = align_c
        cell.border    = b
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Объединяем M1:Q1 → "Реквизиты"
    ws.merge_cells(
        start_row=1, start_column=REQUISITES_START,
        end_row=1,   end_column=REQUISITES_END,
    )
    req_cell = ws.cell(1, REQUISITES_START, "Реквизиты")
    req_cell.font      = font1
    req_cell.fill      = fill1
    req_cell.alignment = align_c
    req_cell.border    = b

    # Строка 2
    for col_idx, h1, h2, width in MO_COLUMNS:
        cell = ws.cell(2, col_idx, h2)
        cell.font      = font2
        cell.fill      = fill2
        cell.alignment = align_c
        cell.border    = b

    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(MO_COLUMNS))}2"


def _write_mo_row(ws, row_idx: int, record: dict) -> None:
    """Записывает одну строку данных МО."""
    fill_alt = PatternFill("solid", start_color=COLOR_ALT_ROW)
    b = _thin_border()
    font = Font(name="Arial", size=10)

    col_map = {h2: col_idx for col_idx, h1, h2, w in MO_COLUMNS}

    for col_idx, h1, h2, width in MO_COLUMNS:
        value = record.get(h2, "")
        cell  = ws.cell(row_idx, col_idx, value)
        cell.font      = font
        cell.border    = b
        cell.alignment = Alignment(vertical="center", wrap_text=False)
        if row_idx % 2 == 0:
            cell.fill = fill_alt

    ws.row_dimensions[row_idx].height = 18


def _next_id(ws) -> int:
    """Определяет следующий ID (максимальный существующий + 1)."""
    max_id = 0
    for row in ws.iter_rows(min_row=3, max_col=1, values_only=True):
        try:
            val = int(row[0] or 0)
            max_id = max(max_id, val)
        except (ValueError, TypeError):
            pass
    return max_id + 1


def _save_file(wb, filename: str) -> str:
    """Сохраняет в latest/ и archive/."""
    latest  = config.OUTPUT_DIR / "latest" / filename
    arc_dir = config.OUTPUT_DIR / "archive" / datetime.now().strftime("%Y-%m-%d")
    arc_dir.mkdir(parents=True, exist_ok=True)
    wb.save(latest)
    wb.save(arc_dir / filename)
    return str(latest)


def _make_filename(prefix: str = "mo") -> str:
    return f"{prefix}_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"


# ==============================
# ИНСТРУМЕНТЫ ДЛЯ АГЕНТА
# ==============================

@tool
def read_excel_tool(file_path: str) -> str:
    """
    Читает Excel файл МО и возвращает ТОЛЬКО строки где нужно заполнить данные.
    Пустой считается строка где заполнены SUB_RF, MUN_R_NAME, MUN_NAME
    но пусты реквизиты (REQUISITES_INN) или контакты (TEL_OSN, EMAIL_OSN).
    Используй в начале работы если пользователь прислал файл.
    Возвращает номера строк и данные только незаполненных записей.
    """
    try:
        path = Path(file_path)
        if not path.exists():
            return f"Файл не найден: {file_path}"

        wb = load_workbook(str(path), data_only=True)
        ws = wb.active

        # Читаем заголовки из строки 2
        headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
        col_map = {h: i + 1 for i, h in enumerate(headers) if h}

        total_rows = ws.max_row - 2
        empty_rows = []

        for row_idx in range(3, ws.max_row + 1):
            def val(col_name):
                col = col_map.get(col_name)
                if not col:
                    return ""
                v = ws.cell(row_idx, col).value
                return str(v).strip() if v else ""

            # Строка считается требующей обработки если:
            # есть название МО но нет ИНН или нет телефона/email
            has_name = val("MUN_NAME") or val("ADM_NAME")
            has_inn  = val("REQUISITES_INN")
            has_tel  = val("TEL_OSN")
            has_email = val("EMAIL_OSN")

            if has_name and not (has_inn and has_tel and has_email):
                empty_rows.append({
                    "excel_row":    row_idx,
                    "record_num":   row_idx - 2,
                    "SUB_RF":       val("SUB_RF"),
                    "MUN_R_NAME":   val("MUN_R_NAME"),
                    "MUN_NAME":     val("MUN_NAME"),
                    "ADM_NAME":     val("ADM_NAME"),
                    "REQUISITES_INN":  val("REQUISITES_INN"),
                    "TEL_OSN":      val("TEL_OSN"),
                    "EMAIL_OSN":    val("EMAIL_OSN"),
                    "missing": ", ".join([
                        x for x, v in [
                            ("ИНН",    has_inn),
                            ("Телефон", has_tel),
                            ("Email",  has_email),
                        ] if not v
                    ])
                })

        if not empty_rows:
            return f"Файл: {path.name} | Строк: {total_rows} | Все строки заполнены!"

        lines = [
            f"Файл: {path.name}",
            f"Всего строк: {total_rows} | Требуют заполнения: {len(empty_rows)}",
            "",
            "Незаполненные строки:",
        ]
        for r in empty_rows:
            lines.append(
                f"  Строка {r['record_num']} (Excel {r['excel_row']}): "
                f"{r['SUB_RF']} | {r['MUN_R_NAME']} | {r['MUN_NAME']} "
                f"| Не хватает: {r['missing']}"
            )

        return "\n".join(lines)

    except Exception as e:
        logger.error(f"[excel/read] {e}")
        return f"Не удалось прочитать файл: {e}"


@tool
def write_excel_tool(records_json: str, filename_prefix: str = "mo") -> str:
    """
    Создаёт новый Excel файл с данными МО.
    Формирует файл строго по структуре: двойная шапка, данные с 3-й строки.

    records_json — список словарей в формате JSON.
    Ключи словаря — технические имена колонок из строки 2:
    ID, SUB_RF, MUN_R_NAME, MUN_NAME, ADM_NAME, ADRES, HEAD_FIO,
    POPULATION, EMAIL_OSN, EMAIL_DOP, TEL_OSN, TEL_DOP,
    REQUISITES_INN, REQUISITES_KPP, REQUISITES_OGRN, REQUISITES_OKPO,
    REQUISITES_OKTMO, STATUS

    ID заполнять не нужно — проставляется автоматически.
    """
    try:
        records = json.loads(records_json)
        if not records:
            return "Нет данных для сохранения"

        wb = Workbook()
        ws = wb.active
        ws.title = "МО"
        ws.sheet_view.showGridLines = True

        _create_mo_header(ws)

        for i, rec in enumerate(records, start=1):
            rec["ID"] = i
            _write_mo_row(ws, i + 2, rec)

        filename = _make_filename(filename_prefix)
        path     = _save_file(wb, filename)

        return f"Файл создан: {path} | Записей: {len(records)}"

    except json.JSONDecodeError:
        return "Ошибка: records_json должен быть валидным JSON"
    except Exception as e:
        logger.error(f"[excel/write] {e}")
        return f"Не удалось создать файл: {e}"


@tool
def append_excel_tool(file_path: str, records_json: str) -> str:
    """
    Дополняет существующий файл МО новыми строками.
    Не добавляет дубли — проверяет по колонке REQUISITES_INN.
    Если ИНН уже есть в файле — строка пропускается.

    file_path    — путь к существующему файлу
    records_json — новые записи в формате JSON (те же ключи что в write_excel_tool)
    """
    try:
        path = Path(file_path)
        if not path.exists():
            return f"Файл не найден: {file_path}"

        new_records = json.loads(records_json)
        if not new_records:
            return "Нет новых данных"

        wb = load_workbook(path)
        ws = wb.active

        # Строим индекс существующих ИНН
        headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
        inn_col = next(
            (i + 1 for i, h in enumerate(headers) if h == "REQUISITES_INN"),
            None
        )
        existing_inns = set()
        if inn_col:
            for row in ws.iter_rows(min_row=3, max_col=inn_col, values_only=True):
                if row[inn_col - 1]:
                    existing_inns.add(str(row[inn_col - 1]).strip())

        added = skipped = 0
        for rec in new_records:
            inn = str(rec.get("REQUISITES_INN", "")).strip()
            if inn and inn in existing_inns:
                skipped += 1
                continue

            next_row    = ws.max_row + 1
            rec["ID"]   = _next_id(ws)
            _write_mo_row(ws, next_row, rec)
            if inn:
                existing_inns.add(inn)
            added += 1

        wb.save(path)
        arc_dir = config.OUTPUT_DIR / "archive" / datetime.now().strftime("%Y-%m-%d")
        arc_dir.mkdir(parents=True, exist_ok=True)
        wb.save(arc_dir / path.name)

        return f"Добавлено: {added} | Пропущено дублей: {skipped} | Файл: {path}"

    except json.JSONDecodeError:
        return "Ошибка: records_json должен быть валидным JSON"
    except Exception as e:
        logger.error(f"[excel/append] {e}")
        return f"Не удалось дополнить файл: {e}"


@tool
def update_excel_tool(file_path: str, updates_json: str) -> str:
    """
    Обновляет строки в файле МО и сохраняет результат в output/latest/.
    Ищет строки по номеру Excel строки (поле "excel_row") или по MUN_NAME.
    Незатронутые колонки остаются нетронутыми.
    Входной файл НЕ изменяется — создаётся новый файл в output/latest/.

    file_path    — путь к исходному файлу
    updates_json — список записей JSON. Каждая запись ДОЛЖНА содержать
                   поле "excel_row" (номер строки из read_excel_tool)
                   и поля которые нужно заполнить.
    """
    try:
        src_path = Path(file_path)
        if not src_path.exists():
            return f"Файл не найден: {file_path}"

        updates = json.loads(updates_json)
        if not updates:
            return "Нет данных для обновления"

        # Копируем файл в output/latest/ — не трогаем оригинал
        out_filename = _make_filename("updated")
        out_path     = config.OUTPUT_DIR / "latest" / out_filename
        arc_dir      = config.OUTPUT_DIR / "archive" / datetime.now().strftime("%Y-%m-%d")
        arc_dir.mkdir(parents=True, exist_ok=True)

        import shutil
        shutil.copy2(str(src_path), str(out_path))

        wb = load_workbook(str(out_path))
        ws = wb.active

        # Индекс колонок
        headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
        col_map = {h: i + 1 for i, h in enumerate(headers) if h}

        # Также строим индекс MUN_NAME → excel_row как запасной вариант
        mun_col = col_map.get("MUN_NAME")
        mun_map = {}
        if mun_col:
            for row_idx in range(3, ws.max_row + 1):
                v = ws.cell(row_idx, mun_col).value
                if v:
                    mun_map[str(v).strip().lower()] = row_idx

        updated = 0
        not_found = []

        for rec in updates:
            # Определяем строку — по excel_row или по MUN_NAME
            row_idx = rec.get("excel_row")
            if not row_idx:
                mun = str(rec.get("MUN_NAME", "")).strip().lower()
                row_idx = mun_map.get(mun)
            if not row_idx:
                not_found.append(rec.get("MUN_NAME", "неизвестно"))
                continue

            row_idx = int(row_idx)
            for field, value in rec.items():
                if field in ("excel_row", "record_num"):
                    continue
                if field in col_map and value:
                    # Не перезаписываем уже заполненные ячейки
                    existing = ws.cell(row_idx, col_map[field]).value
                    if not existing:
                        ws.cell(row_idx, col_map[field], value)
            updated += 1

        wb.save(str(out_path))
        shutil.copy2(str(out_path), str(arc_dir / out_filename))

        result = f"Обновлено строк: {updated} | Файл: {out_path}"
        if not_found:
            result += f" | Не найдено строк: {', '.join(not_found[:5])}"
        return result

    except json.JSONDecodeError:
        return "Ошибка: updates_json должен быть валидным JSON"
    except Exception as e:
        logger.error(f"[excel/update] {e}")
        return f"Не удалось обновить файл: {e}"