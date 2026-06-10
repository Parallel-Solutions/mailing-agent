"""
admin_levels.py — заполнение столбцов B/C/D (субъект / район / поселение).

Пост-проход после batch_processor: разбирает название из E и/или юр. адрес из F
на уровни и дозаполняет ТОЛЬКО пустые ячейки B/C/D.

Правила:
  - Ликвидированные (пометка в NOTE/STATUS) пропускаем целиком — ни LLM, ни запись.
  - B (субъект) обязателен для действующих. Источник: сначала юр. адрес (надёжно),
    затем регион из названия. Канон берётся из списка ОКТМО (resolve_region).
  - C (район/округ): из названия (LLM); если C пуст, но есть D — из юр. адреса.
  - D (поселение): только из названия (в адресе его нет).
  - Заполняем только пустое, ничего не перезатираем.

Разбор названия (region/district/settlement) делает LLM пачками; регион и район
из адреса — детерминированно.
"""
from __future__ import annotations

import json
import re

from openpyxl import load_workbook

try:
    from src.parser_new import config
    from src.parser_new.logger import logger
    from src.parser_new.tools.oktmo_tool import resolve_region
except ImportError:  # запуск из каталога parser_new
    import config
    from logger import logger
    from tools.oktmo_tool import resolve_region


# колонки (как в batch_processor.COL)
COL_SUB_RF = 2     # B — субъект
COL_MUN_R = 3      # C — район/округ
COL_MUN = 4        # D — поселение
COL_ADM = 5        # E — название организации
COL_ADRES = 6      # F — юридический адрес
COL_STATUS = 18    # R — статус
COL_NOTE = 19      # S — примечание
DATA_START_ROW = 3

BATCH = 25         # сколько названий за один вызов LLM


_SYSTEM = """Ты разбираешь юридические названия российских администраций, учреждений и ведомств на уровни.
Верни ТОЛЬКО JSON-массив, по одному объекту на каждое название, в ТОМ ЖЕ порядке.
Поля каждого объекта:
  "region"     — субъект РФ в именительном падеже («Костромская область», «Краснодарский край», «Республика Татарстан»), или null, если в названии его нет;
  "district"   — район или муниципальный/городской округ ЦЕЛИКОМ в именительном падеже, или null;
  "settlement" — сельское/городское поселение ЦЕЛИКОМ в именительном падеже, или null.
Правила:
  - всё в именительном падеже;
  - сохраняй полный тип так, как он есть в названии: «муниципальный район», «район», «муниципальный округ», «городской округ», «сельское поселение», «городское поселение»;
  - если уровня в названии нет — ставь null, НЕ выдумывай (субъект и район дозаполним из адреса отдельно);
  - для ведомства/учреждения, где есть только поселение, district оставь null, если района в названии нет.
Никакого текста кроме JSON."""

_FEWSHOT_IN = [
    "АДМИНИСТРАЦИЯ МУНИЦИПАЛЬНОГО ОБРАЗОВАНИЯ ПРИМОРСКО-АХТАРСКИЙ МУНИЦИПАЛЬНЫЙ ОКРУГ КРАСНОДАРСКОГО КРАЯ",
    "АДМИНИСТРАЦИЯ ИЛЬИНСКОГО СЕЛЬСКОГО ПОСЕЛЕНИЯ КОЛОГРИВСКОГО МУНИЦИПАЛЬНОГО РАЙОНА КОСТРОМСКОЙ ОБЛАСТИ",
    "ДЕПАРТАМЕНТ СТРОИТЕЛЬСТВА КОСТРОМСКОЙ ОБЛАСТИ",
    'МУНИЦИПАЛЬНОЕ КАЗЕННОЕ УЧРЕЖДЕНИЕ "ЮЖНОЕ" ИЛЬИНСКОГО СЕЛЬСКОГО ПОСЕЛЕНИЯ НОВОПОКРОВСКОГО РАЙОНА',
    'МУНИЦИПАЛЬНОЕ УЧРЕЖДЕНИЕ "ЦЕНТРАЛИЗОВАННАЯ БУХГАЛТЕРИЯ ИЛЬИНСКОГО СЕЛЬСКОГО ПОСЕЛЕНИЯ"',
    'ОБЩЕСТВО С ОГРАНИЧЕННОЙ ОТВЕТСТВЕННОСТЬЮ "НАУЧНО-ПРОИЗВОДСТВЕННОЕ ПРЕДПРИЯТИЕ "ГАРАНТ-СЕРВИС"',
]
_FEWSHOT_OUT = [
    {"region": "Краснодарский край", "district": "Приморско-Ахтарский муниципальный округ", "settlement": None},
    {"region": "Костромская область", "district": "Кологривский муниципальный район", "settlement": "Ильинское сельское поселение"},
    {"region": "Костромская область", "district": None, "settlement": None},
    {"region": None, "district": "Новопокровский район", "settlement": "Ильинское сельское поселение"},
    {"region": None, "district": None, "settlement": "Ильинское сельское поселение"},
    {"region": None, "district": None, "settlement": None},
]


_llm = None


def _get_llm():
    global _llm
    if _llm is None:
        from langchain_openai import ChatOpenAI
        _llm = ChatOpenAI(
            model=config.AGENT_MODEL,
            api_key=config.ANTHROPIC_API_KEY,
            base_url=config.LLM_BASE_URL,
            temperature=0,
            max_tokens=4096,
        )
    return _llm


def _empty(v) -> bool:
    return not str(v or "").strip()


def _is_liquidated(note, status) -> bool:
    blob = f"{note or ''} {status or ''}".lower()
    return "ликвид" in blob or "не действ" in blob


# ---------- разбор юр. адреса (детерминированно) ----------

_INDEX_RE = re.compile(r"^\d{5,6}$")


def _addr_parts(address) -> list[str]:
    return [p.strip() for p in str(address or "").split(",") if p.strip()]


def _region_from_address(address) -> str:
    """Регион из юр. адреса — первый компонент после индекса."""
    parts = _addr_parts(address)
    if parts and _INDEX_RE.match(parts[0]):
        parts = parts[1:]
    return parts[0] if parts else ""


def _district_from_address(address) -> str:
    """Район из юр. адреса — компонент со словом 'район' или 'р-н'."""
    for p in _addr_parts(address):
        low = p.lower()
        if "район" in low or "р-н" in low:
            return re.sub(r"\bр-?н\b", "район", p, flags=re.IGNORECASE).strip()
    return ""


def _normalize_region(raw: str) -> str:
    """Чистое имя субъекта в именительном падеже.
    Адрес и LLM уже дают регион как надо («Мурманская область»), поэтому
    resolve_region НЕ используем — он возвращал метку ОКТМО
    «Муниципальные образования Мурманской области», это и был мусор в B."""
    s = re.sub(r"\s+", " ", str(raw or "").strip(" .,;"))
    if not s:
        return ""
    # срежем метку ОКТМО, если вдруг просочилась
    s = re.sub(r"(?i)^муниципальн\w+\s+образовани\w+\s+", "", s).strip()
    # частые сокращения из адресов
    s = re.sub(r"(?i)\bобл\.?(?=$|\s)", "область", s)
    s = re.sub(r"(?i)\bресп\.?(?=\s|$)", "Республика", s)
    return s.strip(" .,;")

# ---------- разбор названия (LLM) ----------

def _extract_json_array(text: str) -> list:
    t = re.sub(r"```(?:json)?|```", "", text).strip()
    try:
        return json.loads(t)
    except json.JSONDecodeError:
        m = re.search(r"\[.*\]", t, re.DOTALL)
        if not m:
            raise
        return json.loads(m.group(0))


def _parse_batch(names: list[str]) -> list[dict]:
    from langchain_core.messages import SystemMessage, HumanMessage

    example = (
        "Пример.\nВход:\n" + json.dumps(_FEWSHOT_IN, ensure_ascii=False)
        + "\nВыход:\n" + json.dumps(_FEWSHOT_OUT, ensure_ascii=False)
    )
    task = "Теперь разбери эти названия:\n" + json.dumps(names, ensure_ascii=False)

    resp = _get_llm().invoke([
        SystemMessage(content=_SYSTEM),
        HumanMessage(content=example + "\n\n" + task),
    ])
    content = resp.content if isinstance(resp.content, str) else str(resp.content)
    data = _extract_json_array(content)
    if not isinstance(data, list) or len(data) != len(names):
        got = len(data) if isinstance(data, list) else "?"
        raise ValueError(f"LLM вернул {got} объектов вместо {len(names)}")
    return data


# ---------- основной проход ----------

def fill_admin_levels(file_path: str) -> dict:
    """Дозаполняет B/C/D. Ликвидированных пропускает. Возвращает статистику."""
    wb = load_workbook(file_path)
    ws = wb.active

    targets, skipped_liq = [], 0
    for row in range(DATA_START_ROW, ws.max_row + 1):
        name = ws.cell(row, COL_ADM).value
        if _empty(name):
            continue
        if _is_liquidated(ws.cell(row, COL_NOTE).value, ws.cell(row, COL_STATUS).value):
            skipped_liq += 1
            continue
        if not _empty(ws.cell(row, COL_SUB_RF).value) \
           and not _empty(ws.cell(row, COL_MUN_R).value) \
           and not _empty(ws.cell(row, COL_MUN).value):
            continue
        targets.append((row, str(name).strip()))

    if not targets:
        logger.info(f"[levels] Нечего заполнять (ликвидированных пропущено: {skipped_liq})")
        return {"processed": 0, "filled_b": 0, "filled_c": 0, "filled_d": 0, "skipped_liquidated": skipped_liq}

    logger.info(f"[levels] Разбираю {len(targets)} названий пачками по {BATCH} "
                f"(ликвидированных пропущено: {skipped_liq})")

    fb = fc = fd = 0
    total_batches = (len(targets) + BATCH - 1) // BATCH
    for i in range(0, len(targets), BATCH):
        chunk = targets[i:i + BATCH]
        names = [n for _, n in chunk]
        logger.info(f"[levels] Пачка {i // BATCH + 1}/{total_batches} ({len(names)} названий)...")
        try:
            parsed = _parse_batch(names)
        except Exception as e:
            logger.warning(f"[levels] Пачка {i // BATCH + 1} не разобрана: {e}")
            continue

        for (row, _name), p in zip(chunk, parsed):
            if not isinstance(p, dict):
                continue
            region_name = str(p.get("region") or "").strip()
            district_name = str(p.get("district") or "").strip()
            settlement_name = str(p.get("settlement") or "").strip()
            address = ws.cell(row, COL_ADRES).value

            # D — поселение (только из названия)
            if settlement_name and _empty(ws.cell(row, COL_MUN).value):
                ws.cell(row, COL_MUN).value = settlement_name
                fd += 1

            # C — район/округ: из названия; если пусто, но есть D — из адреса
            if _empty(ws.cell(row, COL_MUN_R).value):
                c_val = district_name
                if not c_val and not _empty(ws.cell(row, COL_MUN).value):
                    c_val = _district_from_address(address)
                if c_val:
                    ws.cell(row, COL_MUN_R).value = c_val
                    fc += 1

            if _empty(ws.cell(row, COL_SUB_RF).value):
                b_val = _normalize_region(_region_from_address(address) or region_name)
                if b_val:
                    ws.cell(row, COL_SUB_RF).value = b_val
                    fb += 1

    wb.save(file_path)
    stats = {"processed": len(targets), "filled_b": fb, "filled_c": fc,
             "filled_d": fd, "skipped_liquidated": skipped_liq}
    logger.info(f"[levels] Готово: {stats}")
    return stats