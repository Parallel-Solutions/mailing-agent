"""Campaign CRUD, recipients, schedule, launch, pause/resume/cancel."""

from __future__ import annotations

import csv
import io
import re
import uuid
from datetime import datetime, timedelta, timezone
from typing import Any

from sqlalchemy import delete, func, select

from src.campaigns.schedule_planner import plan_batches
from src.campaigns.state import (
    CAMPAIGN_STATUSES,
    CampaignStateConflict,
    active_work_counts,
    allowed_actions,
    recipient_metrics,
    recipient_metrics_many,
    reconcile_campaign_state,
    transition_campaign_status,
)
from src.campaigns.suppression_service import is_email_suppressed_for_import
from src.security.company_access import apply_owner_filter, can_access_owner
from src.infra.db import session_scope
from src.infra.models import (
    Campaign,
    CampaignBatch,
    CampaignRecipient,
    CampaignSchedule,
    DeliveryAttempt,
    BackgroundTask,
    SmtpMailbox,
)

EMAIL_RE = re.compile(r"^[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}$")

CORE_RECIPIENT_COLUMNS = ("company", "contact_name", "email", "email_fallback", "region")

_RECIPIENT_ROW_RESERVED = frozenset(
    {
        "email",
        "email_fallback",
        "company",
        "contact_name",
        "contact",
        "region",
        "source",
        "excluded",
        "row_index",
        "extra",
        "validation_status",
    }
)

_CORE_FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "company": (
        "company",
        "компания",
        "organization",
        "adm_name",
        "mun_name",
        "полное название администрации",
        "муниципальное образование",
        "полное наименование",
        "сокращенное наименование",
        "наименование",
        "организация",
    ),
    "contact_name": (
        "contact",
        "contact_name",
        "контакт",
        "head_fio",
        "глава мо",
        "руководитель",
        "фio руководителя",
        "фио руководителя",
    ),
    "email": ("email", "e-mail", "почта", "email_osn", "эл. адрес (основной)"),
    "email_fallback": ("email_fallback", "email2", "email_dop", "эл. адрес (доп)"),
    "region": ("region", "регион", "sub_rf", "субъект рф"),
}

def _now() -> datetime:
    return datetime.now(timezone.utc)


def _schedule_requires_immediate_start(schedule: CampaignSchedule, *, now: datetime | None = None) -> bool:
    """Start now when there is no valid future slot (past time, outside window, wrong weekday)."""
    if schedule.send_immediately or schedule.start_at is None:
        return True
    from src.campaigns.schedule_planner import is_schedule_start_allowed

    clock = now or _now()
    start_at = schedule.start_at
    if start_at.tzinfo is None:
        start_at = start_at.replace(tzinfo=timezone.utc)
    else:
        start_at = start_at.astimezone(timezone.utc)
    if start_at <= clock:
        return True
    return not is_schedule_start_allowed(
        start_at,
        timezone_name=str(schedule.timezone or "Europe/Moscow"),
        weekdays=list(schedule.weekdays or []),
        time_windows=list(schedule.time_windows or []),
    )


def _new_id() -> str:
    return str(uuid.uuid4())


def _min_positive(*values: int) -> int:
    """Return the tightest positive limit; 0 means unlimited."""
    positives = [int(value) for value in values if int(value or 0) > 0]
    return min(positives) if positives else 0


def _connection_rate_limits(session: Any, mailbox_id: str | None) -> tuple[int, int]:
    if not mailbox_id:
        return 0, 0
    row = session.get(SmtpMailbox, mailbox_id)
    if row is None:
        return 0, 0
    return int(row.max_per_hour or 0), int(row.max_per_day or 0)


def _pool_rate_limits(session: Any, connection_ids: list[str]) -> tuple[int, int]:
    hour_limits: list[int] = []
    day_limits: list[int] = []
    for connection_id in connection_ids:
        hour, day = _connection_rate_limits(session, connection_id)
        if hour > 0:
            hour_limits.append(hour)
        if day > 0:
            day_limits.append(day)
    return (sum(hour_limits) if hour_limits else 0, sum(day_limits) if day_limits else 0)


def _campaign_connection_ids(row: Campaign) -> list[str]:
    from src.campaigns.connection_service import campaign_connection_ids

    return campaign_connection_ids(row)


def _apply_sender_fields(row: Campaign, data: dict[str, Any]) -> None:
    from src.campaigns.connection_service import normalize_connection_ids, resolve_connection

    if "connection_ids" in data and data["connection_ids"] is not None:
        ids = normalize_connection_ids(list(data["connection_ids"]))
    elif "smtp_mailbox_id" in data and data.get("smtp_mailbox_id"):
        ids = normalize_connection_ids(None, str(data["smtp_mailbox_id"]))
    else:
        return

    row.connection_ids = ids
    if ids:
        row.smtp_mailbox_id = ids[0]
        connection = resolve_connection(ids[0], row.owner_username)
        row.transport = connection.transport
    else:
        row.smtp_mailbox_id = None
        row.transport = "smtp"


def _effective_rate_limits(
    *,
    schedule_max_per_hour: int,
    schedule_max_per_day: int,
    connection_max_per_hour: int,
    connection_max_per_day: int,
) -> tuple[int, int]:
    return (
        _min_positive(schedule_max_per_hour, connection_max_per_hour),
        _min_positive(schedule_max_per_day, connection_max_per_day),
    )


def _validate_email(value: str, email_fallback: str = "") -> str:
    from src.campaigns.recipient_email_service import validate_email_field

    return validate_email_field(value, email_fallback)


def extract_recipient_columns(rows: list[dict[str, Any]]) -> list[str]:
    columns: list[str] = list(CORE_RECIPIENT_COLUMNS)
    seen = set(columns)
    for row in rows:
        for key in row:
            normalized = str(key or "").strip().lower()
            if not normalized or normalized in _RECIPIENT_ROW_RESERVED or normalized in seen:
                continue
            seen.add(normalized)
            columns.append(normalized)
        for key in (row.get("extra") or {}).keys():
            normalized = str(key or "").strip().lower()
            if normalized and normalized not in seen:
                seen.add(normalized)
                columns.append(normalized)
    return columns


def _invalidate_variable_mapping(camp: Campaign) -> None:
    draft = dict(camp.draft_payload or {})
    draft["mapping_confirmed"] = False
    draft.pop("mapping_confirmed_at", None)
    camp.draft_payload = draft


def campaign_to_dict(
    row: Campaign,
    *,
    include_draft: bool = True,
    metrics: dict[str, Any] | None = None,
    has_active_work: bool | None = None,
) -> dict[str, Any]:
    campaign_metrics = metrics or {
        "success_count": int(row.sent_count or 0),
        "skipped_count": 0,
        "failed_recipient_count": 0,
        "processed_count": int(row.sent_count or 0),
        "pending_count": max(0, int(row.total_count or 0) - int(row.sent_count or 0)),
        "attempt_count": int(row.sent_count or 0) + int(row.error_count or 0),
        "attempt_error_count": int(row.error_count or 0),
        "progress": (
            round(100.0 * int(row.sent_count or 0) / int(row.total_count or 0), 1)
            if row.total_count
            else 0.0
        ),
        "success_rate": (
            round(100.0 * int(row.sent_count or 0) / int(row.total_count or 0), 1)
            if row.total_count
            else 0.0
        ),
    }
    payload = {
        "id": row.id,
        "owner_username": row.owner_username,
        "job_id": row.job_id,
        "name": row.name,
        "work_type": row.work_type,
        "document_mode": row.document_mode,
        "mail_subject": row.mail_subject,
        "description": row.description,
        "send_scenario": row.send_scenario,
        "tags": list(row.tags or []),
        "internal_comment": row.internal_comment,
        "status": row.status,
        "smtp_mailbox_id": row.smtp_mailbox_id,
        "connection_ids": list(_campaign_connection_ids(row)),
        "transport": row.transport,
        "email_template_id": row.email_template_id,
        "kp_template_id": row.kp_template_id,
        "contract_template_id": row.contract_template_id,
        "audience_id": row.audience_id,
        "email_chain_id": row.email_chain_id,
        "sent_count": row.sent_count,
        "total_count": row.total_count,
        "error_count": row.error_count,
        "success_count": int(campaign_metrics["success_count"]),
        "skipped_count": int(campaign_metrics["skipped_count"]),
        "failed_recipient_count": int(campaign_metrics["failed_recipient_count"]),
        "processed_count": int(campaign_metrics["processed_count"]),
        "pending_count": int(campaign_metrics["pending_count"]),
        "attempt_count": int(campaign_metrics["attempt_count"]),
        "attempt_error_count": int(campaign_metrics["attempt_error_count"]),
        "success_rate": float(campaign_metrics["success_rate"]),
        "allowed_actions": allowed_actions(
            row,
            campaign_metrics,
            has_active_work=has_active_work,
        ),
        "archived": bool(row.archived),
        "launched_at": row.launched_at.isoformat() if row.launched_at else None,
        "completed_at": row.completed_at.isoformat() if row.completed_at else None,
        "created_at": row.created_at.isoformat() if row.created_at else None,
        "updated_at": row.updated_at.isoformat() if row.updated_at else None,
        "progress": float(campaign_metrics["progress"]),
    }
    if include_draft:
        payload["draft_payload"] = dict(row.draft_payload or {})
    return payload


def create_campaign(owner_username: str, data: dict[str, Any] | None = None) -> dict[str, Any]:
    data = data or {}
    campaign_id = _new_id()
    with session_scope() as session:
        job_id = f"job-{campaign_id.replace('-', '')[:12]}"
        from src.jobs.access import assign_job_owner
        from src.jobs.storage import resolve_job_paths
        from src.security.auth import Principal

        resolve_job_paths(job_id).ensure_dirs()
        assign_job_owner(job_id, Principal(owner_username, "default", "user"), overwrite=True)

        row = Campaign(
            id=campaign_id,
            owner_username=owner_username,
            job_id=job_id,
            name=str(data.get("name") or "Черновик рассылки"),
            work_type=str(data.get("work_type") or ""),
            document_mode=str(data.get("document_mode") or "kp"),
            mail_subject=str(data.get("mail_subject") or ""),
            description=str(data.get("description") or ""),
            send_scenario=str(data.get("send_scenario") or "consent_then_materials"),
            tags=list(data.get("tags") or []),
            internal_comment=str(data.get("internal_comment") or ""),
            status="draft",
            smtp_mailbox_id=data.get("smtp_mailbox_id"),
            connection_ids=[],
            transport=str(data.get("transport") or "smtp"),
            draft_payload=dict(data.get("draft_payload") or data),
        )
        session.add(row)
        session.flush()
        _apply_sender_fields(row, data)
        session.flush()
        schedule = CampaignSchedule(
            id=_new_id(),
            campaign_id=campaign_id,
            send_immediately=True,
            timezone="Europe/Moscow",
            weekdays=[0, 1, 2, 3, 4],
            time_windows=[{"start": "09:00", "end": "18:00"}],
            batch_size=25,
            interval_seconds=300,
        )
        session.add(schedule)
        session.flush()
        return campaign_to_dict(row, metrics=recipient_metrics(session, row))


def get_campaign_by_job_id(job_id: str) -> dict[str, Any] | None:
    """Resolve campaign by job_id (internal use after job access is verified)."""
    with session_scope() as session:
        row = session.scalar(select(Campaign).where(Campaign.job_id == job_id).limit(1))
        if row is None:
            return None
        payload = campaign_to_dict(
            row,
            include_draft=False,
            metrics=recipient_metrics(session, row),
        )
        payload["layout_error_count"] = _layout_error_count(session, row.id)
        return payload


def list_delivery_attempts(
    campaign_id: str,
    *,
    page: int = 1,
    per_page: int = 50,
) -> dict[str, Any]:
    """Paginated delivery attempts for statistics (no owner filter — caller verifies access)."""
    page = max(1, page)
    per_page = min(max(1, per_page), 200)
    with session_scope() as session:
        total = (
            session.scalar(
                select(func.count())
                .select_from(DeliveryAttempt)
                .where(DeliveryAttempt.campaign_id == campaign_id)
            )
            or 0
        )
        start = (page - 1) * per_page
        rows = session.execute(
            select(DeliveryAttempt, CampaignRecipient)
            .join(CampaignRecipient, CampaignRecipient.id == DeliveryAttempt.recipient_id)
            .where(DeliveryAttempt.campaign_id == campaign_id)
            .order_by(DeliveryAttempt.created_at.desc(), DeliveryAttempt.id.desc())
            .offset(start)
            .limit(per_page)
        ).all()
        items = [
            {
                "id": attempt.id,
                "recipient_id": attempt.recipient_id,
                "batch_id": attempt.batch_id,
                "attempt_number": attempt.attempt_number,
                "status": attempt.status,
                "delivery_email": attempt.delivery_email,
                "provider_message_id": attempt.provider_message_id,
                "error": attempt.error,
                "company": recipient.company,
                "contact_name": recipient.contact_name,
                "email": recipient.email,
                "created_at": attempt.created_at.isoformat() if attempt.created_at else None,
                "updated_at": attempt.updated_at.isoformat() if attempt.updated_at else None,
            }
            for attempt, recipient in rows
        ]
        return {
            "items": items,
            "pagination": {
                "page": page,
                "per_page": per_page,
                "total": int(total),
                "pages": max(1, (int(total) + per_page - 1) // per_page),
            },
        }


def get_campaign(campaign_id: str, owner_username: str, *, visible_owners: frozenset[str] | None = None) -> dict[str, Any] | None:
    with session_scope() as session:
        row = session.get(Campaign, campaign_id)
        if row is None:
            return None
        if not can_access_owner(visible_owners, row.owner_username):
            return None
        metrics = recipient_metrics(session, row)
        work = active_work_counts(session, campaign_id)
        payload = campaign_to_dict(
            row,
            metrics=metrics,
            has_active_work=bool(work["active_batches"] or work["active_tasks"]),
        )
        payload["layout_error_count"] = _layout_error_count(session, campaign_id)
        return payload


def _layout_error_count(session, campaign_id: str) -> int:
    extras = session.scalars(
        select(CampaignRecipient.extra).where(CampaignRecipient.campaign_id == campaign_id)
    ).all()
    count = 0
    for extra in extras:
        payload = dict(extra or {})
        if str(payload.get("layout_error_code") or "") == "kp_font_compact":
            count += 1
    return count


def list_campaigns(
    owner_username: str,
    *,
    visible_owners: frozenset[str] | None = None,
    status: str | None = None,
    q: str | None = None,
    limit: int = 50,
    offset: int = 0,
    include_archived: bool = False,
) -> dict[str, Any]:
    with session_scope() as session:
        stmt = select(Campaign)
        stmt = apply_owner_filter(stmt, Campaign.owner_username, visible_owners)
        if not include_archived:
            stmt = stmt.where(Campaign.archived.is_(False))
        if status:
            stmt = stmt.where(Campaign.status == status)
        if q:
            like = f"%{q.strip()}%"
            stmt = stmt.where(Campaign.name.ilike(like))
        total = session.scalar(select(func.count()).select_from(stmt.subquery())) or 0
        rows = session.scalars(stmt.order_by(Campaign.updated_at.desc()).limit(limit).offset(offset)).all()
        metrics_by_id = recipient_metrics_many(session, rows)
        return {
            "items": [
                campaign_to_dict(
                    row,
                    include_draft=False,
                    metrics=metrics_by_id[str(row.id)],
                )
                for row in rows
            ],
            "total": int(total),
        }


def update_campaign(
    campaign_id: str,
    owner_username: str,
    data: dict[str, Any],
    *,
    visible_owners: frozenset[str] | None = None,
) -> dict[str, Any] | None:
    with session_scope() as session:
        row = session.get(Campaign, campaign_id)
        if row is None:
            return None
        if not can_access_owner(visible_owners, row.owner_username):
            return None
        if row.status not in {"draft", "scheduled", "paused"} and data.get("force") is not True:
            # Allow metadata updates on draft-like states; still allow draft_payload merge always for drafts
            if row.status not in {"draft"}:
                pass

        old_email_chain_id = row.email_chain_id
        if "email_chain_id" in data and data["email_chain_id"] is not None:
            from src.campaigns.chain_service import _ensure_chain_access
            from src.infra.models import EmailChainRecord

            chain_row = session.get(EmailChainRecord, data["email_chain_id"])
            _ensure_chain_access(chain_row, owner_username, visible_owners=visible_owners)
        for field in (
            "name",
            "work_type",
            "document_mode",
            "mail_subject",
            "description",
            "send_scenario",
            "internal_comment",
            "transport",
            "email_template_id",
            "kp_template_id",
            "contract_template_id",
            "audience_id",
            "email_chain_id",
        ):
            if field in data and data[field] is not None:
                setattr(row, field, data[field])
        if "connection_ids" in data or "smtp_mailbox_id" in data:
            _apply_sender_fields(row, data)
        if "tags" in data:
            row.tags = list(data.get("tags") or [])
        template_fields = ("email_template_id", "kp_template_id", "contract_template_id")
        template_changed = any(field in data and data[field] is not None for field in template_fields)
        chain_changed = "email_chain_id" in data and data["email_chain_id"] != old_email_chain_id
        if "draft_payload" in data and isinstance(data["draft_payload"], dict):
            merged = dict(row.draft_payload or {})
            merged.update(data["draft_payload"])
            row.draft_payload = merged
        # Merge top-level known fields into draft for autosave recovery
        draft = dict(row.draft_payload or {})
        for key in ("name", "work_type", "document_mode", "mail_subject", "description", "send_scenario", "tags", "internal_comment"):
            if key in data:
                draft[key] = data[key]
        row.draft_payload = draft
        if template_changed or chain_changed:
            _invalidate_variable_mapping(row)
        row.updated_at = _now()
        session.flush()
        return campaign_to_dict(row, metrics=recipient_metrics(session, row))


def reset_campaign_draft(
    campaign_id: str,
    owner_username: str,
    *,
    visible_owners: frozenset[str] | None = None,
) -> dict[str, Any] | None:
    with session_scope() as session:
        row = session.get(Campaign, campaign_id)
        if row is None:
            return None
        if not can_access_owner(visible_owners, row.owner_username):
            return None
        if row.status != "draft":
            raise ValueError("Cannot reset campaign in status " + row.status)

        session.execute(
            delete(CampaignRecipient).where(CampaignRecipient.campaign_id == campaign_id)
        )
        row.total_count = 0
        row.name = "Черновик рассылки"
        row.email_chain_id = None
        row.email_template_id = None
        row.kp_template_id = None
        row.contract_template_id = None
        row.smtp_mailbox_id = None
        row.audience_id = None
        row.work_type = ""
        row.document_mode = "kp"
        row.mail_subject = ""
        row.description = ""
        row.send_scenario = "consent_then_materials"
        row.tags = []
        row.internal_comment = ""
        row.draft_payload = {}

        schedule = session.scalar(
            select(CampaignSchedule).where(CampaignSchedule.campaign_id == campaign_id)
        )
        if schedule is None:
            schedule = CampaignSchedule(id=_new_id(), campaign_id=campaign_id)
            session.add(schedule)
        schedule.send_immediately = True
        schedule.start_at = None
        schedule.timezone = "Europe/Moscow"
        schedule.weekdays = [0, 1, 2, 3, 4]
        schedule.time_windows = [{"start": "09:00", "end": "18:00"}]
        schedule.batch_size = 25
        schedule.interval_seconds = 300
        schedule.pause_between_messages_ms = 0
        schedule.max_per_hour = 0
        schedule.max_per_day = 0
        schedule.on_error = "skip"
        schedule.max_retries = 0
        schedule.preview = {}
        schedule.updated_at = _now()

        row.updated_at = _now()
        session.flush()
        return campaign_to_dict(row, metrics=recipient_metrics(session, row))


def duplicate_campaign(campaign_id: str, owner_username: str, *, visible_owners: frozenset[str] | None = None) -> dict[str, Any] | None:
    source = get_campaign(campaign_id, owner_username, visible_owners=visible_owners)
    if not source:
        return None
    created = create_campaign(
        owner_username,
        {
            "name": f"{source['name']} (копия)",
            "work_type": source["work_type"],
            "document_mode": source["document_mode"],
            "mail_subject": source["mail_subject"],
            "description": source["description"],
            "send_scenario": source["send_scenario"],
            "tags": source["tags"],
            "internal_comment": source["internal_comment"],
            "smtp_mailbox_id": source["smtp_mailbox_id"],
            "connection_ids": source.get("connection_ids") or [],
            "transport": source["transport"],
            "draft_payload": source.get("draft_payload") or {},
        },
    )
    # Copy recipients
    with session_scope() as session:
        recipients = session.scalars(
            select(CampaignRecipient).where(CampaignRecipient.campaign_id == campaign_id)
        ).all()
        for idx, rec in enumerate(recipients):
            session.add(
                CampaignRecipient(
                    campaign_id=created["id"],
                    row_index=idx,
                    company=rec.company,
                    contact_name=rec.contact_name,
                    email=rec.email,
                    email_fallback=rec.email_fallback,
                    region=rec.region,
                    source=rec.source,
                    validation_status=rec.validation_status,
                    extra=dict(rec.extra or {}),
                    excluded=rec.excluded,
                )
            )
        camp = session.get(Campaign, created["id"])
        if camp:
            camp.total_count = len(recipients)
            camp.email_template_id = source.get("email_template_id")
            camp.kp_template_id = source.get("kp_template_id")
            camp.contract_template_id = source.get("contract_template_id")
    return get_campaign(created["id"], owner_username, visible_owners=visible_owners)


def archive_campaign(campaign_id: str, owner_username: str, *, visible_owners: frozenset[str] | None = None) -> dict[str, Any] | None:
    with session_scope() as session:
        row = session.get(Campaign, campaign_id)
        if row is None or not can_access_owner(visible_owners, row.owner_username):
            return None
        row.archived = True
        row.updated_at = _now()
        session.flush()
        return campaign_to_dict(row, metrics=recipient_metrics(session, row))


def list_recipients(
    campaign_id: str,
    owner_username: str,
    *,
    visible_owners: frozenset[str] | None = None,
    limit: int = 50,
    offset: int = 0,
    q: str | None = None,
    only_excluded: bool | None = None,
) -> dict[str, Any]:
    with session_scope() as session:
        camp = session.get(Campaign, campaign_id)
        if camp is None or not can_access_owner(visible_owners, camp.owner_username):
            return {"items": [], "total": 0}
        stmt = select(CampaignRecipient).where(CampaignRecipient.campaign_id == campaign_id)
        if q:
            like = f"%{q.strip()}%"
            stmt = stmt.where(
                (CampaignRecipient.email.ilike(like))
                | (CampaignRecipient.company.ilike(like))
                | (CampaignRecipient.contact_name.ilike(like))
            )
        if only_excluded is True:
            stmt = stmt.where(CampaignRecipient.excluded.is_(True))
        elif only_excluded is False:
            stmt = stmt.where(CampaignRecipient.excluded.is_(False))
        total = session.scalar(select(func.count()).select_from(stmt.subquery())) or 0
        rows = session.scalars(stmt.order_by(CampaignRecipient.row_index).limit(limit).offset(offset)).all()
        items = [
            {
                "id": r.id,
                "row_index": r.row_index,
                "company": r.company,
                "contact_name": r.contact_name,
                "email": r.email,
                "email_fallback": r.email_fallback,
                "region": r.region,
                "source": r.source,
                "validation_status": r.validation_status,
                "extra": dict(r.extra or {}),
                "layout_error_code": str(dict(r.extra or {}).get("layout_error_code") or "") or None,
                "excluded": r.excluded,
                "send_status": r.send_status,
                "last_error": r.last_error,
            }
            for r in rows
        ]
        return {"items": items, "total": int(total)}


def replace_recipients(
    campaign_id: str,
    owner_username: str,
    recipients: list[dict[str, Any]],
    *,
    visible_owners: frozenset[str] | None = None,
    recipient_columns: list[str] | None = None,
) -> dict[str, Any]:
    with session_scope() as session:
        camp = session.get(Campaign, campaign_id)
        if camp is None or not can_access_owner(visible_owners, camp.owner_username):
            raise PermissionError("Campaign not found")
        if camp.status not in {"draft", "paused", "scheduled"}:
            raise ValueError("Cannot replace recipients for campaign in status " + camp.status)

        # Bulk delete + flush before inserts so re-import does not collide on
        # unique (campaign_id, row_index) when UOW would emit INSERT before DELETE.
        session.execute(
            delete(CampaignRecipient).where(CampaignRecipient.campaign_id == campaign_id)
        )
        session.flush()
        seen_emails: set[str] = set()
        added = 0
        duplicates = 0
        invalid = 0
        for idx, item in enumerate(recipients):
            from src.campaigns.recipient_email_service import normalize_import_emails, primary_email_key

            item = normalize_import_emails(item)
            email = str(item.get("email") or "").strip().lower()
            email_fallback = str(item.get("email_fallback") or "").strip().lower()
            status = _validate_email(email, email_fallback)
            if status == "invalid":
                invalid += 1
            dedup_key = primary_email_key(email, email_fallback)
            if dedup_key and dedup_key in seen_emails:
                duplicates += 1
                continue
            if dedup_key:
                seen_emails.add(dedup_key)
            suppressed = is_email_suppressed_for_import(email) if email else False
            extra = dict(item.get("extra") or {})
            for key, value in item.items():
                normalized = str(key or "").strip().lower()
                if not normalized or normalized in _RECIPIENT_ROW_RESERVED or normalized in extra:
                    continue
                extra[normalized] = value
            session.add(
                CampaignRecipient(
                    campaign_id=campaign_id,
                    row_index=added,
                    company=str(item.get("company") or ""),
                    contact_name=str(item.get("contact_name") or item.get("contact") or ""),
                    email=email,
                    email_fallback=email_fallback,
                    region=str(item.get("region") or ""),
                    source=str(item.get("source") or "import"),
                    validation_status=status,
                    extra=extra,
                    excluded=bool(item.get("excluded") or status != "valid" or suppressed),
                )
            )
            added += 1
        camp.total_count = added
        columns = list(recipient_columns or extract_recipient_columns(recipients[:500]))
        draft = dict(camp.draft_payload or {})
        draft["recipient_columns"] = columns
        draft["mapping_confirmed"] = False
        draft.pop("mapping_confirmed_at", None)
        camp.draft_payload = draft
        camp.updated_at = _now()
        session.flush()
        return {
            "total": added,
            "duplicates_skipped": duplicates,
            "invalid": invalid,
            "recipient_columns": columns,
        }


def update_recipient(
    campaign_id: str,
    recipient_id: int,
    owner_username: str,
    data: dict[str, Any],
    *,
    visible_owners: frozenset[str] | None = None,
) -> dict[str, Any] | None:
    with session_scope() as session:
        camp = session.get(Campaign, campaign_id)
        if camp is None or not can_access_owner(visible_owners, camp.owner_username):
            return None
        row = session.get(CampaignRecipient, recipient_id)
        if row is None or row.campaign_id != campaign_id:
            return None
        for field in ("company", "contact_name", "email", "email_fallback", "region", "source"):
            if field in data and data[field] is not None:
                setattr(row, field, data[field] if field not in {"email", "email_fallback"} else str(data[field]).strip().lower())
        if "excluded" in data:
            row.excluded = bool(data["excluded"])
        if "extra" in data and isinstance(data["extra"], dict):
            row.extra = dict(data["extra"])
        row.validation_status = _validate_email(row.email, row.email_fallback)
        if row.validation_status != "valid":
            row.excluded = True
        session.flush()
        return {
            "id": row.id,
            "email": row.email,
            "validation_status": row.validation_status,
            "excluded": row.excluded,
        }


def delete_recipients(
    campaign_id: str,
    recipient_ids: list[int],
    owner_username: str,
    *,
    visible_owners: frozenset[str] | None = None,
) -> int:
    with session_scope() as session:
        camp = session.get(Campaign, campaign_id)
        if camp is None or not can_access_owner(visible_owners, camp.owner_username):
            return 0
        deleted = 0
        for rid in recipient_ids:
            row = session.get(CampaignRecipient, rid)
            if row and row.campaign_id == campaign_id:
                session.delete(row)
                deleted += 1
        remaining = session.scalar(
            select(func.count()).select_from(CampaignRecipient).where(CampaignRecipient.campaign_id == campaign_id)
        ) or 0
        camp.total_count = int(remaining)
        return deleted


def parse_recipients_csv(content: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    columns = [str(name or "").strip().lower() for name in (reader.fieldnames or []) if str(name or "").strip()]
    rows: list[dict[str, Any]] = []
    for raw in reader:
        normalized = {str(k or "").strip().lower(): (v or "").strip() for k, v in raw.items()}
        used_keys: set[str] = set()

        def pick(field: str) -> str:
            for alias in _CORE_FIELD_ALIASES[field]:
                if alias in normalized and normalized[alias]:
                    used_keys.add(alias)
                    return normalized[alias]
            return ""

        extra = {
            key: value
            for key, value in normalized.items()
            if key and key not in used_keys and value
        }
        rows.append(
            {
                "company": pick("company"),
                "contact_name": pick("contact_name"),
                "email": pick("email"),
                "email_fallback": pick("email_fallback"),
                "region": pick("region"),
                "source": "csv",
                "extra": extra,
            }
        )
    if not columns:
        columns = extract_recipient_columns(rows)
    return rows, columns


def parse_recipients_xlsx(content: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    first_raw = [str(c or "").strip() for c in next(rows_iter, [])]
    second_raw = [str(c or "").strip() for c in next(rows_iter, [])]
    first = [value.lower() for value in first_raw]
    second = [value.lower() for value in second_raw]

    def _is_mo_tech_header(headers: list[str]) -> bool:
        keys = set(headers)
        return "email_osn" in keys or ("mun_name" in keys and "sub_rf" in keys)

    # Parser MO template: row1 human labels, row2 technical keys (EMAIL_OSN, …).
    if _is_mo_tech_header(second):
        headers = second
        data_rows: list[tuple[Any, ...]] = list(rows_iter)
    elif _is_mo_tech_header(first):
        headers = first
        data_rows = [tuple(second_raw)] + list(rows_iter) if any(second_raw) else list(rows_iter)
    else:
        headers = first
        data_rows = [tuple(second_raw)] + list(rows_iter) if any(second_raw) else list(rows_iter)

    mapping = {h: i for i, h in enumerate(headers)}

    def cell(row: tuple[Any, ...], field: str) -> tuple[str, set[str]]:
        used: set[str] = set()
        for name in _CORE_FIELD_ALIASES[field]:
            idx = mapping.get(name)
            if idx is not None and idx < len(row):
                value = str(row[idx] or "").strip()
                if value:
                    used.add(name)
                    return value, used
        return "", used

    result: list[dict[str, Any]] = []
    for row in data_rows:
        if not row:
            continue
        used_headers: set[str] = set()
        email, used = cell(row, "email")
        used_headers |= used
        company, used = cell(row, "company")
        used_headers |= used
        if not email and not company:
            continue
        contact_name, used = cell(row, "contact_name")
        used_headers |= used
        email_fallback, used = cell(row, "email_fallback")
        used_headers |= used
        region, used = cell(row, "region")
        used_headers |= used
        extra: dict[str, str] = {}
        for header, idx in mapping.items():
            if not header or header in used_headers or idx >= len(row):
                continue
            value = str(row[idx] or "").strip()
            if value:
                extra[header] = value
        result.append(
            {
                "company": company,
                "contact_name": contact_name,
                "email": email,
                "email_fallback": email_fallback,
                "region": region,
                "source": "xlsx",
                "extra": extra,
            }
        )
    file_columns = [header for header in headers if header]
    merged_columns = list(CORE_RECIPIENT_COLUMNS)
    seen = set(merged_columns)
    for column in file_columns + [key for row in result for key in (row.get("extra") or {})]:
        if column and column not in seen:
            seen.add(column)
            merged_columns.append(column)
    return result, merged_columns


def upsert_schedule(
    campaign_id: str,
    owner_username: str,
    data: dict[str, Any],
    *,
    visible_owners: frozenset[str] | None = None,
) -> dict[str, Any]:
    with session_scope() as session:
        camp = session.get(Campaign, campaign_id)
        if camp is None or not can_access_owner(visible_owners, camp.owner_username):
            raise PermissionError("Campaign not found")
        schedule = session.scalar(select(CampaignSchedule).where(CampaignSchedule.campaign_id == campaign_id))
        if schedule is None:
            schedule = CampaignSchedule(id=_new_id(), campaign_id=campaign_id)
            session.add(schedule)

        if "send_immediately" in data:
            schedule.send_immediately = bool(data["send_immediately"])
        if "start_at" in data:
            value = data["start_at"]
            if value in (None, ""):
                schedule.start_at = None
            elif isinstance(value, datetime):
                schedule.start_at = value if value.tzinfo else value.replace(tzinfo=timezone.utc)
            else:
                schedule.start_at = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
            if schedule.start_at is not None:
                start = schedule.start_at if schedule.start_at.tzinfo else schedule.start_at.replace(tzinfo=timezone.utc)
                now = _now()
                if start.astimezone(timezone.utc) < now:
                    schedule.start_at = now
                schedule.send_immediately = False
        for field in ("timezone", "on_error"):
            if field in data and data[field] is not None:
                setattr(schedule, field, str(data[field]))
        for field in ("batch_size", "interval_seconds", "pause_between_messages_ms", "max_per_hour", "max_per_day", "max_retries"):
            if field in data and data[field] is not None:
                setattr(schedule, field, int(data[field]))
        if "weekdays" in data:
            schedule.weekdays = list(data["weekdays"] or [])
        if "time_windows" in data:
            schedule.time_windows = list(data["time_windows"] or [])

        active_count = session.scalar(
            select(func.count())
            .select_from(CampaignRecipient)
            .where(
                CampaignRecipient.campaign_id == campaign_id,
                CampaignRecipient.excluded.is_(False),
            )
        ) or 0
        conn_hour, conn_day = _pool_rate_limits(session, _campaign_connection_ids(camp))
        max_per_hour, max_per_day = _effective_rate_limits(
            schedule_max_per_hour=schedule.max_per_hour,
            schedule_max_per_day=schedule.max_per_day,
            connection_max_per_hour=conn_hour,
            connection_max_per_day=conn_day,
        )
        preview = plan_batches(
            recipient_count=int(active_count),
            batch_size=schedule.batch_size,
            interval_seconds=schedule.interval_seconds,
            start_at=schedule.start_at,
            send_immediately=schedule.send_immediately,
            timezone_name=schedule.timezone,
            weekdays=list(schedule.weekdays or []),
            time_windows=list(schedule.time_windows or []),
            max_per_hour=max_per_hour,
            max_per_day=max_per_day,
        )
        schedule.preview = preview
        schedule.updated_at = _now()
        session.flush()
        return schedule_to_dict(schedule)


def schedule_to_dict(schedule: CampaignSchedule) -> dict[str, Any]:
    return {
        "id": schedule.id,
        "campaign_id": schedule.campaign_id,
        "send_immediately": schedule.send_immediately,
        "start_at": schedule.start_at.isoformat() if schedule.start_at else None,
        "timezone": schedule.timezone,
        "weekdays": list(schedule.weekdays or []),
        "time_windows": list(schedule.time_windows or []),
        "batch_size": schedule.batch_size,
        "interval_seconds": schedule.interval_seconds,
        "pause_between_messages_ms": schedule.pause_between_messages_ms,
        "max_per_hour": schedule.max_per_hour,
        "max_per_day": schedule.max_per_day,
        "on_error": schedule.on_error,
        "max_retries": schedule.max_retries,
        "preview": dict(schedule.preview or {}),
    }


def get_schedule(campaign_id: str, owner_username: str, *, visible_owners: frozenset[str] | None = None) -> dict[str, Any] | None:
    with session_scope() as session:
        camp = session.get(Campaign, campaign_id)
        if camp is None or not can_access_owner(visible_owners, camp.owner_username):
            return None
        schedule = session.scalar(select(CampaignSchedule).where(CampaignSchedule.campaign_id == campaign_id))
        if schedule is None:
            return None
        return schedule_to_dict(schedule)


def _resolve_send_mode(camp: Campaign) -> str:
    if camp.send_scenario == "consent_then_materials":
        return "consent_request"
    if camp.send_scenario == "email_chain":
        return "chain_root"
    if camp.send_scenario == "materials_now":
        return "email"
    return "materials"


def _sender_batch_task_payload(
    *,
    campaign_id: str,
    batch_id: str,
    camp: Campaign,
    schedule: CampaignSchedule,
) -> dict[str, Any]:
    return {
        "campaign_id": campaign_id,
        "batch_id": batch_id,
        "connection_ids": _campaign_connection_ids(camp),
        "smtp_mailbox_id": camp.smtp_mailbox_id,
        "transport": camp.transport,
        "mail_subject": camp.mail_subject,
        "send_mode": _resolve_send_mode(camp),
        "campaign_name": camp.name,
        "pause_between_messages_ms": schedule.pause_between_messages_ms,
        "max_retries": schedule.max_retries,
        "on_error": schedule.on_error,
    }


MAX_SENDER_BATCH_WORKER_RECOVERIES = 5
SENDER_BATCH_WORKER_RECOVERY_BACKOFF_SECONDS = 30


def enqueue_sender_batch_task(
    session: Any,
    *,
    campaign_id: str,
    camp: Campaign,
    batch: CampaignBatch,
    schedule: CampaignSchedule,
    owner_username: str,
    available_at: datetime,
    idempotency_suffix: str,
) -> str:
    """Enqueue a sender_batch task and attach task_id to the batch row."""
    from src.workers.task_queue import enqueue_task

    task, _created = enqueue_task(
        task_type="sender_batch",
        job_id=camp.job_id or campaign_id,
        owner_username=owner_username,
        payload=_sender_batch_task_payload(
            campaign_id=campaign_id,
            batch_id=batch.id,
            camp=camp,
            schedule=schedule,
        ),
        available_at=available_at,
        idempotency_key=f"sender_batch:{batch.id}:{idempotency_suffix}",
        active_key=f"sender_batch:{batch.id}:{idempotency_suffix}",
        max_attempts=max(1, int(schedule.max_retries or 3)),
    )
    batch.task_id = str(task.get("id") or "")
    return batch.task_id


def validate_campaign_for_launch(
    campaign_id: str,
    owner_username: str,
    *,
    visible_owners: frozenset[str] | None = None,
    deep: bool = False,
) -> dict[str, Any]:
    with session_scope() as session:
        camp = session.get(Campaign, campaign_id)
        if camp is None or not can_access_owner(visible_owners, camp.owner_username):
            return {
                "ok": False,
                "errors": ["Рассылка не найдена"],
                "warnings": [],
                "template_issues": [],
            }
        errors: list[str] = []
        warnings: list[str] = []
        if not (camp.name or "").strip():
            errors.append("Укажите название рассылки")
        from src.campaigns.connection_service import validate_connection_ids

        connection_error = validate_connection_ids(_campaign_connection_ids(camp), camp.owner_username)
        if connection_error:
            errors.append(connection_error)
        active = session.scalar(
            select(func.count())
            .select_from(CampaignRecipient)
            .where(CampaignRecipient.campaign_id == campaign_id, CampaignRecipient.excluded.is_(False))
        ) or 0
        excluded = session.scalar(
            select(func.count())
            .select_from(CampaignRecipient)
            .where(CampaignRecipient.campaign_id == campaign_id, CampaignRecipient.excluded.is_(True))
        ) or 0
        if active <= 0:
            errors.append("Нет получателей для отправки")

        if camp.send_scenario == "email_chain":
            from src.campaigns.chain_service import get_email_chain, validate_chain

            chain_validation = validate_chain(get_email_chain(camp), strict=False)
            if not chain_validation["ok"]:
                errors.extend(chain_validation["errors"])
            warnings.extend(chain_validation.get("warnings") or [])
        elif not camp.email_template_id and not (camp.draft_payload or {}).get("email_body"):
            warnings.append("Шаблон письма не выбран — будет использован текст по умолчанию")

        from src.campaigns.template_text_review_service import partition_review_messages
        from src.campaigns.variable_match_service import (
            empty_variable_validation_errors,
            mapping_validation_errors,
            substitution_validation_issues,
            template_text_cache_validation_errors,
        )

        mapping_errors = mapping_validation_errors(camp)
        errors.extend(mapping_errors)
        if not mapping_errors:
            errors.extend(empty_variable_validation_errors(camp))
        errors.extend(template_text_cache_validation_errors(camp))
        template_issues = substitution_validation_issues(
            camp,
            deep=deep,
            advisory=False,
            include_placeholder_issues=True,
            strict_preview=True,
        )
        template_errors, template_warnings = partition_review_messages(template_issues)
        errors.extend(template_errors)
        warnings.extend(template_warnings)
        schedule = session.scalar(select(CampaignSchedule).where(CampaignSchedule.campaign_id == campaign_id))
        draft = dict(camp.draft_payload or {})
        schedule_payload = schedule_to_dict(schedule) if schedule else None
        campaign_payload = campaign_to_dict(
            camp,
            metrics=recipient_metrics(session, camp),
        )
        send_scenario = camp.send_scenario

    return {
        "ok": len(errors) == 0,
        "errors": errors,
        "warnings": warnings,
        "template_issues": template_issues,
        "active_recipients": int(active),
        "mapping_confirmed": bool(draft.get("mapping_confirmed")),
        "excluded_recipients": int(excluded),
        "schedule": schedule_payload,
        "campaign": campaign_payload,
    }
def launch_campaign(
    campaign_id: str,
    owner_username: str,
    *,
    visible_owners: frozenset[str] | None = None,
    force_now: bool = False,
) -> dict[str, Any]:
    # Reject invalid lifecycle actions before expensive/deep launch validation.
    # The status is checked again under a row lock below to close the race.
    with session_scope() as session:
        existing_campaign = session.get(Campaign, campaign_id)
        if (
            existing_campaign is not None
            and can_access_owner(visible_owners, existing_campaign.owner_username)
            and existing_campaign.status != "draft"
        ):
            raise CampaignStateConflict(
                "Only a draft campaign can be launched. Duplicate a completed campaign to send it again."
            )

    validation = validate_campaign_for_launch(
        campaign_id,
        owner_username,
        visible_owners=visible_owners,
        deep=True,
    )
    if not validation["ok"]:
        raise ValueError("; ".join(validation["errors"]))

    with session_scope() as session:
        camp = session.get(Campaign, campaign_id, with_for_update=True)
        assert camp is not None
        if camp.status != "draft":
            raise CampaignStateConflict(
                "Only a draft campaign can be launched. Duplicate a completed campaign to send it again."
            )
        schedule = session.scalar(select(CampaignSchedule).where(CampaignSchedule.campaign_id == campaign_id))
        assert schedule is not None

        # Clear future/stuck batches if re-launch from paused/scheduled/running.
        from src.workers.task_queue import enqueue_task, request_cancel

        existing = session.scalars(
            select(CampaignBatch).where(
                CampaignBatch.campaign_id == campaign_id,
                CampaignBatch.status.in_(
                    ["pending", "paused", "running", "failed", "completed_with_errors", "cancelled"]
                ),
            )
        ).all()
        for batch in existing:
            if batch.task_id:
                try:
                    request_cancel(batch.task_id)
                except Exception:
                    pass
            session.delete(batch)

        recipients = session.scalars(
            select(CampaignRecipient)
            .where(
                CampaignRecipient.campaign_id == campaign_id,
                CampaignRecipient.excluded.is_(False),
                CampaignRecipient.send_status.in_(["pending", "failed"]),
            )
            .order_by(CampaignRecipient.row_index)
        ).all()
        recipient_ids = [r.id for r in recipients]
        intended_recipient_count = int(
            session.scalar(
                select(func.count())
                .select_from(CampaignRecipient)
                .where(
                    CampaignRecipient.campaign_id == campaign_id,
                    CampaignRecipient.excluded.is_(False),
                )
            )
            or 0
        )
        conn_hour, conn_day = _pool_rate_limits(session, _campaign_connection_ids(camp))
        launch_now = _now()
        immediate = force_now or _schedule_requires_immediate_start(schedule, now=launch_now)
        # immediate launch bypasses calendar windows / schedule pacing, but connection limits still apply.
        schedule_hour = 0 if immediate else schedule.max_per_hour
        schedule_day = 0 if immediate else schedule.max_per_day
        max_per_hour, max_per_day = _effective_rate_limits(
            schedule_max_per_hour=schedule_hour,
            schedule_max_per_day=schedule_day,
            connection_max_per_hour=conn_hour,
            connection_max_per_day=conn_day,
        )
        preview = plan_batches(
            recipient_count=len(recipient_ids),
            batch_size=schedule.batch_size,
            interval_seconds=0 if immediate else schedule.interval_seconds,
            start_at=None if immediate else schedule.start_at,
            send_immediately=True if immediate else schedule.send_immediately,
            timezone_name=schedule.timezone,
            weekdays=[] if immediate else list(schedule.weekdays or []),
            time_windows=[] if immediate else list(schedule.time_windows or []),
            max_per_hour=max_per_hour,
            max_per_day=max_per_day,
            now=launch_now,
        )
        schedule.preview = preview

        created_batches: list[dict[str, Any]] = []
        offset = 0
        for plan in preview.get("batches") or []:
            size = int(plan["size"])
            chunk = recipient_ids[offset : offset + size]
            offset += size
            batch_id = _new_id()
            scheduled_at = datetime.fromisoformat(str(plan["scheduled_at"]).replace("Z", "+00:00"))
            scheduled_at = max(scheduled_at, launch_now)
            batch = CampaignBatch(
                id=batch_id,
                campaign_id=campaign_id,
                batch_index=int(plan["batch_index"]),
                scheduled_at=scheduled_at,
                size=len(chunk),
                status="pending",
                recipient_ids=chunk,
            )
            session.add(batch)
            session.flush()

            pre_gen_at = launch_now if immediate else max(launch_now, scheduled_at - timedelta(hours=1))
            enqueue_task(
                task_type="campaign_pre_generate",
                job_id=camp.job_id or campaign_id,
                owner_username=owner_username,
                payload={
                    "campaign_id": campaign_id,
                    "batch_id": batch_id,
                    "recipient_ids": chunk,
                },
                available_at=pre_gen_at,
                idempotency_key=f"campaign_pre_generate:{batch_id}",
                active_key=f"campaign_pre_generate:{batch_id}",
                max_attempts=max(1, int(schedule.max_retries or 3)),
            )

            # Idempotent task per batch
            task, _created = enqueue_task(
                task_type="sender_batch",
                job_id=camp.job_id or campaign_id,
                owner_username=owner_username,
                payload=_sender_batch_task_payload(
                    campaign_id=campaign_id,
                    batch_id=batch_id,
                    camp=camp,
                    schedule=schedule,
                ),
                available_at=scheduled_at,
                idempotency_key=f"sender_batch:{batch_id}",
                active_key=f"sender_batch:{batch_id}",
                max_attempts=max(1, int(schedule.max_retries or 3)),
            )
            batch.task_id = str(task.get("id") or "")
            created_batches.append(
                {
                    "id": batch_id,
                    "batch_index": batch.batch_index,
                    "scheduled_at": scheduled_at.isoformat(),
                    "size": batch.size,
                    "task_id": batch.task_id,
                }
            )

        target_status = "running" if immediate else "scheduled"
        if target_status == "running" or immediate:
            # If first batch is in the future, keep scheduled
            first_at = preview.get("first_batch_at")
            if first_at:
                first_dt = datetime.fromisoformat(str(first_at).replace("Z", "+00:00"))
                if first_dt > _now() + __import__("datetime").timedelta(seconds=5):
                    target_status = "scheduled"
        camp.sent_count = 0
        camp.error_count = 0
        camp.completed_at = None
        camp.launched_at = _now()
        camp.total_count = intended_recipient_count
        transition_campaign_status(
            session,
            camp,
            target_status,
            reason="campaign_launch",
            actor=owner_username,
        )
        if not created_batches:
            reconcile_campaign_state(session, camp, repair=True, actor="launch_reconciler")
        session.flush()

        # Heal missing JobOwner so statistics scoping matches Campaign ownership.
        if camp.job_id:
            from src.jobs.access import assign_job_owner
            from src.security.auth import Principal

            assign_job_owner(
                camp.job_id,
                Principal(camp.owner_username, "default", "user"),
                overwrite=False,
            )

        return {
            "campaign": campaign_to_dict(
                camp,
                metrics=recipient_metrics(session, camp),
            ),
            "batches": created_batches,
            "preview": preview,
        }


def pause_campaign(campaign_id: str, owner_username: str, *, visible_owners: frozenset[str] | None = None) -> dict[str, Any]:
    with session_scope() as session:
        camp = session.get(Campaign, campaign_id, with_for_update=True)
        if camp is None or not can_access_owner(visible_owners, camp.owner_username):
            raise PermissionError("Campaign not found")
        if camp.status not in {"running", "scheduled"}:
            raise CampaignStateConflict(
                f"Campaign cannot be paused from status {camp.status}"
            )
        transition_campaign_status(
            session,
            camp,
            "paused",
            reason="campaign_pause",
            actor=owner_username,
        )
        batches = session.scalars(
            select(CampaignBatch).where(
                CampaignBatch.campaign_id == campaign_id,
                CampaignBatch.status == "pending",
            )
        ).all()
        from src.workers.task_queue import request_cancel

        for batch in batches:
            batch.status = "paused"
            if batch.task_id:
                try:
                    request_cancel(batch.task_id)
                except Exception:
                    pass
        session.flush()
        return campaign_to_dict(camp, metrics=recipient_metrics(session, camp))


def resume_campaign(campaign_id: str, owner_username: str, *, visible_owners: frozenset[str] | None = None) -> dict[str, Any]:
    with session_scope() as session:
        camp = session.get(Campaign, campaign_id, with_for_update=True)
        if camp is None or not can_access_owner(visible_owners, camp.owner_username):
            raise PermissionError("Campaign not found")
        if camp.status != "paused":
            raise CampaignStateConflict("Campaign is not paused")
        if camp.completed_at is not None:
            raise CampaignStateConflict(
                "A completed campaign cannot be resumed. Duplicate it to send again."
            )

        schedule = session.scalar(select(CampaignSchedule).where(CampaignSchedule.campaign_id == campaign_id))
        if schedule is None:
            raise ValueError("Campaign schedule not found")

        batches = session.scalars(
            select(CampaignBatch)
            .where(
                CampaignBatch.campaign_id == campaign_id,
                CampaignBatch.status.in_(["paused", "failed", "running"]),
            )
            .order_by(CampaignBatch.batch_index)
        ).all()
        if not batches:
            raise CampaignStateConflict("Campaign has no unfinished batches to resume")
        now = _now()
        for batch in batches:
            scheduled_at = batch.scheduled_at if batch.scheduled_at > now else now
            batch.status = "pending"
            batch.scheduled_at = scheduled_at
            batch.started_at = None
            batch.completed_at = None
            batch.error = None
            enqueue_sender_batch_task(
                session,
                campaign_id=campaign_id,
                camp=camp,
                batch=batch,
                schedule=schedule,
                owner_username=owner_username,
                available_at=scheduled_at,
                idempotency_suffix=f"resume:{int(now.timestamp())}",
            )
        transition_campaign_status(
            session,
            camp,
            "running",
            reason="campaign_resume",
            actor=owner_username,
            at=now,
        )
        session.flush()
        return campaign_to_dict(camp, metrics=recipient_metrics(session, camp))


def cancel_campaign(campaign_id: str, owner_username: str, *, visible_owners: frozenset[str] | None = None) -> dict[str, Any]:
    with session_scope() as session:
        camp = session.get(Campaign, campaign_id, with_for_update=True)
        if camp is None or not can_access_owner(visible_owners, camp.owner_username):
            raise PermissionError("Campaign not found")
        if camp.status not in {"running", "scheduled", "paused"}:
            raise CampaignStateConflict(
                f"Campaign cannot be cancelled from status {camp.status}"
            )
        transition_campaign_status(
            session,
            camp,
            "cancelled",
            reason="campaign_cancel",
            actor=owner_username,
        )
        from src.workers.task_queue import request_cancel

        batches = session.scalars(
            select(CampaignBatch).where(
                CampaignBatch.campaign_id == campaign_id,
                CampaignBatch.status.in_(["pending", "paused", "running"]),
            )
        ).all()
        for batch in batches:
            if batch.status != "running":
                batch.status = "cancelled"
            if batch.task_id:
                try:
                    request_cancel(batch.task_id)
                except Exception:
                    pass
        session.flush()
        return campaign_to_dict(camp, metrics=recipient_metrics(session, camp))


def list_batches(campaign_id: str, owner_username: str, *, visible_owners: frozenset[str] | None = None) -> list[dict[str, Any]]:
    with session_scope() as session:
        camp = session.get(Campaign, campaign_id)
        if camp is None or not can_access_owner(visible_owners, camp.owner_username):
            return []
        rows = session.scalars(
            select(CampaignBatch).where(CampaignBatch.campaign_id == campaign_id).order_by(CampaignBatch.batch_index)
        ).all()
        task_ids = [str(batch.task_id) for batch in rows if batch.task_id]
        tasks_by_id = {
            str(task.id): task
            for task in (
                session.scalars(
                    select(BackgroundTask).where(BackgroundTask.id.in_(task_ids))
                ).all()
                if task_ids
                else []
            )
        }
        active_tasks = session.scalars(
            select(BackgroundTask)
            .where(
                BackgroundTask.status.in_(("queued", "running", "retry")),
                BackgroundTask.cancel_requested_at.is_(None),
            )
        ).all()
        now = _now()
        active_tasks.sort(
            key=lambda task: (
                0 if task.status == "running" else 1,
                0 if task.available_at <= now else 1,
                (
                    -int(task.priority or 0)
                    if task.available_at <= now
                    else task.available_at
                ),
                (
                    task.created_at
                    if task.available_at <= now
                    else -int(task.priority or 0)
                ),
                task.created_at,
            )
        )
        waiting_positions: dict[str, int] = {}
        waiting_index = 0
        for task in active_tasks:
            if task.status == "running":
                continue
            waiting_index += 1
            waiting_positions[str(task.id)] = waiting_index
        running_task_ids = {
            str(task.id) for task in active_tasks if task.status == "running"
        }
        recipient_status_by_id = {
            int(recipient_id): str(send_status or "pending")
            for recipient_id, send_status in session.execute(
                select(CampaignRecipient.id, CampaignRecipient.send_status).where(
                    CampaignRecipient.campaign_id == campaign_id
                )
            ).all()
        }

        def _batch_metrics(batch: CampaignBatch) -> dict[str, int]:
            statuses = [
                recipient_status_by_id.get(int(recipient_id), "pending")
                for recipient_id in list(batch.recipient_ids or [])
            ]
            success = sum(1 for status in statuses if status in {"sent", "in_chain"})
            skipped = sum(1 for status in statuses if status == "skipped")
            failed = sum(1 for status in statuses if status == "failed")
            processed = success + skipped + failed
            return {
                "processed": processed,
                "skipped": skipped,
                "failed": failed,
                "remaining": max(0, int(batch.size or 0) - processed),
            }

        result: list[dict[str, Any]] = []
        for batch in rows:
            metrics = _batch_metrics(batch)
            task = tasks_by_id.get(str(batch.task_id or ""))
            task_status = str(task.status) if task is not None else ""
            queue_position = waiting_positions.get(str(batch.task_id or ""))
            if camp.status == "paused" or batch.status == "paused":
                wait_reason = "Рассылка приостановлена"
            elif task is not None and task.status == "running":
                wait_reason = "Идёт отправка"
            elif task is not None and task.status == "retry":
                wait_reason = "Ожидает повторной попытки после ошибки"
            elif task is not None and task.available_at and task.available_at > now:
                wait_reason = "Ожидает запланированного времени"
            elif task is not None and task.status == "queued":
                wait_reason = "Ожидает свободного отправщика"
            elif batch.status == "completed":
                wait_reason = "Отправка завершена"
            elif batch.status == "cancelled":
                wait_reason = "Отправка отменена"
            elif batch.status in {"failed", "completed_with_errors"}:
                wait_reason = "Завершено с ошибками"
            else:
                wait_reason = "Ожидает запуска"
            result.append(
                {
                    "id": batch.id,
                    "batch_index": batch.batch_index,
                    "scheduled_at": batch.scheduled_at.isoformat() if batch.scheduled_at else None,
                    "size": batch.size,
                    "sent_count": batch.sent_count,
                    "error_count": batch.error_count,
                    "processed_count": metrics["processed"],
                    "skipped_count": metrics["skipped"],
                    "failed_recipient_count": metrics["failed"],
                    "remaining": metrics["remaining"],
                    "status": batch.status,
                    "task_id": batch.task_id,
                    "task_status": task_status or None,
                    "queue_position": queue_position,
                    "is_current": bool(
                        batch.task_id and str(batch.task_id) in running_task_ids
                    ),
                    "available_at": (
                        task.available_at.isoformat()
                        if task is not None and task.available_at
                        else batch.scheduled_at.isoformat()
                        if batch.scheduled_at
                        else None
                    ),
                    "attempt": int(task.attempt or 0) if task is not None else 0,
                    "max_attempts": int(task.max_attempts or 0) if task is not None else 0,
                    "wait_reason": wait_reason,
                    "error": batch.error,
                    "started_at": batch.started_at.isoformat() if batch.started_at else None,
                    "completed_at": batch.completed_at.isoformat() if batch.completed_at else None,
                }
            )
        return result


def cancel_batch(campaign_id: str, batch_id: str, owner_username: str, *, visible_owners: frozenset[str] | None = None) -> dict[str, Any]:
    with session_scope() as session:
        camp = session.get(Campaign, campaign_id)
        if camp is None or not can_access_owner(visible_owners, camp.owner_username):
            raise PermissionError("Campaign not found")
        batch = session.get(CampaignBatch, batch_id)
        if batch is None or batch.campaign_id != campaign_id:
            raise PermissionError("Batch not found")
        if batch.status not in {"pending", "paused"}:
            raise ValueError("Only future batches can be cancelled")
        batch.status = "cancelled"
        if batch.task_id:
            from src.workers.task_queue import request_cancel

            request_cancel(batch.task_id)
        session.flush()
        return {"id": batch.id, "status": batch.status}


def active_sending(owner_username: str, *, visible_owners: frozenset[str] | None = None) -> dict[str, Any] | None:
    with session_scope() as session:
        stmt = select(Campaign).where(Campaign.status.in_(["running", "scheduled", "paused"]))
        stmt = apply_owner_filter(stmt, Campaign.owner_username, visible_owners)
        camps = session.scalars(stmt.order_by(Campaign.updated_at.desc())).all()
        for camp in camps:
            schedule = session.scalar(select(CampaignSchedule).where(CampaignSchedule.campaign_id == camp.id))
            next_batch = session.scalars(
                select(CampaignBatch)
                .where(CampaignBatch.campaign_id == camp.id, CampaignBatch.status == "pending")
                .order_by(CampaignBatch.scheduled_at)
                .limit(1)
            ).first()
            running_batch = session.scalars(
                select(CampaignBatch)
                .where(CampaignBatch.campaign_id == camp.id, CampaignBatch.status == "running")
                .limit(1)
            ).first()
            queued = session.scalar(
                select(func.count())
                .select_from(CampaignBatch)
                .where(CampaignBatch.campaign_id == camp.id, CampaignBatch.status == "pending")
            ) or 0
            metrics = recipient_metrics(session, camp)
            remaining = int(metrics["pending_count"])
            # Finished / stuck-paused campaigns with nothing left to send stay out of the dashboard card.
            if remaining == 0 and int(queued) == 0 and running_batch is None:
                continue
            return {
                "campaign_id": camp.id,
                "name": camp.name,
                "status": camp.status,
                "sent_count": camp.sent_count,
                "total_count": camp.total_count,
                "processed_count": int(metrics["processed_count"]),
                "skipped_count": int(metrics["skipped_count"]),
                "failed_recipient_count": int(metrics["failed_recipient_count"]),
                "remaining": remaining,
                "queued_batches": int(queued),
                "sending_now": running_batch.size if running_batch else 0,
                "next_batch_size": next_batch.size if next_batch else 0,
                "next_batch_at": next_batch.scheduled_at.isoformat() if next_batch else None,
                "batch_size": schedule.batch_size if schedule else 0,
                "interval_seconds": schedule.interval_seconds if schedule else 0,
                "max_per_hour": schedule.max_per_hour if schedule else 0,
                "max_per_day": schedule.max_per_day if schedule else 0,
                "progress": float(metrics["progress"]),
                "success_rate": float(metrics["success_rate"]),
            }
        return None


def record_delivery_attempt(
    *,
    campaign_id: str,
    recipient_id: int,
    batch_id: str | None,
    status: str,
    error: str | None = None,
    provider_message_id: str | None = None,
    delivery_email: str | None = None,
    attempt_number: int | None = None,
) -> bool:
    """Return False if already recorded (idempotent skip)."""
    with session_scope() as session:
        if attempt_number is None:
            latest = session.scalar(
                select(DeliveryAttempt)
                .where(
                    DeliveryAttempt.campaign_id == campaign_id,
                    DeliveryAttempt.recipient_id == recipient_id,
                )
                .order_by(DeliveryAttempt.attempt_number.desc())
                .limit(1)
            )
            if status == "sending":
                if latest is not None and latest.status == "sending":
                    attempt_number = latest.attempt_number
                elif latest is not None and latest.status in {"sent", "delivered"}:
                    return False
                else:
                    attempt_number = int(latest.attempt_number if latest else 0) + 1
            else:
                attempt_number = int(latest.attempt_number if latest else 1)

        key = f"{campaign_id}:{recipient_id}:{attempt_number}"
        existing = session.scalar(select(DeliveryAttempt).where(DeliveryAttempt.idempotency_key == key))
        if existing and existing.status in {"sent", "delivered"} and status not in {"sending"}:
            return False
        if existing is None:
            session.add(
                DeliveryAttempt(
                    campaign_id=campaign_id,
                    recipient_id=recipient_id,
                    batch_id=batch_id,
                    attempt_number=attempt_number,
                    status=status,
                    error=error,
                    provider_message_id=provider_message_id,
                    delivery_email=delivery_email,
                    idempotency_key=key,
                )
            )
        else:
            existing.status = status
            existing.error = error
            existing.provider_message_id = provider_message_id
            if delivery_email:
                existing.delivery_email = delivery_email
            if batch_id:
                existing.batch_id = batch_id
            existing.updated_at = _now()
        return True
