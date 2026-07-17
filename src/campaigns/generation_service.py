"""Bridge CampaignFlow campaigns to the existing document generator workspace."""

from __future__ import annotations

import hashlib
import json
import shutil
from copy import deepcopy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from sqlalchemy import select

from src.campaigns import template_service
from src.generator.generation.document_builder import (
    CONTRACT_TEMPLATE_FILENAME,
    KP_TEMPLATE_FILENAME,
    KP_TEMPLATE_PDF_FILENAME,
    document_mode_kinds,
    normalize_document_mode,
)
from src.generator.generation.generator_agent import GENERATOR_STATE, get_generator_status
from src.generator.philologist.philologist_agent import PHILOLOGIST_STATE, get_philologist_status
from src.infra.db import session_scope
from src.infra.models import Campaign, CampaignRecipient
from src.jobs.clients_store import import_clients_from_xlsx
from src.jobs.json_store import read_json, write_json_atomic
from src.jobs.state import save_agent_state
from src.jobs.storage import resolve_job_paths
from src.parser.excel_writer import COLUMNS


MANIFEST_FILENAME = "campaign-generation.json"


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _manifest_path(job_id: str) -> Path:
    return resolve_job_paths(job_id).root_dir / "state" / MANIFEST_FILENAME


def _load_manifest(job_id: str) -> dict[str, Any]:
    result = read_json(_manifest_path(job_id), default={})
    return result.data if result.ok and isinstance(result.data, dict) else {}


def _load_campaign_rows(
    campaign_id: str,
    owner_username: str,
    *,
    is_admin: bool = False,
) -> tuple[Campaign, list[CampaignRecipient]]:
    with session_scope() as session:
        campaign = session.get(Campaign, campaign_id)
        if campaign is None or (not is_admin and campaign.owner_username != owner_username):
            raise FileNotFoundError("Рассылка не найдена")
        recipients = session.scalars(
            select(CampaignRecipient)
            .where(
                CampaignRecipient.campaign_id == campaign_id,
                CampaignRecipient.excluded.is_(False),
            )
            .order_by(CampaignRecipient.row_index)
        ).all()
        session.expunge(campaign)
        for recipient in recipients:
            session.expunge(recipient)
        return campaign, list(recipients)


def _signature(campaign: Campaign, recipients: list[CampaignRecipient]) -> str:
    payload = {
        "campaign_id": campaign.id,
        "document_mode": normalize_document_mode(campaign.document_mode),
        "work_type": campaign.work_type or "",
        "kp_template_id": campaign.kp_template_id or "",
        "contract_template_id": campaign.contract_template_id or "",
        "recipients": [
            {
                "id": recipient.id,
                "row_index": recipient.row_index,
                "company": recipient.company,
                "contact_name": recipient.contact_name,
                "email": recipient.email,
                "email_fallback": recipient.email_fallback,
                "region": recipient.region,
                "extra": recipient.extra or {},
            }
            for recipient in recipients
        ],
    }
    encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _extra_value(extra: dict[str, Any], key: str, *aliases: str) -> Any:
    normalized = {str(name).strip().upper(): value for name, value in extra.items()}
    for candidate in (key, *aliases):
        value = normalized.get(candidate.upper())
        if value not in (None, ""):
            return value
    return ""


def _recipient_row(recipient: CampaignRecipient) -> dict[str, Any]:
    extra = dict(recipient.extra or {})
    company = recipient.company or _extra_value(extra, "MUN_NAME", "ADM_NAME", "COMPANY")
    row: dict[str, Any] = {
        key: _extra_value(extra, key)
        for _, key in COLUMNS
        if key
    }
    row.update(
        {
            "ID": str(recipient.id),
            "SUB_RF": _extra_value(extra, "SUB_RF") or recipient.region,
            "MUN_R_NAME": _extra_value(extra, "MUN_R_NAME"),
            "MUN_NAME": _extra_value(extra, "MUN_NAME") or company,
            "ADM_NAME": _extra_value(extra, "ADM_NAME") or company,
            "HEAD_FIO": _extra_value(extra, "HEAD_FIO") or recipient.contact_name,
            "EMAIL_OSN": recipient.email,
            "EMAIL_DOP": recipient.email_fallback,
            "STATUS": "",
        }
    )
    for key, value in extra.items():
        technical_key = str(key).strip().upper()
        if technical_key and technical_key not in row:
            row[technical_key] = value
    return row


def _write_data_xlsx(path: Path, recipients: list[CampaignRecipient]) -> None:
    rows = [_recipient_row(recipient) for recipient in recipients]
    ordered_keys = [key for _, key in COLUMNS if key]
    for row in rows:
        for key in row:
            if key not in ordered_keys:
                ordered_keys.append(key)
    labels = {key: human for human, key in COLUMNS if key}

    workbook = Workbook()
    worksheet = workbook.active
    for column, key in enumerate(ordered_keys, start=1):
        worksheet.cell(row=1, column=column, value=labels.get(key, key))
        worksheet.cell(row=2, column=column, value=key)
    for row_number, row in enumerate(rows, start=3):
        for column, key in enumerate(ordered_keys, start=1):
            value = row.get(key, "")
            cell = worksheet.cell(row=row_number, column=column, value="" if value is None else str(value))
            cell.number_format = "@"
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _copy_selected_templates(campaign: Campaign) -> list[dict[str, str]]:
    paths = resolve_job_paths(campaign.job_id)
    paths.templates_dir.mkdir(parents=True, exist_ok=True)
    copied: list[dict[str, str]] = []
    kinds = document_mode_kinds(campaign.document_mode)

    for stale_name in (KP_TEMPLATE_FILENAME, KP_TEMPLATE_PDF_FILENAME, CONTRACT_TEMPLATE_FILENAME):
        (paths.templates_dir / stale_name).unlink(missing_ok=True)

    if "kp" in kinds:
        if not campaign.kp_template_id:
            raise ValueError("Выберите шаблон КП")
        item = template_service.get_template_file(campaign.kp_template_id, campaign.owner_username)
        if item is None:
            raise ValueError("Файл выбранного шаблона КП не найден")
        suffix = Path(str(item["filename"])).suffix.lower()
        if suffix not in {".docx", ".pdf"}:
            raise ValueError("Шаблон КП должен быть в формате DOCX или PDF")
        destination = paths.templates_dir / (KP_TEMPLATE_PDF_FILENAME if suffix == ".pdf" else KP_TEMPLATE_FILENAME)
        destination.write_bytes(item["content"])
        copied.append({"kind": "kp", "filename": str(item["filename"]), "stored_as": destination.name})

    if "contract" in kinds:
        if not campaign.contract_template_id:
            raise ValueError("Выберите шаблон договора")
        item = template_service.get_template_file(campaign.contract_template_id, campaign.owner_username)
        if item is None:
            raise ValueError("Файл выбранного шаблона договора не найден")
        if Path(str(item["filename"])).suffix.lower() != ".docx":
            raise ValueError("Шаблон договора должен быть в формате DOCX")
        destination = paths.templates_dir / CONTRACT_TEMPLATE_FILENAME
        destination.write_bytes(item["content"])
        copied.append({"kind": "contract", "filename": str(item["filename"]), "stored_as": destination.name})
    return copied


def _reset_generation(job_id: str) -> None:
    generator = get_generator_status(job_id, include_details=False)
    philologist = get_philologist_status(job_id, include_details=False)
    if str(generator.get("status")) == "running" or str(philologist.get("status")) in {"running", "finalizing"}:
        raise RuntimeError("Подготовка документов уже выполняется. Дождитесь завершения или остановите её.")
    paths = resolve_job_paths(job_id)
    for directory in (paths.output_dir, paths.batch_docx_dir, paths.batch_pdf_dir):
        shutil.rmtree(directory, ignore_errors=True)
        directory.mkdir(parents=True, exist_ok=True)
    shutil.rmtree(paths.root_dir / "state" / "template_preview", ignore_errors=True)
    save_agent_state("generator", deepcopy(GENERATOR_STATE), job_id)
    save_agent_state("philologist", deepcopy(PHILOLOGIST_STATE), job_id)


def prepare_campaign_generation(
    campaign_id: str,
    owner_username: str,
    *,
    is_admin: bool = False,
) -> dict[str, Any]:
    campaign, recipients = _load_campaign_rows(campaign_id, owner_username, is_admin=is_admin)
    if not campaign.job_id:
        raise ValueError("У рассылки не создано рабочее пространство")
    if not recipients:
        raise ValueError("Добавьте хотя бы одного получателя")
    signature = _signature(campaign, recipients)
    previous = _load_manifest(campaign.job_id)
    if previous.get("signature") != signature:
        _reset_generation(campaign.job_id)

    paths = resolve_job_paths(campaign.job_id)
    paths.ensure_dirs()
    _write_data_xlsx(paths.data_xlsx, recipients)
    import_clients_from_xlsx(campaign.job_id, paths.data_xlsx)
    templates = _copy_selected_templates(campaign)
    manifest = {
        "campaign_id": campaign.id,
        "job_id": campaign.job_id,
        "signature": signature,
        "document_mode": normalize_document_mode(campaign.document_mode),
        "work_type": campaign.work_type or "",
        "recipient_count": len(recipients),
        "templates": templates,
        "prepared_at": _now_iso(),
    }
    write_json_atomic(_manifest_path(campaign.job_id), manifest)
    return generation_status(campaign_id, owner_username, is_admin=is_admin)


def _compact_status(job_id: str, document_mode: str) -> dict[str, Any]:
    try:
        from src.web.documents_service import compact_documents_status

        return compact_documents_status(job_id, document_mode)
    except RuntimeError:
        generator = get_generator_status(job_id, include_details=False)
        philologist = get_philologist_status(job_id, include_details=False)
        generator_done = generator.get("status") == "completed"
        philologist_done = philologist.get("status") == "completed"
        status = "completed" if generator_done and philologist_done else str(generator.get("status") or "idle")
        return {
            "status": status,
            "output_ready": status == "completed" and int(generator.get("output_file_count") or 0) > 0,
            "progress_percent": 100 if status == "completed" else 0,
            "stage_text": str(generator.get("stage_text") or generator.get("summary_text") or ""),
            "generator": generator,
            "philologist": philologist,
        }


def _recipient_documents_ready(
    recipient_id: int | str,
    *,
    job_id: str,
    document_mode: str,
) -> bool:
    from src.generator.delivery.sender_agent import _resolve_output_folder, _resolve_pdf_attachments

    row_id = str(recipient_id)
    output_dir = resolve_job_paths(job_id).output_dir
    folder, folder_error = _resolve_output_folder(row_id, output_dir=output_dir)
    if folder_error:
        return False
    _, attachment_error = _resolve_pdf_attachments(folder, attachment_mode=document_mode)
    return attachment_error is None


def ensure_campaign_workspace(
    campaign_id: str,
    owner_username: str,
    *,
    is_admin: bool = False,
) -> str:
    """Sync data.xlsx, template files, and manifest before send-time generation."""
    result = prepare_campaign_generation(campaign_id, owner_username, is_admin=is_admin)
    return str(result.get("job_id") or "")


def ensure_recipient_documents(
    *,
    campaign_id: str,
    recipient_id: int,
    owner_username: str,
    job_id: str,
    document_mode: str | None = None,
    work_type: str | None = None,
    is_admin: bool = False,
) -> None:
    """Generate personalized documents for one recipient if output is missing."""
    campaign, _ = _load_campaign_rows(campaign_id, owner_username, is_admin=is_admin)
    effective_mode = normalize_document_mode(document_mode or campaign.document_mode)
    effective_work_type = work_type if work_type is not None else (campaign.work_type or None)
    row_id = str(recipient_id)

    if _recipient_documents_ready(recipient_id, job_id=job_id, document_mode=effective_mode):
        return

    from src.generator.generation.generator_agent import run_generator_agent

    generator_result = run_generator_agent(
        row_ids=[row_id],
        job_id=job_id,
        document_mode=effective_mode,
        work_type=effective_work_type,
        auto_run_philologist=True,
    )
    status = str(generator_result.get("status") or "")
    if status != "completed":
        summary = str(generator_result.get("summary_text") or "Не удалось сформировать документы")
        raise RuntimeError(summary)

    row_result = next(
        (item for item in (generator_result.get("results") or []) if str(item.get("id")) == row_id),
        None,
    )
    if row_result and str(row_result.get("status") or "") == "error":
        raise RuntimeError(str(row_result.get("error") or "Не удалось сформировать документы"))

    if not _recipient_documents_ready(recipient_id, job_id=job_id, document_mode=effective_mode):
        raise RuntimeError(f"Документы для получателя {recipient_id} не готовы после генерации")


def ensure_recipient_documents_for_job(
    *,
    job_id: str,
    row_id: str,
    owner_username: str | None = None,
    attachment_mode: str | None = None,
    work_type: str | None = None,
) -> None:
    """Resolve campaign by job_id and ensure documents exist before materials send."""
    with session_scope() as session:
        campaign = session.scalar(
            select(Campaign)
            .where(Campaign.job_id == job_id)
            .order_by(Campaign.updated_at.desc())
            .limit(1)
        )
        if campaign is None:
            raise RuntimeError(f"Рассылка для job_id={job_id} не найдена")
        session.expunge(campaign)

    owner = owner_username or campaign.owner_username
    ensure_campaign_workspace(campaign.id, owner)
    ensure_recipient_documents(
        campaign_id=campaign.id,
        recipient_id=int(row_id),
        owner_username=owner,
        job_id=job_id,
        document_mode=attachment_mode or campaign.document_mode,
        work_type=work_type or campaign.work_type,
    )


def generation_status(
    campaign_id: str,
    owner_username: str,
    *,
    is_admin: bool = False,
) -> dict[str, Any]:
    campaign, recipients = _load_campaign_rows(campaign_id, owner_username, is_admin=is_admin)
    if not campaign.job_id:
        return {"prepared": False, "stale": False, "ready": False, "status": "not_prepared"}
    manifest = _load_manifest(campaign.job_id)
    prepared = bool(manifest)
    stale = prepared and manifest.get("signature") != _signature(campaign, recipients)
    documents = _compact_status(campaign.job_id, campaign.document_mode) if prepared else {"status": "idle"}
    ready = bool(prepared and not stale and documents.get("output_ready"))
    status = "stale" if stale else ("ready" if ready else str(documents.get("status") or "idle"))
    return {
        "campaign_id": campaign.id,
        "job_id": campaign.job_id,
        "document_mode": normalize_document_mode(campaign.document_mode),
        "work_type": campaign.work_type or "",
        "prepared": prepared,
        "stale": stale,
        "ready": ready,
        "status": status,
        "manifest": manifest,
        "documents": documents,
    }

