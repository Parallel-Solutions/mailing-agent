# -*- coding: utf-8 -*-
"""
Ищет адреса из файла с невалидными почтами в столбце I трёх датасетов,
печатает совпадения в консоль и ВЫДЕЛЯЕТ найденные строки жёлтым цветом
прямо в исходных файлах.

ВАЖНО:
  - Перед изменением каждого файла создаётся резервная копия (..._backup.xlsx).
  - Файлы нужно закрыть в Excel, иначе Windows не даст их перезаписать.
  - Нужен openpyxl:  pip install openpyxl
"""

import os
from shutil import copy2
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ---- Пути к файлам -------------------------------------------------------

SEARCH_FILES = [
    r"C:\Users\User\Downloads\Почты\Датасет по МОКам в РФ на 10.06.26.xlsx",
    r"C:\Users\User\Downloads\Почты\Датасет по всем районам в регионах РФ_1060.xlsx",
    r"C:\Users\User\Downloads\Почты\1. Общий датасет (10817 МО, (Все сельские и городские поселения - действующуе).xlsx",
]

INVALID_FILE = r"C:\Users\User\Downloads\невалидные почты (1) (1).xlsx"

# ---- Настройки -----------------------------------------------------------

MAKE_BACKUP = True          # создавать резервную копию перед изменением
HIGHLIGHT = True            # выделять найденные строки жёлтым

# Столбцы I и J (1 = A, поэтому I = 9, J = 10). Ищем только по I.
COL_I = 9
DATA_START_ROW = 3          # в датасетах данные с 3-й строки
INVALID_START_ROW = 2       # в файле невалидных — со 2-й строки
INVALID_COL = 1             # столбец A

YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def norm(value):
    """Приводим адрес к единому виду для сравнения: строка, без пробелов, нижний регистр."""
    if value is None:
        return ""
    return str(value).strip().lower()


def emails_in_cell(value):
    """Возвращает список адресов из ячейки (на случай нескольких через запятую)."""
    if value is None:
        return []
    return [norm(part) for part in str(value).split(",") if norm(part)]


def load_invalid_emails(path):
    """Читаем невалидные почты из столбца A (по одной в ячейке). Файл только на чтение."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    emails = set()
    for row in ws.iter_rows(min_row=INVALID_START_ROW, min_col=INVALID_COL, max_col=INVALID_COL):
        email = norm(row[0].value if row else None)
        if email:
            emails.add(email)
    wb.close()
    return emails


def backup_path_for(path):
    """Путь резервной копии: рядом с оригиналом, с припиской _backup."""
    base, ext = os.path.splitext(path)
    return base + "_backup" + ext


def process_file(path, invalid_emails):
    """Ищем в столбце I, печатаем совпадения и красим найденные строки жёлтым."""
    print(f"\n=== Файл: {path} ===")

    # --- резервная копия ---
    if HIGHLIGHT and MAKE_BACKUP:
        bpath = backup_path_for(path)
        if os.path.exists(bpath):
            print(f"  (резервная копия уже существует: {bpath})")
        else:
            copy2(path, bpath)
            print(f"  Создана резервная копия: {bpath}")

    # Открываем на запись (без read_only и без data_only, чтобы сохранить формулы/стили).
    wb = load_workbook(path)
    ws = wb.active
    max_col = ws.max_column

    found = 0
    changed = False

    for row_num in range(DATA_START_ROW, ws.max_row + 1):
        value_i = ws.cell(row=row_num, column=COL_I).value
        matched = False

        for email in emails_in_cell(value_i):
            if email in invalid_emails:
                print(f"  строка {row_num}, столбец I: {email}")
                found += 1
                matched = True

        if matched and HIGHLIGHT:
            # красим всю строку по ширине использованных столбцов
            for c in range(1, max_col + 1):
                ws.cell(row=row_num, column=c).fill = YELLOW
            changed = True

    if HIGHLIGHT and changed:
        wb.save(path)
        print(f"  --- Совпадений: {found}. Строки выделены жёлтым, файл сохранён. ---")
    else:
        print(f"  --- Совпадений: {found}. ---")

    wb.close()
    return found


def main():
    invalid_emails = load_invalid_emails(INVALID_FILE)
    print(f"Загружено невалидных почт для поиска: {len(invalid_emails)}")

    total = 0
    for path in SEARCH_FILES:
        try:
            total += process_file(path, invalid_emails)
        except FileNotFoundError:
            print(f"\n!!! Файл не найден: {path}")
        except PermissionError:
            print(f"\n!!! Не удалось записать файл (закрыт ли он в Excel?): {path}")
        except Exception as e:
            print(f"\n!!! Ошибка при обработке {path}: {e}")

    print(f"\n===== ВСЕГО совпадений во всех файлах: {total} =====")


if __name__ == "__main__":
    main()