from __future__ import annotations

import shutil
import unittest
import uuid
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from src.generator.delivery.auto_call_export import build_auto_call_phone_numbers, write_auto_call_csv
from src.generator.delivery.manager_stats import export_report
from src.generator.delivery.phone_normalize import collect_normalized_phones, normalize_phone_for_auto_call
from src.generator.delivery.sender_report import build_sender_delivery_report_xlsx
from tests.bootstrap import bootstrap_test_runtime


class ReportExportTests(unittest.TestCase):
    def setUp(self) -> None:
        bootstrap_test_runtime(reset_db=True)
        self.tmpdir = Path.cwd() / "tmp" / f"test-report-export-{uuid.uuid4().hex}"
        self.tmpdir.mkdir(parents=True, exist_ok=True)

    def tearDown(self) -> None:
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_xlsx_report_has_five_sheets(self) -> None:
        delivery_rows = [
            {
                "row_id": "1",
                "mun_name": "ООО Тест",
                "recipient": "a@example.com",
                "provider": "rusender",
                "provider_status": "hard_bounced",
                "recipient_role": "primary",
                "accepted_status": "accepted",
                "provider_status_label": "Не доставлено",
                "delivery_response": "user unknown",
                "outcome": "Ошибка",
                "sent_at": "2026-05-01 10:00:00",
                "sent_at_timestamp": "2026-05-01T10:00:00",
                "subject": "Тест",
                "work_type": "stp_mo",
                "campaign_name": "Май",
                "email_id": "1",
                "message_id": "1",
                "checked_at": "2026-05-01 11:00:00",
                "comment": "",
                "recipient_role_label": "Основной",
            }
        ]
        output_path = self.tmpdir / "sender_delivery_report.xlsx"
        with patch("src.generator.delivery.sender_report._build_delivery_rows", return_value=(delivery_rows, "")), patch(
            "src.generator.delivery.sender_report._build_consent_rows", return_value=[]
        ), patch("src.generator.delivery.sender_report.resolve_job_paths") as resolve_paths:
            resolve_paths.return_value.root_dir = self.tmpdir
            resolve_paths.return_value.sent_mail_log_path = self.tmpdir / "sent_mail_log.jsonl"
            report_path = build_sender_delivery_report_xlsx("job-report", refresh=False)
        workbook = load_workbook(report_path)
        self.assertEqual(
            workbook.sheetnames,
            ["Статистика", "Журнал отправки", "Согласия", "Проблемные адреса", "Действия менеджера"],
        )

    def test_csv_export_creates_file_and_history(self) -> None:
        rows = [
            {
                "organization": "ООО Тест",
                "email": "a@example.com",
                "provider": "rusender",
                "manager_status": {"label": "Доставлено"},
                "interest": {"label": "Средний"},
                "next_action": {"label": "Ожидать статус"},
                "sent_at": "2026-05-01",
                "last_event_at": "2026-05-01",
            }
        ]
        with patch("src.generator.delivery.manager_stats._load_delivery_for_jobs", return_value=rows), patch(
            "src.generator.delivery.manager_stats._load_consents_for_jobs", return_value=[]
        ), patch("src.generator.delivery.manager_stats._reports_dir", return_value=self.tmpdir / "reports"), patch(
            "src.generator.delivery.manager_stats.normalize_job_id", return_value="job-report"
        ):
            (self.tmpdir / "reports").mkdir(parents=True, exist_ok=True)
            result = export_report(
                "job-report",
                report_type="delivery_summary",
                fmt="csv",
                author="tester",
            )
        path = Path(result["path"])
        self.assertTrue(path.exists())
        self.assertEqual(path.suffix, ".csv")


class AutoCallExportTests(unittest.TestCase):
    def setUp(self) -> None:
        self.tmpdir = Path.cwd() / "tmp" / f"test-auto-call-export-{uuid.uuid4().hex}"
        self.tmpdir.mkdir(parents=True, exist_ok=True)

    def tearDown(self) -> None:
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_normalize_phone_for_auto_call(self) -> None:
        self.assertEqual(normalize_phone_for_auto_call("+7 (343) 939-69-79"), "73439396979")
        self.assertEqual(normalize_phone_for_auto_call("8-343-939-69-79"), "73439396979")
        self.assertEqual(normalize_phone_for_auto_call("9156848204"), "79156848204")
        self.assertIsNone(normalize_phone_for_auto_call(""))
        self.assertIsNone(normalize_phone_for_auto_call("123"))

    def test_collect_normalized_phones_deduplicates(self) -> None:
        phones = collect_normalized_phones("73439396979", "8 (343) 939-69-79", "TEL_DOP: 79156848204; 79156848204")
        self.assertEqual(phones, ["73439396979", "79156848204"])

    def test_write_auto_call_csv(self) -> None:
        output_path = self.tmpdir / "phones.csv"
        write_auto_call_csv(output_path, ["73439396979", "79156848204"])
        self.assertEqual(output_path.read_text(encoding="utf-8-sig").splitlines(), ["phone_number", "73439396979", "79156848204"])

    def test_build_auto_call_phone_numbers_from_rows(self) -> None:
        rows = [{"TEL_OSN": "73439396979"}, {"TEL_DOP": "79156848204"}]
        data_path = self.tmpdir / "data.xlsx"
        data_path.write_bytes(b"stub")
        with patch("src.generator.delivery.auto_call_export._resolve_sender_data_xlsx_path", return_value=data_path), patch(
            "src.generator.delivery.auto_call_export.load_rows",
            return_value=(None, None, rows),
        ):
            phones = build_auto_call_phone_numbers("job-call")
        self.assertEqual(phones, ["73439396979", "79156848204"])


class AutoCallReportExportTests(unittest.TestCase):
    def setUp(self) -> None:
        bootstrap_test_runtime(reset_db=True)
        self.tmpdir = Path.cwd() / "tmp" / f"test-auto-call-report-{uuid.uuid4().hex}"
        self.tmpdir.mkdir(parents=True, exist_ok=True)

    def tearDown(self) -> None:
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_auto_call_csv_export(self) -> None:
        rows = [
            {"TEL_OSN": "+7 (343) 939-69-79", "TEL_DOP": "8-343-939-69-79"},
            {"TEL_OSN": "9156848204", "TEL_DOP": ""},
            {"TEL_OSN": "invalid", "TEL_DOP": ""},
        ]
        data_path = self.tmpdir / "data.xlsx"
        data_path.write_bytes(b"stub")
        with patch("src.generator.delivery.auto_call_export._resolve_sender_data_xlsx_path", return_value=data_path), patch(
            "src.generator.delivery.auto_call_export.load_rows",
            return_value=(None, None, rows),
        ), patch("src.generator.delivery.manager_stats._reports_dir", return_value=self.tmpdir / "reports"), patch(
            "src.generator.delivery.manager_stats.normalize_job_id", return_value="job-call"
        ):
            (self.tmpdir / "reports").mkdir(parents=True, exist_ok=True)
            result = export_report(
                "job-call",
                report_type="auto_call_contacts",
                fmt="csv",
                author="tester",
            )
        path = Path(result["path"])
        content = path.read_text(encoding="utf-8-sig")
        self.assertEqual(content.splitlines(), ["phone_number", "73439396979", "79156848204"])


if __name__ == "__main__":
    unittest.main()
