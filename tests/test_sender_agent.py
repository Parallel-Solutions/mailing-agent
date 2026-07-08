from __future__ import annotations

import io
import json
from datetime import datetime
from pathlib import Path
from types import SimpleNamespace
import unittest
from unittest.mock import patch
from urllib.error import HTTPError, URLError
from urllib.request import Request

from openpyxl import load_workbook

from src.generator.delivery import mailopost_events
from src.generator.delivery import sender_agent
from src.generator.delivery import sender_report
from src.web import sender_service
from src.jobs.job_docs import append_event
from tests.bootstrap import reset_test_database


class SenderAgentScalabilityTests(unittest.TestCase):
    def _tmp_dir(self, name: str) -> Path:
        tmpdir = Path.cwd() / "tmp" / "test_sender_agent" / name
        tmpdir.mkdir(parents=True, exist_ok=True)
        return tmpdir

    def test_state_rows_snapshot_keeps_recent_entries(self) -> None:
        rows = [{"id": index} for index in range(sender_agent.SENDER_STATE_ROWS_LIMIT + 25)]

        snapshot = sender_agent._state_rows_snapshot(rows)

        self.assertEqual(len(snapshot), sender_agent.SENDER_STATE_ROWS_LIMIT)
        self.assertEqual(snapshot[0]["id"], 25)
        self.assertEqual(snapshot[-1]["id"], sender_agent.SENDER_STATE_ROWS_LIMIT + 24)

    def test_should_flush_sender_workbook_every_batch_and_on_finish(self) -> None:
        self.assertFalse(
            sender_agent._should_flush_sender_workbook(dirty=False, processed_rows=25, total_rows=100)
        )
        self.assertTrue(
            sender_agent._should_flush_sender_workbook(
                dirty=True,
                processed_rows=sender_agent.SENDER_WORKBOOK_SAVE_EVERY,
                total_rows=100,
            )
        )
        self.assertTrue(
            sender_agent._should_flush_sender_workbook(dirty=True, processed_rows=100, total_rows=100)
        )

    def test_run_sender_enters_job_lock_before_work(self) -> None:
        events: list[tuple[str, str | None]] = []

        class FakeLock:
            def __init__(self, job_id: str | None) -> None:
                self.job_id = job_id

            def __enter__(self):
                events.append(("enter", self.job_id))
                return self

            def __exit__(self, exc_type, exc, tb) -> None:
                events.append(("exit", self.job_id))

        job_paths = SimpleNamespace(
            uses_legacy_layout=True,
            templates_dir=Path("missing-templates"),
            output_dir=Path("missing-output"),
            sent_mail_log_path=Path("missing-log.jsonl"),
        )

        with patch.object(sender_agent, "_sender_run_lock", side_effect=lambda job_id: FakeLock(job_id)), patch.object(
            sender_agent, "resolve_job_paths", return_value=job_paths
        ), patch.object(
            sender_agent, "_resolve_sender_data_xlsx_path", return_value=Path("missing-data.xlsx")
        ), patch.object(
            sender_agent, "_load_sender_state", side_effect=lambda job_id=None: dict(sender_agent.SENDER_STATE)
        ), patch.object(
            sender_agent, "_save_sender_state", side_effect=lambda state, job_id=None: state
        ), patch.object(
            sender_agent, "clear_sender_stop_request", return_value=None
        ), patch.object(
            sender_agent, "_collect_excel_stats", return_value={"total": 0, "sent": 0, "error": 0, "pending": 0}
        ):
            result = sender_agent.run_sender(dry_run=True, job_id="job-lock")

        self.assertEqual(events, [("enter", "job-lock"), ("exit", "job-lock")])
        self.assertEqual(result["status"], "error")
        self.assertIn("data.xlsx", result["summary_text"])

    def test_unisender_parallel_workers_only_for_real_unisender_send(self) -> None:
        with patch.object(sender_agent.settings, "sender_unisender_concurrency", 7):
            self.assertEqual(sender_agent._unisender_parallel_workers(dry_run=False, transport="unisender"), 7)
            self.assertEqual(sender_agent._unisender_parallel_workers(dry_run=True, transport="unisender"), 1)
            self.assertEqual(sender_agent._unisender_parallel_workers(dry_run=False, transport="smtp"), 1)

    def test_mail_footer_html_uses_versioned_image(self) -> None:
        with patch.object(
            sender_agent.settings,
            "mail_signature_image_url",
            "https://offer.parresh.ru/public/mail-signature.png?v=konstantin",
        ):
            body = sender_agent._append_mail_footer_text("Здравствуйте!")
            html = sender_agent._htmlify_mail_body(body, include_unsubscribe=False)

        self.assertIn("<img ", html)
        self.assertIn("&sig=", html)
        self.assertNotIn("Черкашина", html)
        self.assertNotIn("Крашенинников", html)

    def test_mail_template_is_required_when_no_body_override(self) -> None:
        self.assertFalse(hasattr(sender_agent, "DEFAULT_MAIL_BODY"))
        with self.assertRaisesRegex(RuntimeError, "Не загружен шаблон письма"):
            sender_agent._read_mail_template(None)
        tmpdir = self._tmp_dir("missing_template")
        missing_path = tmpdir / "mail_template.txt"
        if missing_path.exists():
            missing_path.unlink()
        with self.assertRaisesRegex(RuntimeError, "Не загружен шаблон письма"):
            sender_agent._read_mail_template(missing_path)

    def test_empty_mail_template_is_rejected(self) -> None:
        tmpdir = self._tmp_dir("empty_template")
        template_path = tmpdir / "mail_template.txt"
        template_path.write_text("   \n", encoding="utf-8")

        with self.assertRaisesRegex(RuntimeError, "Загруженный шаблон письма пустой"):
            sender_agent._read_mail_template(template_path)

    def test_mail_body_uses_uploaded_template(self) -> None:
        tmpdir = self._tmp_dir("uploaded_template")
        template_path = tmpdir / "mail_template.txt"
        template_path.write_text("Здравствуйте, [наименование муниципального образования]!", encoding="utf-8")

        body = sender_agent._build_mail_body({"MUN_NAME": "Тестовое МО"}, mail_template_path=template_path)

        self.assertIn("Здравствуйте, Тестовое МО!", body)
        self.assertNotIn("Направляем для рассмотрения", body)

    def test_sent_mail_recipients_are_scoped_to_current_send_run(self) -> None:
        item = {"send_run_id": "send-new", "sent_at": "2026-06-02T12:00:00"}

        self.assertTrue(
            sender_agent._sent_log_item_in_send_scope(
                item,
                send_run_id="send-new",
                send_run_started_at="2026-06-02T12:00:00",
            )
        )
        self.assertFalse(
            sender_agent._sent_log_item_in_send_scope(
                item,
                send_run_id="send-old",
                send_run_started_at="2026-06-02T12:00:00",
            )
        )

    def test_unisender_analytics_filters_current_send_run_items(self) -> None:
        items = [
            {"send_run_id": "send-old", "sent_at": "2026-06-01T10:00:00", "row_id": "1"},
            {"send_run_id": "send-new", "sent_at": "2026-06-02T10:00:00", "row_id": "2"},
        ]

        scoped = sender_report._filter_items_by_send_scope(
            items,
            send_run_id="send-new",
            send_run_started_at="2026-06-02T09:00:00",
        )

        self.assertEqual([item["row_id"] for item in scoped], ["2"])

    def test_unisender_analytics_filters_by_complete_sender_state(self) -> None:
        items = [
            {"row_id": "1", "recipient": "one@example.com"},
            {"row_id": "1", "recipient": "old-one@example.com"},
            {"row_id": "2", "recipient": "two@example.com"},
        ]
        state = {
            "total_rows": 2,
            "rows": [
                {"id": "1", "sent_recipients": ["one@example.com"]},
                {"id": "2", "recipient": "two@example.com"},
            ],
        }

        with patch.object(sender_report, "_load_sender_state", return_value=state):
            scoped = sender_report._filter_items_by_current_sender_state("job-current", items)

        self.assertEqual(
            [(item["row_id"], item["recipient"]) for item in scoped],
            [("1", "one@example.com"), ("2", "two@example.com")],
        )

    def test_unisender_analytics_does_not_filter_by_partial_sender_state(self) -> None:
        items = [
            {"row_id": "1", "recipient": "one@example.com"},
            {"row_id": "2", "recipient": "two@example.com"},
        ]
        state = {"total_rows": 3, "rows": [{"id": "1", "sent_recipients": ["one@example.com"]}]}

        with patch.object(sender_report, "_load_sender_state", return_value=state):
            scoped = sender_report._filter_items_by_current_sender_state("job-current", items)

        self.assertEqual(scoped, items)

    def test_sender_analytics_does_not_filter_by_scoped_consent_followup_state(self) -> None:
        items = [
            {"row_id": "1", "recipient": "one@example.com"},
            {"row_id": "2", "recipient": "two@example.com"},
        ]
        state = {
            "selection_scoped": True,
            "total_rows": 1,
            "rows": [{"id": "2", "recipient": "two@example.com"}],
        }

        with patch.object(sender_report, "_load_sender_state", return_value=state):
            scoped = sender_report._filter_items_by_current_sender_state("job-current", items)

        self.assertEqual(scoped, items)

    def test_sender_analytics_keeps_campaign_after_materials_followup_state(self) -> None:
        items = [
            {"row_id": "1", "recipient": "one@example.com"},
            {"row_id": "77", "recipient": "terbuny@example.com"},
        ]
        state = {
            "send_mode": "materials",
            "total_rows": 1,
            "rows": [{"id": "77", "sent_recipients": ["terbuny@example.com"]}],
            "send_run_id": "send-materials",
            "send_run_started_at": "2026-06-29T11:57:00",
        }
        consent_records = [{"row_id": str(index), "recipient": f"person{index}@example.com"} for index in range(1, 4)]

        with patch.object(sender_report, "_load_sender_state", return_value=state), patch.object(
            sender_report, "load_consent_records", return_value=consent_records
        ):
            scoped = sender_report._filter_items_by_current_sender_state("job-current", items)
            send_run_id, send_run_started_at = sender_report._current_send_scope("job-current")

        self.assertEqual(scoped, items)
        self.assertEqual((send_run_id, send_run_started_at), ("", ""))

    def test_sender_analytics_uses_first_sent_time_and_campaign_name(self) -> None:
        delivery_rows = [
            {
                "provider": "rusender",
                "provider_status": "delivered",
                "work_type": "stp_mo",
                "sent_at_timestamp": "2026-07-02T13:47:32+03:00",
                "checked_at": "2026-07-02T10:48:00",
            },
            {
                "provider": "rusender",
                "provider_status": "delivered",
                "work_type": "stp_mo",
                "sent_at_timestamp": "2026-07-02T13:55:00+03:00",
                "checked_at": "2026-07-02T10:56:00",
            },
        ]
        now = datetime(2026, 7, 2, 14, 30, 0, tzinfo=sender_report.MOSCOW_TZ)

        with patch.object(sender_report, "_build_delivery_rows", return_value=(delivery_rows, "")), patch.object(
            sender_report, "load_consent_records", return_value=[]
        ), patch.object(sender_report, "_load_sender_state", return_value={"work_type": "stp_mo", "campaign_name": "июльская рассылка"}), patch.object(
            sender_report, "_now_moscow", return_value=now
        ):
            analytics = sender_report.build_sender_delivery_analytics("job-current", refresh=False)

        self.assertEqual(analytics["generated_at"], "2026-07-02T14:30:00+03:00")
        self.assertEqual(analytics["generated_at_label"], "2026-07-02 14:30:00")
        self.assertEqual(analytics["campaign_started_at"], "2026-07-02T13:47:32+03:00")
        self.assertEqual(analytics["campaign_started_at_label"], "2026-07-02 13:47:32")
        self.assertEqual(analytics["stats_time_label"], "2026-07-02 13:47:32")
        self.assertEqual(analytics["campaign"]["title"], "июльская рассылка")
        self.assertEqual(analytics["campaign"]["custom_name"], "июльская рассылка")
        self.assertEqual(analytics["work_type_label"], "СТП МО")
        self.assertEqual(sender_report._format_moscow_datetime("2026-07-02T10:47:32"), "2026-07-02 13:47:32")

    def test_sender_analytics_and_report_include_consents(self) -> None:
        delivery_rows = [
            {
                "row_id": "1",
                "mun_name": "МО",
                "recipient": "one@example.com",
                "recipient_role": "primary",
                "recipient_role_label": "Основной",
                "sent_at": "2026-06-29T10:00:00",
                "subject": "Тема",
                "accepted_status": "sent",
                "provider": "mailopost",
                "provider_status": "delivered",
                "provider_status_label": "Доставлено",
                "delivery_response": "250 2.0.0 queued as OK",
                "outcome": "Успешно",
                "email_id": "message-1",
                "message_id": "message-1",
                "checked_at": "2026-06-29T10:01:00",
                "comment": "",
            }
        ]
        consent_records = [
            {
                "row_id": "1",
                "mun_name": "МО",
                "recipient": "one@example.com",
                "status": "confirmed",
                "request_sent_at": "2026-06-29T10:00:00",
                "confirmed_at": "2026-06-29T10:02:00",
                "materials_status": "sent",
                "materials_sent_at": "2026-06-29T10:03:00",
                "transport": "mailopost",
                "attachment_mode": "kp",
            }
        ]

        with patch.object(sender_report, "_build_delivery_rows", return_value=(delivery_rows, "")), patch.object(
            sender_report, "load_consent_records", return_value=consent_records
        ):
            analytics = sender_report.build_sender_delivery_analytics("job-consent")

        self.assertEqual(analytics["summary"]["consents"]["confirmed"], 1)
        self.assertIn("confirmed_consents", {card["id"] for card in analytics["cards"]})

        tmp_dir = Path("tmp") / "sender-report-consents-test"
        tmp_dir.mkdir(parents=True, exist_ok=True)
        with patch.object(
            sender_report,
            "resolve_job_paths",
            return_value=SimpleNamespace(root_dir=tmp_dir),
        ), patch.object(sender_report, "_build_delivery_rows", return_value=(delivery_rows, "")), patch.object(
            sender_report, "load_consent_records", return_value=consent_records
        ):
            report_path = sender_report.build_sender_delivery_report_xlsx("job-consent", refresh=False)
            workbook = load_workbook(report_path, read_only=True)
            try:
                self.assertIn("Согласия", workbook.sheetnames)
                journal_sheet = workbook["Журнал отправки"]
                self.assertEqual(journal_sheet[3][8].value, "Причина недоставки")
                self.assertEqual(journal_sheet[4][8].value, "250 2.0.0 queued as OK")
                self.assertEqual(journal_sheet[3][14].value, "Тип адреса")
                self.assertEqual(journal_sheet[4][14].value, "Основной")
                consent_sheet = workbook["Согласия"]
                self.assertEqual(consent_sheet[3][3].value, "Статус согласия")
                self.assertEqual(consent_sheet[4][3].value, "Согласие получено")
            finally:
                workbook.close()

    def test_compact_sender_status_shows_table_totals_after_scoped_material_send(self) -> None:
        state = {
            "status": "completed",
            "mode": "send",
            "send_mode": "materials",
            "selection_scoped": True,
            "processed_rows": 1,
            "ready_rows": 1,
            "sent_rows": 1,
            "error_rows": 0,
            "total_rows": 1,
            "remaining_rows": 0,
            "stats": {"total": 2, "sent": 2, "error": 0, "pending": 0},
        }

        compact = sender_service.compact_sender_status(state)

        self.assertEqual(compact["total_rows"], 2)
        self.assertEqual(compact["sent_rows"], 2)
        self.assertEqual(compact["processed_rows"], 2)
        self.assertEqual(compact["ready_rows"], 2)
        self.assertEqual(compact["stats"], {"total": 2, "sent": 2, "error": 0, "pending": 0})

    def test_compact_sender_status_uses_campaign_log_after_scoped_consent_fallback(self) -> None:
        reset_test_database()
        campaign_name = "СТП_Регионы на букву К_146"
        items = [
            {
                "row_id": str(row_id),
                "recipient": f"primary{row_id}@example.com",
                "send_mode": "consent_request",
                "recipient_role": "primary",
                "campaign_name": campaign_name,
            }
            for row_id in range(1, 147)
        ]
        items.extend(
            {
                "row_id": str(row_id),
                "recipient": f"fallback{row_id}@example.com",
                "send_mode": "consent_request",
                "recipient_role": "fallback",
                "campaign_name": campaign_name,
            }
            for row_id in (3, 11, 13, 16, 29, 139, 140)
        )
        for item in items:
            append_event("job-campaign", "sent_mail_log", item)
        state = {
            "job_id": "job-campaign",
            "status": "completed",
            "mode": "send",
            "send_mode": "consent_request",
            "campaign_name": campaign_name,
            "selection_scoped": True,
            "processed_rows": 7,
            "ready_rows": 7,
            "sent_rows": 7,
            "error_rows": 0,
            "total_rows": 7,
            "remaining_rows": 0,
            "stats": {"total": 146, "sent": 146, "error": 0, "pending": 0},
        }

        compact = sender_service.compact_sender_status(state)

        self.assertEqual(compact["total_rows"], 146)
        self.assertEqual(compact["sent_rows"], 146)
        self.assertEqual(compact["processed_rows"], 146)
        self.assertEqual(compact["ready_rows"], 146)
        self.assertEqual(compact["stats"], {"total": 146, "sent": 146, "error": 0, "pending": 0})
        self.assertTrue(compact["campaign_scope_applied"])
        self.assertEqual(compact["campaign_email_sends"], 153)
        self.assertEqual(compact["summary_text"], "Запросы согласия отправлены. Отправлено: 146.")

    def test_compact_sender_status_keeps_running_scoped_consent_progress(self) -> None:
        state = {
            "job_id": "job-campaign",
            "status": "running",
            "mode": "send",
            "send_mode": "consent_request",
            "campaign_name": "СТП_Регионы на букву К_146",
            "selection_scoped": True,
            "processed_rows": 13,
            "ready_rows": 13,
            "sent_rows": 13,
            "error_rows": 0,
            "total_rows": 15,
            "remaining_rows": 2,
            "stats": {"total": 48, "sent": 33, "error": 0, "pending": 15},
        }

        compact = sender_service.compact_sender_status(state)

        self.assertEqual(compact["total_rows"], 15)
        self.assertEqual(compact["sent_rows"], 13)
        self.assertEqual(compact["processed_rows"], 13)
        self.assertEqual(compact["stats"], {"total": 15, "sent": 13, "error": 0, "pending": 2})
        self.assertFalse(compact["campaign_scope_applied"])


    def test_materials_send_requires_confirmed_consent_even_when_caller_omits_flag(self) -> None:
        row = {
            "ID": "1",
            "_row_index": 2,
            "MUN_NAME": "Test municipality",
            "EMAIL_OSN": "recipient@example.com",
            "EMAIL_DOP": "",
            "STATUS": "",
        }

        def load_state(job_id: str | None = None) -> dict:
            state = dict(sender_agent.SENDER_STATE)
            state["rows"] = []
            state["stats"] = {}
            return state

        with patch.object(sender_agent, "_load_sender_state", side_effect=load_state), patch.object(
            sender_agent, "_save_sender_state", side_effect=lambda state, job_id=None: state
        ), patch.object(
            sender_agent, "_resolve_sender_data_xlsx_path", return_value=Path(__file__)
        ), patch.object(
            sender_agent, "_collect_excel_stats", return_value={"total": 1, "sent": 0, "error": 0, "pending": 1}
        ), patch.object(
            sender_agent, "load_rows", return_value=(SimpleNamespace(close=lambda: None), SimpleNamespace(), [row])
        ), patch.object(
            sender_agent, "_build_output_folder_index", return_value=({}, {})
        ), patch.object(
            sender_agent, "has_confirmed_consent", return_value=False
        ) as has_confirmed_consent, patch.object(
            sender_agent, "_resolve_output_folder"
        ) as resolve_output_folder, patch.object(
            sender_agent, "count_tasks_for_agent", return_value={}
        ), patch.object(
            sender_agent, "get_tasks_for_agent", return_value=[]
        ), patch.object(
            sender_agent, "get_recent_events", return_value=[]
        ), patch.object(
            sender_agent.settings, "email_validation_mode", "syntax"
        ):
            result = sender_agent.run_sender(
                dry_run=True,
                send_mode="materials",
                job_id="job-consent",
            )

        self.assertEqual(result["error_rows"], 1)
        self.assertEqual(result["rows"][0]["result"], "blocked_no_consent")
        self.assertIn("согласия", result["rows"][0]["error"])
        has_confirmed_consent.assert_called_once_with(
            job_id="job-consent",
            row_id="1",
            recipient="recipient@example.com",
            attachment_mode="kp",
        )
        resolve_output_folder.assert_not_called()

    def test_materials_after_consent_sends_to_confirmed_recipient(self) -> None:
        row = {
            "ID": "37",
            "_row_index": 2,
            "MUN_NAME": "Test municipality",
            "EMAIL_OSN": "primary@example.com",
            "EMAIL_DOP": "",
            "STATUS": "",
        }
        sent_recipients: list[str] = []

        def load_state(job_id: str | None = None) -> dict:
            state = dict(sender_agent.SENDER_STATE)
            state["rows"] = []
            state["stats"] = {}
            return state

        def fake_send(row, recipients, attachments, subject, **kwargs):
            sent_recipients.extend(recipients)
            return {
                "recipient": recipients[0],
                "recipients": list(recipients),
                "attempts": [{"recipient": recipients[0], "status": "sent", "error": ""}],
                "error": "",
                "warning": "",
            }

        def fake_has_consent(**kwargs):
            return kwargs.get("recipient") == "consent@example.com"

        with patch.object(sender_agent, "_load_sender_state", side_effect=load_state), patch.object(
            sender_agent, "_save_sender_state", side_effect=lambda state, job_id=None: state
        ), patch.object(
            sender_agent, "clear_sender_stop_request", return_value=None
        ), patch.object(
            sender_agent, "_resolve_sender_data_xlsx_path", return_value=Path(__file__)
        ), patch.object(
            sender_agent, "_collect_excel_stats", return_value={"total": 1, "sent": 0, "error": 0, "pending": 1}
        ), patch.object(
            sender_agent, "load_rows", return_value=(SimpleNamespace(close=lambda: None), SimpleNamespace(), [row])
        ), patch.object(
            sender_agent, "_build_output_folder_index", return_value=({}, {})
        ), patch.object(
            sender_agent, "_resolve_output_folder", return_value=(Path("output/37"), None)
        ), patch.object(
            sender_agent, "_resolve_pdf_attachments", return_value=(["kp.pdf"], None)
        ), patch.object(
            sender_agent, "has_confirmed_consent", side_effect=fake_has_consent
        ) as has_confirmed_consent, patch.object(
            sender_agent, "_send_with_transport", side_effect=fake_send
        ), patch.object(
            sender_agent, "_append_sent_mail_log", return_value=None
        ), patch.object(
            sender_agent, "_persist_row_status", return_value=""
        ), patch.object(
            sender_agent, "_should_flush_sender_workbook", return_value=False
        ), patch.object(
            sender_agent, "_load_sent_mail_recipients", return_value={}
        ), patch.object(
            sender_agent, "count_tasks_for_agent", return_value={}
        ), patch.object(
            sender_agent, "get_tasks_for_agent", return_value=[]
        ), patch.object(
            sender_agent, "get_recent_events", return_value=[]
        ), patch.object(
            sender_agent.settings, "email_validation_mode", "syntax"
        ):
            result = sender_agent.run_sender(
                dry_run=False,
                send_mode="materials",
                target_recipient="consent@example.com",
                require_confirmed_consent=True,
                job_id="job-consent",
            )

        self.assertEqual(sent_recipients, ["consent@example.com"])
        self.assertEqual(result["rows"][0]["recipient"], "consent@example.com")
        self.assertEqual(result["rows"][0]["email_strategy"], "consent_recipient")
        has_confirmed_consent.assert_called_once_with(
            job_id="job-consent",
            row_id="37",
            recipient="consent@example.com",
            attachment_mode="kp",
        )

    def test_allowed_send_recipients_includes_primary_and_extra_emails(self) -> None:
        decision = sender_agent._choose_recipient(
            {
                "EMAIL_OSN": "one@example.com",
                "EMAIL_DOP": "two@example.com; one@example.com, bad-email, three@example.com",
            }
        )

        self.assertEqual(decision["recipient"], "one@example.com")
        self.assertEqual(
            sender_agent._allowed_send_recipients(decision),
            ["one@example.com", "two@example.com", "three@example.com"],
        )
        self.assertEqual(
            sender_agent._allowed_send_recipients(
                decision,
                recipient_strategy=sender_agent.RECIPIENT_STRATEGY_PRIMARY_THEN_FALLBACK,
            ),
            ["one@example.com"],
        )
        self.assertEqual(
            sender_agent._fallback_send_recipients(
                decision,
                recipient_strategy=sender_agent.RECIPIENT_STRATEGY_PRIMARY_THEN_FALLBACK,
            ),
            ["two@example.com"],
        )
        self.assertEqual(
            sender_agent._consent_candidate_recipients(
                decision,
                recipient_strategy=sender_agent.RECIPIENT_STRATEGY_PRIMARY_THEN_FALLBACK,
            ),
            ["one@example.com", "two@example.com"],
        )
        self.assertEqual(decision["fallback_candidates"], ["two@example.com", "three@example.com"])
        self.assertEqual(decision["invalid_emails"], ["bad-email"])
        self.assertIn("дополнительные", decision["decision_reason"])

    def test_primary_then_fallback_uses_only_one_extra_when_primary_missing(self) -> None:
        decision = sender_agent._choose_recipient(
            {
                "EMAIL_OSN": "",
                "EMAIL_DOP": "two@example.com; three@example.com",
            }
        )

        self.assertEqual(decision["recipient"], "two@example.com")
        self.assertEqual(
            sender_agent._allowed_send_recipients(
                decision,
                recipient_strategy=sender_agent.RECIPIENT_STRATEGY_PRIMARY_THEN_FALLBACK,
            ),
            ["two@example.com"],
        )
        self.assertEqual(
            sender_agent._fallback_send_recipients(
                decision,
                recipient_strategy=sender_agent.RECIPIENT_STRATEGY_PRIMARY_THEN_FALLBACK,
            ),
            [],
        )
        self.assertEqual(
            sender_agent._consent_candidate_recipients(
                decision,
                recipient_strategy=sender_agent.RECIPIENT_STRATEGY_PRIMARY_THEN_FALLBACK,
            ),
            ["two@example.com"],
        )


    def test_email_validation_filters_invalid_primary_and_keeps_fallback(self) -> None:
        def fake_result(email: str, is_valid: bool) -> sender_agent.EmailValidationResult:
            return sender_agent.EmailValidationResult(
                email=email,
                normalized_email=email.lower(),
                domain=email.split("@", 1)[-1],
                is_valid=is_valid,
                reason_code="ok_domain" if is_valid else "domain_not_found",
                reason="" if is_valid else "Email не прошёл проверку: домен bad.invalid не найден.",
                checked_at="2026-07-02T12:00:00",
                details={"mode": "domain"},
            )

        decision = sender_agent._choose_recipient(
            {
                "EMAIL_OSN": "person@bad.invalid",
                "EMAIL_DOP": "backup@example.com",
            }
        )
        candidates = [
            *sender_agent._allowed_send_recipients(
                decision,
                recipient_strategy=sender_agent.RECIPIENT_STRATEGY_PRIMARY_THEN_FALLBACK,
            ),
            *sender_agent._fallback_send_recipients(
                decision,
                recipient_strategy=sender_agent.RECIPIENT_STRATEGY_PRIMARY_THEN_FALLBACK,
            ),
        ]

        with patch.object(
            sender_agent,
            "validate_email_address",
            side_effect=lambda email, **kwargs: fake_result(email, email == "backup@example.com"),
        ):
            valid_recipients, attempts = sender_agent._filter_validated_recipients(candidates, {})

        self.assertEqual(valid_recipients, ["backup@example.com"])
        self.assertEqual(attempts[0]["recipient"], "person@bad.invalid")
        self.assertIn("не прошёл проверку", attempts[0]["error"])

    def test_preflight_validation_warning_keeps_successful_fallback_send(self) -> None:
        invalid_attempt = {
            "recipient": "person@bad.invalid",
            "status": "error",
            "error": "Email не прошёл проверку: домен bad.invalid не найден.",
        }
        send_result = {
            "recipient": "backup@example.com",
            "recipients": ["backup@example.com"],
            "attempts": [{"recipient": "backup@example.com", "status": "sent", "error": ""}],
            "error": "",
            "warning": "",
        }

        merged = sender_agent._merge_send_result_with_preflight_attempts(send_result, [invalid_attempt])

        self.assertEqual(merged["recipient"], "backup@example.com")
        self.assertEqual([attempt["recipient"] for attempt in merged["attempts"]], ["person@bad.invalid", "backup@example.com"])
        self.assertIn("Пропущены email", merged["warning"])
    def test_recipient_role_for_log_marks_primary_and_fallback(self) -> None:
        decision = sender_agent._choose_recipient(
            {
                "EMAIL_OSN": "one@example.com",
                "EMAIL_DOP": "two@example.com; three@example.com",
            }
        )

        self.assertEqual(sender_agent._recipient_role_for_log(decision, "one@example.com"), "primary")
        self.assertEqual(sender_agent._recipient_role_for_log(decision, "two@example.com"), "fallback")
        self.assertEqual(sender_agent._recipient_role_for_log(decision, "unknown@example.com"), "unknown")

    def test_delivery_failure_dispatches_next_fallback_recipient(self) -> None:
        sent_items = [
            {
                "sent_at": "2026-06-29T10:00:00",
                "row_id": "1",
                "recipient": "one@example.com",
                "transport": "mailopost",
                "send_mode": "consent_request",
                "attachment_mode": "kp",
                "work_type": "territorial_zone_boundaries",
                "recipient_strategy": sender_agent.RECIPIENT_STRATEGY_PRIMARY_THEN_FALLBACK,
                "sender_email": "override@example.com",
                "provider": {"message_id": "m1", "provider": "mailopost"},
            }
        ]
        events = [
            {
                "message_id": "m1",
                "row_id": "1",
                "recipient": "one@example.com",
                "provider_status": "hard_bounced",
                "received_at": "2026-06-29T10:01:00",
            }
        ]
        rows = {
            "1": {
                "ID": "1",
                "EMAIL_OSN": "one@example.com",
                "EMAIL_DOP": "two@example.com; three@example.com",
            }
        }
        calls: list[dict] = []

        def fake_run_sender(**kwargs):
            calls.append(kwargs)
            return {"status": "completed", "summary_text": "ok"}

        with patch.object(sender_agent, "_load_sender_state", return_value={"status": "completed"}), patch.object(
            sender_agent, "_load_sent_mail_log_items", return_value=sent_items
        ), patch.object(sender_agent, "_load_sender_rows_by_id", return_value=rows), patch.object(
            sender_agent, "_load_delivery_events", return_value=events
        ), patch.object(sender_agent, "run_sender", side_effect=fake_run_sender):
            result = sender_agent.process_delivery_fallbacks(job_id="job-1", provider="mailopost")

        self.assertEqual(result["status"], "dispatched")
        self.assertEqual(len(calls), 1)
        self.assertEqual(calls[0]["row_ids"], ["1"])
        self.assertEqual(calls[0]["recipient_strategy"], sender_agent.RECIPIENT_STRATEGY_PRIMARY_THEN_FALLBACK)
        self.assertEqual(calls[0]["sender_email"], "override@example.com")
        self.assertEqual(calls[0]["transport"], "mailopost")

    def test_delivery_failure_does_not_replay_old_primary_bounce_after_fallback_sent(self) -> None:
        sent_items = [
            {
                "sent_at": "2026-06-29T10:00:00",
                "row_id": "1",
                "recipient": "one@example.com",
                "transport": "mailopost",
                "send_mode": "consent_request",
                "attachment_mode": "kp",
                "recipient_strategy": sender_agent.RECIPIENT_STRATEGY_PRIMARY_THEN_FALLBACK,
                "sender_email": "override@example.com",
                "provider": {"message_id": "m1", "provider": "mailopost"},
            },
            {
                "sent_at": "2026-06-29T10:05:00",
                "row_id": "1",
                "recipient": "two@example.com",
                "transport": "mailopost",
                "send_mode": "consent_request",
                "attachment_mode": "kp",
                "recipient_strategy": sender_agent.RECIPIENT_STRATEGY_PRIMARY_THEN_FALLBACK,
                "provider": {"message_id": "m2", "provider": "mailopost"},
            },
        ]
        events = [
            {
                "message_id": "m1",
                "row_id": "1",
                "recipient": "one@example.com",
                "provider_status": "hard_bounced",
                "received_at": "2026-06-29T10:01:00",
            }
        ]
        rows = {
            "1": {
                "ID": "1",
                "EMAIL_OSN": "one@example.com",
                "EMAIL_DOP": "two@example.com; three@example.com",
            }
        }

        with patch.object(sender_agent, "_load_sender_state", return_value={"status": "completed"}), patch.object(
            sender_agent, "_load_sent_mail_log_items", return_value=sent_items
        ), patch.object(sender_agent, "_load_sender_rows_by_id", return_value=rows), patch.object(
            sender_agent, "_load_delivery_events", return_value=events
        ), patch.object(sender_agent, "run_sender") as run_sender_mock:
            result = sender_agent.process_delivery_fallbacks(job_id="job-1", provider="mailopost")

        self.assertEqual(result["status"], "no_fallback_needed")
        run_sender_mock.assert_not_called()

    def test_delivery_failure_does_not_dispatch_second_extra_fallback(self) -> None:
        sent_items = [
            {
                "sent_at": "2026-06-29T10:00:00",
                "row_id": "1",
                "recipient": "one@example.com",
                "transport": "mailopost",
                "send_mode": "consent_request",
                "attachment_mode": "kp",
                "recipient_strategy": sender_agent.RECIPIENT_STRATEGY_PRIMARY_THEN_FALLBACK,
                "provider": {"message_id": "m1", "provider": "mailopost"},
            },
            {
                "sent_at": "2026-06-29T10:05:00",
                "row_id": "1",
                "recipient": "two@example.com",
                "transport": "mailopost",
                "send_mode": "consent_request",
                "attachment_mode": "kp",
                "recipient_strategy": sender_agent.RECIPIENT_STRATEGY_PRIMARY_THEN_FALLBACK,
                "provider": {"message_id": "m2", "provider": "mailopost"},
            },
        ]
        events = [
            {
                "message_id": "m2",
                "row_id": "1",
                "recipient": "two@example.com",
                "provider_status": "hard_bounced",
                "received_at": "2026-06-29T10:06:00",
            }
        ]
        rows = {
            "1": {
                "ID": "1",
                "EMAIL_OSN": "one@example.com",
                "EMAIL_DOP": "two@example.com; three@example.com",
            }
        }

        with patch.object(sender_agent, "_load_sender_state", return_value={"status": "completed"}), patch.object(
            sender_agent, "_load_sent_mail_log_items", return_value=sent_items
        ), patch.object(sender_agent, "_load_sender_rows_by_id", return_value=rows), patch.object(
            sender_agent, "_load_delivery_events", return_value=events
        ), patch.object(sender_agent, "run_sender") as run_sender_mock:
            result = sender_agent.process_delivery_fallbacks(job_id="job-1", provider="mailopost")

        self.assertEqual(result["status"], "no_fallback_needed")
        run_sender_mock.assert_not_called()

    def test_primary_then_fallback_sequence_stops_after_first_success(self) -> None:
        sent: list[str] = []
        waits: list[str] = []

        def fake_send_one(recipient: str) -> dict:
            sent.append(recipient)
            if recipient == "one@example.com":
                return {
                    "recipient": None,
                    "recipients": [],
                    "attempts": [{"recipient": recipient, "status": "error", "error": "bounce"}],
                    "error": "bounce",
                    "warning": "",
                }
            return {
                "recipient": recipient,
                "recipients": [recipient],
                "attempts": [{"recipient": recipient, "status": "sent", "error": ""}],
                "error": "",
                "warning": "",
            }

        result = sender_agent._send_recipient_sequence_until_success(
            ["one@example.com", "two@example.com", "three@example.com"],
            fake_send_one,
            wait_between_recipients=lambda: waits.append("wait") or True,
        )

        self.assertEqual(sent, ["one@example.com", "two@example.com"])
        self.assertEqual(waits, ["wait"])
        self.assertEqual(result["recipient"], "two@example.com")
        self.assertEqual([attempt["status"] for attempt in result["attempts"]], ["error", "sent"])

    def test_consent_requests_are_sent_separately_to_extra_emails(self) -> None:
        prepared: list[str] = []
        sent_bodies: dict[str, str] = {}
        wait_calls: list[str] = []

        def fake_prepare(**kwargs):
            recipient = kwargs["recipient"]
            prepared.append(recipient)
            return {"consent_url": f"https://example.test/consent/{recipient}"}

        def fake_send(row, recipients, attachments, subject, **kwargs):
            self.assertEqual(len(recipients), 1)
            recipient = recipients[0]
            self.assertEqual(attachments, [])
            sent_bodies[recipient] = kwargs.get("body_override") or ""
            return {
                "recipient": recipient,
                "recipients": [recipient],
                "attempts": [{"recipient": recipient, "status": "sent", "error": ""}],
                "error": "",
                "warning": "",
            }

        def fake_wait() -> bool:
            wait_calls.append("wait")
            return True

        recipients = ["one@example.com", "two@example.com"]
        with patch.object(sender_agent, "prepare_consent_request", side_effect=fake_prepare), patch.object(
            sender_agent, "_send_with_transport", side_effect=fake_send
        ):
            result = sender_agent._send_consent_requests_with_transport(
                {"ID": "1", "MUN_NAME": "Test municipality"},
                recipients,
                "Subject",
                transport="rusender",
                job_id="job-test",
                send_run_id="send-test",
                attachment_mode="kp",
                work_type="mngp_settlements",
                wait_between_recipients=fake_wait,
            )

        self.assertEqual(prepared, recipients)
        self.assertEqual(result["recipients"], recipients)
        self.assertEqual(set(result["consent_urls"]), set(recipients))
        self.assertEqual(wait_calls, ["wait"])
        self.assertIn("/consent/one@example.com", sent_bodies["one@example.com"])
        self.assertIn("/consent/two@example.com", sent_bodies["two@example.com"])

    def test_smtp_transport_waits_between_each_recipient(self) -> None:
        sent: list[str] = []
        wait_calls: list[str] = []

        def fake_smtp(row, recipient, attachments, subject, **kwargs) -> str:
            sent.append(recipient)
            return ""

        def fake_wait() -> bool:
            wait_calls.append("wait")
            return True

        recipients = ["one@example.com", "two@example.com", "three@example.com"]
        with patch.object(sender_agent, "_send_via_smtp", side_effect=fake_smtp):
            result = sender_agent._send_with_transport(
                {"ID": "1", "MUN_NAME": "Test municipality"},
                recipients,
                [],
                "Subject",
                transport="smtp",
                wait_between_recipients=fake_wait,
            )

        self.assertEqual(sent, recipients)
        self.assertEqual(wait_calls, ["wait", "wait"])
        self.assertEqual(result["recipients"], recipients)
        self.assertEqual([attempt["status"] for attempt in result["attempts"]], ["sent", "sent", "sent"])

    def test_smtp_transport_stops_before_next_recipient_when_delay_is_stopped(self) -> None:
        sent: list[str] = []

        def fake_smtp(row, recipient, attachments, subject, **kwargs) -> str:
            sent.append(recipient)
            return ""

        with patch.object(sender_agent, "_send_via_smtp", side_effect=fake_smtp):
            result = sender_agent._send_with_transport(
                {"ID": "1", "MUN_NAME": "Test municipality"},
                ["one@example.com", "two@example.com"],
                [],
                "Subject",
                transport="smtp",
                wait_between_recipients=lambda: False,
            )

        self.assertEqual(sent, ["one@example.com"])
        self.assertEqual(result["recipients"], ["one@example.com"])
        self.assertEqual(result["attempts"][-1]["recipient"], "two@example.com")
        self.assertIn("паузы", result["error"])

    def test_sender_delay_uses_random_range_when_configured(self) -> None:
        with (
            patch.object(sender_agent.settings, "sender_delay_seconds", 180.0),
            patch.object(sender_agent.settings, "sender_delay_min_seconds", 179.0),
            patch.object(sender_agent.settings, "sender_delay_max_seconds", 247.0),
            patch.object(sender_agent.random, "uniform", return_value=211.0) as uniform_mock,
        ):
            delay = sender_agent._sender_delay_seconds()

        self.assertEqual(delay, 211.0)
        uniform_mock.assert_called_once_with(179.0, 247.0)

    def test_sender_delay_falls_back_to_fixed_delay_without_range(self) -> None:
        with (
            patch.object(sender_agent.settings, "sender_delay_seconds", 180.0),
            patch.object(sender_agent.settings, "sender_delay_min_seconds", 0.0),
            patch.object(sender_agent.settings, "sender_delay_max_seconds", 0.0),
        ):
            self.assertEqual(sender_agent._sender_delay_seconds(), 180.0)

    def test_wait_smtp_sender_delay_persists_label(self) -> None:
        state: dict[str, object] = {}
        with (
            patch.object(sender_agent, "_sender_delay_seconds", return_value=211.0),
            patch.object(sender_agent, "_wait_sender_delay", return_value=True) as wait_mock,
            patch.object(sender_agent, "_save_sender_state") as save_mock,
        ):
            self.assertTrue(sender_agent._wait_smtp_sender_delay(state, "job-test"))

        self.assertEqual(state["smtp_delay_seconds"], 211)
        self.assertEqual(state["smtp_delay_label"], "3 мин 31 сек")
        self.assertIn("3 мин 31 сек", str(state["summary_text"]))
        save_mock.assert_called_once_with(state, "job-test")
        wait_mock.assert_called_once_with(211.0, state, "job-test")

    def test_rusender_uses_bearer_authorization_for_current_keys(self) -> None:
        captured: dict[str, Request] = {}

        def fake_request(request: Request, *, timeout: float) -> str:
            captured["request"] = request
            return json.dumps({"uuid": "message-1"})

        with patch.object(sender_agent.settings, "rusender_api_key", "rs_ck_v1_secret"), patch.object(
            sender_agent.settings,
            "rusender_sender_email",
            "sender@example.com",
        ), patch.object(sender_agent, "_run_rusender_request", side_effect=fake_request):
            result = sender_agent._send_via_rusender(
                {"ID": "1", "MUN_NAME": "Тестовое МО"},
                "recipient@example.com",
                [],
                "Тема",
                body_override="Текст письма",
                job_id="job-test",
                send_run_id="send-test",
                send_mode="consent_request",
                attachment_mode="kp",
            )

        request = captured["request"]
        self.assertEqual(request.get_header("Authorization"), "Bearer rs_ck_v1_secret")
        self.assertIsNone(request.get_header("X-api-key"))
        self.assertEqual(result["message_id"], "message-1")

    def test_rusender_uses_x_api_key_for_legacy_jwt_keys(self) -> None:
        captured: dict[str, Request] = {}

        def fake_request(request: Request, *, timeout: float) -> str:
            captured["request"] = request
            return json.dumps({"uuid": "message-1"})

        legacy_key = "eyJhbGciOiJIUzI1NiJ9.test"
        with patch.object(sender_agent.settings, "rusender_api_key", legacy_key), patch.object(
            sender_agent.settings,
            "rusender_sender_email",
            "sender@example.com",
        ), patch.object(sender_agent, "_run_rusender_request", side_effect=fake_request):
            result = sender_agent._send_via_rusender(
                {"ID": "1", "MUN_NAME": "Тестовое МО"},
                "recipient@example.com",
                [],
                "Тема",
                body_override="Текст письма",
                job_id="job-test",
                send_run_id="send-test",
                send_mode="consent_request",
                attachment_mode="kp",
            )

        request = captured["request"]
        self.assertIsNone(request.get_header("Authorization"))
        self.assertEqual(request.get_header("X-api-key"), legacy_key)
        self.assertEqual(result["message_id"], "message-1")

    def test_mailopost_builds_json_send_request(self) -> None:
        captured: dict[str, Request] = {}

        def fake_request(request: Request, *, timeout: float) -> str:
            captured["request"] = request
            return json.dumps({"id": 123, "status": "queued"})

        with patch.object(sender_agent.settings, "mailopost_api_token", "mailopost-token"), patch.object(
            sender_agent.settings,
            "mailopost_sender_email",
            "sender@example.com",
        ), patch.object(sender_agent.settings, "mailopost_sender_name", "ООО «ПР»"), patch.object(
            sender_agent.settings,
            "mailopost_api_base_url",
            "https://api.mailopost.ru/v1",
        ), patch.object(sender_agent, "_run_mailopost_request", side_effect=fake_request):
            result = sender_agent._send_via_mailopost(
                {"ID": "1", "MUN_NAME": "Тестовое МО"},
                "recipient@example.com",
                [],
                "Тема",
                body_override="Текст письма",
                job_id="job-test",
                send_run_id="send-test",
                send_mode="consent_request",
                attachment_mode="kp",
            )

        request = captured["request"]
        self.assertEqual(request.full_url, "https://api.mailopost.ru/v1/email/messages")
        self.assertEqual(request.get_header("Authorization"), "Bearer mailopost-token")
        payload = json.loads((request.data or b"").decode("utf-8"))
        self.assertEqual(payload["from_email"], "sender@example.com")
        self.assertEqual(payload["from_name"], "ООО «ПР»")
        self.assertEqual(payload["to"], "recipient@example.com")
        self.assertEqual(payload["smtp_headers"]["X-Mailing-Agent-Job"], "job-test")
        self.assertEqual(result["provider"], "mailopost")
        self.assertEqual(result["message_id"], "123")

    def test_mailopost_uses_sender_email_override(self) -> None:
        captured: dict[str, Request] = {}

        def fake_request(request: Request, *, timeout: float) -> str:
            captured["request"] = request
            return json.dumps({"id": 123, "status": "queued"})

        with patch.object(sender_agent.settings, "mailopost_api_token", "mailopost-token"), patch.object(
            sender_agent.settings,
            "mailopost_sender_email",
            "env-sender@example.com",
        ), patch.object(sender_agent.settings, "mailopost_api_base_url", "https://api.mailopost.ru/v1"), patch.object(
            sender_agent, "_run_mailopost_request", side_effect=fake_request
        ):
            sender_agent._send_via_mailopost(
                {"ID": "1", "MUN_NAME": "Тестовое МО"},
                "recipient@example.com",
                [],
                "Тема",
                body_override="Текст письма",
                job_id="job-test",
                send_run_id="send-test",
                send_mode="consent_request",
                attachment_mode="kp",
                sender_email="override@example.com",
            )

        payload = json.loads((captured["request"].data or b"").decode("utf-8"))
        self.assertEqual(payload["from_email"], "override@example.com")
    def test_mailopost_delivery_failure_aliases_are_normalized(self) -> None:
        self.assertEqual(mailopost_events._status_from_event("not_delivered"), "err_delivery_failed")
        self.assertEqual(mailopost_events._status_from_event("failed"), "err_delivery_failed")
        self.assertEqual(mailopost_events._status_from_event("bounce"), "hard_bounced")

    def test_mailopost_not_delivered_dispatches_next_fallback_recipient(self) -> None:
        sent_items = [
            {
                "sent_at": "2026-06-29T10:00:00",
                "row_id": "1",
                "recipient": "one@example.com",
                "transport": "mailopost",
                "send_mode": "consent_request",
                "attachment_mode": "kp",
                "recipient_strategy": sender_agent.RECIPIENT_STRATEGY_PRIMARY_THEN_FALLBACK,
                "sender_email": "override@example.com",
                "provider": {"message_id": "m1", "provider": "mailopost"},
            }
        ]
        events = [
            {
                "message_id": "m1",
                "row_id": "1",
                "recipient": "one@example.com",
                "provider_status": "not_delivered",
                "received_at": "2026-06-29T10:01:00",
            }
        ]
        rows = {
            "1": {
                "ID": "1",
                "EMAIL_OSN": "one@example.com",
                "EMAIL_DOP": "two@example.com",
            }
        }
        calls: list[dict] = []

        def fake_run_sender(**kwargs):
            calls.append(kwargs)
            return {"status": "completed", "summary_text": "ok"}

        with patch.object(sender_agent, "_load_sender_state", return_value={"status": "completed"}), patch.object(
            sender_agent, "_load_sent_mail_log_items", return_value=sent_items
        ), patch.object(sender_agent, "_load_sender_rows_by_id", return_value=rows), patch.object(
            sender_agent, "_load_delivery_events", return_value=events
        ), patch.object(sender_agent, "run_sender", side_effect=fake_run_sender):
            result = sender_agent.process_delivery_fallbacks(job_id="job-1", provider="mailopost")

        self.assertEqual(result["status"], "dispatched")
        self.assertEqual(len(calls), 1)
        self.assertEqual(calls[0]["row_ids"], ["1"])
        self.assertEqual(calls[0]["recipient_strategy"], sender_agent.RECIPIENT_STRATEGY_PRIMARY_THEN_FALLBACK)
        self.assertEqual(calls[0]["sender_email"], "override@example.com")

    def test_mailopost_retry_after_uses_header_and_message(self) -> None:
        header_error = HTTPError("https://api.mailopost.ru/v1/email/messages", 429, "Too Many", {"Retry-After": "12"}, None)
        self.assertEqual(
            sender_agent._mailopost_retry_after_seconds(header_error, "", "Too many messages."),
            17.0,
        )

        message_error = HTTPError("https://api.mailopost.ru/v1/email/messages", 429, "Too Many", {}, None)
        self.assertEqual(
            sender_agent._mailopost_retry_after_seconds(
                message_error,
                "{\"message\":\"Too many messages. Try again in 7 seconds.\"}",
                "Too many messages. Try again in 7 seconds.",
            ),
            12.0,
        )

    def test_mailopost_transport_waits_after_rate_limit_and_retries_same_recipient(self) -> None:
        calls: list[str] = []
        waits: list[tuple[float, str]] = []

        def fake_mailopost(row, recipient, attachments, subject, **kwargs) -> dict[str, str]:
            calls.append(recipient)
            if len(calls) == 1:
                raise sender_agent.MailoPostRateLimitError(
                    "Too many messages. Try again in 7 seconds.",
                    7,
                )
            return {"provider": "mailopost", "message_id": "message-1", "recipient": recipient}

        def wait_after_rate_limit(seconds: float, message: str) -> bool:
            waits.append((seconds, message))
            return True

        with patch.object(sender_agent, "_send_via_mailopost", side_effect=fake_mailopost):
            result = sender_agent._send_with_transport(
                {"ID": "1", "MUN_NAME": "Test municipality"},
                ["recipient@example.com"],
                [],
                "Subject",
                transport="mailopost",
                wait_after_rate_limit=wait_after_rate_limit,
            )

        self.assertEqual(calls, ["recipient@example.com", "recipient@example.com"])
        self.assertEqual(waits, [(7.0, "Too many messages. Try again in 7 seconds.")])
        self.assertEqual(result["recipient"], "recipient@example.com")
        self.assertEqual(result["attempts"][0]["status"], "sent")

    def test_mailopost_transport_stops_when_rate_limit_wait_is_stopped(self) -> None:
        calls: list[str] = []

        def fake_mailopost(row, recipient, attachments, subject, **kwargs) -> dict[str, str]:
            calls.append(recipient)
            raise sender_agent.MailoPostRateLimitError(
                "Too many messages. Try again in 30 seconds.",
                30,
            )

        with patch.object(sender_agent, "_send_via_mailopost", side_effect=fake_mailopost):
            result = sender_agent._send_with_transport(
                {"ID": "1", "MUN_NAME": "Test municipality"},
                ["recipient@example.com"],
                [],
                "Subject",
                transport="mailopost",
                wait_after_rate_limit=lambda seconds, message: False,
            )

        self.assertEqual(calls, ["recipient@example.com"])
        self.assertIsNone(result["recipient"])
        self.assertEqual(result["attempts"][0]["status"], "error")
        self.assertIn("ожидания лимита MailoPost", result["error"])

    def test_provider_idempotency_key_is_deterministic_for_same_context(self) -> None:
        first = sender_agent._build_provider_idempotency_key(
            provider="rusender",
            job_id="job-1",
            send_run_id="send-1",
            row_id="42",
            recipient="Recipient@Example.com",
            send_mode="materials",
            attachment_mode="kp",
        )
        second = sender_agent._build_provider_idempotency_key(
            provider="rusender",
            job_id="job-1",
            send_run_id="send-1",
            row_id="42",
            recipient="recipient@example.com",
            send_mode="materials",
            attachment_mode="kp",
        )

        self.assertEqual(first, second)
        self.assertRegex(first, r"^mailing-agent:rusender:[0-9a-f]{40}$")

    def test_provider_idempotency_key_changes_for_recipient_or_mode(self) -> None:
        base_kwargs = {
            "provider": "unisender_go",
            "job_id": "job-1",
            "send_run_id": "send-1",
            "row_id": "42",
            "recipient": "one@example.com",
            "send_mode": "materials",
            "attachment_mode": "kp",
        }

        base = sender_agent._build_provider_idempotency_key(**base_kwargs)

        self.assertNotEqual(
            base,
            sender_agent._build_provider_idempotency_key(**{**base_kwargs, "recipient": "two@example.com"}),
        )
        self.assertNotEqual(
            base,
            sender_agent._build_provider_idempotency_key(**{**base_kwargs, "send_mode": "consent_request"}),
        )
        self.assertNotEqual(
            base,
            sender_agent._build_provider_idempotency_key(**{**base_kwargs, "attachment_mode": "all"}),
        )

    def test_sent_mail_log_promotes_provider_idempotency_key(self) -> None:
        self.addCleanup(lambda: reset_test_database())

        warning = sender_agent._append_sent_mail_log(
            row={"ID": "42", "MUN_NAME": "Test municipality"},
            recipient="recipient@example.com",
            attachments=[],
            subject="Subject",
            transport="rusender",
            provider={
                "provider": "rusender",
                "message_id": "message-1",
                "idempotency_key": "mailing-agent:rusender:stable",
            },
            job_id="job-test",
            send_run_id="send-1",
            recipient_role="fallback",
            send_run_started_at="2026-06-21T10:00:00",
            campaign_name="июльская рассылка",
        )

        self.assertIsNone(warning)
        from src.jobs.job_docs import read_events

        record = read_events("job-test", "sent_mail_log")[0]
        self.assertEqual(record["provider_idempotency_key"], "mailing-agent:rusender:stable")
        self.assertEqual(record["provider"]["idempotency_key"], "mailing-agent:rusender:stable")
        self.assertEqual(record["recipient_role"], "fallback")
        self.assertEqual(record["campaign_name"], "июльская рассылка")

    def test_unisender_analytics_filters_items_outside_current_data(self) -> None:
        items = [
            {"row_id": "1", "recipient": "one@example.com"},
            {"row_id": "2", "recipient": "two@example.com"},
            {"row_id": "2", "recipient": "old-two@example.com"},
            {"row_id": "101", "recipient": "old@example.com"},
        ]

        with patch.object(
            sender_report,
            "resolve_job_paths",
            return_value=SimpleNamespace(data_xlsx=Path(__file__)),
        ), patch.object(
            sender_report,
            "load_rows",
            return_value=(
                None,
                None,
                [
                    {"ID": "1", "EMAIL_OSN": "one@example.com"},
                    {"ID": "2", "EMAIL_OSN": "two@example.com"},
                ],
            ),
        ):
            scoped = sender_report._filter_items_by_current_data("job-current", items)

        self.assertEqual([item["row_id"] for item in scoped], ["1", "2"])
        self.assertEqual([item["recipient"] for item in scoped], ["one@example.com", "two@example.com"])

    def test_delivery_report_infers_recipient_role_from_current_data(self) -> None:
        with patch.object(
            sender_report,
            "resolve_job_paths",
            return_value=SimpleNamespace(data_xlsx=Path(__file__)),
        ), patch.object(
            sender_report,
            "load_rows",
            return_value=(
                None,
                None,
                [
                    {
                        "ID": "1",
                        "EMAIL_OSN": "one@example.com",
                        "EMAIL_DOP": "two@example.com; three@example.com",
                    }
                ],
            ),
        ):
            roles = sender_report._current_data_recipient_roles("job-current")

        self.assertEqual(roles[("1", "one@example.com")], "primary")
        self.assertEqual(roles[("1", "two@example.com")], "fallback")
        self.assertEqual(roles[("1", "three@example.com")], "fallback")

    def test_unisender_analytics_deduplicates_latest_row_recipient_item(self) -> None:
        items = [
            {"row_id": "1", "recipient": "one@example.com", "provider_job_id": "old"},
            {"row_id": "1", "recipient": "one@example.com", "provider_job_id": "new"},
        ]

        deduped = sender_report._dedupe_latest_log_items(items)

        self.assertEqual(deduped, [items[1]])

    def test_unisender_go_event_prefers_row_recipient_over_provider_job_fallback(self) -> None:
        events = {
            "provider_job:new-provider-id": {"provider_status": "delivered"},
            "row_email:1:test@example.com": {"provider_status": "hard_bounced"},
        }
        item = {
            "row_id": "1",
            "recipient": "test@example.com",
            "provider_job_id": "new-provider-id",
        }

        matched = sender_report._match_unisender_go_event(item, events)

        self.assertEqual(matched, {"provider_status": "hard_bounced"})

    def test_unisender_go_event_uses_provider_job_fallback_without_recipient(self) -> None:
        events = {
            "provider_job:new-provider-id": {"provider_status": "delivered"},
        }
        item = {
            "row_id": "1",
            "recipient": "",
            "provider_job_id": "new-provider-id",
        }

        matched = sender_report._match_unisender_go_event(item, events)

        self.assertEqual(matched, {"provider_status": "delivered"})

    def test_unisender_go_event_prefers_provider_job_and_recipient(self) -> None:
        events = {
            "provider_job:shared-provider-id": {"provider_status": "hard_bounced"},
            "provider_job_email:shared-provider-id:first@example.com": {"provider_status": "delivered"},
            "provider_job_email:shared-provider-id:second@example.com": {"provider_status": "err_user_unknown"},
        }

        first = sender_report._match_unisender_go_event(
            {
                "row_id": "1",
                "recipient": "first@example.com",
                "provider_job_id": "shared-provider-id",
            },
            events,
        )
        second = sender_report._match_unisender_go_event(
            {
                "row_id": "1",
                "recipient": "second@example.com",
                "provider_job_id": "shared-provider-id",
            },
            events,
        )

        self.assertEqual(first, {"provider_status": "delivered"})
        self.assertEqual(second, {"provider_status": "err_user_unknown"})

    def test_unisender_go_technical_spam_skip_is_not_hard_bounce(self) -> None:
        with patch.object(
            sender_report,
            "_build_delivery_rows",
            return_value=(
                [
                    {
                        "provider": "unisender_go",
                        "provider_status": "err_spam_skipped",
                        "checked_at": "2026-06-04T10:00:00",
                    }
                ],
                "",
            ),
        ):
            analytics = sender_report.build_sender_delivery_analytics("job-current", refresh=False)

        self.assertEqual(analytics["summary"]["hard_bounced"], 0)
        self.assertEqual(analytics["summary"]["errors"], 1)

    def test_sender_analytics_uses_rusender_provider_label_and_events(self) -> None:
        with patch.object(
            sender_report,
            "_build_delivery_rows",
            return_value=(
                [
                    {
                        "provider": "rusender",
                        "provider_status": "delivered",
                        "checked_at": "2026-06-18T12:10:00",
                    },
                    {
                        "provider": "rusender",
                        "provider_status": "clicked",
                        "checked_at": "2026-06-18T12:11:00",
                    },
                ],
                "",
            ),
        ):
            analytics = sender_report.build_sender_delivery_analytics("job-current", refresh=False)

        self.assertEqual(analytics["provider"], "rusender")
        self.assertEqual(analytics["provider_label"], "RuSender")
        self.assertEqual(analytics["summary"]["delivered"], 2)
        self.assertEqual(analytics["summary"]["clicked"], 1)
        self.assertEqual(analytics["provider_events_count"], 2)
        self.assertEqual(analytics["cards"][0]["title"], "Передано в RuSender")

    def test_sender_analytics_splits_delivered_by_recipient_role(self) -> None:
        with patch.object(
            sender_report,
            "_build_delivery_rows",
            return_value=(
                [
                    {
                        "provider": "mailopost",
                        "provider_status": "delivered",
                        "recipient_role": "primary",
                        "checked_at": "2026-06-29T10:00:00",
                    },
                    {
                        "provider": "mailopost",
                        "provider_status": "opened",
                        "recipient_role": "fallback",
                        "checked_at": "2026-06-29T10:01:00",
                    },
                    {
                        "provider": "mailopost",
                        "provider_status": "hard_bounced",
                        "recipient_role": "fallback",
                        "checked_at": "2026-06-29T10:02:00",
                    },
                ],
                "",
            ),
        ):
            analytics = sender_report.build_sender_delivery_analytics("job-current", refresh=False)

        cards = {card["id"]: card for card in analytics["cards"]}
        self.assertEqual(analytics["summary"]["recipient_roles"]["accepted"]["primary"], 1)
        self.assertEqual(analytics["summary"]["recipient_roles"]["accepted"]["fallback"], 2)
        self.assertEqual(analytics["summary"]["recipient_roles"]["delivered"]["primary"], 1)
        self.assertEqual(analytics["summary"]["recipient_roles"]["delivered"]["fallback"], 1)
        self.assertEqual(cards["delivered_primary"]["value"], 1)
        self.assertEqual(cards["delivered_fallback"]["value"], 1)
        self.assertEqual(cards["delivered_fallback"]["percent"], 50.0)

    def test_run_unisender_request_retries_temporary_network_error(self) -> None:
        attempts = {"count": 0}

        def fake_urlopen(request: Request, timeout: float = 60):
            attempts["count"] += 1
            if attempts["count"] < 3:
                raise URLError("temporary network issue")

            class _Response:
                def __enter__(self):
                    return self

                def __exit__(self, exc_type, exc, tb):
                    return False

                def read(self):
                    return b'{"status":"success"}'

            return _Response()

        with patch.object(sender_agent, "urlopen", side_effect=fake_urlopen), patch.object(
            sender_agent, "_sleep_sender_retry", return_value=None
        ):
            raw = sender_agent._run_unisender_request(
                Request("https://example.com"),
                timeout=5,
                request_label="UniSender Go",
            )

        self.assertEqual(raw, '{"status":"success"}')
        self.assertEqual(attempts["count"], 3)

    def test_run_unisender_request_keeps_last_http_error_body(self) -> None:
        def raise_http_error(request: Request, timeout: float = 60):
            raise HTTPError(
                url="https://example.com",
                code=503,
                msg="Service Unavailable",
                hdrs=None,
                fp=io.BytesIO(b'{"message":"temporary outage"}'),
            )

        with patch.object(sender_agent, "urlopen", side_effect=raise_http_error), patch.object(
            sender_agent, "_sleep_sender_retry", return_value=None
        ):
            with self.assertRaises(HTTPError) as caught:
                sender_agent._run_unisender_request(
                    Request("https://example.com"),
                    timeout=5,
                    request_label="UniSender Go",
                )

        self.assertEqual(caught.exception.code, 503)
        self.assertEqual(getattr(caught.exception, "raw_body", ""), '{"message":"temporary outage"}')


if __name__ == "__main__":
    unittest.main()
