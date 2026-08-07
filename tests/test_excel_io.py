from __future__ import annotations

import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

from src.generator.generation.excel_io import load_rows


class ExcelIoTests(unittest.TestCase):
    def test_load_rows_stops_after_large_blank_tail(self) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["Description", "Email"])
        worksheet.append(["ADM_NAME", "EMAIL_OSN"])
        worksheet.append(["Visible recipient", "visible@example.com"])
        worksheet.cell(row=5000, column=1, value="Stray tail value")
        worksheet.cell(row=5000, column=2, value="tail@example.com")

        with patch(
            "src.generator.generation.excel_io.load_workbook",
            return_value=workbook,
        ):
            loaded_workbook, _worksheet, rows = load_rows(
                Path("unused.xlsx")
            )
        loaded_workbook.close()

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["ADM_NAME"], "Visible recipient")
        self.assertEqual(rows[0]["EMAIL_OSN"], "visible@example.com")


if __name__ == "__main__":
    unittest.main()
