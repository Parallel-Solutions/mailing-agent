from __future__ import annotations

import shutil
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from openpyxl import load_workbook

from src.web import load_test_service


class LoadTestServiceTests(unittest.TestCase):
    def setUp(self) -> None:
        self.tmp_dir = Path("tmp_test_load_test_service")
        self.tmp_dir.mkdir(exist_ok=True)

    def tearDown(self) -> None:
        if self.tmp_dir.exists():
            shutil.rmtree(self.tmp_dir)

    def _paths(self, job_id: str | None):
        root = self.tmp_dir / (job_id or "legacy")
        paths = SimpleNamespace(
            job_id=job_id,
            root_dir=root,
            data_xlsx=root / "input" / "data.xlsx",
            base_xlsx=root / "input" / "base.xlsx",
            templates_dir=root / "templates",
            output_dir=root / "output",
            batch_docx_dir=root / "_batch_docx",
            batch_pdf_dir=root / "_batch_pdf",
            sent_mail_log_path=root / "sent_mail_log.jsonl",
            uses_legacy_layout=job_id is None,
        )

        def ensure_dirs() -> None:
            paths.data_xlsx.parent.mkdir(parents=True, exist_ok=True)
            paths.templates_dir.mkdir(parents=True, exist_ok=True)
            paths.output_dir.mkdir(parents=True, exist_ok=True)
            paths.batch_docx_dir.mkdir(parents=True, exist_ok=True)
            paths.batch_pdf_dir.mkdir(parents=True, exist_ok=True)

        paths.ensure_dirs = ensure_dirs
        return paths

    def test_create_documents_load_test_job_writes_workbook_and_parser_state(self) -> None:
        source_paths = self._paths("source-job")
        source_paths.ensure_dirs()
        (source_paths.templates_dir / "kp_template_source.docx").write_bytes(b"kp")
        (source_paths.templates_dir / "contract_template_source.docx").write_bytes(b"contract")
        saved_states: list[tuple[str, dict, str | None]] = []

        with patch.object(load_test_service, "create_job_id", return_value="job-loadtest"), patch.object(
            load_test_service, "resolve_job_paths", side_effect=self._paths
        ), patch.object(load_test_service, "save_agent_state", side_effect=lambda agent, state, job_id=None: saved_states.append((agent, state, job_id))):
            result = load_test_service.create_documents_load_test_job(
                row_count=500,
                source_job_id="source-job",
                seed=123,
            )
            is_load_test = load_test_service.is_load_test_job("job-loadtest")

        target_paths = self._paths("job-loadtest")
        workbook = load_workbook(target_paths.data_xlsx, read_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        try:
            self.assertEqual(result["job_id"], "job-loadtest")
            self.assertEqual(result["row_count"], 500)
            self.assertEqual(worksheet.max_row, 502)
            self.assertEqual(worksheet.cell(row=2, column=1).value, "ID")
            self.assertEqual(worksheet.cell(row=3, column=9).value, "loadtest0001@example.test")
        finally:
            workbook.close()
        self.assertTrue((target_paths.templates_dir / "kp_template_source.docx").exists())
        self.assertTrue((target_paths.templates_dir / "contract_template_source.docx").exists())
        self.assertEqual(result["missing_templates"], [])
        self.assertTrue(result["generator_ready"])
        self.assertTrue(result["load_test"])
        self.assertTrue(result["send_disabled"])
        self.assertTrue(is_load_test)
        self.assertEqual(saved_states[0][0], "parser")
        self.assertEqual(saved_states[0][1]["status"], "completed")


if __name__ == "__main__":
    unittest.main()
