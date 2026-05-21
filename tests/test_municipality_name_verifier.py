import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from src.generator.verification.minjust_municipality_lookup import MinjustMunicipalityResult
from src.generator.verification.oktmo_municipality_lookup import OktmoMunicipalityResult
from src.generator.verification.municipality_name_verifier import (
    OfficialSiteLookup,
    _build_official_site_queries,
    _candidate_is_safe_for_autoreplace,
    _needs_abbreviation_resolution,
    _official_site_match_from_search_result,
    _search_name_variants,
    _should_use_official_site_followup,
    extract_municipality_name_from_administration,
    verify_municipality_name,
    verify_municipality_names_in_workbook,
)


class FakeMinjustLookup:
    disabled_reason = ""

    def __init__(self, confirmed_name: str = "") -> None:
        self.confirmed_name = confirmed_name
        self.calls: list[tuple[dict, str]] = []

    def confirm(self, row: dict, candidate_name: str):
        self.calls.append((row, candidate_name))
        if self.confirmed_name:
            return MinjustMunicipalityResult(
                name=self.confirmed_name,
                source_url="https://pravo-search.minjust.ru/bigs/portal.html#search=test",
            )
        return None


class FakeOktmoLookup:
    disabled_reason = ""

    def __init__(self, confirmed_name: str = "") -> None:
        self.confirmed_name = confirmed_name
        self.calls: list[tuple[dict, str]] = []

    def confirm(self, row: dict, candidate_name: str):
        self.calls.append((row, candidate_name))
        if self.confirmed_name:
            return OktmoMunicipalityResult(
                name=self.confirmed_name,
                source_url="https://rosstat.gov.ru/opendata/7708234640-oktmo",
                oktmo_code="80651101000",
                source_name="город Туймазы",
                subject_name="Республика Башкортостан",
                parent_name="Туймазинский муниципальный район",
            )
        return None


class FakeOfficialSiteLookup:
    disabled_reason = ""

    def __init__(self, matched_name: str = "", url: str = "https://adm.example.ru/") -> None:
        self.matched_name = matched_name
        self.url = url
        self.calls: list[tuple[dict, list[str]]] = []

    def confirm(self, row: dict, candidate_names: list[str]):
        self.calls.append((row, candidate_names))
        if not self.matched_name:
            return None
        return type("Match", (), {
            "url": self.url,
            "title": "Официальный сайт администрации",
            "content": self.matched_name,
            "score": 20,
            "matched_name": self.matched_name,
            "evidence": f"найдено название «{self.matched_name}»",
        })()


class MunicipalityNameVerifierTests(unittest.TestCase):
    def test_extracts_quoted_official_name_from_administration(self) -> None:
        result = verify_municipality_name(
            {
                "MUN_NAME": "Городское поселение Яблоновский",
                "ADM_NAME": 'Администрация муниципального образования "Яблоновское городское поселение"',
            }
        )

        self.assertEqual(result.status, "kept")
        self.assertEqual(result.confidence, "medium")
        self.assertEqual(result.official_name, "Яблоновское городское поселение")
        self.assertFalse(result.should_replace)

    def test_restores_readable_case_for_uppercase_official_name(self) -> None:
        result = verify_municipality_name(
            {
                "MUN_NAME": "Городское поселение Энем",
                "ADM_NAME": 'АДМИНИСТРАЦИЯ МУНИЦИПАЛЬНОГО ОБРАЗОВАНИЯ "ЭНЕМСКОЕ ГОРОДСКОЕ ПОСЕЛЕНИЕ"',
            }
        )

        self.assertEqual(result.official_name, "Энемское городское поселение")

    def test_keeps_administration_spelling_when_oktmo_confirms_candidate(self) -> None:
        result = verify_municipality_name(
            {
                "MUN_NAME": "Городское поселение Энем",
                "ADM_NAME": 'АДМИНИСТРАЦИЯ МУНИЦИПАЛЬНОГО ОБРАЗОВАНИЯ "ЭНЕМСКОЕ ГОРОДСКОЕ ПОСЕЛЕНИЕ"',
            },
            oktmo_lookup=FakeOktmoLookup("Городское поселение Энем"),
        )

        self.assertTrue(result.should_replace)
        self.assertEqual(result.official_name, "Энемское городское поселение")

    def test_restores_readable_case_for_city_type_name(self) -> None:
        result = verify_municipality_name(
            {
                "MUN_NAME": "Городское поселение город Белебей",
                "ADM_NAME": 'АДМИНИСТРАЦИЯ МУНИЦИПАЛЬНОГО ОБРАЗОВАНИЯ "ГОРОДСКОЕ ПОСЕЛЕНИЕ ГОРОД БЕЛЕБЕЙ"',
            }
        )

        self.assertEqual(result.official_name, "Городское поселение город Белебей")

    def test_rebuilds_city_settlement_from_contextual_quoted_locality(self) -> None:
        result = verify_municipality_name(
            {
                "MUN_NAME": "Городское поселение Поселок Онохой",
                "ADM_NAME": (
                    'АДМИНИСТРАЦИЯ МУНИЦИПАЛЬНОГО ОБРАЗОВАНИЯ '
                    'ГОРОДСКОГО ПОСЕЛЕНИЯ "ПОСЕЛОК ОНОХОЙ"'
                ),
            }
        )

        self.assertEqual(result.status, "verified")
        self.assertEqual(result.confidence, "medium")
        self.assertEqual(result.official_name, "Городское поселение поселок Онохой")
        self.assertFalse(result.should_replace)

    def test_rebuilds_city_settlement_from_contextual_quoted_adjective(self) -> None:
        result = verify_municipality_name(
            {
                "MUN_NAME": "Городское поселение Бабушкинское",
                "ADM_NAME": (
                    'АДМИНИСТРАЦИЯ МУНИЦИПАЛЬНОГО ОБРАЗОВАНИЯ '
                    'ГОРОДСКОГО ПОСЕЛЕНИЯ "БАБУШКИНСКОЕ" КАБАНСКОГО РАЙОНА'
                ),
            }
        )

        self.assertEqual(result.status, "verified")
        self.assertEqual(result.confidence, "medium")
        self.assertEqual(result.official_name, "Городское поселение Бабушкинское")

    def test_unwraps_local_administration_from_quoted_name(self) -> None:
        result = verify_municipality_name(
            {
                "MUN_NAME": "Сельское поселение Джулат",
                "ADM_NAME": (
                    'МУНИЦИПАЛЬНОЕ УЧРЕЖДЕНИЕ '
                    '"МЕСТНАЯ АДМИНИСТРАЦИЯ СЕЛЬСКОГО ПОСЕЛЕНИЯ ДЖУЛАТ" '
                    'ТЕРСКОГО МУНИЦИПАЛЬНОГО РАЙОНА'
                ),
            }
        )

        self.assertEqual(result.status, "verified")
        self.assertEqual(result.confidence, "medium")
        self.assertEqual(result.official_name, "Сельское поселение Джулат")
        self.assertFalse(result.should_replace)

    def test_does_not_replace_settlement_with_quoted_district(self) -> None:
        result = verify_municipality_name(
            {
                "MUN_NAME": "Городское поселение Поселок Заиграево",
                "ADM_NAME": 'АДМИНИСТРАЦИЯ МУНИЦИПАЛЬНОГО ОБРАЗОВАНИЯ "ЗАИГРАЕВСКИЙ РАЙОН"',
            }
        )

        self.assertEqual(result.status, "kept")
        self.assertFalse(result.should_replace)
        self.assertEqual(result.official_name, "Городское поселение Поселок Заиграево")

    def test_extracts_unquoted_city_settlement_candidate_from_administration(self) -> None:
        candidate = extract_municipality_name_from_administration(
            "АДМИНИСТРАЦИЯ ГОРОДСКОГО ПОСЕЛЕНИЯ ГОРОД ТУЙМАЗЫ "
            "МУНИЦИПАЛЬНОГО РАЙОНА ТУЙМАЗИНСКИЙ РАЙОН РЕСПУБЛИКИ БАШКОРТОСТАН"
        )

        self.assertEqual(candidate, "Городское поселение город Туймазы")

    def test_extracts_genitive_rural_settlement_name_from_administration(self) -> None:
        candidate = extract_municipality_name_from_administration(
            "АДМИНИСТРАЦИЯ МАРЬИНСКОГО СЕЛЬСКОГО ПОСЕЛЕНИЯ "
            "КОРСАКОВСКОГО РАЙОНА ОРЛОВСКОЙ ОБЛАСТИ"
        )

        self.assertEqual(candidate, "Сельское поселение Марьинское")

    def test_extracts_genitive_rural_settlement_name_without_district_tail(self) -> None:
        candidate = extract_municipality_name_from_administration(
            "АДМИНИСТРАЦИЯ КРАСНОПОЛЯНСКОГО СЕЛЬСКОГО ПОСЕЛЕНИЯ "
            "БАЙКАЛОВСКОГО МУНИЦИПАЛЬНОГО РАЙОНА СВЕРДЛОВСКОЙ ОБЛАСТИ"
        )

        self.assertEqual(candidate, "Сельское поселение Краснополянское")

    def test_keeps_unquoted_administration_candidate_without_minjust_confirmation(self) -> None:
        result = verify_municipality_name(
            {
                "MUN_NAME": "Городское поселение Туймазы",
                "ADM_NAME": (
                    "АДМИНИСТРАЦИЯ ГОРОДСКОГО ПОСЕЛЕНИЯ ГОРОД ТУЙМАЗЫ "
                    "МУНИЦИПАЛЬНОГО РАЙОНА ТУЙМАЗИНСКИЙ РАЙОН РЕСПУБЛИКИ БАШКОРТОСТАН"
                ),
            }
        )

        self.assertEqual(result.status, "kept")
        self.assertEqual(result.confidence, "medium")
        self.assertEqual(result.source, "ADM_NAME")
        self.assertEqual(result.official_name, "Городское поселение город Туймазы")
        self.assertFalse(result.should_replace)

    def test_replaces_unquoted_administration_candidate_after_oktmo_confirmation(self) -> None:
        lookup = FakeOktmoLookup("Городское поселение город Туймазы")
        result = verify_municipality_name(
            {
                "MUN_NAME": "Городское поселение Туймазы",
                "ADM_NAME": (
                    "АДМИНИСТРАЦИЯ ГОРОДСКОГО ПОСЕЛЕНИЯ ГОРОД ТУЙМАЗЫ "
                    "МУНИЦИПАЛЬНОГО РАЙОНА ТУЙМАЗИНСКИЙ РАЙОН РЕСПУБЛИКИ БАШКОРТОСТАН"
                ),
            },
            oktmo_lookup=lookup,
        )

        self.assertEqual(result.status, "verified")
        self.assertEqual(result.confidence, "high")
        self.assertEqual(result.source, "oktmo+ADM_NAME")
        self.assertEqual(result.official_name, "Городское поселение город Туймазы")
        self.assertTrue(result.should_replace)
        self.assertEqual(lookup.calls[0][1], "Городское поселение город Туймазы")

    def test_replaces_current_name_after_oktmo_confirmation_when_administration_is_district_only(self) -> None:
        lookup = FakeOktmoLookup("Городское поселение город Благовещенск")
        result = verify_municipality_name(
            {
                "MUN_NAME": "Городское поселение Благовещенск",
                "ADM_NAME": "АДМИНИСТРАЦИЯ МУНИЦИПАЛЬНОГО РАЙОНА БЛАГОВЕЩЕНСКИЙ РАЙОН",
            },
            oktmo_lookup=lookup,
        )

        self.assertEqual(result.status, "verified")
        self.assertEqual(result.confidence, "high")
        self.assertEqual(result.source, "oktmo")
        self.assertEqual(result.official_name, "Городское поселение город Благовещенск")
        self.assertTrue(result.should_replace)
        self.assertEqual(lookup.calls[0][1], "Городское поселение Благовещенск")

    def test_replaces_unquoted_administration_candidate_after_minjust_confirmation(self) -> None:
        lookup = FakeMinjustLookup("Городское поселение город Туймазы")
        result = verify_municipality_name(
            {
                "MUN_NAME": "Городское поселение Туймазы",
                "ADM_NAME": (
                    "АДМИНИСТРАЦИЯ ГОРОДСКОГО ПОСЕЛЕНИЯ ГОРОД ТУЙМАЗЫ "
                    "МУНИЦИПАЛЬНОГО РАЙОНА ТУЙМАЗИНСКИЙ РАЙОН РЕСПУБЛИКИ БАШКОРТОСТАН"
                ),
            },
            minjust_lookup=lookup,
        )

        self.assertEqual(result.status, "verified")
        self.assertEqual(result.confidence, "high")
        self.assertEqual(result.source, "minjust")
        self.assertEqual(result.official_name, "Городское поселение город Туймазы")
        self.assertTrue(result.should_replace)

    def test_confirms_rural_administration_matches_current_municipality(self) -> None:
        result = verify_municipality_name(
            {
                "MUN_NAME": "Дубровское сельское поселение",
                "ADM_NAME": "ДУБРОВСКАЯ СЕЛЬСКАЯ АДМИНИСТРАЦИЯ",
            }
        )

        self.assertEqual(result.status, "verified")
        self.assertEqual(result.confidence, "medium")
        self.assertEqual(result.official_name, "Дубровское сельское поселение")
        self.assertFalse(result.should_replace)

    def test_confirms_tskoe_tskaya_adjective_pair(self) -> None:
        result = verify_municipality_name(
            {
                "MUN_NAME": "Стекляннорадицкое сельское поселение",
                "ADM_NAME": "СТЕКЛЯННОРАДИЦКАЯ СЕЛЬСКАЯ АДМИНИСТРАЦИЯ",
            }
        )

        self.assertEqual(result.status, "verified")
        self.assertEqual(result.confidence, "medium")

    def test_confirms_genitive_rural_settlement_administration(self) -> None:
        result = verify_municipality_name(
            {
                "MUN_NAME": "Сельское поселение Вертячинское",
                "ADM_NAME": "АДМИНИСТРАЦИЯ ВЕРТЯЧИНСКОГО СЕЛЬСКОГО ПОСЕЛЕНИЯ",
            }
        )

        self.assertEqual(result.status, "verified")
        self.assertEqual(result.confidence, "medium")

    def test_confirms_selsovet_administration(self) -> None:
        result = verify_municipality_name(
            {
                "MUN_NAME": "Сельское поселение Кирюшкинский сельсовет",
                "ADM_NAME": "АДМИНИСТРАЦИЯ КИРЮШКИНСКОГО СЕЛЬСОВЕТА БУГУРУСЛАНСКОГО РАЙОНА",
            }
        )

        self.assertEqual(result.status, "verified")
        self.assertEqual(result.confidence, "medium")

    def test_confirms_municipal_obrazovanie_dash_administration(self) -> None:
        result = verify_municipality_name(
            {
                "MUN_NAME": "Задубровское сельское поселение",
                "ADM_NAME": "АДМИНИСТРАЦИЯ МУНИЦИПАЛЬНОГО ОБРАЗОВАНИЯ - ЗАДУБРОВСКОЕ СЕЛЬСКОЕ ПОСЕЛЕНИЕ",
            }
        )

        self.assertEqual(result.status, "verified")
        self.assertEqual(result.confidence, "medium")

    def test_confirms_municipal_obrazovanie_rural_settlement_noun(self) -> None:
        result = verify_municipality_name(
            {
                "MUN_NAME": "Сельское поселение Ваеги",
                "ADM_NAME": "АДМИНИСТРАЦИЯ МУНИЦИПАЛЬНОГО ОБРАЗОВАНИЯ СЕЛЬСКОГО ПОСЕЛЕНИЯ ВАЕГИ",
            }
        )

        self.assertEqual(result.status, "verified")
        self.assertEqual(result.confidence, "medium")

    def test_accepts_official_like_site_search_result(self) -> None:
        match = _official_site_match_from_search_result(
            {
                "title": "Администрация Дубровского сельского поселения",
                "url": "https://adm-dubrovskoe.ru/",
                "content": "Официальный сайт администрации Дубровского сельского поселения",
            },
            {
                "MUN_NAME": "Дубровское сельское поселение",
                "ADM_NAME": "ДУБРОВСКАЯ СЕЛЬСКАЯ АДМИНИСТРАЦИЯ",
            },
        )

        self.assertIsNotNone(match)

    def test_rejects_registry_search_result_as_official_site(self) -> None:
        match = _official_site_match_from_search_result(
            {
                "title": "Дубровская сельская администрация — Rusprofile",
                "url": "https://www.rusprofile.ru/id/123",
                "content": "Карточка организации",
            },
            {
                "MUN_NAME": "Дубровское сельское поселение",
                "ADM_NAME": "ДУБРОВСКАЯ СЕЛЬСКАЯ АДМИНИСТРАЦИЯ",
            },
        )

        self.assertIsNone(match)

    def test_accepts_candidate_name_from_official_site_as_strong_confirmation(self) -> None:
        result = verify_municipality_name(
            {
                "MUN_NAME": "Городское поселение Туймазы",
                "ADM_NAME": (
                    "АДМИНИСТРАЦИЯ ГОРОДСКОГО ПОСЕЛЕНИЯ ГОРОД ТУЙМАЗЫ "
                    "МУНИЦИПАЛЬНОГО РАЙОНА ТУЙМАЗИНСКИЙ РАЙОН РЕСПУБЛИКИ БАШКОРТОСТАН"
                ),
            }
        )

        self.assertEqual(result.status, "kept")
        self.assertFalse(result.should_replace)

        with tempfile.TemporaryDirectory(dir=Path.cwd()) as tmp_dir:
            path = Path(tmp_dir) / "official-site.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.cell(row=2, column=1).value = "ID"
            worksheet.cell(row=2, column=2).value = "MUN_NAME"
            worksheet.cell(row=2, column=3).value = "ADM_NAME"
            worksheet.cell(row=2, column=4).value = "SUB_RF"
            worksheet.cell(row=2, column=5).value = "MUN_R_NAME"
            worksheet.cell(row=3, column=1).value = 1
            worksheet.cell(row=3, column=2).value = "Городское поселение Туймазы"
            worksheet.cell(row=3, column=3).value = (
                "АДМИНИСТРАЦИЯ ГОРОДСКОГО ПОСЕЛЕНИЯ ГОРОД ТУЙМАЗЫ "
                "МУНИЦИПАЛЬНОГО РАЙОНА ТУЙМАЗИНСКИЙ РАЙОН РЕСПУБЛИКИ БАШКОРТОСТАН"
            )
            worksheet.cell(row=3, column=4).value = "Республика Башкортостан"
            worksheet.cell(row=3, column=5).value = "Туймазинский район"
            workbook.save(path)
            workbook.close()

            stats = verify_municipality_names_in_workbook(
                path,
                use_oktmo=False,
                use_minjust=False,
                use_official_sites=True,
                official_site_lookup=FakeOfficialSiteLookup("Городское поселение город Туймазы"),
            )

            self.assertEqual(stats["updated_rows"], 1)
            self.assertEqual(stats["official_site_found_rows"], 1)
            updated = load_workbook(path)
            sheet = updated.active
            headers = {sheet.cell(row=2, column=i).value: i for i in range(1, sheet.max_column + 1)}
            self.assertEqual(sheet.cell(row=3, column=headers["MUN_NAME"]).value, "Городское поселение город Туймазы")
            self.assertNotIn("MUN_NAME_VERIFICATION_SOURCE", headers)
            self.assertEqual(stats["replacement_samples"][0]["source"], "ADM_NAME+official_site")
            self.assertEqual(stats["replacement_samples"][0]["confidence"], "high")
            updated.close()

    def test_workbook_verification_reuses_result_for_duplicate_rows(self) -> None:
        tmp_root = Path("C:/tmp")
        tmp_root.mkdir(exist_ok=True)
        with tempfile.TemporaryDirectory(dir=tmp_root) as tmp_dir:
            path = Path(tmp_dir) / "duplicate-rows.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            headers = ["ID", "MUN_NAME", "ADM_NAME", "SUB_RF", "MUN_R_NAME"]
            for column_index, header in enumerate(headers, start=1):
                worksheet.cell(row=2, column=column_index).value = header
            for row_index in range(3, 103):
                worksheet.cell(row=row_index, column=1).value = row_index - 2
                worksheet.cell(row=row_index, column=2).value = "Городское поселение Энем"
                worksheet.cell(row=row_index, column=3).value = (
                    'АДМИНИСТРАЦИЯ МУНИЦИПАЛЬНОГО ОБРАЗОВАНИЯ "ЭНЕМСКОЕ ГОРОДСКОЕ ПОСЕЛЕНИЕ"'
                )
                worksheet.cell(row=row_index, column=4).value = "Республика Адыгея"
                worksheet.cell(row=row_index, column=5).value = "Тахтамукайский район"
            workbook.save(path)
            workbook.close()

            lookup = FakeOktmoLookup("Городское поселение Энем")
            stats = verify_municipality_names_in_workbook(
                path,
                use_oktmo=True,
                oktmo_lookup=lookup,
                use_minjust=False,
            )

            self.assertEqual(stats["total_rows"], 100)
            self.assertEqual(stats["unique_verification_keys"], 1)
            self.assertEqual(stats["cached_verification_rows"], 99)
            self.assertEqual(len(lookup.calls), 1)
            self.assertEqual(stats["updated_rows"], 100)
            updated = load_workbook(path)
            sheet = updated.active
            header_map = {sheet.cell(row=2, column=i).value: i for i in range(1, sheet.max_column + 1)}
            self.assertEqual(sheet.cell(row=3, column=header_map["MUN_NAME"]).value, "Энемское городское поселение")
            self.assertEqual(sheet.cell(row=102, column=header_map["MUN_NAME"]).value, "Энемское городское поселение")
            updated.close()

    def test_official_site_can_confirm_current_name_without_replacement(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path.cwd()) as tmp_dir:
            path = Path(tmp_dir) / "official-site-current.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.cell(row=2, column=1).value = "ID"
            worksheet.cell(row=2, column=2).value = "MUN_NAME"
            worksheet.cell(row=2, column=3).value = "ADM_NAME"
            worksheet.cell(row=3, column=1).value = 1
            worksheet.cell(row=3, column=2).value = "Дубровское сельское поселение"
            worksheet.cell(row=3, column=3).value = "ДУБРОВСКАЯ СЕЛЬСКАЯ АДМИНИСТРАЦИЯ"
            workbook.save(path)
            workbook.close()

            stats = verify_municipality_names_in_workbook(
                path,
                use_oktmo=False,
                use_minjust=False,
                use_official_sites=True,
                official_site_lookup=FakeOfficialSiteLookup("Дубровское сельское поселение"),
            )

            self.assertEqual(stats["updated_rows"], 0)
            self.assertEqual(stats["official_site_checked_rows"], 1)
            updated = load_workbook(path)
            sheet = updated.active
            headers = {sheet.cell(row=2, column=i).value: i for i in range(1, sheet.max_column + 1)}
            self.assertEqual(sheet.cell(row=3, column=headers["MUN_NAME"]).value, "Дубровское сельское поселение")
            self.assertNotIn("MUN_NAME_VERIFICATION_STATUS", headers)
            self.assertEqual(stats["verified_rows"], 1)
            updated.close()

    def test_yandex_lookup_parses_search_and_page_html(self) -> None:
        search_html = (
            '<html><body>'
            '<a href="https://adm-tuymazy.ru/">Администрация городского поселения</a>'
            '</body></html>'
        )
        page_html = (
            "<html><head><title>Официальный сайт администрации городского поселения город Туймазы</title></head>"
            "<body>Городское поселение город Туймазы</body></html>"
        )

        def fetcher(url: str, timeout: float, verify_ssl: bool) -> str:
            if "yandex.ru/search" in url:
                return search_html
            return page_html

        lookup = OfficialSiteLookup(fetcher=fetcher)
        match = lookup.confirm(
            {
                "MUN_NAME": "Городское поселение Туймазы",
                "ADM_NAME": (
                    "АДМИНИСТРАЦИЯ ГОРОДСКОГО ПОСЕЛЕНИЯ ГОРОД ТУЙМАЗЫ "
                    "МУНИЦИПАЛЬНОГО РАЙОНА ТУЙМАЗИНСКИЙ РАЙОН РЕСПУБЛИКИ БАШКОРТОСТАН"
                ),
                "SUB_RF": "Республика Башкортостан",
                "MUN_R_NAME": "Туймазинский район",
            },
            ["Городское поселение город Туймазы"],
        )

        self.assertIsNotNone(match)
        assert match is not None
        self.assertEqual(match.matched_name, "Городское поселение город Туймазы")
        self.assertIn("официальный", match.evidence.lower())

    def test_official_site_queries_include_ustav_and_municipality_pages(self) -> None:
        queries = _build_official_site_queries(
            {
                "MUN_NAME": "Городское поселение Туймазы",
                "ADM_NAME": "АДМИНИСТРАЦИЯ ГОРОДСКОГО ПОСЕЛЕНИЯ ГОРОД ТУЙМАЗЫ",
                "SUB_RF": "Республика Башкортостан",
                "MUN_R_NAME": "Туймазинский район",
            },
            ["Городское поселение город Туймазы"],
        )

        self.assertTrue(any("устав муниципального образования" in item.lower() for item in queries))
        self.assertTrue(any("о муниципальном образовании" in item.lower() for item in queries))

    def test_official_site_match_scores_legal_municipality_page_higher(self) -> None:
        match = _official_site_match_from_search_result(
            {
                "url": "https://adm.example.ru/ustav/o-munitsipalnom-obrazovanii",
                "title": "Устав муниципального образования",
                "content": "Официальный сайт муниципального образования Городское поселение город Туймазы",
            },
            {
                "MUN_NAME": "Городское поселение Туймазы",
                "ADM_NAME": "АДМИНИСТРАЦИЯ ГОРОДСКОГО ПОСЕЛЕНИЯ ГОРОД ТУЙМАЗЫ",
            },
            candidate_names=["Городское поселение город Туймазы"],
        )

        self.assertIsNotNone(match)
        assert match is not None
        self.assertGreaterEqual(match.score, 10)
        self.assertIn("устава", match.evidence.lower())

    def test_search_name_variants_expand_common_settlement_abbreviations(self) -> None:
        variants = _search_name_variants("рп Ордынское")

        self.assertIn("рп Ордынское", variants)
        self.assertIn("рабочий поселок Ордынское", variants)

        variants = _search_name_variants("пгт Кормиловка")
        self.assertIn("поселок городского типа Кормиловка", variants)

    def test_official_site_match_handles_abbreviation_variants(self) -> None:
        match = _official_site_match_from_search_result(
            {
                "url": "https://adm-ordynskoe.ru/ustav",
                "title": "Устав муниципального образования",
                "content": "Городское поселение рабочий поселок Ордынское официальный сайт администрации",
            },
            {
                "MUN_NAME": "рп Ордынское",
                "ADM_NAME": "АДМИНИСТРАЦИЯ Р.П. ОРДЫНСКОЕ",
            },
            candidate_names=["рп Ордынское"],
        )

        self.assertIsNotNone(match)
        assert match is not None
        self.assertEqual(match.matched_name, "рабочий поселок Ордынское")

    def test_abbreviation_resolution_helpers_allow_same_toponym(self) -> None:
        self.assertTrue(_needs_abbreviation_resolution("пгт Духовницкое рп"))
        self.assertTrue(
            _candidate_is_safe_for_autoreplace(
                "Городское поселение рабочий поселок Духовницкое",
                "пгт Духовницкое рп",
            )
        )

    def test_official_site_followup_is_required_for_abbreviation_row(self) -> None:
        verification = verify_municipality_name(
            {
                "MUN_NAME": "пгт Духовницкое рп",
                "ADM_NAME": "АДМИНИСТРАЦИЯ Р.П. ДУХОВНИЦКОЕ",
            }
        )

        self.assertTrue(
            _should_use_official_site_followup(
                {"MUN_NAME": "пгт Духовницкое рп", "ADM_NAME": "АДМИНИСТРАЦИЯ Р.П. ДУХОВНИЦКОЕ"},
                verification,
            )
        )

    def test_official_site_followup_is_skipped_for_strong_autoreplace(self) -> None:
        verification = verify_municipality_name(
            {
                "MUN_NAME": "Городское поселение Туймазы",
                "ADM_NAME": (
                    "АДМИНИСТРАЦИЯ ГОРОДСКОГО ПОСЕЛЕНИЯ ГОРОД ТУЙМАЗЫ "
                    "МУНИЦИПАЛЬНОГО РАЙОНА ТУЙМАЗИНСКИЙ РАЙОН РЕСПУБЛИКИ БАШКОРТОСТАН"
                ),
            },
            oktmo_lookup=FakeOktmoLookup("Городское поселение город Туймазы"),
        )

        self.assertFalse(
            _should_use_official_site_followup(
                {
                    "MUN_NAME": "Городское поселение Туймазы",
                    "ADM_NAME": (
                        "АДМИНИСТРАЦИЯ ГОРОДСКОГО ПОСЕЛЕНИЯ ГОРОД ТУЙМАЗЫ "
                        "МУНИЦИПАЛЬНОГО РАЙОНА ТУЙМАЗИНСКИЙ РАЙОН РЕСПУБЛИКИ БАШКОРТОСТАН"
                    ),
                },
                verification,
            )
        )

    def test_expands_abbreviated_locality_name_without_external_lookup(self) -> None:
        result = verify_municipality_name(
            {
                "MUN_NAME": "пгт Духовницкое рп",
                "ADM_NAME": "",
            }
        )

        self.assertEqual(result.status, "verified")
        self.assertEqual(result.confidence, "high")
        self.assertEqual(result.source, "normalization")
        self.assertEqual(result.official_name, "поселок городского типа Духовницкое рабочий поселок")
        self.assertTrue(result.should_replace)

    def test_updates_workbook_with_expanded_locality_abbreviation(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(r"C:\tmp")) as tmp_dir:
            path = Path(tmp_dir) / "abbrev-data.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.cell(row=2, column=1).value = "ID"
            worksheet.cell(row=2, column=2).value = "MUN_NAME"
            worksheet.cell(row=2, column=3).value = "ADM_NAME"
            worksheet.cell(row=3, column=1).value = 1
            worksheet.cell(row=3, column=2).value = "рп Ордынское"
            worksheet.cell(row=3, column=3).value = ""
            workbook.save(path)
            workbook.close()

            stats = verify_municipality_names_in_workbook(path)

            self.assertEqual(stats["updated_rows"], 1)
            updated = load_workbook(path)
            sheet = updated.active
            headers = {sheet.cell(row=2, column=i).value: i for i in range(1, sheet.max_column + 1)}
            self.assertEqual(sheet.cell(row=3, column=headers["MUN_NAME"]).value, "рабочий поселок Ордынское")
            self.assertNotIn("MUN_NAME_VERIFICATION_SOURCE", headers)
            self.assertEqual(stats["replacement_samples"][0]["source"], "normalization")
            updated.close()

    def test_updates_workbook_and_preserves_original_name(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(r"C:\tmp")) as tmp_dir:
            path = Path(tmp_dir) / "data.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.cell(row=2, column=1).value = "ID"
            worksheet.cell(row=2, column=2).value = "MUN_NAME"
            worksheet.cell(row=2, column=3).value = "ADM_NAME"
            worksheet.cell(row=3, column=1).value = 1
            worksheet.cell(row=3, column=2).value = "Городское поселение Яблоновский"
            worksheet.cell(row=3, column=3).value = (
                'Администрация муниципального образования "Яблоновское городское поселение"'
            )
            workbook.save(path)
            workbook.close()

            stats = verify_municipality_names_in_workbook(path)

            self.assertEqual(stats["updated_rows"], 1)
            self.assertEqual(len(stats["replacements"]), 1)
            self.assertEqual(stats["replacements"][0]["from"], "Городское поселение Яблоновский")
            self.assertEqual(stats["replacements"][0]["to"], "Яблоновское городское поселение")
            updated = load_workbook(path)
            sheet = updated.active
            headers = {sheet.cell(row=2, column=i).value: i for i in range(1, sheet.max_column + 1)}
            self.assertEqual(sheet.cell(row=3, column=headers["MUN_NAME"]).value, "Яблоновское городское поселение")
            self.assertNotIn("MUN_NAME_SOURCE_ORIGINAL", headers)
            self.assertNotIn("MUN_NAME_VERIFICATION_STATUS", headers)
            updated.close()


if __name__ == "__main__":
    unittest.main()
