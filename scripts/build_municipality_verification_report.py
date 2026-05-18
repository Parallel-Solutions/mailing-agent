from __future__ import annotations

import argparse
import shutil
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from src.generator.verification.municipality_name_verifier import (
    ADM_NAME_COLUMN,
    CONFIDENCE_COLUMN,
    HEADER_ROW,
    MUN_NAME_COLUMN,
    OFFICIAL_COLUMN,
    ORIGINAL_COLUMN,
    REASON_COLUMN,
    SOURCE_COLUMN,
    STATUS_COLUMN,
    URL_COLUMN,
    verify_municipality_names_in_workbook,
)
from src.generator.verification.oktmo_municipality_lookup import OktmoMunicipalityLookup
from src.utils.config import settings


BLUE = "1F4E78"
BORDER = Side(style="thin", color="D9E2F3")


def _short(value: Any, max_len: int = 500) -> str:
    text = "" if value is None else str(value)
    text = " ".join(text.split())
    return text if len(text) <= max_len else text[: max_len - 1] + "…"


def _header_map(worksheet) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for column_index in range(1, worksheet.max_column + 1):
        header = worksheet.cell(row=HEADER_ROW, column=column_index).value
        if header not in (None, ""):
            mapping[str(header)] = column_index
    return mapping


def _cell(worksheet, row_index: int, header_map: dict[str, int], name: str) -> Any:
    column = header_map.get(name)
    if not column:
        return ""
    return worksheet.cell(row=row_index, column=column).value


def _add_title(sheet, title: str, subtitle: str = "") -> None:
    sheet["A1"] = title
    sheet["A1"].font = Font(bold=True, size=16, color=BLUE)
    if subtitle:
        sheet["A2"] = subtitle
        sheet["A2"].font = Font(size=10, color="666666")


def _write_table(sheet, start_row: int, headers: list[str], rows: list[list[Any]], *, name: str) -> None:
    for column_index, header in enumerate(headers, 1):
        cell = sheet.cell(start_row, column_index, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index, row in enumerate(rows, start_row + 1):
        for column_index, value in enumerate(row, 1):
            cell = sheet.cell(row_index, column_index, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    max_row = start_row + max(len(rows), 1)
    max_column = len(headers)
    for row in sheet.iter_rows(min_row=start_row, max_row=max_row, min_col=1, max_col=max_column):
        for cell in row:
            cell.border = Border(bottom=BORDER)

    if rows:
        table = Table(displayName=name, ref=f"A{start_row}:{get_column_letter(max_column)}{max_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )
        sheet.add_table(table)


def _autosize(sheet, widths: dict[str, int]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _style_workbook(workbook: Workbook) -> None:
    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A4"
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_report(*, verified_path: Path, report_path: Path, verification_stats: dict[str, Any]) -> Path:
    workbook = load_workbook(verified_path)
    worksheet = workbook[workbook.sheetnames[0]]
    mapping = _header_map(worksheet)

    rows_all: list[list[Any]] = []
    changed_rows: list[list[Any]] = []
    status_counter: Counter[str] = Counter()
    source_counter: Counter[str] = Counter()

    for row_index in range(HEADER_ROW + 1, worksheet.max_row + 1):
        row_id = _cell(worksheet, row_index, mapping, "ID")
        adm_name = _cell(worksheet, row_index, mapping, ADM_NAME_COLUMN)
        mun_name_after = _cell(worksheet, row_index, mapping, MUN_NAME_COLUMN)
        original_name = _cell(worksheet, row_index, mapping, ORIGINAL_COLUMN)
        official_name = _cell(worksheet, row_index, mapping, OFFICIAL_COLUMN)
        status = _cell(worksheet, row_index, mapping, STATUS_COLUMN)
        confidence = _cell(worksheet, row_index, mapping, CONFIDENCE_COLUMN)
        source = _cell(worksheet, row_index, mapping, SOURCE_COLUMN)
        reason = _cell(worksheet, row_index, mapping, REASON_COLUMN)
        source_url = _cell(worksheet, row_index, mapping, URL_COLUMN)

        if not any(value not in (None, "") for value in [row_id, adm_name, mun_name_after, original_name, official_name, status]):
            continue

        changed = bool(original_name != mun_name_after and official_name)
        status_counter[_short(status, 120)] += 1
        source_counter[_short(source, 120)] += 1

        report_row = [
            row_id,
            _short(original_name, 240),
            _short(mun_name_after, 240),
            _short(adm_name, 500),
            _short(official_name, 240),
            _short(status, 120),
            _short(confidence, 80),
            _short(source, 120),
            "да" if changed else "нет",
            _short(reason, 900),
            _short(source_url, 240),
        ]
        rows_all.append(report_row)
        if changed:
            changed_rows.append(report_row)

    workbook.close()

    report_wb = Workbook()
    ws_summary = report_wb.active
    ws_summary.title = "Сводка"
    ws_changed = report_wb.create_sheet("Изменения МО")
    ws_all = report_wb.create_sheet("Все решения МО")

    _add_title(
        ws_summary,
        "Журнал проверки названий МО",
        f"Источник: {verified_path}",
    )

    summary_rows = [
        ["Всего строк", verification_stats.get("total_rows", 0)],
        ["Подтверждено", verification_stats.get("verified_rows", 0)],
        ["Заменено в MUN_NAME", verification_stats.get("updated_rows", 0)],
        ["Оставлено без изменений", verification_stats.get("kept_rows", 0)],
        ["Не удалось определить", verification_stats.get("missing_rows", 0)],
        ["Проверено по оф. сайтам", verification_stats.get("official_site_checked_rows", 0)],
        ["Подходящих сайтов найдено", verification_stats.get("official_site_found_rows", 0)],
    ]
    _write_table(ws_summary, 4, ["Показатель", "Значение"], summary_rows, name="MunicipalitySummary")

    status_rows = [[key, value] for key, value in status_counter.most_common()]
    _write_table(ws_summary, 15, ["Статус решения", "Количество"], status_rows, name="MunicipalityStatuses")

    source_rows = [[key, value] for key, value in source_counter.most_common()]
    _write_table(ws_summary, 15 + max(len(status_rows), 1) + 4, ["Источник решения", "Количество"], source_rows, name="MunicipalitySources")

    headers = [
        "ID строки",
        "MUN_NAME было",
        "MUN_NAME стало",
        "ADM_NAME",
        "Официальное название",
        "Статус",
        "Уверенность",
        "Источник",
        "Изменено",
        "Причина",
        "URL источника",
    ]
    _add_title(ws_changed, "Изменения MUN_NAME", "Только строки, где MUN_NAME был автоматически изменен.")
    _write_table(ws_changed, 3, headers, changed_rows, name="MunicipalityChanges")

    _add_title(ws_all, "Все решения по МО", "Полный журнал проверки MUN_NAME с учетом ADM_NAME и внешней верификации.")
    _write_table(ws_all, 3, headers, rows_all, name="MunicipalityAllDecisions")

    _style_workbook(report_wb)
    _autosize(ws_summary, {"A": 34, "B": 18})
    _autosize(ws_changed, {"A": 10, "B": 28, "C": 28, "D": 60, "E": 30, "F": 14, "G": 12, "H": 20, "I": 12, "J": 80, "K": 40})
    _autosize(ws_all, {"A": 10, "B": 28, "C": 28, "D": 60, "E": 30, "F": 14, "G": 12, "H": 20, "I": 12, "J": 80, "K": 40})

    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_wb.save(report_path)
    return report_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Run municipality-name verification and build a separate change journal.")
    parser.add_argument("source", type=Path, help="Source XLSX file.")
    parser.add_argument("--verified-output", type=Path, required=True, help="Path for the verified XLSX copy.")
    parser.add_argument("--report-output", type=Path, help="Path for the separate report XLSX.")
    args = parser.parse_args()

    args.verified_output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(args.source, args.verified_output)

    oktmo_lookup = (
        OktmoMunicipalityLookup(
            csv_path=Path(settings.municipality_oktmo_csv_path) if settings.municipality_oktmo_csv_path else None,
            verify_ssl=settings.municipality_oktmo_verify_ssl,
        )
        if settings.municipality_oktmo_lookup_enabled
        else None
    )
    stats = verify_municipality_names_in_workbook(
        args.verified_output,
        use_official_sites=settings.municipality_official_sites_enabled,
        use_oktmo=settings.municipality_oktmo_lookup_enabled,
        oktmo_lookup=oktmo_lookup,
        use_minjust=settings.municipality_minjust_lookup_enabled,
    )
    if args.report_output:
        build_report(verified_path=args.verified_output, report_path=args.report_output, verification_stats=stats)

    print(f"verified_file={args.verified_output}")
    if args.report_output:
        print(f"report_file={args.report_output}")
    print(f"total_rows={stats.get('total_rows', 0)}")
    print(f"verified_rows={stats.get('verified_rows', 0)}")
    print(f"updated_rows={stats.get('updated_rows', 0)}")
    print(f"kept_rows={stats.get('kept_rows', 0)}")
    print(f"missing_rows={stats.get('missing_rows', 0)}")


if __name__ == "__main__":
    main()
