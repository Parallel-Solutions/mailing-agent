from __future__ import annotations

import argparse
import json
from collections import Counter
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


BLUE = "1F4E78"
LIGHT_BLUE = "D9EAF7"
GREEN = "E2F0D9"
YELLOW = "FFF2CC"
RED = "F4CCCC"
GRAY = "E7E6E6"
BORDER = Side(style="thin", color="D9E2F3")


def read_json(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    with path.open("r", encoding="utf-8-sig") as file:
        data = json.load(file)
    return data if isinstance(data, dict) else {}


def read_jsonl(path: Path, *, limit: int | None = None) -> Iterable[dict[str, Any]]:
    if not path.exists():
        return
    count = 0
    with path.open("r", encoding="utf-8-sig") as file:
        for line in file:
            if limit is not None and count >= limit:
                break
            line = line.strip()
            if not line:
                continue
            try:
                item = json.loads(line)
            except json.JSONDecodeError:
                continue
            if isinstance(item, dict):
                count += 1
                yield item


def short(value: Any, max_len: int = 500) -> str:
    text = "" if value is None else str(value)
    text = " ".join(text.split())
    return text if len(text) <= max_len else text[: max_len - 1] + "…"


def add_title(ws, title: str, subtitle: str | None = None) -> None:
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=16, color=BLUE)
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(size=10, color="666666")


def write_table(ws, start_row: int, headers: list[str], rows: list[list[Any]], *, name: str) -> None:
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(start_row, col_idx, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, row in enumerate(rows, start_row + 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    max_row = start_row + max(len(rows), 1)
    max_col = len(headers)
    for row in ws.iter_rows(min_row=start_row, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = Border(bottom=BORDER)

    if rows:
        ref = f"A{start_row}:{get_column_letter(max_col)}{max_row}"
        table = Table(displayName=name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )
        ws.add_table(table)


def flatten_philologist_documents(data: dict[str, Any], *, limit: int | None = None) -> list[dict[str, Any]]:
    docs = data.get("documents") or []
    out: list[dict[str, Any]] = []
    for doc in docs:
        if limit is not None and len(out) >= limit:
            break
        if not isinstance(doc, dict):
            continue
        base = {
            "document": doc.get("document") or doc.get("file") or doc.get("path") or "",
            "status": doc.get("status") or "",
            "issue_count": doc.get("issue_count") or len(doc.get("issues") or []),
            "fixed_count": doc.get("fixed_count") or len(doc.get("fixes") or doc.get("decisions") or []),
        }
        decisions = doc.get("decisions") or doc.get("fixes") or doc.get("issues") or []
        if not isinstance(decisions, list):
            decisions = []
        if not decisions:
            out.append(base)
            continue
        for decision in decisions:
            if limit is not None and len(out) >= limit:
                break
            if not isinstance(decision, dict):
                continue
            item = dict(base)
            item.update(decision)
            out.append(item)
    return out


def autosize(ws, widths: dict[str, int] | None = None) -> None:
    if widths:
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        return
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 12
        for cell in ws[letter]:
            max_len = max(max_len, min(60, len(str(cell.value or "")) + 2))
        ws.column_dimensions[letter].width = max_len


def style_workbook(wb: Workbook) -> None:
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A4"
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)


def main() -> None:
    parser = argparse.ArgumentParser(description="Build a readable correction report from agent logs.")
    parser.add_argument("source_dir", type=Path)
    parser.add_argument("--output", type=Path, default=None)
    parser.add_argument("--max-philologist-rows", type=int, default=50000)
    parser.add_argument("--max-inflection-rows", type=int, default=20000)
    parser.add_argument("--max-quarantine-rows", type=int, default=20000)
    args = parser.parse_args()

    source_dir = args.source_dir
    output = args.output or (source_dir / "correction_report.xlsx")

    philologist = read_json(source_dir / "philologist.json")
    inflection_rows = list(read_jsonl(source_dir / "inflection_log.jsonl", limit=args.max_inflection_rows))
    quarantine_rows = list(read_jsonl(source_dir / "agent_quarantine.jsonl", limit=args.max_quarantine_rows))
    philologist_rows = flatten_philologist_documents(philologist, limit=args.max_philologist_rows)
    agent_report_path = source_dir / "agent_report.txt"
    agent_report = agent_report_path.read_text(encoding="utf-8-sig", errors="replace") if agent_report_path.exists() else ""

    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Сводка"
    ws_inf = wb.create_sheet("Склонения")
    ws_phil = wb.create_sheet("Правки филолога")
    ws_quar = wb.create_sheet("Карантин")
    ws_text = wb.create_sheet("Текстовый отчет")

    add_title(ws_sum, "Журнал исправлений", f"Источник: {source_dir}")
    summary = [
        ["Статус филолога", philologist.get("status", "")],
        ["Документов всего", philologist.get("total_documents", "")],
        ["Документов проверено", philologist.get("processed_documents", "")],
        ["Документов с замечаниями", philologist.get("documents_with_issues", "")],
        ["Документов исправлено", philologist.get("fixed_documents", "")],
        ["Строк в журнале склонений", len(inflection_rows)],
        ["Строк в карантине", len(quarantine_rows)],
        ["Строк правок филолога в отчете", len(philologist_rows)],
    ]
    write_table(ws_sum, 4, ["Показатель", "Значение"], summary, name="SummaryTable")

    method_counts = Counter(short(r.get("method")) for r in inflection_rows)
    action_counts = Counter(short(r.get("decision_action")) for r in philologist_rows)
    issue_counts = Counter(short(r.get("issue"), 120) for r in philologist_rows if r.get("issue"))
    quarantine_reason_counts = Counter(short(r.get("reason"), 120) for r in quarantine_rows if r.get("reason"))

    start = 4 + len(summary) + 3
    write_table(
        ws_sum,
        start,
        ["Метод склонения", "Количество"],
        [[k, v] for k, v in method_counts.most_common(20)],
        name="InflectionMethodSummary",
    )
    start = start + max(len(method_counts), 1) + 4
    write_table(
        ws_sum,
        start,
        ["Действие филолога", "Количество"],
        [[k, v] for k, v in action_counts.most_common(20)],
        name="PhilologistActionSummary",
    )
    start = start + max(len(action_counts), 1) + 4
    write_table(
        ws_sum,
        start,
        ["Тип замечания", "Количество"],
        [[k, v] for k, v in issue_counts.most_common(30)],
        name="IssueSummary",
    )
    start = start + max(len(issue_counts), 1) + 4
    write_table(
        ws_sum,
        start,
        ["Причина карантина", "Количество"],
        [[k, v] for k, v in quarantine_reason_counts.most_common(30)],
        name="QuarantineSummary",
    )

    add_title(ws_inf, "Журнал склонений", "Что генератор подставил в документы.")
    inf_headers = [
        "row_id",
        "field",
        "source_value",
        "result_value",
        "target_case",
        "method",
        "confidence",
        "warning",
        "reason",
        "filled_sentence",
    ]
    write_table(
        ws_inf,
        4,
        ["ID строки", "Поле", "Было", "Стало", "Падеж", "Метод", "Уверенность", "Предупреждение", "Причина", "Контекст"],
        [[short(r.get(k), 800) for k in inf_headers] for r in inflection_rows],
        name="InflectionLog",
    )

    add_title(ws_phil, "Правки филолога", "Автоматические решения, замечания и источники стратегии.")
    phil_headers = [
        "document",
        "location",
        "fragment",
        "issue",
        "suggestion",
        "mode",
        "decision_action",
        "decision_reason",
        "decision_confidence",
        "strategy_source",
        "rag_recommendation",
        "rag_reason",
    ]
    write_table(
        ws_phil,
        4,
        [
            "Документ",
            "Место",
            "Фрагмент",
            "Замечание",
            "Предложение",
            "Режим",
            "Действие",
            "Причина решения",
            "Уверенность",
            "Источник стратегии",
            "RAG рекомендация",
            "RAG причина",
        ],
        [[short(r.get(k), 900) for k in phil_headers] for r in philologist_rows],
        name="PhilologistFixes",
    )

    add_title(ws_quar, "Карантин", "Случаи, которые нельзя считать автоматически подтвержденными.")
    quar_headers = [
        "row_id",
        "field",
        "source_value",
        "result_value",
        "method",
        "confidence",
        "reason",
        "warning",
        "filled_sentence",
        "next_action",
    ]
    write_table(
        ws_quar,
        4,
        ["ID строки", "Поле", "Было", "Стало", "Метод", "Уверенность", "Причина", "Предупреждение", "Контекст", "Что делать"],
        [[short(r.get(k), 900) for k in quar_headers] for r in quarantine_rows],
        name="QuarantineLog",
    )

    add_title(ws_text, "Исходный текстовый отчет")
    for idx, line in enumerate(agent_report.splitlines() or ["agent_report.txt не найден"], 4):
        ws_text.cell(idx, 1, line)

    style_workbook(wb)
    autosize(ws_sum, {"A": 42, "B": 28})
    autosize(
        ws_inf,
        {"A": 10, "B": 18, "C": 34, "D": 34, "E": 16, "F": 18, "G": 14, "H": 28, "I": 42, "J": 70},
    )
    autosize(
        ws_phil,
        {"A": 42, "B": 16, "C": 34, "D": 44, "E": 34, "F": 18, "G": 18, "H": 50, "I": 14, "J": 22, "K": 26, "L": 44},
    )
    autosize(
        ws_quar,
        {"A": 10, "B": 18, "C": 34, "D": 34, "E": 18, "F": 14, "G": 40, "H": 28, "I": 70, "J": 44},
    )
    autosize(ws_text, {"A": 120})

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    print(output)


if __name__ == "__main__":
    main()
