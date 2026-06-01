"""
recon_emails.py — РАЗВЕДКА перед написанием fix_emails_tool.

Ничего не меняет в файле. Для первых N строк:
  1) берёт ADM_NAME + регион из файла,
  2) ищет официальный сайт через Яндекс Search API (тот же, что уже в проекте),
  3) показывает топ ссылок,
  4) заходит на первую неагрегаторную ссылку и пытается достать почту
     (главная + страница "Контакты"),
  5) печатает, что нашлось.

Цель — посмотреть ГЛАЗАМИ на реальную выдачу: часто ли первая ссылка это
официальный сайт, где лежит почта, попадается ли движок gosweb.gosuslugi.ru.

Запуск:
    uv run python recon_emails.py
"""
from __future__ import annotations

import re
import time

from openpyxl import load_workbook

# инфраструктура проекта
from src.parser_new import config
from src.parser_new.batch_processor import (
    _get_yandex_search, _parse_yandex_xml, HEADERS,
)
import httpx
from bs4 import BeautifulSoup


# ==============================
# НАСТРОЙКИ
# ==============================
PATH = r"C:\Users\User\Downloads\Не ДОСТАВЛЕННЫЕ.xlsx"   # <-- УКАЖИ СВОЙ ПУТЬ
N = 8                 # сколько строк разведать
DATA_START = 3

ADM_NAME_COL  = 5     # E
SUB_RF_COL    = 2     # B
MUN_R_COL     = 3     # C
EMAIL_OSN_COL = 9     # I

# агрегаторы — не считаем их официальным сайтом
SKIP_DOMAINS = [
    "rusprofile.ru", "egrul.ru", "checko.ru", "sbis.ru", "list-org.com",
    "audit-it.ru", "rusprofile", "gosuslugi.ru/structure", "wikipedia",
    "yandex.ru", "google.", "2gis.", "zhkh.", "bus.gov.ru",
]

EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
# мусорные ящики, которые не считаем почтой администрации
JUNK_EMAIL = ["@gosuslugi", "example", "@sentry", "@2x", "noreply"]


def find_official_site(adm_name: str, region: str, mun_r: str) -> list[dict]:
    """Возвращает топ результатов Яндекса (url/title/snippet)."""
    search = _get_yandex_search()
    query = f"{adm_name} {region} официальный сайт"
    time.sleep(0.6)
    xml = search.run(query, format="xml", page=0)
    return _parse_yandex_xml(xml)


def extract_emails_from_url(url: str) -> list[str]:
    """Тянет страницу, достаёт e-mail из текста и из mailto-ссылок."""
    found: list[str] = []
    try:
        time.sleep(1.0)
        resp = httpx.get(url, headers=HEADERS, timeout=12, follow_redirects=True)
        soup = BeautifulSoup(resp.text, "lxml")

        # mailto-ссылки — самый надёжный источник
        for a in soup.select("a[href^='mailto:']"):
            addr = a.get("href", "")[len("mailto:"):].split("?")[0].strip()
            if addr:
                found.append(addr)

        # плюс из текста страницы
        text = soup.get_text(" ", strip=True)
        found.extend(EMAIL_RE.findall(text))
    except Exception as e:
        print(f"      [заход на {url[:50]} не удался: {e}]")
        return []

    # чистим мусор и дубли
    clean, seen = [], set()
    for e in found:
        el = e.lower()
        if any(j in el for j in JUNK_EMAIL):
            continue
        if el not in seen:
            seen.add(el)
            clean.append(e)
    return clean


def find_contacts_link(url: str) -> str | None:
    """Ищет на странице ссылку 'Контакты' и возвращает абсолютный URL."""
    try:
        time.sleep(0.5)
        resp = httpx.get(url, headers=HEADERS, timeout=12, follow_redirects=True)
        soup = BeautifulSoup(resp.text, "lxml")
        for a in soup.find_all("a", href=True):
            if "контакт" in a.get_text().lower():
                href = a["href"]
                if href.startswith("http"):
                    return href
                # относительная ссылка
                from urllib.parse import urljoin
                return urljoin(url, href)
    except Exception:
        pass
    return None


def main() -> None:
    wb = load_workbook(PATH)
    ws = wb.active

    done = 0
    for row in range(DATA_START, ws.max_row + 1):
        adm = str(ws.cell(row, ADM_NAME_COL).value or "").strip()
        region = str(ws.cell(row, SUB_RF_COL).value or "").strip()
        mun_r = str(ws.cell(row, MUN_R_COL).value or "").strip()
        old_email = str(ws.cell(row, EMAIL_OSN_COL).value or "").strip()
        if not adm:
            continue
        done += 1
        if done > N:
            break

        print("\n" + "=" * 70)
        print(f"стр.{row}: {adm[:70]}")
        print(f"  регион: {region} | старый EMAIL_OSN: {old_email!r}")

        try:
            results = find_official_site(adm, region, mun_r)
        except Exception as e:
            print(f"  [Яндекс-поиск не сработал: {e}]")
            continue

        if not results:
            print("  Яндекс ничего не вернул.")
            continue

        print("  топ ссылок Яндекса:")
        for i, r in enumerate(results[:5], 1):
            print(f"    {i}. {r['url'][:75]}")

        # первая неагрегаторная ссылка
        official = None
        for r in results:
            if not any(d in r["url"] for d in SKIP_DOMAINS):
                official = r["url"]
                break

        if not official:
            print("  -> подходящего (неагрегаторного) сайта в топе нет")
            continue

        print(f"  -> пробую сайт: {official[:75]}")
        emails = extract_emails_from_url(official)
        if not emails:
            # пробуем страницу Контактов
            link = find_contacts_link(official)
            if link:
                print(f"     почты на главной нет, иду в Контакты: {link[:70]}")
                emails = extract_emails_from_url(link)

        if emails:
            print(f"  НАЙДЕНО на сайте: {emails[:5]}")
            same = old_email.lower() in [e.lower() for e in emails]
            print(f"  совпадает со старым? {'ДА' if same else 'нет'}")
        else:
            print("  почту на сайте достать не удалось")

    wb.close()
    print("\n" + "=" * 70)
    print("Разведка завершена. Покажи вывод — решим, как писать парсер.")


if __name__ == "__main__":
    main()